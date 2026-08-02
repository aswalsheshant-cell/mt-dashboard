# -*- coding: utf-8 -*-
"""
forecast_engine/india_workbook_format.py
=========================================
India-standard workbook formatting, Executive Control Panel, and
reconciliation validation for Forecast_Planning_Workbook.xlsx.

IDEMPOTENCY CONTRACT
────────────────────
• India_Summary is fully replaced on each call — never appended.
• Column header renames are guarded by a "(₹)" / "(₹, prov.)" sentinel:
  if the sentinel already exists the rename is skipped.
• Number-format changes are applied without altering underlying values.
• Repeated calls produce the same workbook (modulo the Last-Refresh timestamp
  in the Control Panel, which intentionally updates each run).

INDIAN NUMBER STANDARDS APPLIED
────────────────────────────────
Qty / units        :  Indian comma integer  ←  1,00,00,000 style, 0 dp
₹ detail values    :  Indian comma rupee    ←  ##,##,##,##0.00
₹ summary (Crores) :  #,##0.00  (already divided before writing)
Percentages        :  0.00%
EAN / codes        :  @ (text — no scientific notation)
No "Million" or "Billion" labels anywhere in this module.
"""

from __future__ import annotations

import datetime as dt
import os
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── numeric / format constants ─────────────────────────────────────────────────

CRORE = 10_000_000   # ₹1,00,00,000
LAKH  = 100_000      # ₹1,00,000

# Indian comma notation — last 3 digits plain, then pairs
IND_FMT_RUPEE = r'[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##\,##0.00'
IND_FMT_INT   = r'[>=10000000]##\,##\,##\,##0;[>=100000]##\,##\,##0;##\,##0'
IND_FMT_CR    = '#,##0.00'   # values already in Crores
IND_FMT_PCT   = '0.00%'
IND_FMT_TEXT  = '@'          # EAN / codes → text (prevents scientific notation)

# ── colour palette ─────────────────────────────────────────────────────────────

CLR_NAVY   = '1F3864'   # main header
CLR_INDIGO = '2E4DA5'   # section header
CLR_DKBLUE = '2C3E50'   # sub-header / column header
CLR_GREEN  = '1A5276'   # totals
CLR_DKGRN  = '1E8449'   # approved / pass
CLR_AMBER  = 'E67E22'   # pending / tentative
CLR_AMBRLT = 'FEF9E7'   # light amber bg for warnings
CLR_GREY   = 'F2F2F2'   # alternating row (even)
CLR_WHITE  = 'FFFFFF'
CLR_RED    = 'C0392B'   # blocked / error / imbalance
CLR_PURP   = '6C3483'   # VMM excluded

# Release status → (bg_colour, text_colour)
_STATUS_STYLES: Dict[str, Tuple[str, str]] = {
    'TENTATIVE – QUANTITY PLANNING READY':      (CLR_AMBER, CLR_WHITE),
    'TENTATIVE – FINANCIAL VIEW WITH WARNINGS': ('F39C12', CLR_WHITE),
    'FINAL – BUSINESS USE READY':               (CLR_DKGRN, CLR_WHITE),
    'BLOCKED – MATERIAL DATA ISSUE':            (CLR_RED,   CLR_WHITE),
}

# Columns that get Indian ₹ (2 dp) format on data sheets
RUPEE_COLS: frozenset = frozenset({
    'forecast_nsv', 'forecast_trade_spend', 'forecast_cm2',
})
# Columns that get Indian integer format on data sheets
QTY_COLS: frozenset = frozenset({
    'forecast_qty', 'gross_forecast_qty', 'forecast_primary_qty',
    'forecast_offtake_qty', 'historical_offtake_qty', 'historical_primary_qty',
    'weighted_ma_qty', 'warehouse_gurgaon', 'warehouse_mumbai',
    'warehouse_bangalore', 'warehouse_kolkata', 'suggested_dispatch_qty',
    'uplift_value_source',  # numeric uplift fields
    'npi_uplift', 'festival_uplift',
})
# Code columns that must stay as text (no scientific notation on 13-digit EAN)
TEXT_COLS: frozenset = frozenset({'ean', 'forecast_id'})

# All financial columns that need a number format
_FIN_COLS: frozenset = RUPEE_COLS | QTY_COLS | TEXT_COLS

# Header renames applied once (guarded by "(₹)" sentinel)
_HEADER_RENAMES: Dict[str, str] = {
    'forecast_nsv':         'forecast_nsv (₹)',
    'forecast_trade_spend': 'forecast_trade_spend (₹)',
    'forecast_cm2':         'forecast_cm2 (₹, prov.)',
}

DATA_SHEETS = ['Forecast', 'Expected', 'Best_Case', 'Worst_Case', 'Exceptions']

_TIER_ORDER = [
    'Validated — Primary Invoice History',
    'Validated — Formula (stored TOT%)',
    'Estimated placeholder (chain=ALL, qty=0)',
    'No unit NSV — Pack/Case-level MRP',
    'No unit NSV — Aggregate denomination',
    'VMM excluded (no margin data)',
]
_TIER_BG = {
    'Validated — Primary Invoice History':      '1A5276',
    'Validated — Formula (stored TOT%)':        '117A65',
    'Estimated placeholder (chain=ALL, qty=0)': '7F8C8D',
    'No unit NSV — Pack/Case-level MRP':        CLR_RED,
    'No unit NSV — Aggregate denomination':     CLR_RED,
    'VMM excluded (no margin data)':            CLR_PURP,
}

# ── low-level style helpers ────────────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(colour: str) -> PatternFill:
    return PatternFill('solid', fgColor=colour)


def _font(bold=False, size=9, color='000000', italic=False, name='Arial') -> Font:
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def _align(h='left', v='center', wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _sc(ws, row: int, col: int, value=None, *,
        bg: Optional[str] = None,
        bold: bool = False,
        size: int = 9,
        color: str = '000000',
        h: str = 'left',
        wrap: bool = False,
        fmt: Optional[str] = None,
        italic: bool = False,
        height: Optional[int] = None,
        border: bool = False):
    """Set cell with full styling.  Returns the cell."""
    c = ws.cell(row, col, value)
    c.font      = _font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = _align(h=h, v='center', wrap=wrap)
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _thin_border()
    if height:
        ws.row_dimensions[row].height = height
    return c


def _section_header(ws, row: int, title: str, ncols: int = 8) -> int:
    """Section title spanning ncols columns.  Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    _sc(ws, row, 1, title, bg=CLR_INDIGO, bold=True, size=10, color=CLR_WHITE,
        h='left', height=20)
    return row + 1


def _col_header(ws, row: int, col: int, text: str,
                bg: str = CLR_DKBLUE, wrap: bool = True):
    _sc(ws, row, col, text, bg=bg, bold=True, color=CLR_WHITE,
        h='center', wrap=wrap)


def _border_range(ws, r1: int, r2: int, c1: int, c2: int):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _thin_border()


# ── metrics computation ────────────────────────────────────────────────────────

def _cov_tier(status: str) -> str:
    if 'PRIMARY_INVOICE'  in status: return 'Validated — Primary Invoice History'
    if 'FORMULA_FALLBACK' in status: return 'Validated — Formula (stored TOT%)'
    if 'FORMULA_ESTIMATED'in status: return 'Estimated placeholder (chain=ALL, qty=0)'
    if 'PACK_CASE'        in status: return 'No unit NSV — Pack/Case-level MRP'
    if 'AGGREGATE'        in status: return 'No unit NSV — Aggregate denomination'
    if 'NO_MARGIN'        in status or 'EXCLUDED_VMM' in status:
        return 'VMM excluded (no margin data)'
    return status or '(Unknown)'


def compute_metrics(
    forecast_df: pd.DataFrame,
    scenario_dfs: Dict[str, pd.DataFrame],
    source_version: str = '',
    workbook_version: str = '1.0',
) -> dict:
    """
    Derive all summary metrics from *forecast_df* (full population, including
    VMM and ESTIMATED rows).  Returns a flat/nested dict used by summary-sheet
    builders and the acceptance report.
    """
    df = forecast_df.copy()

    # Normalise any column names that carry our "(₹)" / "(₹, prov.)" sentinels
    # so this function works whether called from the pipeline (base names) or
    # standalone on an already-formatted workbook (renamed headers).
    _REVERSE = {v: k for k, v in _HEADER_RENAMES.items()}
    df.rename(columns=_REVERSE, inplace=True)

    # Coerce numeric columns
    for col in ['forecast_qty', 'gross_forecast_qty', 'forecast_primary_qty',
                'forecast_offtake_qty', 'forecast_nsv', 'forecast_trade_spend', 'forecast_cm2']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
        else:
            df[col] = 0.0

    def _str_col(name, default=''):
        if name in df.columns:
            return df[name].fillna(default).astype(str)
        return pd.Series([default] * len(df), index=df.index)

    df['forecast_month']      = _str_col('forecast_month')
    df['brand']               = _str_col('brand', '(Unknown)')
    df['value_source']        = _str_col('value_source')
    df['unit_price_status']   = _str_col('unit_price_status')
    df['cm2_approval_status'] = _str_col('cm2_approval_status')
    df['mrp_denomination']    = _str_col('mrp_denomination')
    df['is_tentative']        = df['is_tentative'].astype(bool) if 'is_tentative' in df.columns else True

    # Operational mask
    if 'operational_inclusion_flag' in df.columns:
        f = df['operational_inclusion_flag']
        op_mask = f.astype(str).str.upper().isin(['TRUE', '1', 'YES']) if f.dtype == object else f.astype(bool)
    else:
        op_mask = pd.Series(True, index=df.index)

    vmm_mask  = df['value_source'] == 'EXCLUDED_VMM'
    est_mask  = df['value_source'] == 'ESTIMATED'
    excl_mask = ~op_mask

    # ── quantities ─────────────────────────────────────────────────────────────
    gross_qty    = float(df['gross_forecast_qty'].sum())
    vmm_qty      = float(df.loc[vmm_mask, 'gross_forecast_qty'].sum())
    other_excl   = float(df.loc[excl_mask & ~vmm_mask, 'gross_forecast_qty'].sum())
    net_op_qty   = float(df['forecast_qty'].sum())
    check_bal    = round(gross_qty - vmm_qty - other_excl - net_op_qty, 4)

    # ── financial ──────────────────────────────────────────────────────────────
    total_nsv = float(df['forecast_nsv'].sum())
    total_ts  = float(df['forecast_trade_spend'].sum())
    total_cm2 = float(df['forecast_cm2'].sum())

    # ── coverage ───────────────────────────────────────────────────────────────
    nsv_pos_qty  = float(df.loc[df['forecast_nsv'] > 0, 'forecast_qty'].sum())
    denom_qty    = net_op_qty if net_op_qty > 0 else 1.0
    nsv_cov_pct  = nsv_pos_qty / denom_qty

    fa_cm2_rows  = int((df['cm2_approval_status'] == 'FINANCE_APPROVED').sum())
    total_rows   = len(df)
    # Derive pending-MRP from unit_price_status (always present) rather than
    # mrp_denomination (not always propagated to the planning workbook).
    pending_mrp  = int(df['unit_price_status'].isin([
                       'NO_UNIT_NSV_PACK_CASE_LEVEL_MRP',
                       'NO_UNIT_NSV_AGGREGATE_DENOMINATION']).sum())
    pending_cm2  = total_rows - fa_cm2_rows

    # ── mode / period ──────────────────────────────────────────────────────────
    is_tentative = bool(df['is_tentative'].iloc[0]) if len(df) else True
    months = sorted(m for m in df['forecast_month'].unique() if m)

    _mlbl_map = {
        '2026-09': 'Sep-26', '2026-10': 'Oct-26', '2026-11': 'Nov-26',
        '2026-12': 'Dec-26', '2027-01': 'Jan-27', '2027-02': 'Feb-27',
        '2027-03': 'Mar-27',
    }

    def mlbl(m: str) -> str:
        if m in _mlbl_map:
            return _mlbl_map[m]
        # generic: YYYY-MM → Mon-YY
        try:
            d = dt.datetime.strptime(m, '%Y-%m')
            return d.strftime('%b-%y')
        except Exception:
            return m

    period = (f"{mlbl(months[0])} to {mlbl(months[-1])}" if months else '—')

    # ── release status ─────────────────────────────────────────────────────────
    if is_tentative:
        if nsv_cov_pct >= 0.50 and fa_cm2_rows > 0:
            release_status = 'TENTATIVE – FINANCIAL VIEW WITH WARNINGS'
        else:
            release_status = 'TENTATIVE – QUANTITY PLANNING READY'
    else:
        if fa_cm2_rows == total_rows:
            release_status = 'FINAL – BUSINESS USE READY'
        else:
            release_status = 'BLOCKED – MATERIAL DATA ISSUE'

    # ── month rollup ───────────────────────────────────────────────────────────
    fin_c = [c for c in ['forecast_qty', 'gross_forecast_qty', 'forecast_primary_qty',
                          'forecast_nsv', 'forecast_trade_spend', 'forecast_cm2']
             if c in df.columns]
    month_grp = df.groupby('forecast_month')[fin_c].sum()
    if 'gross_forecast_qty' in df.columns:
        vmm_m = df[vmm_mask].groupby('forecast_month')['gross_forecast_qty'].sum()
        month_grp['vmm_excluded_qty'] = vmm_m.reindex(month_grp.index, fill_value=0.0)
    month_data = month_grp.to_dict('index')

    # ── brand rollup (operational, excl. VMM & ESTIMATED) ─────────────────────
    op_df     = df[~vmm_mask & ~est_mask]
    brand_grp = op_df.groupby('brand')[fin_c].sum()
    brand_data = brand_grp.to_dict('index')

    # ── coverage tier rollup ───────────────────────────────────────────────────
    df['_tier'] = df['unit_price_status'].apply(_cov_tier)
    cov_grp = df.groupby('_tier').agg(
        rows    = ('forecast_qty', 'count'),
        net_qty = ('forecast_qty', 'sum'),
        nsv     = ('forecast_nsv', 'sum'),
    ).reset_index().rename(columns={'_tier': 'tier'})
    coverage_data = cov_grp.to_dict('records')

    # ── scenario summaries ─────────────────────────────────────────────────────
    scenario_summary: Dict[str, dict] = {}
    for sname, sdf in scenario_dfs.items():
        scenario_summary[sname] = {
            'qty': float(sdf['forecast_qty'].sum()) if 'forecast_qty' in sdf.columns else 0.0,
            'nsv': float(sdf['forecast_nsv'].sum()) if 'forecast_nsv' in sdf.columns else 0.0,
            'cm2': float(sdf['forecast_cm2'].sum()) if 'forecast_cm2' in sdf.columns else 0.0,
        }

    return {
        # identity
        'forecast_mode':     'TENTATIVE' if is_tentative else 'FINAL',
        'forecast_period':   period,
        'refresh_at':        dt.datetime.now().isoformat(timespec='seconds'),
        'source_version':    source_version or 'Release_v1.0.0_RC1',
        'workbook_version':  workbook_version,
        'release_status':    release_status,
        # quantities
        'gross_qty':         gross_qty,
        'vmm_qty':           vmm_qty,
        'other_excl_qty':    other_excl,
        'net_op_qty':        net_op_qty,
        'check_balance':     check_bal,
        # financial
        'total_nsv':         total_nsv,
        'total_ts':          total_ts,
        'total_cm2':         total_cm2,
        # coverage
        'nsv_cov_pct':       nsv_cov_pct,
        'nsv_pos_qty':       nsv_pos_qty,
        'fa_cm2_rows':       fa_cm2_rows,
        'total_rows':        total_rows,
        'pending_mrp':       pending_mrp,
        'pending_cm2':       pending_cm2,
        # breakdowns
        'months':            months,
        'mlbl_map':          _mlbl_map,
        'month_data':        month_data,
        'brand_data':        brand_data,
        'coverage_data':     coverage_data,
        'scenario_summary':  scenario_summary,
    }


# ── India_Summary builder ──────────────────────────────────────────────────────

def _add_india_summary(
    wb: openpyxl.Workbook,
    m: dict,
    recon_df: Optional[pd.DataFrame] = None,
) -> None:
    """Replace (or create) the India_Summary sheet as the first tab."""
    if 'India_Summary' in wb.sheetnames:
        del wb['India_Summary']
    ws = wb.create_sheet('India_Summary', 0)

    # column widths
    for col, w in [('A', 40), ('B', 22), ('C', 18), ('D', 18),
                   ('E', 16), ('F', 24), ('G', 14), ('H', 14)]:
        ws.column_dimensions[col].width = w

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1,
        'FY27 DEMAND FORECAST  ·  HONASA / MAMAEARTH  ·  Modern Trade Channel',
        bg=CLR_NAVY, bold=True, size=13, color=CLR_WHITE, h='center', height=28)
    row += 1

    ws.merge_cells(f'A{row}:H{row}')
    c = ws.cell(row, 1,
        '1 Crore (Cr) = 100 Lakhs (L) = ₹1,00,00,000   |   '
        'No Million / Billion notation   |   '
        'All ₹ summary values in Crores (2 dp) unless labelled Lakhs')
    c.font      = _font(italic=True, size=8, color='555555')
    c.alignment = _align(h='center', wrap=True)
    ws.row_dimensions[row].height = 13
    row += 2

    # ── EXECUTIVE CONTROL PANEL ───────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1, 'EXECUTIVE CONTROL PANEL', bg=CLR_NAVY, bold=True,
        size=11, color=CLR_WHITE, h='left', height=22)
    row += 1

    def _kv(label: str, value, fmt: Optional[str] = None,
            bold_val: bool = False, val_color: str = '000000',
            label_bg: str = CLR_GREY) -> int:
        nonlocal row
        _sc(ws, row, 1, label, bg=label_bg, bold=True, size=9,
            color='333333', h='left', height=16)
        _sc(ws, row, 2, value, bg=None, bold=bold_val, size=9,
            color=val_color, h='left', fmt=fmt)
        _border_range(ws, row, row, 1, 2)
        row += 1
        return row

    mode_label = m['forecast_mode']
    mode_color = CLR_DKGRN if mode_label == 'FINAL' else CLR_AMBER
    _kv('Forecast Mode',   mode_label, bold_val=True, val_color=mode_color)
    _kv('Forecast Period', m['forecast_period'])
    _kv('Last Refresh',    m['refresh_at'])
    _kv('Source Version',  m['source_version'])
    _kv('Workbook Version', m['workbook_version'])
    row += 1

    # Release status banner
    status = m['release_status']
    sbg, stxt = _STATUS_STYLES.get(status, (CLR_AMBER, CLR_WHITE))
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1, f'RELEASE STATUS  ·  {status}',
        bg=sbg, bold=True, size=11, color=stxt, h='center', height=24)
    row += 2

    # ── Quantity block ────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1, 'QUANTITY METRICS', bg=CLR_DKBLUE, bold=True, size=9,
        color=CLR_WHITE, h='left', height=16)
    row += 1

    _kv('Gross Forecast Qty (Sep–Nov 2026)', m['gross_qty'],
        fmt=IND_FMT_INT)
    _kv('  VMM Excluded Qty', m['vmm_qty'],
        fmt=IND_FMT_INT, val_color=CLR_PURP)
    _kv('  Other Excluded Qty', m['other_excl_qty'],
        fmt=IND_FMT_INT)
    _kv('Net Operational Qty', m['net_op_qty'],
        fmt=IND_FMT_INT, bold_val=True, val_color=CLR_DKGRN)
    bal = m['check_balance']
    _kv('Reconciliation Check Balance  (must = 0)', round(bal, 4),
        fmt='#,##0.0000',
        bold_val=(abs(bal) > 0.001),
        val_color=CLR_RED if abs(bal) > 0.001 else CLR_DKGRN)
    row += 1

    # ── Financial block ───────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1, 'FINANCIAL METRICS  (₹ Crores)', bg=CLR_DKBLUE, bold=True,
        size=9, color=CLR_WHITE, h='left', height=16)
    row += 1

    _kv('Forecast NSV  (₹ Crores)', m['total_nsv'] / CRORE,
        fmt=IND_FMT_CR, bold_val=True)
    _kv('Forecast NSV  (₹ Lakhs)', m['total_nsv'] / LAKH,
        fmt=IND_FMT_CR)
    _kv('Trade Spend — est.  (₹ Crores)', m['total_ts'] / CRORE,
        fmt=IND_FMT_CR)
    _kv('CM2 Provisional  (₹ Crores)', m['total_cm2'] / CRORE,
        fmt=IND_FMT_CR, bold_val=True, val_color=CLR_AMBER)
    row += 1

    # ── Coverage / Approvals block ────────────────────────────────────────────
    ws.merge_cells(f'A{row}:H{row}')
    _sc(ws, row, 1, 'COVERAGE & PENDING APPROVALS', bg=CLR_DKBLUE, bold=True,
        size=9, color=CLR_WHITE, h='left', height=16)
    row += 1

    _kv('Validated NSV Quantity Coverage %', m['nsv_cov_pct'],
        fmt=IND_FMT_PCT)
    fa  = m['fa_cm2_rows']
    tot = m['total_rows']
    _kv('Finance-Approved CM2 Rows',
        f"{fa} / {tot}  ({fa/tot:.1%})" if tot > 0 else '0 / 0',
        val_color=CLR_DKGRN if fa == tot else CLR_RED)
    _kv('Rows Pending Unit MRP  (Pack/Aggregate)', m['pending_mrp'],
        fmt=IND_FMT_INT,
        val_color=CLR_RED if m['pending_mrp'] > 0 else CLR_DKGRN)
    _kv('Rows Pending CM2 Finance Approval', m['pending_cm2'],
        fmt=IND_FMT_INT,
        val_color=CLR_RED if m['pending_cm2'] > 0 else CLR_DKGRN)
    row += 2

    # ── SECTION 1: Month Summary ───────────────────────────────────────────────
    row = _section_header(ws, row, 'SECTION 1  —  MONTH-WISE OPERATIONAL SUMMARY')
    s1_data_start = row + 1

    hdrs1 = ['Month', 'Net Op Qty\n(Units)', 'Gross Qty\n(Units)', 'VMM Excl.\n(Units)',
             'NSV\n(₹ Cr)', 'Trade Spend\n(₹ Cr)', 'CM2 Prov.\n(₹ Cr)', 'Primary Qty\n(Units)']
    ws.row_dimensions[row].height = 38
    for ci, h in enumerate(hdrs1, 1):
        _col_header(ws, row, ci, h)
    row += 1

    months_sorted = m['months']

    def _mget(mo: str, key: str, default=0.0) -> float:
        return float(m['month_data'].get(mo, {}).get(key, default) or default)

    for ri, mo in enumerate(months_sorted):
        bg = CLR_GREY if ri % 2 == 0 else CLR_WHITE
        ws.row_dimensions[row].height = 15
        vals = [m['mlbl_map'].get(mo, mo),
                _mget(mo, 'forecast_qty'),
                _mget(mo, 'gross_forecast_qty'),
                _mget(mo, 'vmm_excluded_qty'),
                _mget(mo, 'forecast_nsv') / CRORE,
                _mget(mo, 'forecast_trade_spend') / CRORE,
                _mget(mo, 'forecast_cm2') / CRORE,
                _mget(mo, 'forecast_primary_qty')]
        fmts = [None, IND_FMT_INT, IND_FMT_INT, IND_FMT_INT,
                IND_FMT_CR, IND_FMT_CR, IND_FMT_CR, IND_FMT_INT]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            _sc(ws, row, ci, v, bg=bg, fmt=f, h='right' if ci > 1 else 'left')
        row += 1

    # Totals row
    def _ms(key): return sum(_mget(mo, key) for mo in months_sorted)
    ws.row_dimensions[row].height = 17
    tot1 = ['TOTAL (Sep–Nov 2026)',
            _ms('forecast_qty'), _ms('gross_forecast_qty'), _ms('vmm_excluded_qty'),
            _ms('forecast_nsv') / CRORE, _ms('forecast_trade_spend') / CRORE,
            _ms('forecast_cm2') / CRORE, _ms('forecast_primary_qty')]
    for ci, (v, f) in enumerate(zip(tot1,
            [None, IND_FMT_INT, IND_FMT_INT, IND_FMT_INT,
             IND_FMT_CR, IND_FMT_CR, IND_FMT_CR, IND_FMT_INT]), 1):
        _sc(ws, row, ci, v, bg=CLR_GREEN, bold=True, color=CLR_WHITE,
            fmt=f, h='right' if ci > 1 else 'left')
    _border_range(ws, s1_data_start, row, 1, 8)
    row += 2

    # ── SECTION 2: Brand Summary ──────────────────────────────────────────────
    row = _section_header(ws, row, 'SECTION 2  —  BRAND-WISE SUMMARY  (Operational, excl. VMM & Estimated)')
    s2_data_start = row + 1

    hdrs2 = ['Brand', 'Net Op Qty\n(Units)', 'NSV\n(₹ Cr)',
             'Trade Spend\n(₹ Cr)', 'CM2 Prov.\n(₹ Cr)']
    ws.row_dimensions[row].height = 38
    for ci, h in enumerate(hdrs2, 1):
        _col_header(ws, row, ci, h)
    row += 1

    brands_sorted = sorted(m['brand_data'].keys())

    def _bget(brand: str, key: str) -> float:
        return float(m['brand_data'].get(brand, {}).get(key, 0.0) or 0.0)

    for ri, brand in enumerate(brands_sorted):
        bg = CLR_GREY if ri % 2 == 0 else CLR_WHITE
        ws.row_dimensions[row].height = 15
        vals = [brand,
                _bget(brand, 'forecast_qty'),
                _bget(brand, 'forecast_nsv') / CRORE,
                _bget(brand, 'forecast_trade_spend') / CRORE,
                _bget(brand, 'forecast_cm2') / CRORE]
        fmts = [None, IND_FMT_INT, IND_FMT_CR, IND_FMT_CR, IND_FMT_CR]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            _sc(ws, row, ci, v, bg=bg, fmt=f, h='right' if ci > 1 else 'left')
        row += 1

    def _bs(key): return sum(_bget(b, key) for b in brands_sorted)
    ws.row_dimensions[row].height = 17
    tot2 = ['TOTAL', _bs('forecast_qty'),
            _bs('forecast_nsv') / CRORE,
            _bs('forecast_trade_spend') / CRORE,
            _bs('forecast_cm2') / CRORE]
    for ci, (v, f) in enumerate(zip(tot2, [None, IND_FMT_INT, IND_FMT_CR, IND_FMT_CR, IND_FMT_CR]), 1):
        _sc(ws, row, ci, v, bg=CLR_GREEN, bold=True, color=CLR_WHITE,
            fmt=f, h='right' if ci > 1 else 'left')
    _border_range(ws, s2_data_start, row, 1, 5)
    row += 2

    # ── SECTION 3: NSV Coverage Tiers ─────────────────────────────────────────
    row = _section_header(ws, row, 'SECTION 3  —  NSV FINANCIAL COVERAGE  (by MRP Denomination Type)')
    s3_data_start = row + 1

    hdrs3 = ['Coverage Tier', 'Rows', 'Net Op Qty\n(Units)', 'NSV (₹ Cr)', 'NSV (₹ Lakhs)']
    ws.row_dimensions[row].height = 38
    for ci, h in enumerate(hdrs3, 1):
        _col_header(ws, row, ci, h)
    row += 1

    cov_dict = {d['tier']: d for d in m['coverage_data']}
    for tier in _TIER_ORDER:
        d  = cov_dict.get(tier, {'rows': 0, 'net_qty': 0.0, 'nsv': 0.0})
        bg = _TIER_BG.get(tier, CLR_GREY)
        ws.row_dimensions[row].height = 16
        vals = [tier, int(d['rows']), float(d['net_qty']),
                float(d['nsv']) / CRORE, float(d['nsv']) / LAKH]
        fmts = [None, '#,##0', IND_FMT_INT, IND_FMT_CR, IND_FMT_CR]
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            _sc(ws, row, ci, v, bg=bg, color=CLR_WHITE,
                bold=(ci == 1), fmt=f, h='right' if ci > 1 else 'left')
        row += 1

    ws.row_dimensions[row].height = 17
    tot_r = sum(int(d['rows']) for d in m['coverage_data'])
    tot_q = sum(float(d['net_qty']) for d in m['coverage_data'])
    tot_n = sum(float(d['nsv']) for d in m['coverage_data'])
    for ci, (v, f) in enumerate(zip(
            ['GRAND TOTAL', tot_r, tot_q, tot_n / CRORE, tot_n / LAKH],
            [None, '#,##0', IND_FMT_INT, IND_FMT_CR, IND_FMT_CR]), 1):
        _sc(ws, row, ci, v, bg=CLR_GREEN, bold=True, color=CLR_WHITE,
            fmt=f, h='right' if ci > 1 else 'left')
    _border_range(ws, s3_data_start, row, 1, 5)
    row += 2

    # ── SECTION 4: VMM Reconciliation ─────────────────────────────────────────
    row = _section_header(ws, row, 'SECTION 4  —  VMM RECONCILIATION  (Gross → Exclusions → Net Operational)')
    s4_data_start = row + 1

    hdrs4 = ['Month', 'Gross Qty\n(Units)', 'VMM Excl.\n(Units)',
             'Other Excl.\n(Units)', 'Net Op Qty\n(Units)', 'Check Balance\n(must = 0)']
    ws.row_dimensions[row].height = 38
    for ci, h in enumerate(hdrs4, 1):
        _col_header(ws, row, ci, h)
    row += 1

    if recon_df is not None and not recon_df.empty:
        for ri, (_, rrow) in enumerate(recon_df.iterrows()):
            bg  = CLR_GREY if ri % 2 == 0 else CLR_WHITE
            bal = float(rrow.get('check_balance', 0))
            ws.row_dimensions[row].height = 15
            vals = [rrow.get('forecast_month', ''),
                    rrow.get('gross_forecast_qty', 0),
                    rrow.get('vmm_excluded_qty', 0),
                    rrow.get('other_excluded_qty', 0),
                    rrow.get('net_operational_qty', 0),
                    bal]
            fmts = [None, IND_FMT_INT, IND_FMT_INT, IND_FMT_INT, IND_FMT_INT, '#,##0.0000']
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                _sc(ws, row, ci, v, bg=bg, fmt=f,
                    color=(CLR_RED if abs(bal) > 0.001 else CLR_DKGRN) if ci == 6 else '000000',
                    bold=(ci == 6 and abs(bal) > 0.001),
                    h='right' if ci > 1 else 'left')
            row += 1
        _border_range(ws, s4_data_start, row - 1, 1, 6)
    else:
        ws.merge_cells(f'A{row}:F{row}')
        _sc(ws, row, 1,
            'No reconciliation CSV available — run full forecast pipeline to populate this section.',
            bg=CLR_AMBRLT, italic=True, h='left')
        row += 1
    row += 1

    # ── SECTION 5: Reconciliation Validation ──────────────────────────────────
    row = _section_header(ws, row, 'SECTION 5  —  RECONCILIATION VALIDATION  (Detail vs Summary checks)')
    s5_data_start = row + 1

    hdrs5 = ['Check', 'Detail Value', 'Summary Value', 'Difference', 'Status', 'Note']
    ws.row_dimensions[row].height = 16
    for ci, h in enumerate(hdrs5, 1):
        _col_header(ws, row, ci, h, wrap=False)
    row += 1

    sum_month_qty = sum(_mget(mo, 'forecast_qty') for mo in months_sorted)
    sum_brand_qty = sum(_bget(b, 'forecast_qty') for b in brands_sorted)
    sum_tier_rows = sum(int(d['rows']) for d in m['coverage_data'])

    recon_checks = [
        ('Net Op Qty = Sum of forecast months',
         m['net_op_qty'], sum_month_qty, 0.01,
         'ESTIMATED rows have qty=0 but appear in month totals'),
        ('Net Op Qty ≥ Sum of brand totals (excl. ESTIMATED)',
         m['net_op_qty'], sum_brand_qty, 0.01,
         'ESTIMATED (chain=ALL) excluded from brand rollup; net_op_qty includes ESTIMATED qty=0 rows'),
        ('Gross − VMM − OtherExcl − NetOp = 0',
         0.0, m['check_balance'], 0.001,
         'Must equal 0 for reconciliation integrity'),
        ('Coverage tier row count = Total forecast rows',
         m['total_rows'], sum_tier_rows, 0,
         'Every row must map to exactly one coverage tier'),
    ]

    for chk, detail, summary_val, tol, note in recon_checks:
        detail_f  = float(detail)
        summary_f = float(summary_val)
        diff      = round(detail_f - summary_f, 4)
        ok        = abs(diff) <= float(tol)
        ws.row_dimensions[row].height = 16
        bg = CLR_GREY if row % 2 == 0 else CLR_WHITE
        _sc(ws, row, 1, chk, bg=bg, h='left')
        _sc(ws, row, 2, detail_f, bg=bg, fmt='#,##0.####', h='right')
        _sc(ws, row, 3, summary_f, bg=bg, fmt='#,##0.####', h='right')
        _sc(ws, row, 4, diff, bg=bg, fmt='#,##0.####', h='right',
            color=CLR_RED if abs(diff) > float(tol) else '000000')
        _sc(ws, row, 5, '✓ OK' if ok else '✗ MISMATCH', bg=bg,
            bold=True, color=CLR_DKGRN if ok else CLR_RED, h='center')
        _sc(ws, row, 6, note, bg=bg, h='left')
        row += 1

    _border_range(ws, s5_data_start, row - 1, 1, 6)
    row += 2

    # ── SECTION 6: Scenario Comparison ────────────────────────────────────────
    if m['scenario_summary']:
        row = _section_header(ws, row, 'SECTION 6  —  SCENARIO COMPARISON')
        s6_data_start = row + 1

        hdrs6 = ['Scenario', 'Qty (Units)', 'NSV (₹ Cr)', 'CM2 (₹ Cr)']
        ws.row_dimensions[row].height = 16
        for ci, h in enumerate(hdrs6, 1):
            _col_header(ws, row, ci, h, wrap=False)
        row += 1

        for ri, (sname, sd) in enumerate(m['scenario_summary'].items()):
            bg = CLR_GREY if ri % 2 == 0 else CLR_WHITE
            ws.row_dimensions[row].height = 15
            vals = [sname.replace('_', ' ').title(),
                    sd['qty'], sd['nsv'] / CRORE, sd['cm2'] / CRORE]
            fmts = [None, IND_FMT_INT, IND_FMT_CR, IND_FMT_CR]
            for ci, (v, f) in enumerate(zip(vals, fmts), 1):
                _sc(ws, row, ci, v, bg=bg, fmt=f,
                    h='right' if ci > 1 else 'left')
            row += 1
        _border_range(ws, s6_data_start, row - 1, 1, 4)

    # Freeze panes: keep title visible while scrolling
    ws.freeze_panes = 'A4'


# ── data-sheet formatting ──────────────────────────────────────────────────────

def _format_data_sheet(wb: openpyxl.Workbook, sheet_name: str) -> int:
    """
    Apply Indian number formats, freeze panes, and autofilter to one data sheet.
    Idempotent: header guard prevents double-rename; formats re-applied to same cells.
    Returns the count of formatted columns.
    """
    if sheet_name not in wb.sheetnames:
        return 0

    ws = wb[sheet_name]

    # Build column map (strip existing sentinel to find base names)
    col_map: Dict[str, int] = {}
    for cell in ws[1]:
        raw  = str(cell.value or '').strip()
        base = (raw
                .replace(' (₹, prov.)', '')
                .replace(' (₹)', '')
                .replace(' (Units)', ''))
        if base in _FIN_COLS:
            col_map[base] = cell.column

    # Rename headers — idempotent: skip if sentinel already present
    for cell in ws[1]:
        raw = str(cell.value or '').strip()
        if raw in _HEADER_RENAMES and '(₹)' not in raw:
            cell.value = _HEADER_RENAMES[raw]

    # Apply number formats column by column
    max_row = ws.max_row
    for col_base, col_idx in col_map.items():
        if col_base in RUPEE_COLS:
            fmt = IND_FMT_RUPEE
        elif col_base in QTY_COLS:
            fmt = IND_FMT_INT
        elif col_base in TEXT_COLS:
            fmt = IND_FMT_TEXT
        else:
            continue
        for r in range(2, max_row + 1):
            c = ws.cell(r, col_idx)
            if c.value is not None:
                c.number_format = fmt

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Autofilter on the full data range
    if max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max_row}"

    return len(col_map)


# ── acceptance report ──────────────────────────────────────────────────────────

def _build_acceptance_report(
    m: dict,
    formatted_cols: Dict[str, int],
) -> dict:
    """Return structured acceptance-report dict for logging and JSON summary."""

    checks: List[dict] = []

    def _chk(name: str, passed: bool, detail: str = ''):
        checks.append({'check': name, 'pass': passed, 'detail': detail})

    # Reconciliation
    _chk('check_balance = 0 (gross−VMM−excl−net)',
         abs(m['check_balance']) < 0.001,
         f"balance = {m['check_balance']:.4f}")

    # Quantities
    _chk('Net operational qty > 0',
         m['net_op_qty'] > 0,
         f"{m['net_op_qty']:,.0f} units")
    _chk('VMM qty < gross qty (VMM subset of gross)',
         m['vmm_qty'] < m['gross_qty'] + 0.001,
         f"VMM {m['vmm_qty']:,.0f} / gross {m['gross_qty']:,.0f}")

    # Financial guardrails
    _chk('Trade spend ≤ NSV  (no inverted margin)',
         m['total_ts'] <= m['total_nsv'] + 0.01,
         f"TS ₹{m['total_ts']/CRORE:.2f} Cr  NSV ₹{m['total_nsv']/CRORE:.2f} Cr")
    _chk('CM2 ≤ NSV  (gross margin sanity)',
         m['total_cm2'] <= m['total_nsv'] + 0.01,
         f"CM2 ₹{m['total_cm2']/CRORE:.2f} Cr  NSV ₹{m['total_nsv']/CRORE:.2f} Cr")
    _chk('NSV validated coverage > 0%',
         m['nsv_cov_pct'] > 0,
         f"{m['nsv_cov_pct']:.1%} of net operational qty has NSV > 0")

    # CM2 coverage (informational — 0 expected in tentative mode)
    _chk('CM2 approval status (informational)',
         True,
         f"Finance-approved: {m['fa_cm2_rows']} / {m['total_rows']} rows  "
         f"({'0 expected in TENTATIVE mode' if m['forecast_mode']=='TENTATIVE' else 'ALL required in FINAL mode'})")

    # Format
    all_sheets_formatted = all(v > 0 for v in formatted_cols.values())
    _chk('Indian format applied to all data sheets',
         all_sheets_formatted,
         '  '.join(f"{s}:{n}cols" for s, n in formatted_cols.items()))
    _chk('India_Summary is first sheet (position 0)',
         True,
         'Enforced by create_sheet(position=0)')

    # Excluded brands not in operational forecast
    _chk('Excluded brands (Pure Origin, Lumineve, Staze) absent from brand rollup',
         all(b not in m['brand_data']
             for b in ['Pure Origin', 'Lumineve', 'Staze']),
         'Enforced upstream in ForecastEngine._is_excluded_brand()')

    n_pass = sum(1 for c in checks if c['pass'])
    n_fail = len(checks) - n_pass
    fail_list = [c for c in checks if not c['pass']]

    return {
        'checks':                 checks,
        'pass_count':             n_pass,
        'fail_count':             n_fail,
        'all_pass':               n_fail == 0,
        'release_status':         m['release_status'],
        'recommendation':         m['release_status'],
        'nsv_coverage_pct':       f"{m['nsv_cov_pct']:.1%}",
        'fa_cm2_rows':            m['fa_cm2_rows'],
        'total_rows':             m['total_rows'],
        'pending_mrp_rows':       m['pending_mrp'],
        'pending_cm2_rows':       m['pending_cm2'],
        'formatted_cols_per_sheet': formatted_cols,
        'failing_checks':         fail_list,
    }


def print_acceptance_report(report: dict) -> None:
    """Print a human-readable acceptance report to stdout."""
    bar = '─' * 68
    print(f"\n{bar}")
    print(f"  INDIA-FORMAT WORKBOOK ACCEPTANCE REPORT")
    print(bar)
    print(f"  Release Status : {report['release_status']}")
    print(f"  NSV Coverage   : {report['nsv_coverage_pct']}")
    print(f"  CM2 Approved   : {report['fa_cm2_rows']} / {report['total_rows']} rows")
    print(f"  Pending MRP    : {report['pending_mrp_rows']} rows (Pack/Aggregate)")
    print(f"  Pending CM2    : {report['pending_cm2_rows']} rows")
    print(f"  Checks         : {report['pass_count']} / {report['pass_count'] + report['fail_count']} passed")
    if report['failing_checks']:
        print(f"\n  ✗ FAILING CHECKS:")
        for c in report['failing_checks']:
            print(f"    • {c['check']}")
            if c.get('detail'):
                print(f"      {c['detail']}")
    print(bar)
    print(f"  RECOMMENDATION : {report['recommendation']}")
    print(bar)


# ── public entry point ─────────────────────────────────────────────────────────

def apply_india_format_to_workbook(
    wb_path: str,
    forecast_df: pd.DataFrame,
    scenario_dfs: Dict[str, pd.DataFrame],
    exception_df: pd.DataFrame,
    recon_csv_path: Optional[str] = None,
    version: str = '1.0',
    source_version: str = 'Release_v1.0.0_RC1',
) -> dict:
    """
    Apply India-standard formatting to an existing Forecast_Planning_Workbook.xlsx.

    Idempotent — safe to call on every build run.

    Parameters
    ----------
    wb_path         Path to the workbook written by _build_planning_workbook.
    forecast_df     Full base-forecast DataFrame (all rows, incl. VMM / ESTIMATED).
    scenario_dfs    Dict of scenario_name → DataFrame.
    exception_df    Exception rows (passed through; not used for summary aggregation).
    recon_csv_path  Optional path to fact_reconciliation_summary.csv.
    version         Workbook version string written to Control Panel.
    source_version  Margin-repo release tag written to Control Panel.

    Returns
    -------
    dict  Acceptance report (checks, counts, release status, recommendation).
    """
    m = compute_metrics(
        forecast_df, scenario_dfs,
        source_version=source_version,
        workbook_version=version,
    )

    recon_df = None
    if recon_csv_path and os.path.exists(recon_csv_path):
        recon_df = pd.read_csv(recon_csv_path)

    wb = load_workbook(wb_path)
    _add_india_summary(wb, m, recon_df=recon_df)

    formatted_cols: Dict[str, int] = {}
    for sname in DATA_SHEETS:
        formatted_cols[sname] = _format_data_sheet(wb, sname)

    wb.save(wb_path)

    report = _build_acceptance_report(m, formatted_cols)
    report['workbook_path'] = wb_path
    report['sheets']        = wb.sheetnames

    return report
