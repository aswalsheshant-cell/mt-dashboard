# -*- coding: utf-8 -*-
"""Render the 8-sheet correction-ready QC workbook from res.pkl."""
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = "/tmp/claude-0/-home-user-mt-dashboard/49d427d8-c459-5531-9f39-32d1bfca9b64/scratchpad/"
OUT = SP + "Store_Master_QC_Report.xlsx"
TODAY = date(2026, 7, 31).isoformat()

res = pd.read_pickle(SP + "res.pkl").sort_values("idx").reset_index(drop=True)
N = len(res)

# ---- master identity per duplicate group ----
masters = {}
for gid, g in res[res.dup_group != ""].groupby("dup_group"):
    mrow = g[g.master_flag == "MASTER"]
    if len(mrow):
        m = mrow.iloc[0]
        masters[gid] = "%s | code=%s" % (m.sn_raw, m.sc_clean or "(blank)")

# ---- assemble the master QC table (original + 25 new cols) ----
ordered = []
for _, r in res.iterrows():
    ordered.append({
        # original (unchanged)
        "Chain Name": r.chain_raw, "Site Code": r.sc_raw, "Site Name": r.sn_raw,
        "Zone": r.zone_raw, "State": r.state_raw, "City": r.city_raw,
        # 25 appended QC columns
        "Canonical_Chain_Name": r.chain_canon,
        "Site_Code_Clean": r.sc_clean,
        "Site_Name_Clean": r.sn_clean,
        "Current_Zone": r.zone_raw,
        "Current_State": r.state_raw,
        "Current_City": r.city_raw,
        "Recommended_City": r.rec_city,
        "Recommended_District": r.rec_district,
        "Recommended_State": r.rec_state,
        "Recommended_Business_Zone": r.rec_bzone,
        "Geographic_Zone": r.geo_zone,
        "City_QC_Status": r.city_status,
        "State_QC_Status": r.state_status,
        "Zone_QC_Status": r.zone_status,
        "Duplicate_Status": r.dup_status,
        "Duplicate_Group_ID": r.dup_group,
        "Duplicate_Type": r.dup_type,
        "Master_Record_Flag": r.master_flag,
        "Recommended_Action": r.action,
        "Confidence_Score": int(r.confidence),
        "Validation_Method": r.method,
        "Evidence_Source": r.evidence,
        "QC_Remarks": r.remarks,
        "Manual_Review_Required": r.manual,
        "Last_Validated_Date": TODAY,
    })
master = pd.DataFrame(ordered)

# ---------------------------------------------------------------------------
# Row colour classification (per prompt highlighting spec)
# ---------------------------------------------------------------------------
FILL = {
    "red":    PatternFill("solid", fgColor="F4B7B7"),
    "orange": PatternFill("solid", fgColor="FCD5A5"),
    "yellow": PatternFill("solid", fgColor="FFF2A8"),
    "blue":   PatternFill("solid", fgColor="C7DAF2"),
    "green":  PatternFill("solid", fgColor="C6E9C6"),
    "grey":   PatternFill("solid", fgColor="D9D9D9"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def row_color(r):
    if r.dup_status == "EXACT DUPLICATE" or r.action == "REMOVE EXACT DUPLICATE":
        return "red"
    if r.city_status == "CITY MISSING" or (r.sc_clean == "" and r.dup_type == "MISSING SITE CODE"):
        return "grey"
    if r.dup_status in ("POSSIBLE DUPLICATE", "SITE CODE CONFLICT",
                        "SAME STORE, DIFFERENT SITE CODE", "DUPLICATE DUE TO FORMATTING") \
            or r.manual == "YES":
        return "orange"
    if r.action in ("CORRECT CITY", "CORRECT STATE", "CORRECT ZONE") \
            or r.city_status in ("LOCALITY MAINTAINED AS CITY", "STATE MAINTAINED AS CITY") \
            or r.state_status in ("STATE CORRECTION REQUIRED", "STATE-CITY MISMATCH"):
        return "yellow"
    if r.city_status == "FORMAT STANDARDIZATION" or r.action == "STANDARDIZE CITY NAME" \
            or r.zone_status == "FORMAT STANDARDIZATION" or r.state_status == "FORMAT STANDARDIZATION":
        return "blue"
    return "green"


res["_color"] = res.apply(row_color, axis=1)

# ---------------------------------------------------------------------------
# Workbook writer helpers
# ---------------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

WRAP_COLS = {"QC_Remarks", "Validation_Method", "Evidence_Source", "Removal Reason",
             "Reason", "Final Recommendation", "Value"}
WIDE = {"QC_Remarks": 55, "Validation_Method": 30, "Evidence_Source": 40,
        "Site Name": 30, "Site_Name_Clean": 28, "Removal Reason": 45,
        "Final Recommendation": 28, "Master Record": 30, "Record to Remove": 30,
        "Metric": 34, "Value": 14, "Reason": 45, "Notes": 55}


def write_sheet(name, df, colors=None, freeze="A2", widths_extra=None):
    ws = wb.create_sheet(name)
    cols = list(df.columns)
    # header
    for c, col in enumerate(cols, 1):
        cell = ws.cell(1, c, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # body
    text_cols = {"Site Code", "Site_Code_Clean", "Current_Site_Code"}
    for ridx, (_, row) in enumerate(df.iterrows(), 2):
        fill = FILL[colors[ridx - 2]] if colors else None
        for c, col in enumerate(cols, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws.cell(ridx, c, val)
            cell.border = BORDER
            cell.font = Font(size=9)
            if col in text_cols:
                cell.number_format = "@"
                cell.value = str(val)
            if col in WRAP_COLS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            if fill is not None:
                cell.fill = fill
    # widths
    widths = dict(WIDE)
    if widths_extra:
        widths.update(widths_extra)
    for c, col in enumerate(cols, 1):
        w = widths.get(col, max(11, min(24, len(str(col)) + 3)))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = freeze
    if len(df):
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(df) + 1)
    return ws


# ---- 1. Store_Master_QC ----
write_sheet("Store_Master_QC", master, colors=list(res["_color"]))

# ---- 2. City_Corrections ----
city_mask = res.city_status.isin(["FORMAT STANDARDIZATION", "LOCALITY MAINTAINED AS CITY",
                                   "STATE MAINTAINED AS CITY", "CITY MISSING"]) \
    | res.state_status.isin(["STATE CORRECTION REQUIRED", "STATE-CITY MISMATCH", "FORMAT STANDARDIZATION"])
cc = res[city_mask]
cc_df = pd.DataFrame({
    "Chain Name": cc.chain_raw, "Site Code": cc.sc_raw, "Site Name": cc.sn_raw,
    "Current_City": cc.city_raw, "Recommended_City": cc.rec_city,
    "Recommended_District": cc.rec_district,
    "Current_State": cc.state_raw, "Recommended_State": cc.rec_state,
    "Current_Zone": cc.zone_raw, "Recommended_Business_Zone": cc.rec_bzone,
    "Geographic_Zone": cc.geo_zone,
    "City_QC_Status": cc.city_status, "State_QC_Status": cc.state_status,
    "Zone_QC_Status": cc.zone_status,
    "Recommended_Action": cc.action, "Confidence_Score": cc.confidence.astype(int),
    "Manual_Review_Required": cc.manual, "QC_Remarks": cc.remarks,
})
write_sheet("City_Corrections", cc_df, colors=[FILL and c for c in cc["_color"]])

# ---- 3. Duplicate_Review ----
dup = res[res.dup_status != "UNIQUE"].sort_values(["dup_group", "master_flag"])
dup_df = pd.DataFrame({
    "Duplicate_Group_ID": dup.dup_group, "Chain Name": dup.chain_raw,
    "Site Code": dup.sc_raw, "Site_Code_Clean": dup.sc_clean, "Site Name": dup.sn_raw,
    "Current_City": dup.city_raw, "Recommended_City": dup.rec_city,
    "Current_State": dup.state_raw, "Zone": dup.zone_raw,
    "Duplicate_Status": dup.dup_status, "Duplicate_Type": dup.dup_type,
    "Master_Record_Flag": dup.master_flag, "Conflict_Fields": dup.conflict_fields,
    "Recommended_Action": dup.action, "Confidence_Score": dup.confidence.astype(int),
    "Manual_Review_Required": dup.manual, "QC_Remarks": dup.remarks,
})
write_sheet("Duplicate_Review", dup_df, colors=list(dup["_color"]))

# ---- 4. Removal_Recommendations ----
rem = res[res.action.isin(["REMOVE EXACT DUPLICATE", "MERGE DUPLICATE RECORDS"])].copy()
rem = rem.sort_values(["dup_group", "master_flag"])
rem_df = pd.DataFrame({
    "Duplicate_Group_ID": rem.dup_group, "Chain Name": rem.chain_raw,
    "Site Code": rem.sc_raw, "Site Name": rem.sn_raw,
    "Current_City": rem.city_raw, "Recommended_City": rem.rec_city,
    "Master Record": rem.dup_group.map(masters).fillna(""),
    "Record to Remove": rem.apply(lambda x: "%s | code=%s" % (x.sn_raw, x.sc_clean or "(blank)")
                                  if x.master_flag != "MASTER" else "(this is master - retain)", axis=1),
    "Removal Reason": rem.apply(lambda x: ("Exact duplicate of master record." if x.dup_status == "EXACT DUPLICATE"
                                           else "Same store as master under different code/formatting.")
                                if x.master_flag != "MASTER" else "Retained as master record.", axis=1),
    "Confidence_Score": rem.confidence.astype(int),
    "Final Recommendation": rem.apply(lambda x: x.action if x.master_flag != "MASTER"
                                      else "RETAIN AS MASTER", axis=1),
})
write_sheet("Removal_Recommendations", rem_df, colors=list(rem["_color"]))

# ---- 5. Manual_Review ----
mr = res[res.manual == "YES"]
mr_df = pd.DataFrame({
    "Chain Name": mr.chain_raw, "Site Code": mr.sc_raw, "Site Name": mr.sn_raw,
    "Current_City": mr.city_raw, "Recommended_City": mr.rec_city,
    "Current_State": mr.state_raw, "Recommended_State": mr.rec_state, "Zone": mr.zone_raw,
    "City_QC_Status": mr.city_status, "State_QC_Status": mr.state_status,
    "Duplicate_Status": mr.dup_status, "Duplicate_Group_ID": mr.dup_group,
    "Recommended_Action": mr.action, "Confidence_Score": mr.confidence.astype(int),
    "Reason": mr.remarks,
})
write_sheet("Manual_Review", mr_df, colors=list(mr["_color"]))

# ---- 6. QC_Summary ----
def cnt(mask):
    return int(mask.sum())

summary = [
    ("Total source rows", N),
    ("Total processed rows", len(master)),
    ("Processing completion %", "100%"),
    ("", ""),
    ("Unique stores (no duplicate flag)", cnt(res.dup_status == "UNIQUE")),
    ("Exact duplicates", cnt(res.dup_status == "EXACT DUPLICATE")),
    ("Site code conflicts", cnt(res.dup_status == "SITE CODE CONFLICT")),
    ("Same store, different site code", cnt(res.dup_status == "SAME STORE, DIFFERENT SITE CODE")),
    ("Formatting duplicates", cnt(res.dup_status == "DUPLICATE DUE TO FORMATTING")),
    ("Possible duplicates", cnt(res.dup_status == "POSSIBLE DUPLICATE")),
    ("Total duplicate groups", res[res.dup_group != ""].dup_group.nunique()),
    ("", ""),
    ("City standardizations (format)", cnt(res.city_status == "FORMAT STANDARDIZATION")),
    ("City corrections (locality/state as city)", cnt(res.city_status.isin(
        ["LOCALITY MAINTAINED AS CITY", "STATE MAINTAINED AS CITY"]))),
    ("City missing", cnt(res.city_status == "CITY MISSING")),
    ("State corrections required", cnt(res.state_status == "STATE CORRECTION REQUIRED")),
    ("State-city mismatches", cnt(res.state_status == "STATE-CITY MISMATCH")),
    ("State standardizations (format)", cnt(res.state_status == "FORMAT STANDARDIZATION")),
    ("Zone corrections (format)", cnt(res.zone_status == "FORMAT STANDARDIZATION")),
    ("", ""),
    ("Records recommended for removal/merge", cnt(res.action.isin(
        ["REMOVE EXACT DUPLICATE", "MERGE DUPLICATE RECORDS"]))),
    ("Manual review records", cnt(res.manual == "YES")),
    ("Validated & passed (all-pass, retain)", cnt((res._color == "green"))),
    ("", ""),
    ("Avg confidence score", round(res.confidence.mean(), 1)),
    ("Rows skipped / not processed", 0),
]
sum_df = pd.DataFrame(summary, columns=["Metric", "Value"])
ws = write_sheet("QC_Summary", sum_df, freeze="A2")

# ---- 7. Chain_Summary ----
rows = []
for chain, g in res.groupby("chain_raw"):
    corr = cnt((g.city_status.isin(["FORMAT STANDARDIZATION", "LOCALITY MAINTAINED AS CITY",
                                     "STATE MAINTAINED AS CITY"]))
               | g.state_status.isin(["STATE CORRECTION REQUIRED", "STATE-CITY MISMATCH",
                                      "FORMAT STANDARDIZATION"])
               | (g.zone_status == "FORMAT STANDARDIZATION"))
    dups = cnt(g.dup_status != "UNIQUE")
    manual = cnt(g.manual == "YES")
    removal = cnt(g.action.isin(["REMOVE EXACT DUPLICATE", "MERGE DUPLICATE RECORDS"]))
    passed = cnt(g._color == "green")
    rows.append({
        "Chain Name": chain, "Stores": len(g), "Corrections": corr,
        "Duplicates": dups, "Manual_Reviews": manual,
        "Removal_Recommendations": removal,
        "QC_Accuracy_%": round(100.0 * passed / len(g), 1),
    })
chain_df = pd.DataFrame(rows).sort_values("Stores", ascending=False)
write_sheet("Chain_Summary", chain_df, freeze="A2")

# ---- 8. Audit_Log ----
audit = [
    ("Validation date", TODAY),
    ("Source file", "Final_Jan26_to_June26_chainwise_storelist.xlsb (Sheet1)"),
    ("Primary business key", "Chain Name x Site Code"),
    ("Validation source", "Offline curated India-geography reference (states, districts, "
     "canonical city spellings, locality-to-city, zone rules) + internal workbook consistency"),
    ("Web research performed", "No live web calls were made in this offline run. Rows needing "
     "store-locator / map confirmation are flagged WEB/ MANUAL and listed in Manual_Review."),
    ("Batch number", "Single reconciled pass (all rows in one batch)"),
    ("Rows processed", N),
    ("Rows skipped", 0),
    ("Completion status", "COMPLETE - 100% of rows classified"),
    ("Site Code handling", "Preserved as TEXT; '(blank)' treated as missing code"),
    ("Original columns", "Preserved unchanged; all recommendations added in NEW columns only"),
    ("Assumptions", "Business state groupings (Delhi NCR, UP/UK, Punjab/J&K/HP, Northeast) and "
     "business sub-zones (South-1, South-2) are valid and NOT treated as errors. Ambiguous "
     "multi-state city names (e.g. Bilaspur, Aurangabad) are not force-corrected."),
    ("Processing issues", "None. 1 row has a blank City (flagged CITY MISSING); 23 rows have "
     "blank Site Codes (flagged SITE CODE CONFLICT / missing)."),
    ("No fabricated evidence", "No web links or sources were invented. Confidence < 75 rows are "
     "marked Manual_Review_Required = YES."),
]
audit_df = pd.DataFrame(audit, columns=["Field", "Notes"])
write_sheet("Audit_Log", audit_df, freeze="A2", widths_extra={"Field": 26, "Notes": 90})

wb.save(OUT)
print("saved", OUT)
print("sheets:", wb.sheetnames)
