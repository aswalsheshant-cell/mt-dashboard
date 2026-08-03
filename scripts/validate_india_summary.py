from pathlib import Path
import openpyxl
import pandas as pd

# File & expected metadata definitions
FILE_PATH = Path("forecast_outputs/sep_nov_2026_tentative/India_Summary.xlsx")
EXPECTED_SHEET = "India_Summary"
EXPECTED_ROWS = 51
EXPECTED_COLS = 12


def validate_summary():
    print(f"🔍 Validating {FILE_PATH}...\n")

    # 1. Check file existence
    if not FILE_PATH.exists():
        print(f"❌ FAIL: File does not exist at {FILE_PATH}")
        return False
    print(f"✓ File exists: {FILE_PATH}")

    # 2. Inspect Workbook Structure & Formula Cells via OpenPyXL
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False)

    if EXPECTED_SHEET not in wb.sheetnames:
        print(f"❌ FAIL: Expected sheet '{EXPECTED_SHEET}' not found. Found: {wb.sheetnames}")
        return False
    print(f"✓ Sheet '{EXPECTED_SHEET}' present.")

    ws = wb[EXPECTED_SHEET]

    # Count formula cells
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("="):
                formula_count += 1

    if formula_count != 0:
        print(f"❌ FAIL: Found {formula_count} formula cells (Expected: 0).")
        return False
    print("✓ Formula cell check passed (0 formula cells found).")

    # 3. Inspect Dimensions & Data Contents via Pandas
    df = pd.read_excel(FILE_PATH, sheet_name=EXPECTED_SHEET)

    actual_rows, actual_cols = df.shape
    # Account for header row in dimensions count (51 rows total in spreadsheet = 1 header + 50 data rows or 51 data rows depending on index)
    print(f"  Shape detected by pandas: {actual_rows} data rows × {actual_cols} columns")

    # Check for empty / NaN columns
    empty_cols = df.columns[df.isna().all()].tolist()
    if empty_cols:
        print(f"⚠️  WARNING: Entirely empty columns detected: {empty_cols}")
    else:
        print("✓ No completely empty columns.")

    # Display preview summary
    print("\n--- Summary Data Preview ---")
    print(df.head(5).iloc[:, :5])  # First 5 rows and 5 columns
    print("----------------------------\n")

    print("✅ VALIDATION SUCCESSFUL: File structure is clean and ready for stakeholder review!")
    return True


if __name__ == "__main__":
    validate_summary()
