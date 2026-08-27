#!/usr/bin/env python3
"""
Excel Generator: TGT vs ACH — RKAM (Key Account / Chain) Scorecard
Produces: exports/TGT_vs_ACH_RKAM_<date>.xlsx
"""
import json, re, os
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print("Installing openpyxl..."); os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Load data ────────────────────────────────────────────────────────────────
raw = open('dashboard/data.js').read()
raw = re.sub(r'^.*?window\.DASH\s*=\s*', '', raw, flags=re.DOTALL)
raw = raw.rstrip().rstrip(';')
D = json.loads(raw)

# ── Colours / styles ─────────────────────────────────────────────────────────
NAVY  = "10254A"
GOLD  = "DAA520"
WHITE = "FFFFFF"
LIGHT = "EAF0FB"
GREEN = "1E8449"
RED   = "C0392B"
AMBER = "E67E22"
GRAY  = "7F8C8D"

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, r, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def shade(ws, r, ncols, col=LIGHT):
    for c in range(1, ncols+1):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=col)

def cell_set(ws, r, c, val, bold=False, align="left", color=None, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=color or "000000", name='Calibri', size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = border
    if num_fmt:
        cell.number_format = num_fmt
    return cell

# ── Pull chain data ───────────────────────────────────────────────────────────
by_chain_fy25 = {c['name']: c.get('nsv', 0) for c in D.get('primary', {}).get('by_chain', [])
                 if isinstance(c, dict)}
# FY27 chain breakdown from detail_records
chain_fy26 = {}
chain_fy27 = {}
for rec in D.get('detail_records', []):
    ch = rec.get('Chain', rec.get('chain', ''))
    fy = rec.get('FY', rec.get('fy', ''))
    nsv = float(rec.get('NSV', rec.get('nsv', 0)) or 0)
    if fy in ('FY26', 'fy26', 'FY26'):
        chain_fy26[ch] = chain_fy26.get(ch, 0) + nsv
    elif fy in ('FY27', 'fy27'):
        chain_fy27[ch] = chain_fy27.get(ch, 0) + nsv

# Pipeline ratios from offtake
offtake_by_chain = {}
for rec in D.get('offtake', {}).get('by_chain', []):
    if isinstance(rec, dict):
        offtake_by_chain[rec.get('name', '')] = rec.get('fy26', 0) or 0

# FY27 run-rate (annualise YTD — 4 months Apr–Jul)
MONTHS_YTD = 4
FULL_YEAR   = 12

all_chains = sorted(set(list(chain_fy26.keys()) + list(chain_fy27.keys())))

# Build rows
rows = []
for ch in all_chains:
    fy26_nsv = chain_fy26.get(ch, 0)
    fy27_ytd = chain_fy27.get(ch, 0)
    run_rate  = round(fy27_ytd * FULL_YEAR / MONTHS_YTD, 2) if MONTHS_YTD else 0
    # Target = FY26 × 1.20 (20% growth assumption — replace with actual JBP targets)
    tgt = round(fy26_nsv * 1.20, 2)
    ach_pct = round(fy27_ytd / tgt * 100, 1) if tgt else None
    offtake = offtake_by_chain.get(ch, 0)
    pipeline = round(fy26_nsv / offtake, 2) if offtake else None
    pip_flag = '' if pipeline is None else ('⚠️ HIGH' if pipeline > 1.3 else ('⚠️ LOW' if pipeline < 0.7 else '✅ OK'))
    status = ('✅ On-Track' if (ach_pct or 0) >= 80 else
              '⚠️ Monitor'  if (ach_pct or 0) >= 60 else
              '🔴 Escalate')
    rows.append({
        'chain': ch,
        'fy26': round(fy26_nsv, 2),
        'fy27_ytd': round(fy27_ytd, 2),
        'run_rate': run_rate,
        'tgt': tgt,
        'ach_pct': ach_pct,
        'gap': round(fy27_ytd - tgt, 2),
        'pipeline': pipeline,
        'pip_flag': pip_flag,
        'status': status,
    })

rows.sort(key=lambda x: x['fy26'], reverse=True)

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "RKAM TGT vs ACH"
ws.freeze_panes = 'B3'

# Title row
ws.merge_cells('A1:J1')
title_cell = ws.cell(row=1, column=1,
    value="Modern Trade — RKAM (Key Account) TGT vs ACH Scorecard | FY27 YTD (Apr–Jul'26)")
title_cell.font = Font(bold=True, color=WHITE, size=12, name='Calibri')
title_cell.fill = PatternFill("solid", fgColor=NAVY)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Sub-header
ws.merge_cells('A2:J2')
sub = ws.cell(row=2, column=1,
    value=f"Values in ₹ Lakh  |  TGT = FY26 × 1.20 (20% growth)  |  Pipeline Ratio = FY26 Primary ÷ FY26 Offtake  |  Generated {date.today()}")
sub.font = Font(italic=True, color=GRAY, size=8, name='Calibri')
sub.fill = PatternFill("solid", fgColor="F0F3FA")
sub.alignment = Alignment(horizontal="center")

# Column headers
COLS = ['Chain / RKAM', 'FY26 NSV (₹L)', 'FY27 YTD (₹L)', 'FY27 Run-Rate',
        'FY27 TGT (₹L)', 'ACH %', 'GAP vs TGT', 'Pipeline Ratio', 'Pip. Flag', 'Status']
hdr(ws, 3, len(COLS))
for i, h in enumerate(COLS, 1):
    ws.cell(row=3, column=i).value = h
ws.row_dimensions[3].height = 36

# Data rows
for ri, r in enumerate(rows, 4):
    is_alt = (ri % 2 == 0)
    if is_alt:
        shade(ws, ri, len(COLS))

    cell_set(ws, ri, 1, r['chain'], bold=(r['chain'] not in ('Unmapped Chain',)))
    cell_set(ws, ri, 2, r['fy26'], align="right", num_fmt='#,##0.00')
    cell_set(ws, ri, 3, r['fy27_ytd'], align="right", num_fmt='#,##0.00')
    cell_set(ws, ri, 4, r['run_rate'], align="right", num_fmt='#,##0.00')
    cell_set(ws, ri, 5, r['tgt'], align="right", num_fmt='#,##0.00')

    # ACH % with traffic-light colour
    ach_cell = cell_set(ws, ri, 6,
        f"{r['ach_pct']}%" if r['ach_pct'] is not None else '–', align="center", bold=True)
    if r['ach_pct'] is not None:
        col = GREEN if r['ach_pct'] >= 100 else (AMBER if r['ach_pct'] >= 80 else RED)
        ach_cell.font = Font(bold=True, color=col, name='Calibri', size=9)

    gap_cell = cell_set(ws, ri, 7, r['gap'], align="right", num_fmt='#,##0.00')
    if r['gap'] < 0:
        gap_cell.font = Font(color=RED, bold=True, name='Calibri', size=9)

    pip_val = f"{r['pipeline']:.2f}" if r['pipeline'] is not None else '–'
    cell_set(ws, ri, 8, pip_val, align="center")
    cell_set(ws, ri, 9, r['pip_flag'], align="center")

    status_cell = cell_set(ws, ri, 10, r['status'], align="center", bold=True)
    scol = GREEN if '✅' in r['status'] else (AMBER if '⚠️' in r['status'] else RED)
    status_cell.font = Font(bold=True, color=scol, name='Calibri', size=9)

# Totals row
tot_row = len(rows) + 4
ws.merge_cells(f'A{tot_row}:A{tot_row}')
cell_set(ws, tot_row, 1, 'TOTAL', bold=True)
shade(ws, tot_row, len(COLS), NAVY)
for c in range(1, len(COLS)+1):
    ws.cell(row=tot_row, column=c).font = Font(bold=True, color=WHITE, name='Calibri', size=9)

total_fy26 = sum(r['fy26'] for r in rows)
total_ytd  = sum(r['fy27_ytd'] for r in rows)
total_tgt  = sum(r['tgt'] for r in rows)
total_gap  = sum(r['gap'] for r in rows)
total_ach  = round(total_ytd / total_tgt * 100, 1) if total_tgt else None

ws.cell(row=tot_row, column=2).value = round(total_fy26, 2)
ws.cell(row=tot_row, column=3).value = round(total_ytd, 2)
ws.cell(row=tot_row, column=5).value = round(total_tgt, 2)
ws.cell(row=tot_row, column=6).value = f"{total_ach}%" if total_ach else '–'
ws.cell(row=tot_row, column=7).value = round(total_gap, 2)
for c in [2,3,4,5,7]:
    ws.cell(row=tot_row, column=c).number_format = '#,##0.00'
    ws.cell(row=tot_row, column=c).alignment = Alignment(horizontal="right")

# Column widths
widths = [32, 16, 16, 16, 16, 10, 14, 15, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Notes sheet ───────────────────────────────────────────────────────────────
wn = wb.create_sheet("Notes & Methodology")
notes = [
    ("Report", "RKAM TGT vs ACH Scorecard — MT Dashboard"),
    ("Data Source", "dashboard/data.js (window.DASH.detail_records + primary + offtake)"),
    ("FY27 YTD Period", "Apr 2026 – Jul 2026 (4 months)"),
    ("TGT Basis", "FY26 Actuals × 1.20 (20% growth target). Replace with actual JBP targets when available."),
    ("Run-Rate", "FY27 YTD × (12 / 4) = annualised projection"),
    ("Pipeline Ratio", "FY26 Primary NSV ÷ FY26 Secondary Offtake. Healthy: 0.7–1.3"),
    ("Status", "≥80% ACH = ✅ On-Track | 60–79% = ⚠️ Monitor | <60% = 🔴 Escalate"),
    ("Unmapped Chain", "11,344 records with no chain mapping (₹16,956L). Shown in table; fix requires distributor mapping update."),
    ("Negative NSV", "1,444 records with negative NSV (returns/credit notes) are included in actuals."),
    ("Generated By", f"scripts/generate_excel_rkam_tgt_vs_ach.py | {date.today()}"),
]
wn.column_dimensions['A'].width = 20
wn.column_dimensions['B'].width = 70
for ri, (k, v) in enumerate(notes, 1):
    kc = wn.cell(row=ri, column=1, value=k)
    kc.font = Font(bold=True, name='Calibri', size=9)
    vc = wn.cell(row=ri, column=2, value=v)
    vc.font = Font(name='Calibri', size=9)
    vc.alignment = Alignment(wrap_text=True)

# ── Save ──────────────────────────────────────────────────────────────────────
Path('exports').mkdir(exist_ok=True)
out = f"exports/TGT_vs_ACH_RKAM_{date.today().strftime('%Y%m%d')}.xlsx"
wb.save(out)
print(f"✅ Excel saved: {out}")
print(f"   Chains: {len(rows)}")
print(f"   FY26 Total NSV: ₹{total_fy26:,.2f} L")
print(f"   FY27 YTD NSV:   ₹{total_ytd:,.2f} L")
print(f"   Overall ACH%:   {total_ach}%")
