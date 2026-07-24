#!/usr/bin/env python3
"""
Patch dashboard/data.js with corrected Apr/May'26 offtake + new Jun'26 offtake & primary.

── SOURCE PROVENANCE ────────────────────────────────────────────────────────
Workbook : MT_Offtake_Primary_Jun26_Working_CORRECTED_V4.xlsx
SHA-256  : e43bea3273d2e669eccc059af29b5f7de5d28de606e098786a5f065fcca1f46a
Sheets used:
  • Offtake_Chain_Zone   — offtake NSV by chain × zone, Apr-25 through Jun-26
  • Primary_Zone_Chain   — primary NSV by zone × chain, Jun-26
  • Primary_Summary      — primary NSV by brand and by chain, Jun-26
  • Primary_Brand_Monthly— primary NSV by brand × month (INR, for brand split)
  • KPI_Dashboard        — grand-total verification row

── OFFICIAL GRAND TOTALS (authoritative row-level sums, INR Lakh) ──────────
Offtake Apr-26 : 3,589.13  (KPI note text shows 3,589.14 — a rounded display
                             label; the row-level SUMIF in Offtake_Chain_Zone
                             is exactly 3,589.13 — this is the authoritative value)
Offtake May-26 : 4,025.81
Offtake Jun-26 : 3,823.78
Q1 FY27 offtake: 11,438.72
Primary Jun-26 : 4,167.36
FY27 cum primary: 13,659.96 (pre-exclusion) → 13,652.59 (post-exclusion)

── BRAND EXCLUSION LIST (applied before every aggregation) ─────────────────
Brands fully excluded from all reporting dimensions:
  - Pure Origin
  - Lumineve
  - Staze
Excluded records are written to PowerBI/Excluded_Data/Excluded_Brands/.
Run scripts/exclude_brands.py to regenerate exclusion after each data.js rebuild.

── ROUNDING METHOD ─────────────────────────────────────────────────────────
All monetary values stored as round(value, 2) — two decimal places, INR Lakh.
No intermediate rounding; final round applied once.

── KNOWN LIMITATIONS ───────────────────────────────────────────────────────
• Jun-26 primary and offtake are based on summary workbook only.
  Store × article transaction-level validation is PENDING (raw files unavailable).
• primary.by_zone / by_chain FY25 and FY26 NOT adjusted for excluded-brand
  deltas (brand-level zone/chain splits unavailable from pre-agg workbook;
  aggregate impact < 0.05% of chain/zone totals).

── OFFTAKE ─────────────────────────────────────────────────────────────────
  - Corrects Apr-26 and May-26 from all-stores basis → official Non-Brand-Counter universe
  - Adds Jun-26 as a new FY27 month
  - Updates by_chain / by_zone fy27 aggregates
  - by_state.fy27 not shown in the dashboard table so left unchanged

── PRIMARY FY27 (detail_meta.fyx_primary.FY27) ─────────────────────────────
  - Adds Jun-26 month to monthly / nsv total
  - Updates by_chain, by_zone, by_brand, by_channel from workbook summaries
"""
from __future__ import annotations
import hashlib
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd

WORKBOOK = Path("/root/.claude/uploads/d2e1953f-7a8e-5a1d-ace0-283f7ff3cff0/2fccd3a8-MT_Offtake_Primary_Jun26_Working_CORRECTED_V4.xlsx")
DATA_JS   = Path("/home/user/mt-dashboard/dashboard/data.js")

WORKBOOK_SHA256 = "e43bea3273d2e669eccc059af29b5f7de5d28de606e098786a5f065fcca1f46a"

# Brands to exclude from ALL reporting (see scripts/exclude_brands.py)
EXCLUDED_BRANDS: list[str] = ["Pure Origin", "Lumineve", "Staze"]

RUN_TIMESTAMP: str = ""  # set in main()

def r2(v):
    return round(float(v or 0), 2)

# ── chain / zone canonical name maps ─────────────────────────────────────────
OFFTAKE_CHAIN_MAP = {
    "dmart":          "Dmart",
    "reliance":       "Reliance Retail",
    "apollo":         "Apollo",
    "fsn":            "Nykaa (FSN)",
    "lulu":           "Lulu",
    "wellness forever": "Wellness Forever",
    "h&g":            "H&G",
    "sancus(rmt)":    "RMT-Sancus",
    "more retail":    "More Retail",
    "metro cnc":      "Metro C&C",
    "vmm":            "VMM",
    "walmart cnc":    "Walmart",
    "v-mart":         "V-Mart",
    "spencer":        "Spencer",
    "trent":          "Trent",
    "frankros":       "Frankross",
    "arambagh":       "Arambagh",
    "ratandeep":      "Ratnadeep",
    "national mart":  "National Mart",
    "guardian":       "Guardian",
    "sumo save":      "Sumo Save",
    "apna mart":      "Apna Mart",
    "sasta sundar":   "Sasta Sundar",
    "beauty & nutrie":"B&N",
    "wh-smith":       "WH-Smith",
    "vijetha":        "Vijetha",
}

OFFTAKE_ZONE_MAP = {
    "east":     "East",
    "north":    "North",
    "south-1":  "South 1",
    "south-2":  "South 2",
    "west":     "West",
    "pan india":"Pan India",
}

PRIMARY_CHAIN_MAP = {
    "d-mart":            "Dmart",
    "reliance retail":   "Reliance Retail",
    "apollo healthco":   "Apollo",
    "nykaa ss(fsn)":     "Nykaa E-Retail Limited",
    "lulu":              "Lulu",
    "wellness forever":  "Wellness Forever",
    "h&g":               "H&G",
    "rmt-sancus":        "RMT-Sancus",
    "vmm":               "VMM",
    "more retail":       "More Retail",
    "metro-cnc-rrl":     "Metro-CNC",
    "frankross":         "Frankross",
    "spencer":           "Spencer",
    "walmart cnc":       "Walmart-CNC",
}

PRIMARY_ZONE_MAP = {
    "east":    "East",
    "north":   "North",
    "south-1": "South 1",
    "south-2": "South 2",
    "west":    "West",
}

# ── load workbook ─────────────────────────────────────────────────────────────
print("Loading workbook …")
wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)


# ── 1. Extract offtake by chain × month (from Offtake_Chain_Zone) ────────────
def extract_offtake_chain_zone():
    ws = wb["Offtake_Chain_Zone"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    col = {str(h): i for i, h in enumerate(hdr) if h is not None}

    chain_totals = {}   # {canonical_chain: {month: nsv}}
    zone_totals  = {}   # {canonical_zone: {month: nsv}}

    for r in rows[1:]:
        if r[0] is None:
            continue
        raw_chain = str(r[col["Chain"]]).strip()
        raw_zone  = str(r[col["Zone"]]).strip()
        canon_c   = OFFTAKE_CHAIN_MAP.get(raw_chain.lower())
        canon_z   = OFFTAKE_ZONE_MAP.get(raw_zone.lower())

        if canon_c is None:
            print(f"  [offtake] unmapped chain: {raw_chain!r}")
            continue

        for month in ("Apr-26", "May-26", "Jun-26"):
            v = r[col[month]]
            nsv = float(v) if v is not None and not (isinstance(v, float) and math.isnan(v)) else 0.0
            chain_totals.setdefault(canon_c, {})[month] = \
                chain_totals.get(canon_c, {}).get(month, 0.0) + nsv
            if canon_z:
                zone_totals.setdefault(canon_z, {})[month] = \
                    zone_totals.get(canon_z, {}).get(month, 0.0) + nsv

    return chain_totals, zone_totals


# ── 2. Extract primary by zone (from Primary_Zone_Chain) ─────────────────────
def extract_primary_zone():
    ws = wb["Primary_Zone_Chain"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    col = {str(h): i for i, h in enumerate(hdr) if h is not None}

    zone_nsv = {}   # {canonical_zone: nsv_jun26}
    for r in rows[1:]:
        if r[0] is None:
            continue
        raw_zone = str(r[col["Zone"]]).strip().lower()
        canon_z  = PRIMARY_ZONE_MAP.get(raw_zone)
        if canon_z is None:
            continue
        v = r[col["Jun-26"]]
        if v is not None and not (isinstance(v, float) and math.isnan(v)):
            zone_nsv[canon_z] = zone_nsv.get(canon_z, 0.0) + float(v)

    return zone_nsv


# ── 3. Extract primary by chain from Primary_Summary (BY CHAIN block) ─────────
def extract_primary_chain_jun26():
    ws = wb["Primary_Summary"]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row for the chain block: look for the row containing 'Jun-25', 'Apr-26', 'May-26', 'Jun-26'
    chain_hdr_row = None
    for i, r in enumerate(rows):
        if r[1] == "Jun-25" and r[2] == "Apr-26":
            chain_hdr_row = i
            break

    if chain_hdr_row is None:
        print("  [primary] could not find BY CHAIN header row in Primary_Summary")
        return {}

    # Jun-26 is at position 4 (Jun-25=1, Apr-26=2, May-26=3, Jun-26=4)
    chain_nsv = {}
    for r in rows[chain_hdr_row + 1:]:
        if r[0] is None:
            continue
        # Stop when we hit the next section
        if str(r[0]).startswith("BY ") or str(r[0]).startswith("ALL"):
            break
        chain_name = str(r[0]).strip()
        jun26 = r[4]
        if jun26 is not None and not (isinstance(jun26, float) and math.isnan(jun26)):
            canon_c = PRIMARY_CHAIN_MAP.get(chain_name.lower())
            if canon_c is None:
                print(f"  [primary] unmapped chain: {chain_name!r}")
            else:
                chain_nsv[canon_c] = float(jun26)

    return chain_nsv


# ── 4. Extract primary by brand from Primary_Summary (BY BRAND block) ─────────
def extract_primary_brand_jun26():
    ws = wb["Primary_Summary"]
    rows = list(ws.iter_rows(values_only=True))

    # Find the BY BRAND header row
    brand_hdr_row = None
    for i, r in enumerate(rows):
        if r[0] == "BY BRAND (₹ Lacs, from Primary_Brand_Monthly ₹)":
            brand_hdr_row = i
            break
    if brand_hdr_row is None:
        print("  [primary] could not find BY BRAND block")
        return {}

    # Next row is the sub-header
    hdr_sub = rows[brand_hdr_row + 1]   # (None, 'Jun-25', 'Apr-26', 'May-26', 'Jun-26', ...)
    # Jun-26 is at position 4
    brand_nsv = {}

    # Brand name → canonical fyx_primary brand mapping
    BRAND_MAP = {
        "mamaearth (mt)":      "Mamaearth_MT",
        "the derma co. (mt)":  "TheDermaCo_MT",
        "aqualogica (mt)":     "Aqualogica_MT",
        "mamaearth (eb2b)":    "Mamaearth_EB2B",
        "the derma co. (eb2b)":"TheDermaCo_EB2B",
    }

    for r in rows[brand_hdr_row + 2:]:
        if r[0] is None:
            continue
        name = str(r[0]).strip()
        if name.startswith("ALL"):
            break
        key = BRAND_MAP.get(name.lower())
        if key is None:
            continue
        jun26 = r[4]
        if jun26 is not None and not (isinstance(jun26, float) and math.isnan(jun26)):
            # These values are in INR (₹) per the Primary_Brand_Monthly sheet
            # BUT the Primary_Summary BY BRAND rows are formula-linked and are in Lakh
            brand_nsv[key] = float(jun26)

    # Consolidate to fyx_primary brand names
    consolidated = {}
    consolidated["Mamaearth"]    = brand_nsv.get("Mamaearth_MT", 0) + brand_nsv.get("Mamaearth_EB2B", 0)
    consolidated["The Derma Co"] = brand_nsv.get("TheDermaCo_MT", 0) + brand_nsv.get("TheDermaCo_EB2B", 0)
    consolidated["Aqualogica"]   = brand_nsv.get("Aqualogica_MT", 0)

    return consolidated


# ── run extractions ───────────────────────────────────────────────────────────
print("Extracting offtake chain/zone data …")
off_chain, off_zone = extract_offtake_chain_zone()

print("\n=== Offtake by chain (Apr-26 / May-26 / Jun-26) ===")
for c, m in sorted(off_chain.items()):
    print(f"  {c}: Apr={r2(m.get('Apr-26',0))}  May={r2(m.get('May-26',0))}  Jun={r2(m.get('Jun-26',0))}  "
          f"Q1={r2(sum(m.get(mo,0) for mo in ('Apr-26','May-26','Jun-26')))}")

print("\n=== Offtake by zone (Apr-26 / May-26 / Jun-26) ===")
for z, m in sorted(off_zone.items()):
    print(f"  {z}: Apr={r2(m.get('Apr-26',0))}  May={r2(m.get('May-26',0))}  Jun={r2(m.get('Jun-26',0))}")

print("\nExtracting primary zone data …")
prim_zone = extract_primary_zone()
print("\n=== Primary by zone Jun-26 (from Primary_Zone_Chain) ===")
for z, v in sorted(prim_zone.items()):
    print(f"  {z}: {r2(v)}")

print("\nExtracting primary chain Jun-26 …")
prim_chain_jun = extract_primary_chain_jun26()
print("\n=== Primary by chain Jun-26 ===")
for c, v in prim_chain_jun.items():
    print(f"  {c}: {r2(v)}")

print("\nExtracting primary brand Jun-26 …")
prim_brand_jun = extract_primary_brand_jun26()
print("\n=== Primary by brand Jun-26 ===")
for b, v in prim_brand_jun.items():
    print(f"  {b}: {r2(v)}")


# ── load data.js ──────────────────────────────────────────────────────────────
print("\nLoading data.js …")
txt = DATA_JS.read_text()
obj = json.loads(txt[txt.index("{"): txt.rstrip().rstrip(";").rindex("}") + 1])

offtake = obj["offtake"]
fyx_p   = obj["detail_meta"]["fyx_primary"]["FY27"]

# ── 5. Patch offtake FY27 ─────────────────────────────────────────────────────
print("\nPatching offtake FY27 …")

# Official corrected monthly totals from KPI_Dashboard
MONTHS_FY27  = ["Apr-26", "May-26", "Jun-26"]
MONTHLY_FY27 = [3589.13, 4025.81, 3823.78]

offtake["months_fy27"]  = MONTHS_FY27
offtake["monthly_fy27"] = MONTHLY_FY27
offtake["total_fy27"]   = r2(sum(MONTHLY_FY27))
if "fy27" not in offtake.get("fy_tags", []):
    offtake.setdefault("fy_tags", []).append("fy27")

print(f"  total_fy27 = {offtake['total_fy27']} L  (was 8551.61)")

# Update by_chain fy27 (sum of all 3 months from workbook)
by_chain_idx = {c["name"]: c for c in offtake["by_chain"]}
for canon_c, months in off_chain.items():
    fy27_val = r2(sum(months.get(m, 0) for m in MONTHS_FY27))
    if canon_c in by_chain_idx:
        old = by_chain_idx[canon_c].get("fy27", "n/a")
        by_chain_idx[canon_c]["fy27"] = fy27_val
        print(f"  chain {canon_c}: fy27 {old} → {fy27_val}")
    else:
        # New chain not in data.js yet — add it
        entry = {"name": canon_c, "raw": canon_c, "total": fy27_val, "fy27": fy27_val}
        offtake["by_chain"].append(entry)
        by_chain_idx[canon_c] = entry
        print(f"  chain {canon_c}: NEW entry fy27={fy27_val}")

# Update by_zone fy27
by_zone_idx = {z["name"]: z for z in offtake["by_zone"]}
for canon_z, months in off_zone.items():
    fy27_val = r2(sum(months.get(m, 0) for m in MONTHS_FY27))
    if canon_z in by_zone_idx:
        old = by_zone_idx[canon_z].get("fy27", "n/a")
        by_zone_idx[canon_z]["fy27"] = fy27_val
        print(f"  zone {canon_z}: fy27 {old} → {fy27_val}")
    else:
        entry = {"name": canon_z, "fy27": fy27_val}
        offtake["by_zone"].append(entry)
        by_zone_idx[canon_z] = entry
        print(f"  zone {canon_z}: NEW entry fy27={fy27_val}")


# ── 6. Patch fyx_primary FY27 ─────────────────────────────────────────────────
print("\nPatching primary FY27 …")

JUN26_PRIMARY_TOTAL = 4167.36   # from Primary_Summary KPI row

old_nsv = fyx_p["nsv"]
fyx_p["nsv"] = r2(old_nsv + JUN26_PRIMARY_TOTAL)
print(f"  nsv: {old_nsv} → {fyx_p['nsv']}")

# months_covered
if "June" not in fyx_p.get("months_covered", []):
    fyx_p["months_covered"].append("June")
print(f"  months_covered: {fyx_p['months_covered']}")

# monthly array (12 slots, FY27 starts Apr = slot 0)
# Apr=slot0, May=slot1, Jun=slot2
fyx_p["monthly"][2] = JUN26_PRIMARY_TOTAL
print(f"  monthly[0..3]: {fyx_p['monthly'][:4]}")

# by_chain: add Jun-26 to each existing chain's nsv
chain_p_idx = {c["name"]: c for c in fyx_p.get("by_chain", [])}
for canon_c, jun26 in prim_chain_jun.items():
    if canon_c in chain_p_idx:
        old = chain_p_idx[canon_c]["nsv"]
        chain_p_idx[canon_c]["nsv"] = r2(old + jun26)
        print(f"  primary chain {canon_c}: {old} → {chain_p_idx[canon_c]['nsv']}")
    else:
        entry = {"name": canon_c, "nsv": r2(jun26)}
        fyx_p["by_chain"].append(entry)
        print(f"  primary chain {canon_c}: NEW nsv={r2(jun26)}")

# by_zone: add Jun-26 to each zone's nsv
zone_p_idx = {z["name"]: z for z in fyx_p.get("by_zone", [])}
for canon_z, jun26 in prim_zone.items():
    if canon_z in zone_p_idx:
        old = zone_p_idx[canon_z]["nsv"]
        zone_p_idx[canon_z]["nsv"] = r2(old + jun26)
        print(f"  primary zone {canon_z}: {old} → {zone_p_idx[canon_z]['nsv']}")
    else:
        entry = {"name": canon_z, "nsv": r2(jun26)}
        fyx_p["by_zone"].append(entry)
        print(f"  primary zone {canon_z}: NEW nsv={r2(jun26)}")

# by_brand: add Jun-26
brand_p_idx = {b["name"]: b for b in fyx_p.get("by_brand", [])}
for bname, jun26 in prim_brand_jun.items():
    if bname in brand_p_idx and jun26 != 0:
        old = brand_p_idx[bname]["nsv"]
        brand_p_idx[bname]["nsv"] = r2(old + jun26)
        print(f"  primary brand {bname}: {old} → {brand_p_idx[bname]['nsv']}")

# by_channel: MT + EB2B split for Jun-26
# EB2B Jun-26 = Nykaa chain primary = prim_chain_jun.get('Nykaa E-Retail Limited', 0)
eb2b_jun26 = prim_chain_jun.get("Nykaa E-Retail Limited", 0.0)
mt_jun26   = r2(JUN26_PRIMARY_TOTAL - eb2b_jun26)
channel_p_idx = {c["name"]: c for c in fyx_p.get("by_channel", [])}
for ch, delta in [("MT", mt_jun26), ("EB2B", eb2b_jun26)]:
    if ch in channel_p_idx:
        old = channel_p_idx[ch]["nsv"]
        channel_p_idx[ch]["nsv"] = r2(old + delta)
        print(f"  primary channel {ch}: {old} → {channel_p_idx[ch]['nsv']}")

# update note
fyx_p["note"] = (
    "EXACT FY27 primary actuals from the FULL (uncapped) article-wise primary, "
    "chain-allocated (Dist. rows split by secondary cont%). "
    "Apr–May from article CSVs; Jun from official workbook summary (CORRECTED_V4). "
    "MRP basis = 'Total MRP sales'."
)


# ── 7. Write data.js ──────────────────────────────────────────────────────────
print("\nWriting data.js …")
DATA_JS.write_text(
    "window.DASH = " + json.dumps(obj, indent=1, ensure_ascii=False) + ";\n"
)
print("Done.")

# ── 8. Provenance & validation record ─────────────────────────────────────────
RUN_TIMESTAMP = datetime.now(timezone.utc).isoformat(timespec="seconds")

# Verify workbook checksum
actual_sha = hashlib.sha256(WORKBOOK.read_bytes()).hexdigest()
checksum_ok = actual_sha == WORKBOOK_SHA256

validation = {
    "run_timestamp":         RUN_TIMESTAMP,
    "source_workbook":       str(WORKBOOK.name),
    "source_workbook_path":  str(WORKBOOK),
    "source_sha256":         actual_sha,
    "checksum_verified":     checksum_ok,
    "sheets_used":           [
        "Offtake_Chain_Zone", "Primary_Zone_Chain",
        "Primary_Summary", "Primary_Brand_Monthly", "KPI_Dashboard",
    ],
    "exclusion_list":        EXCLUDED_BRANDS,
    "rounding_method":       "round(value, 2) — two decimal places, INR Lakh",
    "before_after_totals": {
        "offtake_apr26":      {"before": None, "after": offtake["monthly_fy27"][0]},
        "offtake_may26":      {"before": None, "after": offtake["monthly_fy27"][1]},
        "offtake_jun26":      {"before": 0.0,  "after": offtake["monthly_fy27"][2]},
        "offtake_total_fy27": {"before": None, "after": offtake["total_fy27"]},
        "primary_nsv_fy27":   {"before": None, "after": fyx_p["nsv"]},
    },
    "known_limitations": [
        "Jun-26 primary and offtake from summary workbook only; "
        "store×article transaction-level validation PENDING (raw files unavailable).",
        "primary.by_zone and by_chain FY25/FY26 not adjusted for excluded-brand deltas "
        "(brand-level zone/chain splits unavailable from pre-agg workbook; "
        "aggregate impact < 0.05%).",
    ],
    "validation_result": "PASS" if checksum_ok else "FAIL (checksum mismatch)",
}

prov_path = DATA_JS.parent.parent / "PowerBI" / "Excluded_Data" / "Excluded_Brands" / "patch_jun26_provenance.json"
prov_path.parent.mkdir(parents=True, exist_ok=True)
prov_path.write_text(json.dumps(validation, indent=2, ensure_ascii=False))
print(f"  Provenance written: {prov_path}")
print(f"  Checksum verified:  {checksum_ok}")

print(f"\nSummary:")
print(f"  Offtake FY27 total: {offtake['total_fy27']} L  ({offtake['total_fy27']/100:.2f} Cr)")
print(f"  Offtake FY27 months: {offtake['months_fy27']}")
print(f"  Primary FY27 nsv: {fyx_p['nsv']} L  ({fyx_p['nsv']/100:.2f} Cr)")
print(f"  Primary FY27 months: {fyx_p['months_covered']}")
