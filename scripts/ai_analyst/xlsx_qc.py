"""
Phase 7 — Workbook QC scanner (items 10 & 13).

Scans an .xlsx for the problems a leadership dashboard accumulates — formula
errors, broken (#REF!) references, hidden sheets, merged cells in backend data
sheets, duplicate keys, blank/unmapped mapping values, inconsistent date
formats, and anomalous / near-duplicate filter labels (e.g. a stray "METock" or
"West" vs "WEST") — and emits a QC_Check table with the exact columns requested:

  Check name | Tool/method | Tab | Source value | Dashboard value | Difference |
  Status (Pass/Fail/Pending) | Remarks | Action required

The reader uses only the standard library (zipfile + xml), so it runs offline
with no dependency. If openpyxl is available it can also write the QC_Check
sheet back into the workbook without disturbing the existing design; otherwise
the results render to CSV/HTML/Markdown via report.Report.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape as _xml_escape  # noqa: F401 (used by tests/builders)
from typing import Dict, List, Optional, Sequence, Tuple

from ai_analyst.report import Report

_ERROR_LITERALS = {"#REF!", "#DIV/0!", "#N/A", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _col_letters(ref: str) -> str:
    return "".join(ch for ch in ref if ch.isalpha())


@dataclass
class Sheet:
    name: str
    hidden: bool = False
    merged: List[str] = field(default_factory=list)
    errors: List[Tuple[str, str]] = field(default_factory=list)  # (cell ref, error value)
    headers: Dict[str, str] = field(default_factory=dict)         # col letter -> header text
    _rows: List[Dict[str, str]] = field(default_factory=list)     # col letter -> value (data rows)

    def column(self, header: str) -> List[str]:
        """Values under a header (case/space-insensitive match)."""
        want = _norm(header)
        col = next((c for c, h in self.headers.items() if _norm(h) == want), None)
        if col is None:
            return []
        return [r.get(col, "") for r in self._rows]


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


# --------------------------------------------------------------------------
# stdlib .xlsx reader
# --------------------------------------------------------------------------
def read_workbook(path: str) -> Dict[str, Sheet]:
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        shared = _read_shared_strings(z) if "xl/sharedStrings.xml" in names else []
        wb_sheets, rels = _read_workbook_index(z)
        result: Dict[str, Sheet] = {}
        for name, state, rid in wb_sheets:
            target = rels.get(rid)
            if not target:
                continue
            part = "xl/" + target if not target.startswith("xl/") else target
            if part not in names:
                part = target.lstrip("/")
            if part not in names:
                continue
            sheet = _read_sheet(z, part, name, shared)
            sheet.hidden = state in ("hidden", "veryHidden")
            result[name] = sheet
        return result


def _read_shared_strings(z: zipfile.ZipFile) -> List[str]:
    root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    out = []
    for si in root:
        # concatenate all <t> descendants (handles rich text runs)
        out.append("".join(t.text or "" for t in si.iter() if _local(t.tag) == "t"))
    return out


def _read_workbook_index(z: zipfile.ZipFile):
    root = ET.fromstring(z.read("xl/workbook.xml"))
    sheets = []
    for el in root.iter():
        if _local(el.tag) == "sheet":
            name = el.get("name", "")
            state = el.get("state", "visible")
            rid = next((v for k, v in el.attrib.items() if _local(k) == "id"), None)
            sheets.append((name, state, rid))
    rels = {}
    try:
        rroot = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        for r in rroot:
            rels[r.get("Id")] = r.get("Target")
    except KeyError:
        pass
    return sheets, rels


def _read_sheet(z: zipfile.ZipFile, part: str, name: str, shared: List[str]) -> Sheet:
    root = ET.fromstring(z.read(part))
    sheet = Sheet(name=name)
    rowvals: List[Tuple[int, Dict[str, str]]] = []
    for el in root.iter():
        tag = _local(el.tag)
        if tag == "mergeCell":
            ref = el.get("ref")
            if ref:
                sheet.merged.append(ref)
        elif tag == "row":
            ridx = int(el.get("r", "0") or 0)
            cells: Dict[str, str] = {}
            for c in el:
                if _local(c.tag) != "c":
                    continue
                ref = c.get("r", "")
                ctype = c.get("t", "")
                val = _cell_value(c, ctype, shared)
                if ctype == "e" or (isinstance(val, str) and val in _ERROR_LITERALS):
                    sheet.errors.append((ref, val))
                cells[_col_letters(ref)] = "" if val is None else str(val)
            rowvals.append((ridx, cells))
    rowvals.sort(key=lambda x: x[0])
    if rowvals:
        sheet.headers = dict(rowvals[0][1])
        sheet._rows = [cells for _, cells in rowvals[1:]]
    return sheet


def _cell_value(c, ctype: str, shared: List[str]):
    v_el = next((ch for ch in c if _local(ch.tag) == "v"), None)
    if ctype == "inlineStr":
        return "".join(t.text or "" for t in c.iter() if _local(t.tag) == "t")
    if ctype == "s":
        try:
            return shared[int(v_el.text)] if v_el is not None else ""
        except (ValueError, IndexError, TypeError):
            return ""
    return v_el.text if v_el is not None else ""


# --------------------------------------------------------------------------
# QC rows + checks
# --------------------------------------------------------------------------
QC_COLUMNS = ["Check name", "Tool/method", "Tab", "Source value", "Dashboard value",
              "Difference", "Status", "Remarks", "Action required"]

PASS, FAIL, PENDING, WARN = "Pass", "Fail", "Pending", "Warn"


@dataclass
class QCRow:
    check: str
    tool: str
    tab: str
    source: str = ""
    dashboard: str = ""
    difference: str = ""
    status: str = PASS
    remarks: str = ""
    action: str = ""

    def as_row(self) -> List[str]:
        return [self.check, self.tool, self.tab, self.source, self.dashboard,
                self.difference, self.status, self.remarks, self.action]


_DATA_SHEET_RE = re.compile(r"(raw|data|fact|source|backend)", re.I)


class WorkbookQC:
    def __init__(self, sheets: Dict[str, Sheet]):
        self.sheets = sheets

    @classmethod
    def open(cls, path: str) -> "WorkbookQC":
        return cls(read_workbook(path))

    # -- automatic checks --------------------------------------------------
    def check_errors(self) -> List[QCRow]:
        rows = []
        for s in self.sheets.values():
            for ref, val in s.errors:
                is_ref = "#REF!" in val
                rows.append(QCRow(
                    "Broken reference" if is_ref else "Formula/cell error",
                    "stdlib xlsx scan", s.name, ref, val, "", FAIL,
                    f"{val} at {s.name}!{ref}",
                    "Repair the reference/formula" if is_ref else "Fix upstream input/formula"))
        return rows

    def check_hidden_sheets(self) -> List[QCRow]:
        return [QCRow("Hidden sheet", "stdlib xlsx scan", s.name, "", "", "", PENDING,
                      "Sheet is hidden", "Confirm hidden state is intentional")
                for s in self.sheets.values() if s.hidden]

    def check_merged_in_data(self) -> List[QCRow]:
        rows = []
        for s in self.sheets.values():
            if s.merged and _DATA_SHEET_RE.search(s.name):
                rows.append(QCRow("Merged cells in data sheet", "stdlib xlsx scan", s.name,
                                  f"{len(s.merged)} merged range(s)", "", "", WARN,
                                  "Merged cells break Tables/Power Query/Power BI",
                                  "Unmerge in backend source sheets"))
        return rows

    # -- targeted checks ---------------------------------------------------
    def check_duplicate_keys(self, sheet: str, key_headers: Sequence[str]) -> List[QCRow]:
        s = self.sheets.get(sheet)
        if not s:
            return [self._missing(sheet)]
        cols = [s.column(h) for h in key_headers]
        if any(not c for c in cols):
            return [QCRow("Duplicate key", "stdlib xlsx scan", sheet,
                          ", ".join(key_headers), "", "", PENDING,
                          "One or more key columns not found", "Check key header names")]
        keys = ["|".join(vals) for vals in zip(*cols)]
        seen, dups = set(), 0
        for k in keys:
            if k in seen:
                dups += 1
            seen.add(k)
        return [QCRow("Duplicate key", "stdlib xlsx scan", sheet,
                      f"{len(keys)} rows", f"{len(seen)} unique", str(dups),
                      PASS if dups == 0 else FAIL,
                      f"key = {' + '.join(key_headers)}",
                      "" if dups == 0 else "Deduplicate or fix the key")]

    def check_blank_mapping(self, sheet: str, headers: Sequence[str]) -> List[QCRow]:
        s = self.sheets.get(sheet)
        if not s:
            return [self._missing(sheet)]
        rows = []
        for h in headers:
            vals = s.column(h)
            if not vals:
                rows.append(QCRow("Missing mapping", "stdlib xlsx scan", sheet, h, "", "",
                                  PENDING, "Column not found", "Check header name"))
                continue
            blanks = sum(1 for v in vals if _norm(v) == "")
            rows.append(QCRow("Missing mapping", "stdlib xlsx scan", sheet, h,
                              f"{blanks} blank", str(blanks), PASS if blanks == 0 else WARN,
                              f"{blanks}/{len(vals)} rows blank in '{h}'",
                              "" if blanks == 0 else "Fill mapping or flag as unmapped"))
        return rows

    def check_date_consistency(self, sheet: str, header: str) -> List[QCRow]:
        s = self.sheets.get(sheet)
        if not s:
            return [self._missing(sheet)]
        vals = [v for v in s.column(header) if _norm(v) != ""]
        if not vals:
            return [QCRow("Date format consistency", "stdlib xlsx scan", sheet, header, "", "",
                          PENDING, "Column empty/not found", "Check header name")]
        numeric = sum(1 for v in vals if _is_number(v))
        textual = len(vals) - numeric
        mixed = numeric > 0 and textual > 0
        return [QCRow("Date format consistency", "stdlib xlsx scan", sheet, header,
                      f"{numeric} serial / {textual} text", str(min(numeric, textual)),
                      WARN if mixed else PASS,
                      "Mixed date serials and text" if mixed else "Consistent",
                      "Standardise to one date type (Power Query)" if mixed else "")]

    def check_filter_labels(self, sheet: str, header: str,
                            allowed: Optional[Sequence[str]] = None) -> List[QCRow]:
        s = self.sheets.get(sheet)
        if not s:
            return [self._missing(sheet)]
        vals = [v for v in s.column(header) if _norm(v) != ""]
        if not vals:
            return [QCRow("Filter label", "stdlib xlsx scan", sheet, header, "", "",
                          PENDING, "Column empty/not found", "Check header name")]
        rows: List[QCRow] = []
        # near-duplicate labels (case/whitespace variants of the same thing)
        groups: Dict[str, set] = {}
        for v in vals:
            groups.setdefault(_norm(v), set()).add(v)
        for canon, variants in groups.items():
            if len(variants) > 1:
                rows.append(QCRow("Inconsistent filter label", "stdlib xlsx scan", sheet,
                                  " / ".join(sorted(variants)), canon, str(len(variants)),
                                  WARN, "Same value stored multiple ways",
                                  "Standardise the label (Clean helper column)"))
        # values outside an allowed set (this is the 'METock' catcher)
        if allowed:
            ok = {_norm(a) for a in allowed}
            bad = sorted({v for v in vals if _norm(v) not in ok})
            for v in bad:
                rows.append(QCRow("Unexpected filter value", "stdlib xlsx scan", sheet,
                                  v, ", ".join(allowed), "", FAIL,
                                  f"'{v}' is not an allowed {header} value",
                                  "Fix/relabel or add to mapping master"))
        if not rows:
            rows.append(QCRow("Filter label", "stdlib xlsx scan", sheet, header,
                              f"{len(groups)} clean values", "", PASS, "No anomalies", ""))
        return rows

    def _missing(self, sheet: str) -> QCRow:
        return QCRow("Sheet present", "stdlib xlsx scan", sheet, "", "", "", FAIL,
                     "Sheet not found in workbook", "Check the sheet name")

    # -- orchestration -----------------------------------------------------
    def run(self, duplicate_keys: Optional[Dict[str, Sequence[str]]] = None,
            mappings: Optional[Dict[str, Sequence[str]]] = None,
            dates: Optional[Dict[str, str]] = None,
            filters: Optional[Dict[str, Dict[str, Optional[Sequence[str]]]]] = None
            ) -> List[QCRow]:
        rows: List[QCRow] = []
        rows += self.check_errors()
        rows += self.check_hidden_sheets()
        rows += self.check_merged_in_data()
        for sheet, keys in (duplicate_keys or {}).items():
            rows += self.check_duplicate_keys(sheet, keys)
        for sheet, cols in (mappings or {}).items():
            rows += self.check_blank_mapping(sheet, cols)
        for sheet, col in (dates or {}).items():
            rows += self.check_date_consistency(sheet, col)
        for sheet, spec in (filters or {}).items():
            for col, allowed in spec.items():
                rows += self.check_filter_labels(sheet, col, allowed)
        return rows


def qc_report(rows: List[QCRow], title: str = "QC_Check",
              classification: str = "Confidential - MT Internal") -> Report:
    rep = Report(title, subtitle="Workbook QC summary", classification=classification)
    counts = {}
    for r in rows:
        counts[r.status] = counts.get(r.status, 0) + 1
    rep.add_kpis([(k, str(v)) for k, v in sorted(counts.items())], title="QC status summary")
    rep.add_table(QC_COLUMNS, [r.as_row() for r in rows], title="QC_Check",
                  note="One row per check. Status: Pass / Fail / Warn / Pending.")
    return rep


def write_qc_sheet(path: str, rows: List[QCRow], sheet_name: str = "QC_Check") -> str:
    """Append/replace a QC_Check sheet inside the workbook (needs openpyxl)."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise RuntimeError("Writing the QC_Check sheet into the workbook needs openpyxl "
                           "(`pip install openpyxl`). Without it, export via qc_report().save().")
    wb = openpyxl.load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(QC_COLUMNS)
    for r in rows:
        ws.append(r.as_row())
    wb.save(path)
    return path


def _is_number(s) -> bool:
    try:
        float(s)
        return True
    except (TypeError, ValueError):
        return False
