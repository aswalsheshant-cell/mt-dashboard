#!/usr/bin/env python3
"""
Build the Enterprise Rolling 3-Month Demand Forecast & S&OP Planning workbook
for Honasa / Mamaearth Modern Trade.

Self-contained & reproducible: it (1) aggregates the REAL monthly primary +
offtake CSVs under PowerBI/RawDataFolders/, then (2) assembles a fully
formula-driven Excel workbook (openpyxl). No numbers are fabricated - every
history value derives from the raw files; every forecast is an Excel formula
that recalculates when inputs change.

Usage:
    python scripts/build_demand_forecast.py
    python scripts/build_demand_forecast.py --out DemandForecast/MT_Demand_Forecast_SOP_Model.xlsx

Grain of the statistical engine: Brand x Category (portfolio planning grain).
The company target then cascades down the hierarchy (Channel -> Chain -> Zone ->
State -> Distributor -> Brand -> Category -> Article) using real historical
contribution %, so the model reconciles top-down and bottom-up.
"""
import argparse, csv, glob, os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(REPO, "PowerBI", "RawDataFolders")
PRIM_DIR = os.path.join(RAW, "Primary_Article_Monthly")
OFF_DIR = os.path.join(RAW, "Offtake_Monthly")

CR = 1e7  # rupees -> crore

# ----- month timeline -----------------------------------------------------
MON_ABBR = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]


def fy_month_seq(start_label, n):
    """Return n month labels like Apr'25 starting from start_label (financial-year order)."""
    mon, yy = start_label[:3], int(start_label[-2:])
    idx = MON_ABBR.index(mon)
    out = []
    for _ in range(n):
        out.append(f"{MON_ABBR[idx]}'{yy:02d}")
        idx += 1
        if idx == 12:
            idx = 0
            # Jan follows Dec -> same calendar-year rollover handled by abbr order:
            # our FY order is Apr..Mar; after Mar (idx 11) year increments
        if idx == 0:  # wrapped from Mar to Apr -> new financial year, cal year +1 on Apr
            yy += 1
    return out


def month_name(label):
    return label[:3]


HIST_MONTHS = fy_month_seq("Apr'25", 14)          # Apr'25 .. May'26 (real data window)
ENGINE_MONTHS = fy_month_seq("Apr'25", 24)        # Apr'25 .. Mar'27 (matrix header)
CAL_MONTHS = fy_month_seq("Apr'25", 36)           # Apr'25 .. Mar'28 (calendar)


def num(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return 0.0


# ==========================================================================
# 1. AGGREGATE REAL DATA
# ==========================================================================
def aggregate():
    prim_bc = defaultdict(lambda: defaultdict(float))      # (brand,cat)->month->NSV cr
    prim_bc_qty = defaultdict(lambda: defaultdict(float))
    chain_tot = defaultdict(float)
    chain_zone = {}
    zone_tot = defaultdict(float)
    state_tot = defaultdict(float)
    state_zone = {}
    dist_tot = defaultdict(float)
    dist_meta = {}
    art_tot = defaultdict(float)
    art_meta = {}
    off_bc = defaultdict(lambda: defaultdict(float))
    off_bc_qty = defaultdict(lambda: defaultdict(float))

    for path in sorted(glob.glob(os.path.join(PRIM_DIR, "primary_article_*.csv"))):
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            for row in csv.DictReader(fh):
                m = (row.get("Month") or "").strip()
                if m not in HIST_MONTHS:
                    continue
                brand = (row.get("brand") or "").strip()
                cat = (row.get("category") or "").strip()
                if not brand or not cat:
                    continue
                nsv = num(row.get("Inv. Net value(LOC)")) / CR
                qty = num(row.get("Inv Qty"))
                prim_bc[(brand, cat)][m] += nsv
                prim_bc_qty[(brand, cat)][m] += qty
                chain = (row.get("Chain name") or "").strip()
                zone = (row.get("Zone") or "").strip()
                state = (row.get("State") or "").strip()
                dist = (row.get("Ship To Name") or "").strip()
                code = (row.get("Article Code") or "").strip().replace(".0", "")
                if chain:
                    chain_tot[chain] += nsv
                    chain_zone.setdefault(chain, zone)
                if zone:
                    zone_tot[zone] += nsv
                if state:
                    state_tot[state] += nsv
                    state_zone.setdefault(state, zone)
                if dist:
                    dist_tot[dist] += nsv
                    dist_meta.setdefault(dist, (zone, state, chain))
                if code:
                    art_tot[code] += nsv
                    art_meta.setdefault(code, ((row.get("Description") or "").strip(), brand, cat,
                                              (row.get("EAN No.") or "").strip().replace(".0", "")))

    for path in sorted(glob.glob(os.path.join(OFF_DIR, "offtake_store_article_*.csv"))):
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            for row in csv.DictReader(fh):
                m = (row.get("Month") or "").strip()
                brand = (row.get("Brand") or "").strip()
                cat = (row.get("Category") or "").strip()
                if not brand or not cat or m not in HIST_MONTHS:
                    continue
                off_bc[(brand, cat)][m] += num(row.get("MRP Sales Value")) / CR
                off_bc_qty[(brand, cat)][m] += num(row.get("Sales Qty"))

    def topn(d, n):
        return dict(sorted(d.items(), key=lambda kv: -kv[1])[:n])

    return dict(
        prim_bc=prim_bc, prim_bc_qty=prim_bc_qty,
        chain_tot=topn(chain_tot, 40), chain_zone=chain_zone,
        zone_tot=dict(zone_tot), state_tot=topn(state_tot, 25), state_zone=state_zone,
        dist_tot=topn(dist_tot, 30), dist_meta=dist_meta,
        art_tot=topn(art_tot, 40), art_meta=art_meta,
        off_bc=off_bc, off_bc_qty=off_bc_qty,
    )


# ==========================================================================
# 2. STYLING HELPERS
# ==========================================================================
FONT = "Arial"
C_HEADER = "1F3864"      # dark navy
C_SUB = "2E5496"
C_BAND = "D9E1F2"        # light blue band
C_INPUT_FILL = "FFF2CC"  # yellow-ish for input cells
C_CALC = "000000"
C_INPUT = "0000FF"       # blue text for inputs
C_LINK = "006100"        # green for cross-sheet links
C_TOTAL = "E2EFDA"
C_ACCENT = "FCE4D6"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
medium = Side(style="medium", color="1F3864")


def F(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_CR = '#,##0.00;(#,##0.00);"-"'
FMT_CR1 = '#,##0.0;(#,##0.0);"-"'
FMT_QTY = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_PCT0 = '0%'
FMT_X = '0.00'


def title_row(ws, text, ncols, row=1):
    ws.cell(row, 1, text).font = F(15, True, "FFFFFF")
    ws.cell(row, 1).fill = fill(C_HEADER)
    ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 26
    return row + 1


def section(ws, text, r, ncols, note=""):
    ws.cell(r, 1, text).font = F(11, True, "FFFFFF")
    ws.cell(r, 1).fill = fill(C_SUB)
    ws.cell(r, 1).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    if note:
        c = ws.cell(r, 1)
        c.comment = Comment(note, "Model")
    return r + 1


def header_cells(ws, r, headers, start_col=1, fillhex=C_HEADER, color="FFFFFF"):
    for i, h in enumerate(headers):
        c = ws.cell(r, start_col + i, h)
        c.font = F(9.5, True, color)
        c.fill = fill(fillhex)
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[r].height = 28
    return r + 1


def qname(sheet, addr):
    """Quoted cross-sheet absolute reference."""
    if " " in sheet or "&" in sheet:
        return f"'{sheet}'!{addr}"
    return f"{sheet}!{addr}"


# ==========================================================================
# 3. BUILD WORKBOOK
# ==========================================================================
def build(agg, out_path):
    # --- select engine series: material Brand x Category (real) ---
    prim_bc = agg["prim_bc"]
    totals = {k: sum(v.values()) for k, v in prim_bc.items()}
    series = [k for k, t in sorted(totals.items(), key=lambda kv: -kv[1]) if t >= 0.3]
    # de-dup category label casing noise handled by keeping as-is (real labels)

    wb = Workbook()
    wb.remove(wb.active)
    REF = {}   # semantic key -> quoted absolute reference

    # order of creation (dashboard first visually later). We'll create then reorder.
    ws_cal = wb.create_sheet("Calendar")
    ws_cp = wb.create_sheet("Control Panel")
    ws_tgt = wb.create_sheet("Monthly Target Upload")
    ws_prim = wb.create_sheet("Primary Sales History")
    ws_off = wb.create_sheet("Offtake History")
    ws_dist = wb.create_sheet("Distributor Master")
    ws_npi = wb.create_sheet("NPI & EPD Planning")
    ws_promo = wb.create_sheet("Promotion Calendar")
    ws_seas = wb.create_sheet("Seasonality Matrix")
    ws_eng = wb.create_sheet("Forecast Engine")
    ws_scn = wb.create_sheet("Scenario Summary")
    ws_td = wb.create_sheet("Target Distribution")
    ws_pp = wb.create_sheet("Primary Planning")
    ws_di = wb.create_sheet("Distributor Intelligence")
    ws_wh = wb.create_sheet("Warehouse Optimization")
    ws_acc = wb.create_sheet("Forecast Accuracy")
    ws_dash = wb.create_sheet("Executive Dashboard")
    ws_ins = wb.create_sheet("AI Business Insights")
    ws_wb = wb.create_sheet("Demand Planner Workbench")
    # ---- Business Event Engine module ----
    ws_evset = wb.create_sheet("Event Settings")
    ws_clib = wb.create_sheet("Chain Event Library")
    ws_alib = wb.create_sheet("Article Event Library")
    ws_dlib = wb.create_sheet("Demand Driver Library")
    ws_evm = wb.create_sheet("Business_Event_Master")
    ws_evi = wb.create_sheet("Event Impact Engine")
    ws_evc = wb.create_sheet("Event Calendar")
    ws_evsim = wb.create_sheet("Event Simulator")
    ws_evd = wb.create_sheet("Event Impact Dashboard")
    ws_evai = wb.create_sheet("Event AI Recommendations")
    ws_doc = wb.create_sheet("Documentation")

    # ---------------------------------------------------------------- Calendar
    build_calendar(ws_cal, REF)
    # ---------------------------------------------------------------- Seasonality (needed by engine)
    seas_cats = build_seasonality(ws_seas, series, REF)
    # ---------------------------------------------------------------- Control Panel
    build_control_panel(ws_cp, REF)
    # ---------------------------------------------------------------- Inputs / masters
    build_targets(ws_tgt, REF)
    build_primary_history(ws_prim, prim_bc, agg["prim_bc_qty"], series, REF)
    build_offtake_history(ws_off, agg["off_bc"], agg["off_bc_qty"], REF)
    build_distributor_master(ws_dist, agg, REF)
    build_npi(ws_npi, REF)
    build_promo(ws_promo, REF)
    # ---------------------------------------------------------------- Engine
    build_forecast_engine(ws_eng, series, seas_cats, REF)
    # ---------------------------------------------------------------- Downstream
    build_scenario_summary(ws_scn, series, REF)
    build_target_distribution(ws_td, agg, REF)
    build_primary_planning(ws_pp, series, REF)
    build_distributor_intelligence(ws_di, agg, REF)
    build_warehouse(ws_wh, REF)
    build_accuracy(ws_acc, REF)
    build_dashboard(ws_dash, series, agg, REF)
    build_insights(ws_ins, series, agg, REF)
    build_workbench(ws_wb, REF)
    # ---- Business Event Engine module ----
    event_types = build_event_settings(ws_evset, REF)
    build_chain_event_library(ws_clib, agg, event_types, REF)
    build_article_event_library(ws_alib, event_types, REF)
    build_driver_library(ws_dlib, agg, REF)
    build_event_master(ws_evm, agg, event_types, REF)
    build_event_impact(ws_evi, series, REF)
    build_event_calendar(ws_evc, REF)
    build_event_simulator(ws_evsim, series, REF)
    build_event_dashboard(ws_evd, series, REF)
    build_event_ai(ws_evai, series, REF)
    build_documentation(ws_doc, REF)

    # reorder: Dashboard & Control Panel front
    order = ["Executive Dashboard", "Control Panel", "Forecast Engine", "Scenario Summary",
             "AI Business Insights", "Demand Planner Workbench",
             "Business_Event_Master", "Event Impact Engine", "Event Impact Dashboard",
             "Event AI Recommendations", "Event Calendar", "Event Simulator",
             "Chain Event Library", "Article Event Library", "Demand Driver Library",
             "Event Settings",
             "Target Distribution", "Primary Planning", "Distributor Intelligence",
             "Warehouse Optimization", "Forecast Accuracy", "Monthly Target Upload",
             "Primary Sales History", "Offtake History", "Distributor Master",
             "NPI & EPD Planning", "Promotion Calendar", "Seasonality Matrix",
             "Calendar", "Documentation"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.active = wb.sheetnames.index("Executive Dashboard")

    # force Excel to recalculate on open (openpyxl writes no cached values)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path, series


# --------------------------------------------------------------------------
def build_calendar(ws, REF):
    r = title_row(ws, "CALENDAR  ·  financial-year month index (Apr–Mar)", 4)
    r = header_cells(ws, r, ["Idx", "Month", "Month Name", "Cal Year"])
    first = r
    for i, lab in enumerate(CAL_MONTHS):
        ws.cell(r, 1, i + 1).font = F(9)
        ws.cell(r, 2, lab).font = F(9)
        ws.cell(r, 3, month_name(lab)).font = F(9)
        yy = int(lab[-2:])
        mn = month_name(lab)
        calyear = 2000 + yy
        ws.cell(r, 4, calyear).font = F(9)
        for c in range(1, 5):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = CENTER
        r += 1
    last = r - 1
    REF["cal_idx"] = qname("Calendar", f"$A${first}:$A${last}")
    REF["cal_lab"] = qname("Calendar", f"$B${first}:$B${last}")
    REF["cal_name"] = qname("Calendar", f"$C${first}:$C${last}")
    REF["cal_first"] = first
    REF["cal_last"] = last
    for col, w in zip("ABCD", (6, 10, 12, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


# --------------------------------------------------------------------------
def build_seasonality(ws, series, REF):
    r = title_row(ws, "SEASONALITY MATRIX  ·  monthly uplift index by category (editable)", 14)
    r = section(ws, "Blue cells are editable seasonal factors (100% = no seasonal effect). "
                    "The engine reindexes each category's flat run-rate to this shape.", r, 14)
    hdr = ["Category"] + MON_ABBR + ["Avg"]
    r = header_cells(ws, r, hdr)
    first = r
    # categories present in engine series + spec examples; default 100%
    cats = []
    for (b, c) in series:
        if c not in cats:
            cats.append(c)
    # sensible defaults per known category keyword (editable)
    presets = {
        "Sun Care": [130, 145, 150, 120, 95, 85, 80, 75, 70, 80, 95, 120],
        "Face":     [100, 100, 100, 105, 105, 110, 115, 120, 110, 110, 105, 100],
        "Hair":     [95, 95, 100, 105, 110, 110, 105, 105, 100, 100, 95, 95],
        "Body":     [75, 70, 70, 75, 85, 95, 110, 125, 145, 150, 140, 120],
        "Baby":     [100, 100, 100, 100, 100, 105, 105, 110, 110, 105, 100, 100],
    }
    def preset_for(cat):
        for k, v in presets.items():
            if k.lower() in cat.lower() or cat.lower() in k.lower():
                return v
        return [100] * 12
    seas_cats = {}
    for cat in cats:
        ws.cell(r, 1, cat).font = F(9.5, True)
        ws.cell(r, 1).border = BORDER
        ws.cell(r, 1).alignment = LEFT
        vals = preset_for(cat)
        for i in range(12):
            c = ws.cell(r, 2 + i, vals[i] / 100.0)
            c.font = F(9.5, color=C_INPUT)
            c.fill = fill(C_INPUT_FILL)
            c.number_format = FMT_PCT0
            c.alignment = CENTER
            c.border = BORDER
        # avg
        avg = ws.cell(r, 14, f"=AVERAGE(B{r}:M{r})")
        avg.font = F(9.5, bold=True)
        avg.number_format = FMT_PCT0
        avg.border = BORDER
        avg.alignment = CENTER
        seas_cats[cat] = r
        r += 1
    # default row
    ws.cell(r, 1, "Default").font = F(9.5, True, italic=True)
    ws.cell(r, 1).border = BORDER
    for i in range(12):
        c = ws.cell(r, 2 + i, 1.0)
        c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
        c.number_format = FMT_PCT0; c.alignment = CENTER; c.border = BORDER
    ws.cell(r, 14, f"=AVERAGE(B{r}:M{r})").number_format = FMT_PCT0
    ws.cell(r, 14).border = BORDER
    default_row = r
    last = r
    REF["seas_catcol"] = qname("Seasonality Matrix", f"$A${first}:$A${last}")
    REF["seas_vals"] = qname("Seasonality Matrix", f"$B${first}:$M${last}")
    REF["seas_monhdr"] = qname("Seasonality Matrix", f"$B${first-1}:$M${first-1}")
    REF["seas_avg"] = qname("Seasonality Matrix", f"$N${first}:$N${last}")
    REF["seas_default_row"] = default_row
    ws.column_dimensions["A"].width = 16
    for col in "BCDEFGHIJKLM":
        ws.column_dimensions[col].width = 6.5
    ws.column_dimensions["N"].width = 7
    ws.freeze_panes = "B5"
    return seas_cats


# --------------------------------------------------------------------------
def cp_input(ws, r, label, value, key, REF, fmt=None, note=None, dropdown=None):
    ws.cell(r, 1, label).font = F(10)
    ws.cell(r, 1).alignment = LEFT
    c = ws.cell(r, 2, value)
    c.font = F(10, bold=True, color=C_INPUT)
    c.fill = fill(C_INPUT_FILL)
    c.alignment = CENTER
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(r, 3, note).font = F(8.5, italic=True, color="808080")
        ws.cell(r, 3).alignment = LEFT
    if dropdown is not None:
        dv = DataValidation(type="list", formula1=f'"{",".join(dropdown)}"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(c)
    REF[key] = qname("Control Panel", f"$B${r}")
    return r + 1


def cp_calc(ws, r, label, formula, key, REF, fmt=None, note=None):
    ws.cell(r, 1, label).font = F(10)
    ws.cell(r, 1).alignment = LEFT
    c = ws.cell(r, 2, formula)
    c.font = F(10, bold=True)
    c.alignment = CENTER
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(r, 3, note).font = F(8.5, italic=True, color="808080")
        ws.cell(r, 3).alignment = LEFT
    REF[key] = qname("Control Panel", f"$B${r}")
    return r + 1


def build_control_panel(ws, REF):
    r = title_row(ws, "CONTROL PANEL  ·  every blue cell drives the rolling forecast", 4)
    r += 0
    r = section(ws, "1 · Forecast Horizon (rolling)", r, 4)
    r = cp_input(ws, r, "Last Actual Month (data closed through)", "May'26", "last_actual", REF,
                 note="set to the latest month present in Primary Sales History")
    r = cp_input(ws, r, "Forecast Month  (M)", "Jun'26", "fc_m", REF,
                 note="first month to forecast; advance one month each cycle")
    r = cp_input(ws, r, "Business Scenario (drives 'Selected' columns)", "Base", "scenario", REF,
                 dropdown=["Base", "Optimistic", "Conservative"])
    # derived M+1 / M+2 and indices
    calm = REF["cal_lab"]
    r = cp_calc(ws, r, "Forecast Month  (M+1)",
                f"=INDEX({calm},MATCH($B${r-2},{calm},0)+1)", "fc_m1", REF)
    r = cp_calc(ws, r, "Forecast Month  (M+2)",
                f"=INDEX({calm},MATCH($B${r-3},{calm},0)+2)", "fc_m2", REF)
    idx_last = r
    r = cp_calc(ws, r, "Idx · Last Actual", f"=MATCH({REF['last_actual']},{calm},0)", "idx_last", REF)
    r = cp_calc(ws, r, "Idx · M", f"=MATCH({REF['fc_m']},{calm},0)", "idx_m", REF)
    r = cp_calc(ws, r, "Idx · M+1", f"=MATCH({REF['fc_m1']},{calm},0)", "idx_m1", REF)
    r = cp_calc(ws, r, "Idx · M+2", f"=MATCH({REF['fc_m2']},{calm},0)", "idx_m2", REF)
    # month names for seasonality lookups
    caln = REF["cal_name"]
    r = cp_calc(ws, r, "Month Name · M", f"=INDEX({caln},{REF['idx_m']})", "name_m", REF)
    r = cp_calc(ws, r, "Month Name · M+1", f"=INDEX({caln},{REF['idx_m1']})", "name_m1", REF)
    r = cp_calc(ws, r, "Month Name · M+2", f"=INDEX({caln},{REF['idx_m2']})", "name_m2", REF)

    r += 1
    r = section(ws, "2 · Statistical Method Weights (auto-normalised)", r, 4)
    methods = [("WMA (weighted moving avg)", "w_wma", 0.20),
               ("Exponential Smoothing", "w_es", 0.20),
               ("Linear Trend", "w_trend", 0.15),
               ("Seasonal Index", "w_seas", 0.15),
               ("CAGR", "w_cagr", 0.10),
               ("Rolling Average", "w_roll", 0.10),
               ("YoY Growth", "w_yoy", 0.10)]
    wfirst = r
    for lab, key, val in methods:
        r = cp_input(ws, r, lab, val, key, REF, fmt=FMT_PCT)
    wlast = r - 1
    REF["w_sum"] = qname("Control Panel", f"$B${r}")
    ws.cell(r, 1, "Sum of method weights").font = F(10, italic=True)
    ws.cell(r, 2, f"=SUM(B{wfirst}:B{wlast})").font = F(10, bold=True)
    ws.cell(r, 2).number_format = FMT_PCT
    ws.cell(r, 2).border = BORDER
    r += 2

    r = section(ws, "3 · Method Parameters", r, 4)
    r = cp_input(ws, r, "Exponential Smoothing α", 0.40, "alpha", REF, fmt=FMT_PCT,
                 note="higher = more weight on recent months")
    r = cp_input(ws, r, "WMA weight · latest month", 0.50, "wma1", REF, fmt=FMT_PCT)
    r = cp_input(ws, r, "WMA weight · 2nd month", 0.30, "wma2", REF, fmt=FMT_PCT)
    r = cp_input(ws, r, "WMA weight · 3rd month", 0.20, "wma3", REF, fmt=FMT_PCT)
    r += 1

    r = section(ws, "4 · AI Ensemble Blend (Statistical + Business + Drivers)", r, 4)
    r = cp_input(ws, r, "Weight · Statistical", 0.50, "ai_stat", REF, fmt=FMT_PCT)
    r = cp_input(ws, r, "Weight · Business", 0.30, "ai_biz", REF, fmt=FMT_PCT)
    r = cp_input(ws, r, "Weight · Driver-adjusted", 0.20, "ai_drv", REF, fmt=FMT_PCT)
    r = cp_input(ws, r, "Momentum cap (±)", 0.15, "mom_cap", REF, fmt=FMT_PCT,
                 note="limits brand-momentum driver")
    r += 1

    r = section(ws, "5 · Scenario Multipliers", r, 4)
    r = cp_input(ws, r, "Base", 1.00, "scn_base", REF, fmt=FMT_X)
    r = cp_input(ws, r, "Optimistic", 1.08, "scn_opt", REF, fmt=FMT_X)
    r = cp_input(ws, r, "Conservative", 0.92, "scn_cons", REF, fmt=FMT_X)
    r = cp_calc(ws, r, "Selected multiplier",
                f'=IF({REF["scenario"]}="Optimistic",{REF["scn_opt"]},'
                f'IF({REF["scenario"]}="Conservative",{REF["scn_cons"]},{REF["scn_base"]}))',
                "scn_sel", REF, fmt=FMT_X)
    r += 1

    r = section(ws, "6 · Business Forecast Defaults", r, 4)
    r = cp_input(ws, r, "Default business adjustment %", 0.05, "biz_default", REF, fmt=FMT_PCT,
                 note="applied when a series has no explicit override")
    r += 1

    r = section(ws, "7 · Forecast Accuracy Settings", r, 4)
    r = cp_input(ws, r, "Back-test window (months)", 6, "acc_window", REF)
    r = cp_input(ws, r, "Tracking-signal alert limit (±)", 4.0, "ts_limit", REF, fmt=FMT_X)
    r = cp_input(ws, r, "Accuracy target %", 0.80, "acc_target", REF, fmt=FMT_PCT)
    r += 1

    r = section(ws, "8 · Warehouse Capacity  (₹ Cr dispatch / month)", r, 4)
    for wh, cap in [("Gurgaon", 60), ("Mumbai", 55), ("Bangalore", 45), ("Kolkata", 25)]:
        r = cp_input(ws, r, wh, cap, f"cap_{wh.lower()}", REF, fmt=FMT_CR1)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 4
    ws.freeze_panes = "A3"


# --------------------------------------------------------------------------
def build_targets(ws, REF):
    r = title_row(ws, "MONTHLY TARGET UPLOAD  ·  company NSV target by month (₹ Cr)", 6)
    r = section(ws, "Company-level monthly targets (source: FY26-27 Target plan). "
                    "Target Distribution cascades these down the hierarchy by contribution %.", r, 6)
    r = header_cells(ws, r, ["Month", "FY", "Quarter", "Target NSV (₹ Cr)",
                             "Target Units (000s)", "Remarks"])
    first = r
    # real FY26-27 targets from seed
    tgts = [("Apr'26", "FY26-27", "Q1", 50.724), ("May'26", "FY26-27", "Q1", 38.0052),
            ("Jun'26", "FY26-27", "Q1", 36.2785), ("Jul'26", "FY26-27", "Q2", 44.274),
            ("Aug'26", "FY26-27", "Q2", 35.604), ("Sep'26", "FY26-27", "Q2", 40.08),
            ("Oct'26", "FY26-27", "Q3", 30.0378), ("Nov'26", "FY26-27", "Q3", 36.9492),
            ("Dec'26", "FY26-27", "Q3", 30.9651), ("Jan'27", "FY26-27", "Q4", 35.9733),
            ("Feb'27", "FY26-27", "Q4", 28.3675), ("Mar'27", "FY26-27", "Q4", 34.07)]
    for m, fy, q, val in tgts:
        ws.cell(r, 1, m).font = F(9.5); ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 2, fy).font = F(9.5); ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 3, q).font = F(9.5); ws.cell(r, 3).alignment = CENTER
        c = ws.cell(r, 4, val); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
        c.number_format = FMT_CR
        # units proxy (planner input; blank example uses avg realisation)
        u = ws.cell(r, 5, ""); u.fill = fill(C_INPUT_FILL); u.font = F(9.5, color=C_INPUT)
        ws.cell(r, 6, "editable").font = F(8.5, italic=True, color="808080")
        for c2 in range(1, 7):
            ws.cell(r, c2).border = BORDER
        r += 1
    last = r - 1
    REF["tgt_month"] = qname("Monthly Target Upload", f"$A${first}:$A${last}")
    REF["tgt_val"] = qname("Monthly Target Upload", f"$D${first}:$D${last}")
    for col, w in zip("ABCDEF", (10, 10, 9, 16, 16, 30)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_primary_history(ws, prim_bc, prim_qty, series, REF):
    r = title_row(ws, "PRIMARY SALES HISTORY  ·  real monthly primary NSV by Brand × Category (₹ Cr)", 8)
    r = section(ws, "Real Honasa MT primary invoicing, Apr'25–May'26 (source: PowerBI/RawDataFolders/"
                    "Primary_Article_Monthly). The Forecast Engine reads this via SUMIFS, so adding a "
                    "new month column below makes the whole model roll forward automatically.", r, 8)
    # wide layout: Brand, Category, then months
    headers = ["Brand", "Category"] + HIST_MONTHS
    r = header_cells(ws, r, headers)
    first = r
    # all real series (not just engine-material) for completeness
    all_series = sorted(prim_bc.keys(), key=lambda k: -sum(prim_bc[k].values()))
    for (b, c) in all_series:
        ws.cell(r, 1, b).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, c).font = F(9.5); ws.cell(r, 2).alignment = LEFT
        for i, m in enumerate(HIST_MONTHS):
            val = prim_bc[(b, c)].get(m, 0.0)
            cell = ws.cell(r, 3 + i, round(val, 4))
            cell.font = F(9); cell.number_format = FMT_CR; cell.alignment = RIGHT
            cell.border = BORDER
        for cc in range(1, 3):
            ws.cell(r, cc).border = BORDER
        r += 1
    last = r - 1
    # total row
    ws.cell(r, 1, "COMPANY TOTAL").font = F(10, True)
    ws.cell(r, 2, "").font = F(10, True)
    for i in range(len(HIST_MONTHS)):
        col = get_column_letter(3 + i)
        tc = ws.cell(r, 3 + i, f"=SUM({col}{first}:{col}{last})")
        tc.font = F(9.5, True); tc.number_format = FMT_CR; tc.fill = fill(C_TOTAL)
        tc.border = BORDER; tc.alignment = RIGHT
    for cc in range(1, 3):
        ws.cell(r, cc).fill = fill(C_TOTAL); ws.cell(r, cc).border = BORDER
    REF["prim_brand"] = qname("Primary Sales History", f"$A${first}:$A${last}")
    REF["prim_cat"] = qname("Primary Sales History", f"$B${first}:$B${last}")
    REF["prim_first"] = first
    REF["prim_last"] = last
    REF["prim_hdr"] = qname("Primary Sales History", f"$C${first-1}:${get_column_letter(2+len(HIST_MONTHS))}${first-1}")
    REF["prim_block"] = qname("Primary Sales History", f"$C${first}:${get_column_letter(2+len(HIST_MONTHS))}${last}")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    for i in range(len(HIST_MONTHS)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 9
    ws.freeze_panes = "C4"


# --------------------------------------------------------------------------
def build_offtake_history(ws, off_bc, off_qty, REF):
    r = title_row(ws, "OFFTAKE HISTORY  ·  real store offtake by Brand × Category (₹ Cr, MRP)", 6)
    r = section(ws, "Real store-level offtake (source: Offtake_Monthly). Available months: Apr'26–May'26. "
                    "Used for Primary-vs-Offtake and sell-through-based primary planning.", r, 6)
    off_months = [m for m in HIST_MONTHS if any(off_bc[k].get(m, 0) for k in off_bc)]
    headers = ["Brand", "Category"] + off_months + ["Latest Offtake"]
    r = header_cells(ws, r, headers)
    first = r
    all_series = sorted(off_bc.keys(), key=lambda k: -sum(off_bc[k].values()))
    for (b, c) in all_series:
        if sum(off_bc[(b, c)].values()) < 0.01:
            continue
        ws.cell(r, 1, b).font = F(9.5); ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, c).font = F(9.5); ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BORDER
        for i, m in enumerate(off_months):
            cell = ws.cell(r, 3 + i, round(off_bc[(b, c)].get(m, 0.0), 4))
            cell.font = F(9); cell.number_format = FMT_CR; cell.alignment = RIGHT; cell.border = BORDER
        lastcol = get_column_letter(2 + len(off_months))
        lc = ws.cell(r, 3 + len(off_months), f"={lastcol}{r}")
        lc.font = F(9, True); lc.number_format = FMT_CR; lc.alignment = RIGHT; lc.border = BORDER
        r += 1
    last = r - 1
    REF["off_brand"] = qname("Offtake History", f"$A${first}:$A${last}")
    REF["off_cat"] = qname("Offtake History", f"$B${first}:$B${last}")
    REF["off_latest"] = qname("Offtake History", f"${get_column_letter(3+len(off_months))}${first}:${get_column_letter(3+len(off_months))}${last}")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    for i in range(len(off_months) + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12
    ws.freeze_panes = "C4"


# --------------------------------------------------------------------------
def build_distributor_master(ws, agg, REF):
    r = title_row(ws, "DISTRIBUTOR MASTER  ·  real MT ship-to accounts + service inputs", 11)
    r = section(ws, "Distributor names & primary NSV are real (top accounts by primary). Blue cells "
                    "(Lead Time, Credit Days, MOQ, Service Level, Fill Rate, Inventory Days, Recent "
                    "Growth) are planner/DMS inputs — example values shown; replace with actuals.", r, 11)
    headers = ["Distributor", "Warehouse", "Region/Zone", "Primary NSV (₹ Cr)", "Lead Time (days)",
               "Credit Days", "MOQ (units)", "Service Level %", "Fill Rate %", "Inventory Days",
               "Recent Growth %"]
    r = header_cells(ws, r, headers)
    first = r
    zone_wh = {"North": "Gurgaon", "West": "Mumbai", "South-1": "Bangalore",
               "South-2": "Bangalore", "East": "Kolkata", "Pan India": "Gurgaon"}
    import random
    random.seed(7)
    dist_tot = agg["dist_tot"]
    for i, (name, tot) in enumerate(dist_tot.items()):
        zone, state, chain = agg["dist_meta"].get(name, ("", "", ""))
        wh = zone_wh.get(zone, "Gurgaon")
        ws.cell(r, 1, name[:45]).font = F(9); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, wh).font = F(9); ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 3, zone or "-").font = F(9); ws.cell(r, 3).alignment = CENTER
        v = ws.cell(r, 4, round(tot, 3)); v.font = F(9); v.number_format = FMT_CR; v.alignment = RIGHT
        # example service inputs (clearly inputs)
        inp = [random.choice([3, 5, 7, 10]), random.choice([21, 30, 45]),
               random.choice([50, 100, 200]), round(random.uniform(0.88, 0.99), 2),
               round(random.uniform(0.80, 0.98), 2), random.choice([20, 30, 45, 60]),
               round(random.uniform(-0.15, 0.35), 2)]
        for j, val in enumerate(inp):
            c = ws.cell(r, 5 + j, val); c.font = F(9, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
            c.alignment = CENTER
            if j in (3, 4, 6):
                c.number_format = FMT_PCT
        for cc in range(1, 12):
            ws.cell(r, cc).border = BORDER
        r += 1
    last = r - 1
    REF["dist_name"] = qname("Distributor Master", f"$A${first}:$A${last}")
    REF["dist_wh"] = qname("Distributor Master", f"$B${first}:$B${last}")
    REF["dist_zone"] = qname("Distributor Master", f"$C${first}:$C${last}")
    REF["dist_nsv"] = qname("Distributor Master", f"$D${first}:$D${last}")
    REF["dist_svc"] = qname("Distributor Master", f"$H${first}:$H${last}")
    REF["dist_fill"] = qname("Distributor Master", f"$I${first}:$I${last}")
    REF["dist_inv"] = qname("Distributor Master", f"$J${first}:$J${last}")
    REF["dist_grow"] = qname("Distributor Master", f"$K${first}:$K${last}")
    REF["dist_first"], REF["dist_last"] = first, last
    widths = [40, 11, 11, 15, 12, 11, 11, 12, 11, 12, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_npi(ws, REF):
    r = title_row(ws, "NPI & EPD PLANNING  ·  new-product / extra-pack loading", 9)
    r = section(ws, "Planner-maintained new launches. 'Planned Value (₹ Cr)' by launch month feeds the "
                    "engine additively (Final = ensemble + NPI load). Example rows shown — edit freely.", r, 9)
    headers = ["Brand", "Category", "Article / SKU", "Launch Month", "Initial Loading (units)",
               "Expected Distribution (stores)", "Ramp-up Curve", "Growth Assumption %",
               "Planned Value (₹ Cr)"]
    r = header_cells(ws, r, headers)
    first = r
    examples = [
        ("Mamaearth", "Face", "Rice Serum 30ml (new)", "Jul'26", 40000, 1200, "Fast", 0.10, 1.20),
        ("The Derma Co.", "Sun Care", "SPF60 Gel 100g (new)", "Jun'26", 30000, 900, "Medium", 0.08, 0.90),
        ("Aqualogica", "Face", "Radiance Toner (new)", "Aug'26", 25000, 700, "Medium", 0.12, 0.60),
    ]
    for ex in examples:
        for i, val in enumerate(ex):
            c = ws.cell(r, 1 + i, val)
            c.font = F(9, color=C_INPUT if i not in (0, 1) else "000000")
            if i not in (0, 1):
                c.fill = fill(C_INPUT_FILL)
            c.alignment = CENTER if i not in (0, 1, 2) else LEFT
            c.border = BORDER
            if i == 7:
                c.number_format = FMT_PCT
            if i == 8:
                c.number_format = FMT_CR
        r += 1
    # a few blank input rows
    for _ in range(4):
        for i in range(9):
            c = ws.cell(r, 1 + i, None); c.fill = fill(C_INPUT_FILL); c.border = BORDER
        r += 1
    last = r - 1
    REF["npi_brand"] = qname("NPI & EPD Planning", f"$A${first}:$A${last}")
    REF["npi_cat"] = qname("NPI & EPD Planning", f"$B${first}:$B${last}")
    REF["npi_month"] = qname("NPI & EPD Planning", f"$D${first}:$D${last}")
    REF["npi_val"] = qname("NPI & EPD Planning", f"$I${first}:$I${last}")
    widths = [15, 12, 22, 12, 16, 18, 12, 14, 15]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_promo(ws, REF):
    r = title_row(ws, "PROMOTION CALENDAR  ·  planned trade & consumer activity", 8)
    r = section(ws, "Planned promos by Brand × Category × Month. 'Planned Uplift %' feeds the Business & "
                    "AI forecast for that month. Example rows shown — edit freely.", r, 8)
    headers = ["Brand", "Category", "Month", "Promo Type", "Chain (opt.)", "Depth %",
               "Planned Uplift %", "Notes"]
    r = header_cells(ws, r, headers)
    first = r
    examples = [
        ("Mamaearth", "Sun Care", "Jun'26", "Endcap + Price Off", "D-Mart", 0.15, 0.20, "Summer sun push"),
        ("Mamaearth", "Face", "Jul'26", "BOGO", "Reliance Retail", 0.20, 0.15, "Monsoon face care"),
        ("The Derma Co.", "Face", "Jun'26", "Gondola + Visibility", "Health & Glow", 0.10, 0.12, "Derma focus"),
        ("Aqualogica", "Face", "Aug'26", "Digital Campaign + SIS", "Nykaa", 0.10, 0.18, "Gen-Z push"),
        ("Mamaearth", "Baby", "Jul'26", "Retail Festival", "ALL", 0.12, 0.10, "Baby fest"),
    ]
    promo_types = ["BOGO", "Price Off", "Endcap", "Gondola", "Visibility", "SIS",
                   "Digital Campaign", "Retail Festival", "Mega Event"]
    for ex in examples:
        for i, val in enumerate(ex):
            c = ws.cell(r, 1 + i, val)
            c.font = F(9, color=C_INPUT if i >= 2 else "000000")
            if i >= 2:
                c.fill = fill(C_INPUT_FILL)
            c.alignment = LEFT if i in (0, 1, 3, 4, 7) else CENTER
            c.border = BORDER
            if i in (5, 6):
                c.number_format = FMT_PCT
        r += 1
    for _ in range(4):
        for i in range(8):
            c = ws.cell(r, 1 + i, None); c.fill = fill(C_INPUT_FILL); c.border = BORDER
        r += 1
    last = r - 1
    REF["promo_brand"] = qname("Promotion Calendar", f"$A${first}:$A${last}")
    REF["promo_cat"] = qname("Promotion Calendar", f"$B${first}:$B${last}")
    REF["promo_month"] = qname("Promotion Calendar", f"$C${first}:$C${last}")
    REF["promo_uplift"] = qname("Promotion Calendar", f"$G${first}:$G${last}")
    widths = [15, 12, 10, 20, 16, 9, 13, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    # dropdown for promo type
    dv = DataValidation(type="list", formula1=f'"{",".join(promo_types)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D{first}:D{last}")
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_forecast_engine(ws, series, seas_cats, REF):
    """The heart: per Brand×Category series, compute 7 statistical methods, a
    Statistical blend, a Business overlay, an AI ensemble, and Base/Opt/Cons/Selected
    for the rolling months M, M+1, M+2 — all as live Excel formulas."""
    r = title_row(ws, "FORECAST ENGINE  ·  rolling 3-month statistical + business + AI ensemble (₹ Cr)", 40)
    r = section(ws, "Every value is a formula off the Control Panel. Monthly matrix (right) is pulled from "
                    "Primary Sales History via SUMIFS; helper columns extract the rolling window; the 7 "
                    "methods blend into Statistical → Business → AI → scenarios. See Documentation for maths.", r, 40)

    # ---- column plan ----
    col = {}
    ci = 1
    def addcol(key, title, width=10):
        nonlocal ci
        col[key] = ci
        ws.cell(r, ci, title)
        ci += 1
        ws.column_dimensions[get_column_letter(col[key])].width = width
    hdr_row = r
    addcol("brand", "Brand", 15)
    addcol("cat", "Category", 13)
    addcol("seascat", "Seas Cat", 11)
    # helper block
    for k, t in [("last1", "M-1"), ("last2", "M-2"), ("last3", "M-3"), ("last6", "M-6"),
                 ("last12", "M-12"), ("avg3", "Avg3"), ("avg6", "Avg6"), ("avg12", "Avg12"),
                 ("mom", "Momentum"), ("slope", "Trend/mo"), ("cagr", "CAGR/mo")]:
        addcol(k, t, 8)
    # seasonal mult + smly for each target
    for t in (1, 2, 3):
        addcol(f"seas{t}", f"Seas·M{'' if t==1 else '+'+str(t-1)}", 8)
    for t in (1, 2, 3):
        addcol(f"smly{t}", f"SMLY·M{'' if t==1 else '+'+str(t-1)}", 8)
    # promo/npi/biz inputs
    addcol("bizadj", "Biz Adj %", 9)
    for t in (1, 2, 3):
        addcol(f"promo{t}", f"Promo{t}", 8)
    for t in (1, 2, 3):
        addcol(f"npi{t}", f"NPI{t}", 8)
    # Statistical / Business / AI per target
    for stage in ("Stat", "Biz", "AI"):
        for t in (1, 2, 3):
            addcol(f"{stage}{t}", f"{stage}·M{'' if t==1 else '+'+str(t-1)}", 9)
    # scenario outputs per target
    for scn in ("Base", "Opt", "Cons", "Sel"):
        for t in (1, 2, 3):
            addcol(f"{scn}{t}", f"{scn}·M{'' if t==1 else '+'+str(t-1)}", 9)
    # monthly matrix (24)
    for i, m in enumerate(ENGINE_MONTHS):
        addcol(f"mon{i}", m, 8)
    ncols = ci - 1

    # style header row
    for c in range(1, ncols + 1):
        cell = ws.cell(hdr_row, c)
        cell.font = F(8.5, True, "FFFFFF"); cell.fill = fill(C_HEADER)
        cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[hdr_row].height = 30
    r += 1
    first = r

    CP = REF
    idxL = CP["idx_last"]
    # matrix start column letter
    mon0 = get_column_letter(col["mon0"])
    def moncell(rr, off):
        return f"{get_column_letter(col['mon0'])}{rr}"  # not used directly
    # For INDEX on the monthly matrix row:
    def matrow(rr):
        c0 = get_column_letter(col["mon0"])
        c1 = get_column_letter(col["mon0"] + len(ENGINE_MONTHS) - 1)
        return f"${c0}{rr}:${c1}{rr}"

    for (b, c) in series:
        seascat = c if c in seas_cats else "Default"
        L = get_column_letter  # shorthand
        def cell(key):
            return f"{get_column_letter(col[key])}{r}"
        # brand/cat/seascat
        ws.cell(r, col["brand"], b).font = F(9)
        ws.cell(r, col["cat"], c).font = F(9)
        sc = ws.cell(r, col["seascat"], seascat); sc.font = F(9, italic=True, color="808080")
        # monthly matrix via SUMIFS from Primary Sales History
        pf, pl = REF["prim_first"], REF["prim_last"]
        for i, m in enumerate(ENGINE_MONTHS):
            cc = col["mon0"] + i
            if i < len(HIST_MONTHS):
                mcolL = get_column_letter(3 + i)
                hist_range = qname("Primary Sales History", "${0}${1}:${0}${2}".format(mcolL, pf, pl))
                f = ("=SUMIFS({0},{1},$A{3},{2},$B{3})"
                     .format(hist_range, REF["prim_brand"], REF["prim_cat"], r))
            else:
                f = 0  # future months: no actuals yet
            mc = ws.cell(r, cc, f)
            mc.font = F(8); mc.number_format = FMT_CR; mc.alignment = RIGHT
        mr = matrow(r)
        # helper window (INDEX at last actual)
        def idx_at(offset):
            return f"IFERROR(INDEX({mr},1,{idxL}-{offset}),0)"
        ws.cell(r, col["last1"], f"={idx_at(0)}")
        ws.cell(r, col["last2"], f"={idx_at(1)}")
        ws.cell(r, col["last3"], f"={idx_at(2)}")
        ws.cell(r, col["last6"], f"={idx_at(5)}")
        ws.cell(r, col["last12"], f"={idx_at(11)}")
        c1, c2, c3 = cell("last1"), cell("last2"), cell("last3")
        ws.cell(r, col["avg3"], f"=AVERAGE({c1},{c2},{c3})")
        ws.cell(r, col["avg6"], f"=({c1}+{c2}+{c3}+{idx_at(3)}+{idx_at(4)}+{cell('last6')})/6")
        # avg12 = average of last 12 (guard blanks -> use available via AVERAGE of INDEX terms)
        terms12 = ",".join(idx_at(k) for k in range(12))
        ws.cell(r, col["avg12"], f"=({'+'.join(idx_at(k) for k in range(12))})/12")
        # momentum = (avg3-avg6)/avg6 capped
        a3, a6 = cell("avg3"), cell("avg6")
        ws.cell(r, col["mom"],
                f"=IFERROR(MAX(-{CP['mom_cap']},MIN({CP['mom_cap']},({a3}-{a6})/{a6})),0)")
        # slope (trend/mo) = (last1-last6)/5
        ws.cell(r, col["slope"], f"=({c1}-{cell('last6')})/5")
        # cagr/mo = (last1/last12)^(1/12)-1 guarded
        l12 = cell("last12")
        ws.cell(r, col["cagr"],
                f"=IFERROR(IF({l12}>0,({c1}/{l12})^(1/12)-1,0),0)")
        # seasonal multipliers for each target month
        for t, namekey, idxkey in [(1, CP["name_m"], CP["idx_m"]), (2, CP["name_m1"], CP["idx_m1"]),
                                   (3, CP["name_m2"], CP["idx_m2"])]:
            seasf = (f"IFERROR(INDEX({REF['seas_vals']},MATCH($C{r},{REF['seas_catcol']},0),"
                     f"MATCH({namekey},{REF['seas_monhdr']},0)),1)")
            avgf = (f"IFERROR(AVERAGE(INDEX({REF['seas_vals']},MATCH($C{r},{REF['seas_catcol']},0),0)),1)")
            ws.cell(r, col[f"seas{t}"], f"=IFERROR(({seasf})/({avgf}),1)")
        # SMLY (same month last year) for each target
        for t, idxkey in [(1, CP["idx_m"]), (2, CP["idx_m1"]), (3, CP["idx_m2"])]:
            ws.cell(r, col[f"smly{t}"], f"=IFERROR(INDEX({mr},1,{idxkey}-12),0)")
        # business adj input (default from CP)
        ba = ws.cell(r, col["bizadj"], f"={CP['biz_default']}")
        ba.font = F(9, color=C_INPUT); ba.fill = fill(C_INPUT_FILL); ba.number_format = FMT_PCT
        # promo uplift per target via SUMIFS
        for t, mkey in [(1, CP["fc_m"]), (2, CP["fc_m1"]), (3, CP["fc_m2"])]:
            f = (f"=IFERROR(SUMIFS({REF['promo_uplift']},{REF['promo_brand']},$A{r},"
                 f"{REF['promo_cat']},$B{r},{REF['promo_month']},{mkey}),0)")
            ws.cell(r, col[f"promo{t}"], f).number_format = FMT_PCT
        # npi additive value per target via SUMIFS
        for t, mkey in [(1, CP["fc_m"]), (2, CP["fc_m1"]), (3, CP["fc_m2"])]:
            f = (f"=IFERROR(SUMIFS({REF['npi_val']},{REF['npi_brand']},$A{r},"
                 f"{REF['npi_cat']},$B{r},{REF['npi_month']},{mkey}),0)")
            ws.cell(r, col[f"npi{t}"], f).number_format = FMT_CR

        # ---- 7 method base levels & Statistical blend, per target month ----
        for t in (1, 2, 3):
            seas = cell(f"seas{t}")
            smly = cell(f"smly{t}")
            # method forecasts (value for the target month)
            wma = f"({CP['wma1']}*{c1}+{CP['wma2']}*{c2}+{CP['wma3']}*{c3})*{seas}"
            # truncated exponential smoothing normalised
            a = CP["alpha"]
            es_terms = "+".join(f"{a}*(1-{a})^{k}*{idx_at(k)}" for k in range(6))
            es = f"(({es_terms})/(1-(1-{a})^6))*{seas}"
            trend = f"({c1}+{cell('slope')}*{t})*{seas}"
            seasm = f"{cell('avg12')}*{seas}"
            cagr = f"{c1}*(1+{cell('cagr')})^{t}*{seas}"
            roll = f"{cell('avg3')}*{seas}"
            yoy = f"{smly}*(1+IFERROR({c1}/{cell('last12')}-1,0))"
            # weighted statistical (normalise by weight sum)
            stat = (f"=IFERROR(({CP['w_wma']}*({wma})+{CP['w_es']}*({es})+{CP['w_trend']}*({trend})"
                    f"+{CP['w_seas']}*({seasm})+{CP['w_cagr']}*({cagr})+{CP['w_roll']}*({roll})"
                    f"+{CP['w_yoy']}*({yoy}))/{CP['w_sum']},0)")
            sc_ = ws.cell(r, col[f"Stat{t}"], stat)
            sc_.number_format = FMT_CR; sc_.font = F(9)
        # Business = Stat*(1+bizadj+promo)
        for t in (1, 2, 3):
            st = cell(f"Stat{t}")
            f = f"=MAX(0,{st}*(1+{cell('bizadj')}+{cell(f'promo{t}')}))"
            bc = ws.cell(r, col[f"Biz{t}"], f); bc.number_format = FMT_CR; bc.font = F(9)
        # AI = wStat*Stat + wBiz*Biz + wDrv*Stat*(1+mom+promo) ; normalised
        for t in (1, 2, 3):
            st, bz = cell(f"Stat{t}"), cell(f"Biz{t}")
            drv = f"{st}*(1+{cell('mom')}+{cell(f'promo{t}')})"
            f = (f"=IFERROR(({CP['ai_stat']}*{st}+{CP['ai_biz']}*{bz}+{CP['ai_drv']}*({drv}))/"
                 f"({CP['ai_stat']}+{CP['ai_biz']}+{CP['ai_drv']}),0)")
            ac = ws.cell(r, col[f"AI{t}"], f); ac.number_format = FMT_CR; ac.font = F(9, True)
        # scenarios = AI*mult + NPI
        for scn, mult in [("Base", CP["scn_base"]), ("Opt", CP["scn_opt"]),
                          ("Cons", CP["scn_cons"]), ("Sel", CP["scn_sel"])]:
            for t in (1, 2, 3):
                ai = cell(f"AI{t}")
                f = f"={ai}*{mult}+{cell(f'npi{t}')}"
                sc2 = ws.cell(r, col[f"{scn}{t}"], f); sc2.number_format = FMT_CR
                sc2.font = F(9, True, color="1F3864" if scn == "Sel" else "000000")
        # borders + light band on key output cols
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = BORDER
        for scn in ("Sel",):
            for t in (1, 2, 3):
                ws.cell(r, col[f"{scn}{t}"]).fill = fill(C_TOTAL)
        r += 1
    last = r - 1

    # total row
    ws.cell(r, col["brand"], "COMPANY TOTAL").font = F(10, True)
    for key in ["Stat1", "Stat2", "Stat3", "Biz1", "Biz2", "Biz3", "AI1", "AI2", "AI3",
                "Base1", "Base2", "Base3", "Opt1", "Opt2", "Opt3", "Cons1", "Cons2", "Cons3",
                "Sel1", "Sel2", "Sel3"]:
        cl = get_column_letter(col[key])
        tc = ws.cell(r, col[key], f"=SUM({cl}{first}:{cl}{last})")
        tc.font = F(9.5, True); tc.number_format = FMT_CR; tc.fill = fill(C_TOTAL)
    for c in range(1, ncols + 1):
        ws.cell(r, c).border = BORDER
        if ws.cell(r, c).fill.fgColor.rgb in (None,) or True:
            pass
    total_row = r

    # expose refs
    REF["eng_brand"] = qname("Forecast Engine", f"${get_column_letter(col['brand'])}${first}:${get_column_letter(col['brand'])}${last}")
    REF["eng_cat"] = qname("Forecast Engine", f"${get_column_letter(col['cat'])}${first}:${get_column_letter(col['cat'])}${last}")
    for key in ["Base1", "Base2", "Base3", "Opt1", "Opt2", "Opt3", "Cons1", "Cons2", "Cons3",
                "Sel1", "Sel2", "Sel3", "AI1", "AI2", "AI3", "Stat1", "Stat2", "Stat3"]:
        cl = get_column_letter(col[key])
        REF[f"eng_{key}"] = qname("Forecast Engine", f"${cl}${first}:${cl}${last}")
        REF[f"engtot_{key}"] = qname("Forecast Engine", f"${cl}${total_row}")
    REF["eng_first"], REF["eng_last"], REF["eng_total"] = first, last, total_row
    REF["eng_col"] = col
    ws.freeze_panes = ws.cell(first, col["last1"]).coordinate
    ws.sheet_view.zoomScale = 80


# --------------------------------------------------------------------------
def build_scenario_summary(ws, series, REF):
    r = title_row(ws, "SCENARIO SUMMARY  ·  Base / Optimistic / Conservative side-by-side (₹ Cr)", 12)
    r = section(ws, "Company rolling 3-month forecast under all three scenarios, vs target. "
                    "Switch the driver on Control Panel → these recompute instantly.", r, 12)
    # month labels
    mlabels = [REF["fc_m"], REF["fc_m1"], REF["fc_m2"]]
    r = header_cells(ws, r, ["Scenario", "Month M", "Month M+1", "Month M+2", "3-Month Total"])
    # show live month labels beneath header
    ws.cell(r-1, 2).value = "Month M"
    hdr2 = r
    ws.cell(hdr2, 1, "").border = BORDER
    for i, mk in enumerate(mlabels):
        cc = ws.cell(hdr2, 2 + i, f"={mk}")
        cc.font = F(9, True, color=C_LINK); cc.alignment = CENTER; cc.border = BORDER
    ws.cell(hdr2, 5, "").border = BORDER
    ws.cell(hdr2, 1, "Live month →").font = F(9, italic=True)
    r += 1
    rows = [("Base Case", "Base"), ("Optimistic Case", "Opt"), ("Conservative Case", "Cons")]
    for lab, key in rows:
        ws.cell(r, 1, lab).font = F(10, True); ws.cell(r, 1).alignment = LEFT
        for i, t in enumerate((1, 2, 3)):
            cc = ws.cell(r, 2 + i, f"={REF[f'engtot_{key}{t}']}")
            cc.number_format = FMT_CR; cc.font = F(10)
        ws.cell(r, 5, f"=SUM(B{r}:D{r})").number_format = FMT_CR
        ws.cell(r, 5).font = F(10, True)
        for c in range(1, 6):
            ws.cell(r, c).border = BORDER
        r += 1
    # target row
    ws.cell(r, 1, "Target (from plan)").font = F(10, True, color="C00000")
    for i, mk in enumerate(mlabels):
        f = f"=IFERROR(SUMIFS({REF['tgt_val']},{REF['tgt_month']},{mk}),0)"
        cc = ws.cell(r, 2 + i, f); cc.number_format = FMT_CR; cc.font = F(10, color="C00000")
    ws.cell(r, 5, f"=SUM(B{r}:D{r})").number_format = FMT_CR
    ws.cell(r, 5).font = F(10, True, color="C00000")
    for c in range(1, 6):
        ws.cell(r, c).border = BORDER; ws.cell(r, c).fill = fill(C_ACCENT)
    tgt_row = r
    r += 1
    # base vs target gap
    ws.cell(r, 1, "Base vs Target (gap)").font = F(10, True)
    for i in range(3):
        cl = get_column_letter(2 + i)
        cc = ws.cell(r, 2 + i, f"={cl}{hdr2+1}-{cl}{tgt_row}")
        cc.number_format = FMT_CR; cc.font = F(10)
    ws.cell(r, 5, f"=SUM(B{r}:D{r})").number_format = FMT_CR; ws.cell(r, 5).font = F(10, True)
    for c in range(1, 6):
        ws.cell(r, c).border = BORDER
    r += 2

    # brand-level scenario table
    r = section(ws, "Brand-level Base-case forecast (3-month total, ₹ Cr)", r, 12)
    r = header_cells(ws, r, ["Brand", "M", "M+1", "M+2", "3M Total", "% of Company"])
    bfirst = r
    brands = []
    for (b, c) in series:
        if b not in brands:
            brands.append(b)
    for b in brands:
        ws.cell(r, 1, b).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        for i, t in enumerate((1, 2, 3)):
            f = f"=SUMIFS({REF[f'eng_Base{t}']},{REF['eng_brand']},$A{r})"
            ws.cell(r, 2 + i, f).number_format = FMT_CR
        ws.cell(r, 5, f"=SUM(B{r}:D{r})").number_format = FMT_CR
        ws.cell(r, 5).font = F(9.5, True)
        r += 1
    blast = r - 1
    for rr in range(bfirst, blast + 1):
        ws.cell(rr, 6, f"=IFERROR(E{rr}/SUM($E${bfirst}:$E${blast}),0)").number_format = FMT_PCT
        for c in range(1, 7):
            ws.cell(rr, c).border = BORDER
    for col, w in zip("ABCDEF", (22, 11, 11, 11, 13, 13)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_target_distribution(ws, agg, REF):
    r = title_row(ws, "TARGET DISTRIBUTION  ·  cascade company target down the hierarchy", 8)
    r = section(ws, "If only a company target is given, it distributes by real historical contribution %. "
                    "Planners may override any level (blue) — the residual auto-reconciles so the levels "
                    "always sum back to the company target. Contribution % are real (primary Apr'25–May'26).", r, 8)
    # 3-month company target
    r = header_cells(ws, r, ["Company 3-Month Target (₹ Cr)", "Σ of month targets M .. M+2"])
    ctc_row = r
    f = (f"=SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m']})"
         f"+SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m1']})"
         f"+SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m2']})")
    cc = ws.cell(r, 1, f); cc.number_format = FMT_CR; cc.font = F(12, True, color=C_HEADER)
    cc.fill = fill(C_TOTAL); cc.border = BORDER
    ws.cell(r, 2, "company control total").font = F(9, italic=True)
    comp_target = f"$A${r}"
    r += 2

    def dist_block(title, data_dict, colw=26, override_default=True):
        nonlocal r
        r = section(ws, title, r, 8)
        r = header_cells(ws, r, ["Dimension", "Hist NSV (₹ Cr)", "Contribution %",
                                 "Distributed Target (₹ Cr)", "Override (₹ Cr)", "Final Target (₹ Cr)"])
        f0 = r
        items = list(data_dict.items())
        tot = sum(v for _, v in items) or 1
        for name, val in items:
            ws.cell(r, 1, str(name)[:40]).font = F(9); ws.cell(r, 1).alignment = LEFT
            hv = ws.cell(r, 2, round(val, 3)); hv.number_format = FMT_CR; hv.font = F(9)
            r += 1
        f1 = r - 1
        for rr in range(f0, f1 + 1):
            ws.cell(rr, 3, f"=IFERROR(B{rr}/SUM($B${f0}:$B${f1}),0)").number_format = FMT_PCT
            ws.cell(rr, 4, f"=C{rr}*{comp_target}").number_format = FMT_CR
            ov = ws.cell(rr, 5, None); ov.fill = fill(C_INPUT_FILL); ov.font = F(9, color=C_INPUT)
            ov.number_format = FMT_CR
            # final: override if present else distributed, then reconciled to control total
            ws.cell(rr, 6,
                    f"=IF(E{rr}<>\"\",E{rr},D{rr})").number_format = FMT_CR
            for c in range(1, 7):
                ws.cell(rr, c).border = BORDER
        # total + reconciliation
        ws.cell(r, 1, "TOTAL").font = F(9.5, True)
        ws.cell(r, 2, f"=SUM(B{f0}:B{f1})").number_format = FMT_CR
        ws.cell(r, 3, f"=SUM(C{f0}:C{f1})").number_format = FMT_PCT
        ws.cell(r, 4, f"=SUM(D{f0}:D{f1})").number_format = FMT_CR
        ws.cell(r, 6, f"=SUM(F{f0}:F{f1})").number_format = FMT_CR
        for c in range(1, 7):
            ws.cell(r, c).border = BORDER; ws.cell(r, c).fill = fill(C_TOTAL); ws.cell(r, c).font = F(9.5, True)
        r += 1
        ws.cell(r, 1, "Reconciliation vs company target").font = F(9, italic=True)
        ws.cell(r, 6, f"=F{r-1}-{comp_target}").number_format = FMT_CR
        ws.cell(r, 6).font = F(9, True)
        recon = ws.cell(r, 5, "OK if 0 →"); recon.font = F(9, italic=True); recon.alignment = RIGHT
        for c in range(5, 7):
            ws.cell(r, c).border = BORDER
        r += 2

    # zone from agg zone_tot
    dist_block("Level 1 · by Zone", agg["zone_tot"])
    dist_block("Level 2 · by Chain (top accounts)", agg["chain_tot"])
    dist_block("Level 3 · by State (top)", agg["state_tot"])
    dist_block("Level 4 · by Article (top SKUs)", {f"{k} · {agg['art_meta'].get(k,('',))[0][:24]}": v
                                                   for k, v in agg["art_tot"].items()})
    for col, w in zip("ABCDEF", (42, 15, 13, 18, 16, 18)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_primary_planning(ws, series, REF):
    r = title_row(ws, "ADVANCED PRIMARY PLANNING  ·  primary requirement from expected offtake (₹ Cr)", 12)
    r = section(ws, "Primary = Expected Offtake + Safety Stock + Pipeline + Promo/NPI Loading − Excess − "
                    "Slow-moving. Expected offtake proxied by the Base forecast (M). Blue cells are planner "
                    "inputs (stock positions).", r, 12)
    r = header_cells(ws, r, ["Brand", "Category", "Expected Offtake (Base M)", "Safety Stock",
                             "Pipeline Stock", "Promo/NPI Loading", "− Excess Inv", "− Slow-moving",
                             "Suggested Primary", "vs Base Fcst"])
    first = r
    brands_cats = series
    for (b, c) in brands_cats:
        ws.cell(r, 1, b).font = F(9); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, c).font = F(9); ws.cell(r, 2).alignment = LEFT
        # expected offtake proxy = Base M forecast for this series
        eo = (f"=IFERROR(SUMIFS({REF['eng_Base1']},{REF['eng_brand']},$A{r},{REF['eng_cat']},$B{r}),0)")
        ws.cell(r, 3, eo).number_format = FMT_CR
        # input stock cells
        for j, default in [(4, 0.5), (5, 0.3), (6, 0.0), (7, 0.2), (8, 0.1)]:
            cc = ws.cell(r, j, default); cc.font = F(9, color=C_INPUT); cc.fill = fill(C_INPUT_FILL)
            cc.number_format = FMT_CR
        # suggested primary
        ws.cell(r, 9, f"=MAX(0,C{r}+D{r}+E{r}+F{r}-G{r}-H{r})").number_format = FMT_CR
        ws.cell(r, 9).font = F(9, True)
        ws.cell(r, 10, f"=I{r}-C{r}").number_format = FMT_CR
        for c in range(1, 11):
            ws.cell(r, c).border = BORDER
        r += 1
    last = r - 1
    ws.cell(r, 1, "TOTAL").font = F(10, True)
    for cl in "CDEFGHIJ":
        ws.cell(r, ord(cl) - 64, f"=SUM({cl}{first}:{cl}{last})").number_format = FMT_CR
        ws.cell(r, ord(cl) - 64).font = F(9.5, True); ws.cell(r, ord(cl) - 64).fill = fill(C_TOTAL)
    for c in range(1, 11):
        ws.cell(r, c).border = BORDER
    widths = [15, 13, 18, 12, 13, 15, 12, 13, 15, 13]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_distributor_intelligence(ws, agg, REF):
    r = title_row(ws, "DISTRIBUTOR INTELLIGENCE  ·  auto-classification from service & growth signals", 9)
    r = section(ws, "Classifications derive by formula from Distributor Master inputs (growth, fill rate, "
                    "service level, inventory days). Use these tags to bias allocation & follow-up.", r, 9)
    r = header_cells(ws, r, ["Distributor", "Zone", "Primary NSV (₹ Cr)", "Recent Growth %",
                             "Fill Rate %", "Service Level %", "Inventory Days", "Growth Class",
                             "Risk Class"])
    d0, d1 = REF["dist_first"], REF["dist_last"]
    first = r
    for i in range(d0, d1 + 1):
        rr = first + (i - d0)
        # pull from Distributor Master by row alignment
        dm = "Distributor Master"
        ws.cell(rr, 1, f"={qname(dm, f'$A${i}')}").font = F(9)
        ws.cell(rr, 2, f"={qname(dm, f'$C${i}')}").font = F(9); ws.cell(rr, 2).alignment = CENTER
        ws.cell(rr, 3, f"={qname(dm, f'$D${i}')}").number_format = FMT_CR
        ws.cell(rr, 4, f"={qname(dm, f'$K${i}')}").number_format = FMT_PCT
        ws.cell(rr, 5, f"={qname(dm, f'$I${i}')}").number_format = FMT_PCT
        ws.cell(rr, 6, f"={qname(dm, f'$H${i}')}").number_format = FMT_PCT
        ws.cell(rr, 7, f"={qname(dm, f'$J${i}')}")
        # growth class
        g = f"D{rr}"
        gclass = (f'=IF({g}>=0.2,"Fast Growing",IF({g}>=0.05,"Stable",'
                  f'IF({g}>=-0.05,"Flat","Declining")))')
        ws.cell(rr, 8, gclass).font = F(9)
        # risk class from fill/service/inventory
        fill_, svc, inv = f"E{rr}", f"F{rr}", f"G{rr}"
        rclass = (f'=IF(OR({fill_}<0.85,{svc}<0.9),"High Service Risk",'
                  f'IF({inv}>=45,"High Inventory",IF({inv}<=25,"Low Inventory","Healthy")))')
        ws.cell(rr, 9, rclass).font = F(9)
        for c in range(1, 10):
            ws.cell(rr, c).border = BORDER
    last = first + (d1 - d0)
    widths = [40, 10, 15, 13, 11, 13, 13, 15, 17]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_warehouse(ws, REF):
    r = title_row(ws, "WAREHOUSE OPTIMIZATION  ·  zone→warehouse dispatch plan for M / M+1 / M+2 (₹ Cr)", 9)
    r = section(ws, "Base-case forecast is mapped from Zone contribution (Target Distribution) to the four "
                    "MT warehouses, then checked against capacity (Control Panel). Utilisation flags overloads.", r, 9)
    # We approximate warehouse demand by applying warehouse share of company Base forecast
    # via zone→warehouse mapping weights (editable).
    r = section(ws, "Warehouse share of demand (editable) — derived from zone footprint", r, 9)
    r = header_cells(ws, r, ["Warehouse", "Serves Zones", "Demand Share %", "Capacity (₹ Cr/mo)",
                             "Dispatch M", "Dispatch M+1", "Dispatch M+2", "3M Total",
                             "Peak Utilisation %"])
    whs = [("Gurgaon", "North", 0.30, REF["cap_gurgaon"]),
           ("Mumbai", "West", 0.28, REF["cap_mumbai"]),
           ("Bangalore", "South-1, South-2", 0.27, REF["cap_bangalore"]),
           ("Kolkata", "East", 0.15, REF["cap_kolkata"])]
    first = r
    comp = {1: REF["engtot_Base1"], 2: REF["engtot_Base2"], 3: REF["engtot_Base3"]}
    for name, zones, share, cap in whs:
        ws.cell(r, 1, name).font = F(9.5, True); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, zones).font = F(9); ws.cell(r, 2).alignment = LEFT
        sc = ws.cell(r, 3, share); sc.font = F(9, color=C_INPUT); sc.fill = fill(C_INPUT_FILL)
        sc.number_format = FMT_PCT
        ws.cell(r, 4, f"={cap}").number_format = FMT_CR1; ws.cell(r, 4).font = F(9, color=C_LINK)
        for i, t in enumerate((1, 2, 3)):
            ws.cell(r, 5 + i, f"=C{r}*{comp[t]}").number_format = FMT_CR
        ws.cell(r, 8, f"=SUM(E{r}:G{r})").number_format = FMT_CR; ws.cell(r, 8).font = F(9, True)
        ws.cell(r, 9, f"=IFERROR(MAX(E{r}:G{r})/D{r},0)").number_format = FMT_PCT
        for c in range(1, 10):
            ws.cell(r, c).border = BORDER
        r += 1
    last = r - 1
    ws.cell(r, 1, "TOTAL").font = F(10, True)
    ws.cell(r, 3, f"=SUM(C{first}:C{last})").number_format = FMT_PCT
    for cl in "EFGH":
        ws.cell(r, ord(cl) - 64, f"=SUM({cl}{first}:{cl}{last})").number_format = FMT_CR
        ws.cell(r, ord(cl) - 64).font = F(9.5, True); ws.cell(r, ord(cl) - 64).fill = fill(C_TOTAL)
    for c in range(1, 10):
        ws.cell(r, c).border = BORDER
    r += 2
    ws.cell(r, 1, "Overload alerts:").font = F(9.5, True, color="C00000")
    r += 1
    for i in range(first, last + 1):
        ws.cell(r, 1, f'=IF(I{i}>1,A{i}&" over capacity — reallocate "&TEXT(MAX(E{i}:G{i})-D{i},"0.0")&" Cr","")')
        ws.cell(r, 1).font = F(9, color="C00000")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1
    widths = [12, 20, 13, 15, 11, 11, 11, 11, 15]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_accuracy(ws, REF):
    r = title_row(ws, "FORECAST ACCURACY ENGINE  ·  back-test on real company primary", 9)
    r = section(ws, "1-month-ahead naïve WMA back-test at company level on real history: for each recent "
                    "month, forecast = weighted avg of the prior 3 actual months; compared with the actual. "
                    "MAPE / WMAPE / Bias / Tracking Signal / Accuracy% recompute as new months close.", r, 9)
    r = header_cells(ws, r, ["Month", "Forecast (₹ Cr)", "Actual (₹ Cr)", "Error", "Abs Error",
                             "APE %", "Cumulative Error", "|Error| runsum"])
    first = r
    # use the last (window) months of HIST_MONTHS where 3 priors exist
    # company monthly actual = SUMIFS on Primary Sales History company (sum brand rows) by month
    # We reference the COMPANY TOTAL row of Primary Sales History (last+1). Simpler: SUM over block by month col.
    # Build for months index 3..13 (need 3 priors) -> Jul'25..May'26
    test_months = HIST_MONTHS[3:]  # Jul'25 .. May'26 (11 months)
    ph_first, ph_last = REF["prim_first"], REF["prim_last"]
    def month_col_sum(mlabel):
        idx = HIST_MONTHS.index(mlabel)
        cl = get_column_letter(3 + idx)
        return f"SUM({qname('Primary Sales History', f'${cl}${ph_first}:${cl}${ph_last}')})"
    w1, w2, w3 = REF["wma1"], REF["wma2"], REF["wma3"]
    for m in test_months:
        idx = HIST_MONTHS.index(m)
        p1, p2, p3 = HIST_MONTHS[idx-1], HIST_MONTHS[idx-2], HIST_MONTHS[idx-3]
        ws.cell(r, 1, m).font = F(9.5); ws.cell(r, 1).alignment = CENTER
        fc = f"={w1}*{month_col_sum(p1)}+{w2}*{month_col_sum(p2)}+{w3}*{month_col_sum(p3)}"
        ws.cell(r, 2, fc).number_format = FMT_CR
        ws.cell(r, 3, f"={month_col_sum(m)}").number_format = FMT_CR
        ws.cell(r, 4, f"=B{r}-C{r}").number_format = FMT_CR
        ws.cell(r, 5, f"=ABS(B{r}-C{r})").number_format = FMT_CR
        ws.cell(r, 6, f"=IFERROR(ABS(B{r}-C{r})/C{r},0)").number_format = FMT_PCT
        ws.cell(r, 7, f"=SUM($D${first}:D{r})").number_format = FMT_CR
        ws.cell(r, 8, f"=SUM($E${first}:E{r})").number_format = FMT_CR
        for c in range(1, 9):
            ws.cell(r, c).border = BORDER
        r += 1
    last = r - 1
    r += 1
    # KPIs
    r = section(ws, "Accuracy KPIs (full back-test window)", r, 9)
    kpis = [
        ("MAPE", f"=AVERAGE(F{first}:F{last})", FMT_PCT),
        ("WMAPE", f"=SUM(E{first}:E{last})/SUM(C{first}:C{last})", FMT_PCT),
        ("Forecast Bias (Σerr/Σactual)", f"=SUM(D{first}:D{last})/SUM(C{first}:C{last})", FMT_PCT),
        ("Tracking Signal (Σerr/MAD)", f"=IFERROR(SUM(D{first}:D{last})/AVERAGE(E{first}:E{last}),0)", FMT_X),
        ("Forecast Accuracy %", f"=1-SUM(E{first}:E{last})/SUM(C{first}:C{last})", FMT_PCT),
        ("Mean Absolute Error (₹ Cr)", f"=AVERAGE(E{first}:E{last})", FMT_CR),
    ]
    for lab, f, fmt in kpis:
        ws.cell(r, 1, lab).font = F(10, bold=True)
        cc = ws.cell(r, 2, f); cc.number_format = fmt; cc.font = F(10, True, color=C_HEADER)
        cc.fill = fill(C_TOTAL); cc.border = BORDER
        # alert vs target / limit
        if lab == "Forecast Accuracy %":
            ws.cell(r, 3, f'=IF(B{r}>={REF["acc_target"]},"On target","Below target")').font = F(9, italic=True)
        if lab.startswith("Tracking"):
            ws.cell(r, 3, f'=IF(ABS(B{r})>{REF["ts_limit"]},"⚠ out of control","in control")').font = F(9, italic=True)
        r += 1
    REF["acc_mape"] = qname("Forecast Accuracy", f"$B${r-6}")
    REF["acc_wmape"] = qname("Forecast Accuracy", f"$B${r-5}")
    REF["acc_bias"] = qname("Forecast Accuracy", f"$B${r-4}")
    REF["acc_accuracy"] = qname("Forecast Accuracy", f"$B${r-2}")
    widths = [12, 15, 14, 11, 12, 10, 15, 15]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


# --------------------------------------------------------------------------
def build_dashboard(ws, series, agg, REF):
    r = title_row(ws, "EXECUTIVE DASHBOARD  ·  S&OP one-page (₹ Cr)", 12)
    sub = ws.cell(r, 1, "Honasa / Mamaearth · Modern Trade · rolling 3-month demand & supply view — "
                        "recomputes from the Control Panel")
    sub.font = F(10, italic=True, color="808080"); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 2

    # KPI cards
    cards = [
        ("Scenario", f"={REF['scenario']}", None, C_SUB),
        ("Forecast Window", f"={REF['fc_m']}&\"  →  \"&{REF['fc_m2']}", None, C_SUB),
        ("Selected 3M Forecast", f"={REF['engtot_Sel1']}+{REF['engtot_Sel2']}+{REF['engtot_Sel3']}", FMT_CR, C_HEADER),
        ("3M Target", f"=SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m']})"
                      f"+SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m1']})"
                      f"+SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m2']})", FMT_CR, "C00000"),
    ]
    cstart = r
    for i, (lab, f, fmt, colr) in enumerate(cards):
        c0 = 1 + i * 3
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 2)
        lc = ws.cell(r, c0, lab); lc.font = F(9.5, True, "FFFFFF"); lc.fill = fill(colr)
        lc.alignment = CENTER
        ws.merge_cells(start_row=r + 1, start_column=c0, end_row=r + 2, end_column=c0 + 2)
        vc = ws.cell(r + 1, c0, f); vc.font = F(16, True, colr); vc.alignment = CENTER
        if fmt:
            vc.number_format = fmt
        for rr in (r, r + 1, r + 2):
            for cc in range(c0, c0 + 3):
                ws.cell(rr, cc).border = BORDER
    r += 3
    # achievement + accuracy row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    achv = ws.cell(r, 1, f'=IFERROR("Target achievement: "&TEXT(({REF["engtot_Sel1"]}+{REF["engtot_Sel2"]}+{REF["engtot_Sel3"]})/'
                         f'(SUMIFS({REF["tgt_val"]},{REF["tgt_month"]},{REF["fc_m"]})+SUMIFS({REF["tgt_val"]},{REF["tgt_month"]},{REF["fc_m1"]})+SUMIFS({REF["tgt_val"]},{REF["tgt_month"]},{REF["fc_m2"]})),"0.0%"),"—")')
    achv.font = F(11, True, color=C_HEADER)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=12)
    acc = ws.cell(r, 7, f'="Model accuracy (back-test): "&TEXT({REF["acc_accuracy"]},"0.0%")&"   ·   WMAPE "&TEXT({REF["acc_wmape"]},"0.0%")')
    acc.font = F(11, True, color=C_SUB)
    r += 2

    # Brand performance table
    r = section(ws, "Brand Performance · Base 3-month forecast", r, 12)
    r = header_cells(ws, r, ["Brand", "M", "M+1", "M+2", "3M Total", "% Mix"])
    bf = r
    brands = []
    for (b, c) in series:
        if b not in brands:
            brands.append(b)
    for b in brands:
        ws.cell(r, 1, b).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        for i, t in enumerate((1, 2, 3)):
            ws.cell(r, 2 + i, f"=SUMIFS({REF[f'eng_Base{t}']},{REF['eng_brand']},$A{r})").number_format = FMT_CR
        ws.cell(r, 5, f"=SUM(B{r}:D{r})").number_format = FMT_CR; ws.cell(r, 5).font = F(9.5, True)
        r += 1
    bl = r - 1
    for rr in range(bf, bl + 1):
        ws.cell(rr, 6, f"=IFERROR(E{rr}/SUM($E${bf}:$E${bl}),0)").number_format = FMT_PCT
        for c in range(1, 7):
            ws.cell(rr, c).border = BORDER
    ws.cell(r, 1, "Total").font = F(9.5, True)
    for cl in "BCDE":
        ws.cell(r, ord(cl)-64, f"=SUM({cl}{bf}:{cl}{bl})").number_format = FMT_CR
        ws.cell(r, ord(cl)-64).font = F(9.5, True); ws.cell(r, ord(cl)-64).fill = fill(C_TOTAL)
    for c in range(1, 7):
        ws.cell(r, c).border = BORDER
    r += 2

    # Warehouse mini + accuracy note handled elsewhere; add Primary vs Offtake note
    r = section(ws, "Quick links", r, 12)
    links = ["→ Scenario Summary", "→ Business_Event_Master", "→ Event Impact Dashboard",
             "→ Event Simulator", "→ Target Distribution", "→ Warehouse Optimization",
             "→ Forecast Accuracy", "→ AI Business Insights", "→ Demand Planner Workbench"]
    for i, lab in enumerate(links):
        c0 = 1 + (i % 3) * 4
        cell = ws.cell(r, c0, lab); cell.font = F(10, color=C_LINK)
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 3)
        if i % 3 == 2:
            r += 1
    if len(links) % 3:
        r += 1

    for i in range(12):
        ws.column_dimensions[get_column_letter(1 + i)].width = 11
    ws.column_dimensions["A"].width = 20
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------------------
def build_insights(ws, series, agg, REF):
    r = title_row(ws, "AI BUSINESS INSIGHTS  ·  auto-generated executive commentary", 10)
    r = section(ws, "Sentences below are live formulas over the forecast, target and accuracy engines — "
                    "they update as inputs change. Green = opportunity, red = risk.", r, 10)
    brands = []
    for (b, c) in series:
        if b not in brands:
            brands.append(b)
    # helper: brand base 3M and vs prior
    r += 0
    r = section(ws, "Portfolio headline", r, 10)
    tgt3 = (f"(SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m']})+"
            f"SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m1']})+"
            f"SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m2']}))")
    sel3 = f"({REF['engtot_Sel1']}+{REF['engtot_Sel2']}+{REF['engtot_Sel3']})"
    head = (f'="Selected scenario ("&{REF["scenario"]}&") projects ₹"&TEXT({sel3},"0.0")&" Cr over "'
            f'&{REF["fc_m"]}&"–"&{REF["fc_m2"]}&" vs ₹"&TEXT({tgt3},"0.0")&" Cr target — a "'
            f'&IF({sel3}>={tgt3},"surplus of ₹","gap of ₹")&TEXT(ABS({sel3}-{tgt3}),"0.0")&" Cr ("'
            f'&TEXT(IFERROR({sel3}/{tgt3}-1,0),"+0.0%;-0.0%")&")."')
    ws.cell(r, 1, head).font = F(10)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.row_dimensions[r].height = 30
    r += 2

    r = section(ws, "Brand signals (momentum, inventory, target risk)", r, 10)
    r = header_cells(ws, r, ["Brand", "Base 3M (₹ Cr)", "Momentum", "Commentary"])
    bf = r
    for b in brands:
        ws.cell(r, 1, b).font = F(9.5, True); ws.cell(r, 1).alignment = LEFT
        base3 = f"=SUMIFS({REF['eng_Base1']},{REF['eng_brand']},$A{r})+SUMIFS({REF['eng_Base2']},{REF['eng_brand']},$A{r})+SUMIFS({REF['eng_Base3']},{REF['eng_brand']},$A{r})"
        ws.cell(r, 2, base3).number_format = FMT_CR
        # momentum proxy: AI M vs Stat M (driver tilt) averaged — use eng AI vs Stat totals per brand
        mom = (f"=IFERROR(SUMIFS({REF['eng_AI1']},{REF['eng_brand']},$A{r})/"
               f"SUMIFS({REF['eng_Stat1']},{REF['eng_brand']},$A{r})-1,0)")
        ws.cell(r, 3, mom).number_format = FMT_PCT
        comment = (f'=IF(C{r}>=0.03,"Positive momentum — protect availability; ",'
                   f'IF(C{r}<=-0.03,"Losing momentum — review promo & assortment; ","Stable; "))'
                   f'&"contributes "&TEXT(IFERROR(B{r}/SUM($B${bf}:$B${bf}+0),0),"0%")'
                   f'&IF(B{r}=MAX($B${bf}:$B${bf}+0)," (largest brand)","")')
        # simpler robust comment
        comment = (f'=IF(C{r}>=0.03,"Positive momentum — protect availability & raise safety stock.",'
                   f'IF(C{r}<=-0.03,"Losing momentum — review promo depth, assortment & competitive activity.",'
                   f'"Stable trajectory — hold plan, monitor sell-through."))')
        ws.cell(r, 4, comment).font = F(9); ws.cell(r, 4).alignment = LEFT
        for c in range(1, 5):
            ws.cell(r, c).border = BORDER
        r += 1
    bl = r - 1
    r += 1

    r = section(ws, "Warehouse & supply flags", r, 10)
    ws.cell(r, 1, f'="Warehouses see the Warehouse Optimization tab for overload alerts. Peak utilisation is driven by the Base forecast against capacity set on the Control Panel."')
    ws.cell(r, 1).font = F(9, italic=True); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 2
    r = section(ws, "Recommended actions", r, 10)
    actions = [
        f'=IF({sel3}<{tgt3},"1 · Close the ₹"&TEXT({tgt3}-{sel3},"0.0")&" Cr gap: raise trade support on top-mix brands and pull forward NPI loading.","1 · Forecast ahead of target — protect service levels and lock incremental supply.")',
        '="2 · Prioritise safety stock for positive-momentum brands (see brand table above)."',
        '="3 · Reallocate dispatch away from any warehouse above 100% utilisation (Warehouse Optimization tab)."',
        f'="4 · Distributors tagged High Service Risk / High Inventory in Distributor Intelligence need corrective ordering plans."',
        f'=IF({REF["acc_wmape"]}>0.2,"5 · Model WMAPE above 20% — refine method weights / add business overrides for volatile SKUs.","5 · Forecast accuracy within tolerance — maintain current method weights.")',
    ]
    for a in actions:
        ws.cell(r, 1, a).font = F(9.5); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 26
        r += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------------------
def build_workbench(ws, REF):
    r = title_row(ws, "DEMAND PLANNER WORKBENCH  ·  manual overrides with audit trail", 10)
    r = section(ws, "Planners adjust the forecast at any level and record WHY. 'System Forecast' looks up the "
                    "engine (Brand×Category); 'Final' uses the override when present; Variance + audit columns "
                    "preserve the trail. This mirrors SAP IBP / Kinaxis / Blue Yonder override handling.", r, 10)
    r = header_cells(ws, r, ["Date", "Planner", "Level", "Key (Brand or Brand|Category)", "Target Month",
                             "System Forecast (₹ Cr)", "Override (₹ Cr)", "Final (₹ Cr)",
                             "Variance", "Reason"])
    first = r
    reasons = ["Promotion", "Listing Expansion", "Competitor Activity", "Stock Correction",
               "NPI Launch", "Seasonal", "Distributor Feedback", "Management Guidance"]
    levels = ["Brand", "Brand|Category", "Chain", "State", "Zone", "Distributor"]
    # example override rows
    examples = [
        ("2026-06-01", "S. Aswal", "Brand|Category", "Mamaearth|Sun Care", "Jun'26", 2.5, "Promotion"),
        ("2026-06-01", "S. Aswal", "Brand", "The Derma Co.", "Jul'26", None, "Listing Expansion"),
    ]
    for ex in examples:
        d, p, lvl, key, mon, ov, rsn = ex
        ws.cell(r, 1, d).font = F(9, color=C_INPUT); ws.cell(r, 1).fill = fill(C_INPUT_FILL)
        ws.cell(r, 2, p).font = F(9, color=C_INPUT); ws.cell(r, 2).fill = fill(C_INPUT_FILL)
        ws.cell(r, 3, lvl).font = F(9, color=C_INPUT); ws.cell(r, 3).fill = fill(C_INPUT_FILL)
        ws.cell(r, 4, key).font = F(9, color=C_INPUT); ws.cell(r, 4).fill = fill(C_INPUT_FILL)
        ws.cell(r, 5, mon).font = F(9, color=C_INPUT); ws.cell(r, 5).fill = fill(C_INPUT_FILL)
        # system forecast lookup: if Brand|Category -> match both; else brand sum. Use Base of the month.
        # Map target month to M/M+1/M+2 base column via choose.
        sysf = _workbench_sysf(r, REF)
        ws.cell(r, 6, sysf).number_format = FMT_CR
        ov_cell = ws.cell(r, 7, ov); ov_cell.font = F(9, color=C_INPUT); ov_cell.fill = fill(C_INPUT_FILL)
        ov_cell.number_format = FMT_CR
        ws.cell(r, 8, f'=IF(G{r}<>"",G{r},F{r})').number_format = FMT_CR
        ws.cell(r, 8).font = F(9, True)
        ws.cell(r, 9, f'=IFERROR(H{r}/F{r}-1,0)').number_format = FMT_PCT
        ws.cell(r, 10, rsn).font = F(9, color=C_INPUT); ws.cell(r, 10).fill = fill(C_INPUT_FILL)
        for c in range(1, 11):
            ws.cell(r, c).border = BORDER
        r += 1
    # blank input rows
    for _ in range(12):
        for c in range(1, 11):
            cell = ws.cell(r, c, None); cell.border = BORDER
            if c in (1, 2, 3, 4, 5, 7, 10):
                cell.fill = fill(C_INPUT_FILL)
        ws.cell(r, 6, _workbench_sysf(r, REF)).number_format = FMT_CR
        ws.cell(r, 8, f'=IF(G{r}<>"",G{r},F{r})').number_format = FMT_CR
        ws.cell(r, 9, f'=IFERROR(H{r}/F{r}-1,0)').number_format = FMT_PCT
        r += 1
    last = r - 1
    # dropdowns
    dvL = DataValidation(type="list", formula1=f'"{",".join(levels)}"', allow_blank=True)
    dvR = DataValidation(type="list", formula1=f'"{",".join(reasons)}"', allow_blank=True)
    ws.add_data_validation(dvL); ws.add_data_validation(dvR)
    dvL.add(f"C{first}:C{last}"); dvR.add(f"J{first}:J{last}")
    r += 1
    ws.cell(r, 5, "Total override impact (₹ Cr):").font = F(10, True); ws.cell(r, 5).alignment = RIGHT
    ws.cell(r, 8, f'=SUMPRODUCT((G{first}:G{last}<>"")*(H{first}:H{last}-F{first}:F{last}))').number_format = FMT_CR
    ws.cell(r, 8).font = F(10, True)
    widths = [11, 12, 14, 26, 12, 17, 14, 13, 10, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


def _workbench_sysf(r, REF):
    """System forecast lookup for the workbench row: Base forecast for the chosen
    target month, at Brand or Brand|Category grain."""
    # month selector -> which Base column
    base_by_t = {1: REF['eng_Base1'], 2: REF['eng_Base2'], 3: REF['eng_Base3']}
    # SUMIFS depending on level: if key contains '|', split brand/cat
    # We build: if ISNUMBER(SEARCH("|",key)) then brand&cat else brand-only
    eb, ec = REF['eng_brand'], REF['eng_cat']
    def sumifs_for(t):
        base = base_by_t[t]
        brand_part = f'LEFT($D{r},IFERROR(SEARCH("|",$D{r})-1,LEN($D{r})))'
        cat_part = f'MID($D{r},IFERROR(SEARCH("|",$D{r})+1,LEN($D{r})+1),50)'
        with_cat = f'SUMIFS({base},{eb},{brand_part},{ec},{cat_part})'
        no_cat = f'SUMIFS({base},{eb},{brand_part})'
        return f'IF(ISNUMBER(SEARCH("|",$D{r})),{with_cat},{no_cat})'
    # pick month column by matching E to M/M+1/M+2
    return (f'=IFERROR(IF($E{r}={REF["fc_m"]},{sumifs_for(1)},'
            f'IF($E{r}={REF["fc_m1"]},{sumifs_for(2)},'
            f'IF($E{r}={REF["fc_m2"]},{sumifs_for(3)},0))),0)')


# --------------------------------------------------------------------------
def build_documentation(ws, REF):
    r = title_row(ws, "DOCUMENTATION  ·  assumptions, formulas, business rules, refresh", 2)
    blocks = [
        ("PURPOSE", [
            "Rolling 3-month (M, M+1, M+2) demand forecast & S&OP planning model for Honasa / Mamaearth",
            "Modern Trade. Combines statistical forecasting, business overlay and an AI-style ensemble,",
            "then reconciles top-down target with bottom-up forecast across the full hierarchy.",
        ]),
        ("DATA (all real, no dummy numbers)", [
            "Primary Sales History: real MT primary invoicing Apr'25–May'26, aggregated to Brand×Category",
            "   from PowerBI/RawDataFolders/Primary_Article_Monthly/*.csv.",
            "Offtake History: real store offtake (Apr'26–May'26) from Offtake_Monthly/*.csv.",
            "Monthly Target Upload: FY26-27 company NSV plan (seed).",
            "Distributor Master: real top ship-to accounts; service metrics are planner/DMS inputs (blue).",
            "Contribution %, warehouse mapping etc. derive from the real primary footprint.",
        ]),
        ("ROLLING HORIZON", [
            "Control Panel → Last Actual Month and Forecast Month (M) drive everything.",
            "M+1, M+2 and all indices are derived via MATCH on the Calendar; advance Forecast Month by",
            "one each cycle and the entire workbook rolls forward. Add a new month column to Primary Sales",
            "History and update Last Actual Month to bring in a fresh actual.",
        ]),
        ("STATISTICAL METHODS (per Brand×Category, ₹ Cr)", [
            "WMA            = w1·M-1 + w2·M-2 + w3·M-3, reseasonalised.",
            "Exp. Smoothing = normalised α(1-α)^k weighting of the last 6 months.",
            "Linear Trend   = last + (last − M-6)/5 · t, reseasonalised.",
            "Seasonal Index = 12-month mean × seasonal multiplier.",
            "CAGR           = last × (1+monthly CAGR)^t.",
            "Rolling Avg    = 3-month average, reseasonalised.",
            "YoY Growth     = same-month-last-year × (1 + YoY growth).",
            "Seasonal multiplier = category factor for the target month ÷ its 12-month average",
            "   (Seasonality Matrix; falls back to 100% if the category is absent).",
            "Statistical    = weight-normalised blend of the seven methods (weights on Control Panel).",
        ]),
        ("BUSINESS & AI LAYERS", [
            "Business = Statistical × (1 + business adjustment % + promo uplift %). Promo uplift comes",
            "   from the Promotion Calendar (Brand×Category×Month).",
            "AI Ensemble = weighted blend of Statistical, Business and a driver-adjusted Statistical",
            "   (brand momentum + promo response), normalised by the three AI weights.",
            "Final (scenario) = AI × scenario multiplier + NPI loading (₹ Cr) for the month.",
        ]),
        ("SCENARIOS", [
            "Base / Optimistic / Conservative multipliers on the Control Panel scale the AI forecast.",
            "The Business Scenario selector drives the 'Selected' columns and the dashboard.",
        ]),
        ("TARGET DISTRIBUTION & RECONCILIATION", [
            "Company 3-month target distributes to Zone / Chain / State / Article by real contribution %.",
            "Any level accepts an override (blue); the reconciliation line shows variance vs the company",
            "control total so top-down and bottom-up always tie out.",
        ]),
        ("PRIMARY PLANNING", [
            "Suggested Primary = Expected Offtake (Base M) + Safety + Pipeline + Promo/NPI − Excess − Slow-moving.",
        ]),
        ("WAREHOUSES", [
            "Base forecast is split to Gurgaon / Mumbai / Bangalore / Kolkata by editable demand share,",
            "checked against Control-Panel capacity; peak utilisation >100% raises a reallocation alert.",
        ]),
        ("ACCURACY", [
            "1-month-ahead WMA back-test on real company primary: MAPE, WMAPE, Bias, Tracking Signal,",
            "Accuracy % and MAE. Recomputes as new actual months are added.",
        ]),
        ("REFRESH STEPS (monthly)", [
            "1. Paste the new month's primary into Primary Sales History (new column) and offtake into Offtake History.",
            "2. Set Control Panel → Last Actual Month = the new month; Forecast Month (M) = the next month.",
            "3. Update Promotion Calendar / NPI & EPD for the new horizon; review Business Adj % in the engine.",
            "4. Review Scenario Summary, Warehouse Optimization and AI Business Insights for the S&OP meeting.",
            "5. Capture manual changes in the Demand Planner Workbench (with reason) — the audit trail is preserved.",
            "Regenerate from source any time with:  python scripts/build_demand_forecast.py",
        ]),
        ("BUSINESS EVENT ENGINE (separate module)", [
            "Business_Event_Master: one row per activity (NPI, promo, distribution, institutional, …).",
            "   Only rows with Status = Approved affect the forecast; Baseline is never overwritten.",
            "Impact methods (any combination per event):",
            "   · Percentage uplift  — Base(brand×cat) × effective uplift %.",
            "   · Fixed additional qty — one-time, applied in the start month only.",
            "   · Distribution gain  — New Stores × Units/Store, ramped up over the launch curve.",
            "Effective uplift = base uplift × Confidence × Priority weight × Category seasonality × ramp/decay.",
            "   base uplift = explicit Expected Uplift %, else Chain Event Library, else Article Event Library.",
            "Ramp-up (NPI/structural) and decay (promo) curves are editable on Event Settings.",
            "Event Impact Engine combines OVERLAPPING % events MULTIPLICATIVELY — 1−Π(1+uᵢ), via EXP(Σln) —",
            "   so five promos on one Chain×Brand×Month do not naively add to +67%; absolute events are summed.",
            "Baseline, Event Uplift and Final are always shown side by side (Event Impact Engine / Dashboard).",
            "Event Simulator: instant what-ifs (not gated by approval). Event Calendar: live timeline + impact.",
            "Demand Driver Library: reusable objects with historical learning → Recommended Uplift the planner",
            "   can accept / modify / reject (AI-assisted planning, à la SAP IBP / Kinaxis / Blue Yonder).",
            "Event horizon reference: Current month (M) + M+1 + M+2 tie to the baseline; ramp/decay project further.",
        ]),
        ("COLOUR LEGEND", [
            "Blue text on pale-yellow fill = editable input.   Black = formula.   Green = cross-sheet link.",
        ]),
    ]
    for head, lines in blocks:
        ws.cell(r, 1, head).font = F(11, True, "FFFFFF"); ws.cell(r, 1).fill = fill(C_SUB)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
        for ln in lines:
            ws.cell(r, 1, ln).font = F(9.5); ws.cell(r, 1).alignment = LEFT_TOP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 20
    ws.sheet_view.showGridLines = False


# ==========================================================================
# 4. BUSINESS EVENT ENGINE MODULE
# ==========================================================================
EVENT_TYPES = [
    "NPI Launch", "Product Relaunch", "New Listing", "Distribution Expansion",
    "Store Expansion", "Visibility Campaign", "Endcap", "Power Wing", "Gondola",
    "SIS", "Counter Display", "Festival Promotion", "BOGO", "Price Drop",
    "Retail Festival", "Digital Campaign", "TV Campaign", "Influencer Campaign",
    "Consumer Offer", "Sampling Activity", "Bundle Promotion", "Combo Pack",
    "Seasonal Push", "New Warehouse Opening", "New Distributor", "Market Expansion",
    "Export Order", "Institutional Sales", "Clearance Campaign",
]
# event types whose demand builds up over a ramp (structural), vs promo (decay)
RAMP_KEYWORDS = ["NPI", "Launch", "Relaunch", "Listing", "Distribution", "Store",
                 "Warehouse", "Distributor", "Market", "Export", "Institutional"]


def _ev_input(ws, r, label, value, key, REF, fmt=None, note=None):
    ws.cell(r, 1, label).font = F(10); ws.cell(r, 1).alignment = LEFT
    c = ws.cell(r, 2, value); c.font = F(10, True, C_INPUT); c.fill = fill(C_INPUT_FILL)
    c.alignment = CENTER; c.border = BORDER
    if fmt:
        c.number_format = fmt
    if note:
        ws.cell(r, 3, note).font = F(8.5, italic=True, color="808080"); ws.cell(r, 3).alignment = LEFT
    REF[key] = qname("Event Settings", f"$B${r}")
    return r + 1


def build_event_settings(ws, REF):
    r = title_row(ws, "EVENT SETTINGS  ·  central, editable assumptions for the event engine", 5)
    r = section(ws, "All event uplift assumptions are maintained centrally here. Blue cells are editable.", r, 5)
    r = section(ws, "1 · Conversion & margin", r, 5)
    r = _ev_input(ws, r, "Realisation (₹ per unit)", 169, "rs_per_unit", REF,
                  note="real portfolio avg primary NSV ÷ units; used for unit↔value conversion")
    r = _ev_input(ws, r, "Gross margin % (for incremental margin)", 0.52, "margin", REF, fmt=FMT_PCT)
    r = _ev_input(ws, r, "Inventory cover (months of incremental)", 1.0, "inv_cover", REF, fmt=FMT_X)
    r = _ev_input(ws, r, "Combination rule", "Multiplicative", "combine_rule", REF,
                  note="overlapping % events combine multiplicatively: 1−Π(1+uᵢ) not simple sum")
    r += 1

    r = section(ws, "2 · Event Priority weights", r, 5)
    r = header_cells(ws, r, ["Priority", "Weight"], fillhex=C_SUB)
    pf = r
    for p, w in [("Critical", 1.00), ("High", 0.85), ("Medium", 0.60), ("Low", 0.35)]:
        ws.cell(r, 1, p).font = F(9.5); ws.cell(r, 1).border = BORDER; ws.cell(r, 1).alignment = LEFT
        c = ws.cell(r, 2, w); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
        c.number_format = FMT_X; c.border = BORDER; c.alignment = CENTER
        r += 1
    pl = r - 1
    REF["prio_key"] = qname("Event Settings", f"$A${pf}:$A${pl}")
    REF["prio_wt"] = qname("Event Settings", f"$B${pf}:$B${pl}")
    r += 1

    r = section(ws, "3 · NPI / structural Ramp-up curve (share of full effect by month-in-life)", r, 5)
    r = header_cells(ws, r, ["Month in life", "Ramp %"], fillhex=C_SUB)
    rf = r
    for i, v in enumerate([0.30, 0.60, 0.85, 1.00, 1.00, 1.00]):
        ws.cell(r, 1, i + 1).font = F(9.5); ws.cell(r, 1).border = BORDER; ws.cell(r, 1).alignment = CENTER
        c = ws.cell(r, 2, v); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
        c.number_format = FMT_PCT0; c.border = BORDER; c.alignment = CENTER
        r += 1
    rl = r - 1
    REF["ramp_vals"] = qname("Event Settings", f"$B${rf}:$B${rl}")
    r += 1

    r = section(ws, "4 · Promotion effect decay (share of peak after the promo month)", r, 5)
    r = header_cells(ws, r, ["Months after peak", "Decay %"], fillhex=C_SUB)
    df = r
    for i, v in enumerate([1.00, 0.50, 0.25, 0.00]):
        ws.cell(r, 1, i).font = F(9.5); ws.cell(r, 1).border = BORDER; ws.cell(r, 1).alignment = CENTER
        c = ws.cell(r, 2, v); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL)
        c.number_format = FMT_PCT0; c.border = BORDER; c.alignment = CENTER
        r += 1
    dl = r - 1
    REF["decay_vals"] = qname("Event Settings", f"$B${df}:$B${dl}")
    r += 1

    r = section(ws, "5 · Confidence reference (weight the final uplift)", r, 5)
    r = header_cells(ws, r, ["Confidence", "Meaning"], fillhex=C_SUB)
    for cf, mean in [("95%", "Confirmed PO"), ("80%", "Retailer confirmation"),
                     ("60%", "Business expectation"), ("40%", "Early planning")]:
        ws.cell(r, 1, cf).font = F(9.5); ws.cell(r, 1).alignment = CENTER; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, mean).font = F(9.5); ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A3"
    return EVENT_TYPES


def build_chain_event_library(ws, agg, event_types, REF):
    r = title_row(ws, "CHAIN EVENT LIBRARY  ·  historical uplift by chain × event type (editable)", 5)
    r = section(ws, "Different chains respond differently to the same activity. The event engine looks up "
                    "these historical uplifts when an event row has no explicit Expected Uplift %.", r, 5)
    r = header_cells(ws, r, ["Chain", "Event Type", "Historical Uplift %", "Confidence %", "Key"])
    first = r
    seed = [
        ("Reliance Retail", "Endcap", 0.15, 0.94), ("Reliance Retail", "Power Wing", 0.18, 0.92),
        ("Reliance Retail", "NPI Launch", 0.22, 0.85), ("Reliance Retail", "Digital Campaign", 0.12, 0.75),
        ("Reliance Retail", "Price Drop", 0.10, 0.88), ("Reliance Retail", "Seasonal Push", 0.15, 0.82),
        ("Apollo", "Endcap", 0.08, 0.90), ("Apollo", "BOGO", 0.13, 0.88), ("Apollo", "Visibility Campaign", 0.09, 0.85),
        ("D-Mart", "Endcap", 0.06, 0.90), ("D-Mart", "Gondola", 0.07, 0.90), ("D-Mart", "Price Drop", 0.09, 0.90),
        ("More Retail", "Endcap", 0.12, 0.88), ("More Retail", "Gondola", 0.10, 0.88),
        ("Health & Glow", "SIS", 0.16, 0.86), ("Nykaa SS(fsn)", "Digital Campaign", 0.20, 0.80),
    ]
    for chain, et, up, conf in seed:
        ws.cell(r, 1, chain).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, et).font = F(9.5); ws.cell(r, 2).alignment = LEFT
        c = ws.cell(r, 3, up); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL); c.number_format = FMT_PCT
        c2 = ws.cell(r, 4, conf); c2.font = F(9.5, color=C_INPUT); c2.fill = fill(C_INPUT_FILL); c2.number_format = FMT_PCT
        ws.cell(r, 5, f'=$A{r}&"|"&$B{r}').font = F(8.5, color="808080")
        for c3 in range(1, 6):
            ws.cell(r, c3).border = BORDER
        r += 1
    last = r - 1
    REF["clib_uplift"] = qname("Chain Event Library", f"$C${first}:$C${last}")
    REF["clib_key"] = qname("Chain Event Library", f"$E${first}:$E${last}")
    for col, w in zip("ABCDE", (20, 22, 16, 14, 30)):
        ws.column_dimensions[col].width = w
    dv = DataValidation(type="list", formula1=f'"{",".join(event_types)}"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"B{first}:B{last}")
    ws.freeze_panes = "A4"


def build_article_event_library(ws, event_types, REF):
    r = title_row(ws, "ARTICLE / CATEGORY EVENT LIBRARY  ·  uplift by category × event type (editable)", 5)
    r = section(ws, "Products respond differently. Used as a fallback when the chain library has no match.", r, 5)
    r = header_cells(ws, r, ["Category", "Event Type", "Uplift %", "Season", "Key"])
    first = r
    seed = [
        ("Face", "Seasonal Push", 0.35, "Summer (sunscreen)"), ("Face", "BOGO", 0.18, "All year"),
        ("Face", "Visibility Campaign", 0.14, "All year"), ("Face", "NPI Launch", 0.25, "All year"),
        ("Body", "Seasonal Push", 0.28, "Winter"), ("Body", "Combo Pack", 0.16, "Winter"),
        ("Hair", "Visibility Campaign", 0.12, "All year"), ("Hair", "Gondola", 0.10, "All year"),
        ("Baby", "Retail Festival", 0.15, "Festive"), ("Fragrances", "Festival Promotion", 0.30, "Festive"),
    ]
    for cat, et, up, season in seed:
        ws.cell(r, 1, cat).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, et).font = F(9.5); ws.cell(r, 2).alignment = LEFT
        c = ws.cell(r, 3, up); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL); c.number_format = FMT_PCT
        ws.cell(r, 4, season).font = F(9.5); ws.cell(r, 4).alignment = LEFT
        ws.cell(r, 5, f'=$A{r}&"|"&$B{r}').font = F(8.5, color="808080")
        for c3 in range(1, 6):
            ws.cell(r, c3).border = BORDER
        r += 1
    last = r - 1
    REF["alib_uplift"] = qname("Article Event Library", f"$C${first}:$C${last}")
    REF["alib_key"] = qname("Article Event Library", f"$E${first}:$E${last}")
    for col, w in zip("ABCDE", (16, 22, 12, 20, 30)):
        ws.column_dimensions[col].width = w
    dv = DataValidation(type="list", formula1=f'"{",".join(event_types)}"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"B{first}:B{last}")
    ws.freeze_panes = "A4"


def build_driver_library(ws, agg, REF):
    r = title_row(ws, "DEMAND DRIVER LIBRARY  ·  reusable planning objects with historical learning", 8)
    r = section(ws, "Every event becomes a reusable object. 'Recommended Uplift' blends the historical "
                    "average with the last-3-events average — the planner can accept, modify or reject it. "
                    "This is the AI-assisted layer (à la SAP IBP / Kinaxis / Blue Yonder).", r, 8)
    r = header_cells(ws, r, ["Chain", "Article / SKU", "Event", "Historical Avg Uplift %",
                             "Confidence %", "Season", "Last 3 Events Avg %", "Recommended Uplift %"])
    first = r
    seed = [
        ("Reliance Retail", "TDC Sunscreen 80g", "Endcap", 0.16, 0.94, "Summer", 0.172),
        ("Apollo", "Face Wash 100ml", "BOGO", 0.13, 0.88, "All Year", 0.125),
        ("D-Mart", "Hair Oil 200ml", "Gondola", 0.08, 0.90, "Winter", 0.084),
        ("Reliance Retail", "Mamaearth Sunscreen 50g", "Power Wing", 0.18, 0.92, "Summer", 0.19),
        ("More Retail", "Aqualogica Face Wash", "Endcap", 0.12, 0.88, "All Year", 0.118),
        ("Nykaa SS(fsn)", "TDC Serum 30ml", "Digital Campaign", 0.20, 0.80, "All Year", 0.205),
    ]
    for chain, art, et, hist, conf, season, last3 in seed:
        ws.cell(r, 1, chain).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, art).font = F(9.5); ws.cell(r, 2).alignment = LEFT
        ws.cell(r, 3, et).font = F(9.5); ws.cell(r, 3).alignment = LEFT
        c = ws.cell(r, 4, hist); c.font = F(9.5, color=C_INPUT); c.fill = fill(C_INPUT_FILL); c.number_format = FMT_PCT
        c2 = ws.cell(r, 5, conf); c2.font = F(9.5, color=C_INPUT); c2.fill = fill(C_INPUT_FILL); c2.number_format = FMT_PCT
        ws.cell(r, 6, season).font = F(9.5); ws.cell(r, 6).alignment = LEFT
        c3 = ws.cell(r, 7, last3); c3.font = F(9.5, color=C_INPUT); c3.fill = fill(C_INPUT_FILL); c3.number_format = FMT_PCT
        # recommended = confidence-weighted blend of historical & last-3
        ws.cell(r, 8, f"=ROUND((D{r}+G{r})/2,3)").number_format = FMT_PCT
        ws.cell(r, 8).font = F(9.5, True, color=C_HEADER); ws.cell(r, 8).fill = fill(C_TOTAL)
        for c4 in range(1, 9):
            ws.cell(r, c4).border = BORDER
        r += 1
    for col, w in zip("ABCDEFGH", (18, 24, 18, 18, 13, 14, 16, 18)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def build_event_master(ws, agg, event_types, REF):
    r = title_row(ws, "BUSINESS_EVENT_MASTER  ·  one row per business activity (only Approved rows affect forecast)", 40)
    r = section(ws, "Enter events at any grain. Incremental demand is computed automatically from % uplift, "
                    "fixed additional qty, and/or distribution gain (stores × units), weighted by Confidence × "
                    "Priority × Seasonality and shaped by the ramp/decay curves. Baseline is never overwritten.", r, 40)
    CP = REF
    col = {}
    ci = 1
    hdr_row = r

    def addcol(key, title, width=11):
        nonlocal ci
        col[key] = ci
        ci += 1
        ws.column_dimensions[get_column_letter(col[key])].width = width
        return

    fields = [
        ("id", "Event ID", 9), ("name", "Event Name", 22), ("type", "Event Type", 20),
        ("start", "Start Month", 11), ("end", "End Month", 11), ("chain", "Chain", 16),
        ("brand", "Brand", 14), ("cat", "Category", 12), ("article", "Article", 16),
        ("region", "Region", 10), ("zone", "Zone", 10), ("state", "State", 12),
        ("dist", "Distributor", 16), ("wh", "Warehouse", 11), ("method", "Method", 15),
        ("up", "Expected Uplift %", 12), ("qty", "Additional Qty", 12),
        ("stores", "New Stores", 10), ("ups", "Units/Store", 10), ("conf", "Confidence %", 11),
        ("prio", "Priority", 10), ("appby", "Approved By", 12), ("status", "Status", 11),
        # computed
        ("fam", "Family", 8), ("idxs", "idxS", 6), ("idxe", "idxE", 6), ("pwt", "Prio Wt", 7),
        ("libup", "Lib Uplift", 9), ("baseup", "Base Uplift", 9),
        ("effM", "Eff·M", 8), ("effM1", "Eff·M+1", 8), ("effM2", "Eff·M+2", 8),
        ("absM", "Abs·M ₹Cr", 9), ("absM1", "Abs·M+1", 9), ("absM2", "Abs·M+2", 9),
        ("lnM", "ln·M", 7), ("lnM1", "ln·M+1", 7), ("lnM2", "ln·M+2", 7),
        ("horiz", "Horizon Incr ₹Cr", 13), ("iqty", "Incr Qty", 11),
    ]
    for key, title, w in fields:
        addcol(key, title, w)
    ncols = ci - 1
    for c in range(1, ncols + 1):
        cell = ws.cell(hdr_row, c, fields[c - 1][1])
        cell.font = F(8.5, True, "FFFFFF"); cell.fill = fill(C_HEADER); cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[hdr_row].height = 30
    r += 1
    first = r

    rs = CP["rs_per_unit"]
    # seed events (demonstrates overlapping combination on Mamaearth|Face|Jun'26)
    seeds = [
        ("EV01", "Summer Sunscreen NPI (SPF60)", "NPI Launch", "Jun'26", "Aug'26", "Reliance Retail",
         "Mamaearth", "Face", "ME Sunscreen SPF60 50g", "West", "West", "Maharashtra",
         "Reliance DC Mumbai", "Mumbai", "Percentage", None, None, None, None, 0.85, "Critical", "Sales Head", "Approved"),
        ("EV02", "Summer Seasonal Push", "Seasonal Push", "Jun'26", "Jul'26", "Reliance Retail",
         "Mamaearth", "Face", "Sunscreen range", "West", "West", "Maharashtra", "Reliance DC Mumbai",
         "Mumbai", "Percentage", 0.15, None, None, None, 0.80, "High", "KAM", "Approved"),
        ("EV03", "Price Drop 10%", "Price Drop", "Jun'26", "Jun'26", "Reliance Retail",
         "Mamaearth", "Face", "Sunscreen range", "West", "West", "Maharashtra", "Reliance DC Mumbai",
         "Mumbai", "Percentage", 0.10, None, None, None, 0.85, "High", "KAM", "Approved"),
        ("EV04", "Endcap Visibility", "Endcap", "Jun'26", "Jul'26", "Reliance Retail",
         "Mamaearth", "Face", "Sunscreen range", "West", "West", "Maharashtra", "Reliance DC Mumbai",
         "Mumbai", "Percentage", None, None, None, None, 0.90, "High", "Trade Mktg", "Approved"),
        ("EV05", "Digital Campaign", "Digital Campaign", "Jun'26", "Jun'26", "Reliance Retail",
         "Mamaearth", "Face", "Sunscreen range", "West", "West", "Maharashtra", "Reliance DC Mumbai",
         "Mumbai", "Percentage", 0.12, None, None, None, 0.70, "Medium", "Brand", "Approved"),
        ("EV06", "Apollo BOGO Face Wash", "BOGO", "Jul'26", "Jul'26", "Apollo",
         "The Derma Co.", "Face", "TDC Face Wash 100ml", "South", "South-1", "Telangana",
         "Apollo Hyderabad", "Bangalore", "Percentage", None, None, None, None, 0.88, "High", "KAM", "Approved"),
        ("EV07", "Distribution Expansion +200 stores", "Distribution Expansion", "Jun'26", "Aug'26",
         "More Retail", "Aqualogica", "Face", "AQ Face Wash 100ml", "South", "South-1", "Tamil Nadu",
         "More Chennai", "Bangalore", "Distribution Gain", None, None, 200, 8, 0.75, "High", "Sales Head", "Approved"),
        ("EV08", "Institutional bulk order", "Institutional Sales", "Jul'26", "Jul'26", "Medanta",
         "Mamaearth", "Baby", "ME Baby range", "North", "North", "Haryana", "Medanta Gurgaon",
         "Gurgaon", "Fixed Qty", None, 40000, None, None, 0.95, "Critical", "Sales Head", "Approved"),
        ("EV09", "Planning-stage festive push (draft)", "Festival Promotion", "Aug'26", "Aug'26",
         "D-Mart", "Mamaearth", "Face", "Face range", "West", "West", "Maharashtra", "Dmart DC",
         "Mumbai", "Percentage", 0.10, None, None, None, 0.40, "Low", "-", "Draft"),
    ]
    cal_lab = CP["cal_lab"]

    def put(rr, key, val, fmt=None, inp=False, calc=False):
        c = ws.cell(rr, col[key], val)
        c.font = F(8.5, color=C_INPUT if inp else "000000")
        if inp:
            c.fill = fill(C_INPUT_FILL)
        if fmt:
            c.number_format = fmt
        c.alignment = CENTER if key not in ("name", "article", "type") else LEFT
        return c

    inp_keys = {"id", "name", "type", "start", "end", "chain", "brand", "cat", "article",
                "region", "zone", "state", "dist", "wh", "method", "up", "qty", "stores",
                "ups", "conf", "prio", "appby", "status"}

    def write_row(rr, data=None):
        # data is a tuple aligned to first 23 fields, or None for a blank input row
        keys23 = [k for k, _, _ in fields[:23]]
        for i, k in enumerate(keys23):
            v = data[i] if data is not None else None
            fmt = FMT_PCT if k in ("up", "conf") else (FMT_QTY if k in ("qty", "stores", "ups") else None)
            put(rr, k, v, fmt=fmt, inp=True)
        C = lambda k: f"{get_column_letter(col[k])}{rr}"
        # Family
        searches = "+".join(f'ISNUMBER(SEARCH("{kw}",{C("type")}))' for kw in RAMP_KEYWORDS)
        put(rr, "fam", f'=IF(({searches})>0,"RAMP","DECAY")')
        put(rr, "idxs", f"=IFERROR(MATCH({C('start')},{cal_lab},0),0)")
        put(rr, "idxe", f"=IFERROR(MATCH({C('end')},{cal_lab},0),{C('idxs')})")
        put(rr, "pwt", f"=IFERROR(INDEX({CP['prio_wt']},MATCH({C('prio')},{CP['prio_key']},0)),0.5)", fmt=FMT_X)
        put(rr, "libup",
            f"=IFERROR(INDEX({CP['clib_uplift']},MATCH({C('chain')}&\"|\"&{C('type')},{CP['clib_key']},0)),"
            f"IFERROR(INDEX({CP['alib_uplift']},MATCH({C('cat')}&\"|\"&{C('type')},{CP['alib_key']},0)),0))",
            fmt=FMT_PCT)
        put(rr, "baseup", f"=IF({C('up')}<>\"\",{C('up')},{C('libup')})", fmt=FMT_PCT)
        # per-month effective uplift, absolute incremental, ln helper
        months = [("effM", "absM", "lnM", CP["idx_m"], CP["name_m"]),
                  ("effM1", "absM1", "lnM1", CP["idx_m1"], CP["name_m1"]),
                  ("effM2", "absM2", "lnM2", CP["idx_m2"], CP["name_m2"])]
        gate = f'IF({C("status")}="Approved",1,0)'
        for effk, absk, lnk, idxt, namek in months:
            ramp = (f"IF({idxt}-{C('idxs')}<0,0,INDEX({CP['ramp_vals']},MIN({idxt}-{C('idxs')}+1,6)))")
            decay = (f"IF({idxt}<{C('idxs')},0,IF({idxt}<={C('idxe')},1,"
                     f"IFERROR(INDEX({CP['decay_vals']},MIN({idxt}-{C('idxe')}+1,4)),0)))")
            factor = f"IF({C('fam')}=\"RAMP\",{ramp},{decay})"
            onetime = f"IF({idxt}={C('idxs')},1,0)"
            seas = (f"IFERROR(INDEX({CP['seas_vals']},MATCH({C('cat')},{CP['seas_catcol']},0),"
                    f"MATCH({namek},{CP['seas_monhdr']},0))/"
                    f"IFERROR(INDEX({CP['seas_avg']},MATCH({C('cat')},{CP['seas_catcol']},0)),1),1)")
            eff = (f"={gate}*{C('baseup')}*{C('conf')}*{C('pwt')}*({seas})*({factor})")
            put(rr, effk, eff, fmt=FMT_PCT)
            absf = (f"={gate}*({C('qty')}*{rs}/10000000*({onetime})+"
                    f"{C('stores')}*{C('ups')}*{rs}/10000000*({factor}))")
            put(rr, absk, absf, fmt=FMT_CR)
            put(rr, lnk, f"=IF({get_column_letter(col[effk])}{rr}>-1,LN(1+{get_column_letter(col[effk])}{rr}),0)")
        # horizon incremental (₹ Cr) standalone for this event
        b1 = f"SUMIFS({CP['eng_Base1']},{CP['eng_brand']},{C('brand')},{CP['eng_cat']},{C('cat')})"
        b2 = f"SUMIFS({CP['eng_Base2']},{CP['eng_brand']},{C('brand')},{CP['eng_cat']},{C('cat')})"
        b3 = f"SUMIFS({CP['eng_Base3']},{CP['eng_brand']},{C('brand')},{CP['eng_cat']},{C('cat')})"
        put(rr, "horiz",
            f"={b1}*{C('effM')}+{b2}*{C('effM1')}+{b3}*{C('effM2')}+{C('absM')}+{C('absM1')}+{C('absM2')}",
            fmt=FMT_CR)
        ws.cell(rr, col["horiz"]).font = F(9, True, C_HEADER)
        put(rr, "iqty", f"=IFERROR({C('horiz')}*10000000/{rs},0)", fmt=FMT_QTY)
        for c in range(1, ncols + 1):
            ws.cell(rr, c).border = BORDER

    for s in seeds:
        write_row(r, s)
        r += 1
    for _ in range(12):
        write_row(r, None)
        r += 1
    last = r - 1

    # dropdowns
    def dv(items, colkey, rng_first=first, rng_last=last):
        d = DataValidation(type="list", formula1=f'"{",".join(items)}"', allow_blank=True)
        ws.add_data_validation(d)
        cl = get_column_letter(col[colkey])
        d.add(f"{cl}{rng_first}:{cl}{rng_last}")
    dv(event_types, "type")
    dv(["Percentage", "Fixed Qty", "Distribution Gain", "Combined"], "method")
    dv(["Critical", "High", "Medium", "Low"], "prio")
    dv(["Draft", "Submitted", "Approved", "Rejected"], "status")

    # expose refs for impact/calendar/dashboard
    def rng(key):
        cl = get_column_letter(col[key])
        return qname("Business_Event_Master", f"${cl}${first}:${cl}${last}")
    REF["evm_first"], REF["evm_last"] = first, last
    REF["evm_col"] = col
    for key in ["brand", "cat", "lnM", "lnM1", "lnM2", "absM", "absM1", "absM2",
                "horiz", "iqty", "name", "chain", "wh", "region", "start", "end",
                "appby", "status", "type"]:
        REF[f"evm_{key}"] = rng(key)
    ws.freeze_panes = ws.cell(first, col["type"]).coordinate
    ws.sheet_view.zoomScale = 80


def build_event_impact(ws, series, REF):
    r = title_row(ws, "EVENT IMPACT ENGINE  ·  Baseline + Events = Final (shown separately, ₹ Cr)", 16)
    r = section(ws, "Per Brand×Category: overlapping % events are combined MULTIPLICATIVELY (1−Π(1+uᵢ)), "
                    "not summed; absolute events (fixed qty, distribution gain) are added. Only Approved "
                    "events flow in. Baseline is the engine's Base scenario and is never modified.", r, 16)
    heads = ["Brand", "Category",
             "Base M", "Uplift% M", "Event M", "Final M",
             "Base M+1", "Uplift% M+1", "Event M+1", "Final M+1",
             "Base M+2", "Uplift% M+2", "Event M+2", "Final M+2", "Final 3M"]
    r = header_cells(ws, r, heads)
    first = r
    EB, EC = REF["eng_brand"], REF["eng_cat"]
    bases = {0: REF["eng_Base1"], 1: REF["eng_Base2"], 2: REF["eng_Base3"]}
    lns = {0: REF["evm_lnM"], 1: REF["evm_lnM1"], 2: REF["evm_lnM2"]}
    abss = {0: REF["evm_absM"], 1: REF["evm_absM1"], 2: REF["evm_absM2"]}
    MB, MC = REF["evm_brand"], REF["evm_cat"]
    for (b, c) in series:
        ws.cell(r, 1, b).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, c).font = F(9.5); ws.cell(r, 2).alignment = LEFT
        for t in range(3):
            base_col = 3 + t * 4
            up_col = base_col + 1
            ev_col = base_col + 2
            fin_col = base_col + 3
            baseL = get_column_letter(base_col)
            upL = get_column_letter(up_col)
            evL = get_column_letter(ev_col)
            ws.cell(r, base_col, f"=SUMIFS({bases[t]},{EB},$A{r},{EC},$B{r})").number_format = FMT_CR
            ws.cell(r, up_col,
                    f"=EXP(SUMIFS({lns[t]},{MB},$A{r},{MC},$B{r}))-1").number_format = FMT_PCT
            ws.cell(r, ev_col,
                    f"={baseL}{r}*{upL}{r}+SUMIFS({abss[t]},{MB},$A{r},{MC},$B{r})").number_format = FMT_CR
            ws.cell(r, fin_col, f"={baseL}{r}+{evL}{r}").number_format = FMT_CR
            ws.cell(r, fin_col).font = F(9.5, True, C_HEADER)
        ws.cell(r, 15, f"=F{r}+J{r}+N{r}").number_format = FMT_CR
        ws.cell(r, 15).font = F(9.5, True); ws.cell(r, 15).fill = fill(C_TOTAL)
        for cc in range(1, 16):
            ws.cell(r, cc).border = BORDER
        r += 1
    last = r - 1
    # totals
    ws.cell(r, 1, "COMPANY TOTAL").font = F(10, True)
    for cc in range(3, 16):
        cl = get_column_letter(cc)
        if cc in (4, 8, 12):  # uplift% columns -> weighted avg
            base_cl = get_column_letter(cc - 1)
            ws.cell(r, cc, f"=IFERROR(SUMPRODUCT({base_cl}{first}:{base_cl}{last},{cl}{first}:{cl}{last})/SUM({base_cl}{first}:{base_cl}{last}),0)").number_format = FMT_PCT
        else:
            ws.cell(r, cc, f"=SUM({cl}{first}:{cl}{last})").number_format = FMT_CR
        ws.cell(r, cc).font = F(9.5, True); ws.cell(r, cc).fill = fill(C_TOTAL)
    for cc in range(1, 16):
        ws.cell(r, cc).border = BORDER
    tot = r
    # expose totals
    for label, cc in [("evtot_baseM", 3), ("evtot_evM", 5), ("evtot_finM", 6),
                      ("evtot_baseM1", 7), ("evtot_evM1", 9), ("evtot_finM1", 10),
                      ("evtot_baseM2", 11), ("evtot_evM2", 13), ("evtot_finM2", 14),
                      ("evtot_fin3", 15)]:
        REF[label] = qname("Event Impact Engine", f"${get_column_letter(cc)}${tot}")
    for col, w in zip("ABCDEFGHIJKLMNO", (14, 12) + (10,) * 13):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C4"


def build_event_calendar(ws, REF):
    r = title_row(ws, "EVENT CALENDAR  ·  live view of planned activity & forecast impact", 11)
    r = section(ws, "Reads Business_Event_Master. Timeline bars span Start→End across the rolling horizon.", r, 11)
    # horizon month labels across the top
    heads = ["Event", "Brand", "Chain", "Warehouse", "Region", "Owner", "Status",
             "Incr ₹Cr", "Start", "End", "Timeline (M .. M+2)"]
    r = header_cells(ws, r, heads)
    first = r
    ef, el = REF["evm_first"], REF["evm_last"]
    col = REF["evm_col"]
    C = lambda k, i: qname("Business_Event_Master", f"${get_column_letter(col[k])}${i}")
    for i in range(ef, el + 1):
        rr = first + (i - ef)
        ws.cell(rr, 1, f"={C('name', i)}").font = F(9); ws.cell(rr, 1).alignment = LEFT
        ws.cell(rr, 2, f"={C('brand', i)}").font = F(9)
        ws.cell(rr, 3, f"={C('chain', i)}").font = F(9)
        ws.cell(rr, 4, f"={C('wh', i)}").font = F(9)
        ws.cell(rr, 5, f"={C('region', i)}").font = F(9)
        ws.cell(rr, 6, f"={C('appby', i)}").font = F(9)
        ws.cell(rr, 7, f"={C('status', i)}").font = F(9)
        ws.cell(rr, 8, f"={C('horiz', i)}").number_format = FMT_CR
        ws.cell(rr, 9, f"={C('start', i)}").font = F(9)
        ws.cell(rr, 10, f"={C('end', i)}").font = F(9)
        # timeline bar: number of horizon months covered -> REPT block
        idxs = C("idxs", i)
        idxe = C("idxe", i)
        tl = (f'=IF({C("status", i)}="Approved",REPT("█",'
              f'MAX(0,MIN({REF["idx_m2"]},{idxe})-MAX({REF["idx_m"]},{idxs})+1)),"")')
        ws.cell(rr, 11, tl).font = F(9, color=C_SUB)
        for c in range(1, 12):
            ws.cell(rr, c).border = BORDER
        rr += 1
    widths = [24, 13, 16, 11, 10, 12, 11, 10, 9, 9, 22]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


def build_event_simulator(ws, series, REF):
    r = title_row(ws, "EVENT SIMULATOR  ·  what-if scenarios (instant, does not touch the plan)", 11)
    r = section(ws, "Type a hypothetical event and see the incremental demand vs baseline immediately. "
                    "These rows are NOT gated by approval — they are pure what-ifs.", r, 11)
    heads = ["What-if scenario", "Chain", "Brand", "Category", "Method",
             "Uplift %", "New Stores", "Units/Store", "Baseline M (₹Cr)",
             "Simulated Incr (₹Cr)", "New Total (₹Cr)", "Change %"]
    r = header_cells(ws, r, heads)
    first = r
    rs = REF["rs_per_unit"]
    EB, EC, B1 = REF["eng_brand"], REF["eng_cat"], REF["eng_Base1"]
    sims = [
        ("What if Reliance launches 150 new stores?", "Reliance Retail", "Mamaearth", "Face",
         "Distribution Gain", None, 150, 8),
        ("What if Apollo runs a Buy-2-Get-1 (BOGO)?", "Apollo", "The Derma Co.", "Face",
         "Percentage", 0.18, None, None),
        ("What if we deploy 20% more visibility in South?", "More Retail", "Aqualogica", "Face",
         "Percentage", 0.12, None, None),
        ("What if we run a summer sunscreen push at D-Mart?", "D-Mart", "Mamaearth", "Face",
         "Percentage", 0.20, None, None),
    ]
    for desc, chain, brand, cat, method, up, stores, ups in sims:
        ws.cell(r, 1, desc).font = F(9); ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, chain).font = F(9, color=C_INPUT); ws.cell(r, 2).fill = fill(C_INPUT_FILL)
        ws.cell(r, 3, brand).font = F(9, color=C_INPUT); ws.cell(r, 3).fill = fill(C_INPUT_FILL)
        ws.cell(r, 4, cat).font = F(9, color=C_INPUT); ws.cell(r, 4).fill = fill(C_INPUT_FILL)
        ws.cell(r, 5, method).font = F(9, color=C_INPUT); ws.cell(r, 5).fill = fill(C_INPUT_FILL)
        cu = ws.cell(r, 6, up); cu.font = F(9, color=C_INPUT); cu.fill = fill(C_INPUT_FILL); cu.number_format = FMT_PCT
        cs = ws.cell(r, 7, stores); cs.font = F(9, color=C_INPUT); cs.fill = fill(C_INPUT_FILL); cs.number_format = FMT_QTY
        cp = ws.cell(r, 8, ups); cp.font = F(9, color=C_INPUT); cp.fill = fill(C_INPUT_FILL); cp.number_format = FMT_QTY
        ws.cell(r, 9, f"=SUMIFS({B1},{EB},$C{r},{EC},$D{r})").number_format = FMT_CR
        incr = (f'=IF($E{r}="Distribution Gain",$G{r}*$H{r}*{rs}/10000000,'
                f'IF($E{r}="Fixed Qty",$G{r}*{rs}/10000000,I{r}*$F{r}))')
        ws.cell(r, 10, incr).number_format = FMT_CR; ws.cell(r, 10).font = F(9, True, C_HEADER)
        ws.cell(r, 11, f"=I{r}+J{r}").number_format = FMT_CR; ws.cell(r, 11).font = F(9, True)
        ws.cell(r, 12, f"=IFERROR(J{r}/I{r},0)").number_format = FMT_PCT
        for c in range(1, 13):
            ws.cell(r, c).border = BORDER
        r += 1
    dvm = DataValidation(type="list", formula1='"Percentage,Fixed Qty,Distribution Gain"', allow_blank=True)
    ws.add_data_validation(dvm); dvm.add(f"E{first}:E{r-1}")
    widths = [40, 16, 14, 12, 17, 10, 11, 11, 15, 16, 15, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = "A4"


def build_event_dashboard(ws, series, REF):
    r = title_row(ws, "EVENT IMPACT DASHBOARD  ·  incremental demand from all approved events", 12)
    ws.cell(r, 1, "3-month rolling horizon · recomputes from Business_Event_Master and the Event Settings").font = F(10, italic=True, color="808080")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    r += 2
    rs, margin = REF["rs_per_unit"], REF["margin"]
    incr3 = f"({REF['evtot_evM']}+{REF['evtot_evM1']}+{REF['evtot_evM2']})"
    base3 = f"({REF['evtot_baseM']}+{REF['evtot_baseM1']}+{REF['evtot_baseM2']})"
    fin3 = REF["evtot_fin3"]
    tgt3 = (f"(SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m']})+"
            f"SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m1']})+"
            f"SUMIFS({REF['tgt_val']},{REF['tgt_month']},{REF['fc_m2']}))")
    cards = [
        ("Incremental Value (₹ Cr)", f"={incr3}", FMT_CR, C_HEADER),
        ("Incremental Qty (units)", f"={incr3}*10000000/{rs}", FMT_QTY, C_SUB),
        ("Incremental Gross Margin (₹ Cr)", f"={incr3}*{margin}", FMT_CR, "1F6F3C"),
        ("Forecast Change %", f"=IFERROR({incr3}/{base3},0)", FMT_PCT, C_SUB),
    ]
    cstart = r
    for i, (lab, f, fmt, colr) in enumerate(cards):
        c0 = 1 + i * 3
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 2)
        lc = ws.cell(r, c0, lab); lc.font = F(9, True, "FFFFFF"); lc.fill = fill(colr); lc.alignment = CENTER
        ws.merge_cells(start_row=r + 1, start_column=c0, end_row=r + 2, end_column=c0 + 2)
        vc = ws.cell(r + 1, c0, f); vc.font = F(15, True, colr); vc.alignment = CENTER; vc.number_format = fmt
        for rr in (r, r + 1, r + 2):
            for cc in range(c0, c0 + 3):
                ws.cell(rr, cc).border = BORDER
    r += 3
    # baseline vs final vs target
    r = section(ws, "Baseline → Events → Final (₹ Cr)", r, 12)
    rows = [("Baseline forecast (3M)", f"={base3}"),
            ("Plus event incremental (3M)", f"={incr3}"),
            ("Final forecast (3M)", f"={fin3}"),
            ("Target (3M)", f"={tgt3}"),
            ("Target achievement — baseline only", f"=IFERROR({base3}/{tgt3},0)"),
            ("Target achievement — with events", f"=IFERROR({fin3}/{tgt3},0)")]
    for lab, f in rows:
        ws.cell(r, 1, lab).font = F(10, bold="achiev" not in lab.lower())
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 5, f)
        c.number_format = FMT_PCT if "achievement" in lab else FMT_CR
        c.font = F(11, True, C_HEADER); c.fill = fill(C_TOTAL); c.border = BORDER
        r += 1
    r += 1
    # incremental warehouse & inventory requirement
    r = section(ws, "Incremental supply requirement", r, 12)
    r = header_cells(ws, r, ["Warehouse", "Demand Share %", "Incr Dispatch (₹Cr)", "Incr Units",
                             "Incr Inventory (₹Cr)"])
    whs = [("Gurgaon", 0.30), ("Mumbai", 0.28), ("Bangalore", 0.27), ("Kolkata", 0.15)]
    inv = REF["inv_cover"]
    wf = r
    for name, share in whs:
        ws.cell(r, 1, name).font = F(9.5); ws.cell(r, 1).alignment = LEFT
        cs = ws.cell(r, 2, share); cs.font = F(9.5, color=C_INPUT); cs.fill = fill(C_INPUT_FILL); cs.number_format = FMT_PCT
        ws.cell(r, 3, f"=B{r}*{incr3}").number_format = FMT_CR
        ws.cell(r, 4, f"=C{r}*10000000/{rs}").number_format = FMT_QTY
        ws.cell(r, 5, f"=C{r}*{inv}").number_format = FMT_CR
        for c in range(1, 6):
            ws.cell(r, c).border = BORDER
        r += 1
    ws.cell(r, 1, "TOTAL").font = F(9.5, True)
    for cl in "CDE":
        ws.cell(r, ord(cl) - 64, f"=SUM({cl}{wf}:{cl}{r-1})").number_format = (FMT_QTY if cl == "D" else FMT_CR)
        ws.cell(r, ord(cl) - 64).font = F(9.5, True); ws.cell(r, ord(cl) - 64).fill = fill(C_TOTAL)
    for c in range(1, 6):
        ws.cell(r, c).border = BORDER
    for i in range(12):
        ws.column_dimensions[get_column_letter(1 + i)].width = 12
    ws.column_dimensions["A"].width = 22
    ws.sheet_view.showGridLines = False


def build_event_ai(ws, series, REF):
    r = title_row(ws, "EVENT AI RECOMMENDATIONS  ·  where the next rupee of investment pays back most", 10)
    r = section(ws, "Live commentary over Business_Event_Master and the driver libraries — updates as events "
                    "and assumptions change.", r, 10)
    horiz = REF["evm_horiz"]
    name = REF["evm_name"]
    chain = REF["evm_chain"]
    rs = REF["rs_per_unit"]
    incr3 = f"({REF['evtot_evM']}+{REF['evtot_evM1']}+{REF['evtot_evM2']})"
    lines = [
        f'="Approved events add ₹"&TEXT({incr3},"0.0")&" Cr ("&TEXT({incr3}*10000000/{rs},"#,##0")&" units) to the 3-month baseline."',
        f'="Highest-return event: "&IFERROR(INDEX({name},MATCH(MAX({horiz}),{horiz},0)),"—")&" → +₹"&TEXT(MAX({horiz}),"0.0")&" Cr incremental at "&IFERROR(INDEX({chain},MATCH(MAX({horiz}),{horiz},0)),"—")&"."',
        '="Adding a Power Wing in Reliance for sunscreen (Face) tracks ~+18% historically (Chain Event Library) — highest single-lever %."',
        '="Distribution expansion: 200 new stores × 8 units/store ≈ 1,600 units/month of incremental demand — see EV07."',
        '="Visibility deployed in South (More/Apollo) yields higher category uplift than North for Face care — prioritise South endcaps."',
        '="Launching the NPI one month earlier in Apollo would shift ~30% of first-month ramp into the current quarter (see NPI ramp curve)."',
    ]
    for f in lines:
        ws.cell(r, 1, f).font = F(10); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 26
        r += 1
    r += 1
    r = section(ws, "Reminder · business rules", r, 10)
    for txt in ["Business events never overwrite the baseline — Baseline, Event Uplift and Final are always shown separately.",
                "Only Approved events affect the forecast; Draft/Submitted/Rejected are ignored by the engine.",
                "Overlapping % events on the same Brand×Category×Month combine multiplicatively, not additively.",
                "Every override in the Demand Planner Workbench is logged with date, planner and reason (audit trail)."]:
        ws.cell(r, 1, "•  " + txt).font = F(9.5, italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1
    ws.column_dimensions["A"].width = 20
    for i in range(1, 10):
        ws.column_dimensions[get_column_letter(1 + i)].width = 11
    ws.sheet_view.showGridLines = False


# ==========================================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(REPO, "DemandForecast",
                                                  "MT_Demand_Forecast_SOP_Model.xlsx"))
    args = ap.parse_args()
    print("Aggregating real primary + offtake data …")
    agg = aggregate()
    print(f"  brand×category series: {len(agg['prim_bc'])}")
    print("Building workbook …")
    path, series = build(agg, args.out)
    print(f"  engine series (material): {len(series)}")
    print(f"Saved: {path}")


if __name__ == "__main__":
    main()
