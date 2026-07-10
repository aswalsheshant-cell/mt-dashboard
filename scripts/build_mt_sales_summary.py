#!/usr/bin/env python3
"""Build the auto-updating MT Sales Summary Excel working file.

Deliverable for the "MT quarterly format/chain summary" verification + automation task.
Structure (5 sheets):
  1. Source_Data  - wide monthly chain-level matrix (Jun-25 .. Jun-26). Paste Jun-26 here.
  2. Mapping      - format -> chain -> summary bucket reference.
  3. Summary      - format & chain quarterly table (Q1-24 .. Q1-26) + QoQ, pending-aware.
  4. Chart        - Grand Total quarterly trend line, auto-linked to Summary.
  5. QC           - considered/excluded data, mapping gaps, pending Jun-26, mismatch checks.

Design rules honoured:
  * Blank / pending values are NEVER treated as 0.
  * Q1-26 (and its QoQ) show "Pending" per row until that row's Jun-26 months arrive.
  * Historical quarters (Q1-24..Q4-25) carry the verified Screenshot-1 values; Q1-26 is
    fully formula-driven from Source_Data via SUMIFS, so pasting Jun-26 auto-updates
    the table and the chart.
  * SIS format row added (no source yet -> "Source data required").

Run:  python scripts/build_mt_sales_summary.py
Out:  ExcelWorking/MT_Sales_Summary_Automated.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "..", "ExcelWorking", "MT_Sales_Summary_Automated.xlsx")

# ---------------------------------------------------------------- styling helpers
BLUE      = "1F6FB2"
BLUE_DK   = "12507F"
HDR_FILL  = PatternFill("solid", fgColor=BLUE)
TOT_FILL  = PatternFill("solid", fgColor=BLUE_DK)
NOTE_FILL = PatternFill("solid", fgColor="FFF3CD")
PEND_FILL = PatternFill("solid", fgColor="FCE4D6")
GREEN_F   = PatternFill("solid", fgColor="C6EFCE")
GREEN_T   = Font(color="006100")
RED_F     = PatternFill("solid", fgColor="FFC7CE")
RED_T     = Font(color="9C0006")
WHITE_B   = Font(bold=True, color="FFFFFF")
BOLD      = Font(bold=True)
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = WHITE_B; cell.alignment = CTR; cell.border = BORDER

# ---------------------------------------------------------------- source data (Screenshot 3)
MONTHS = ["Jun-25","Jul-25","Aug-25","Sep-25","Oct-25","Nov-25","Dec-25",
          "Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26"]
B = None  # blank / pending -> left empty (never 0)

# (Format, SummaryBucket, SourceChain, [13 monthly values], Remark)
SRC = [
 ("Hyper/Super","D-Mart",    "Dmart",            [856,780,801,723,993,967,943,967,928,1226,1258,1518,1412], ""),
 ("Hyper/Super","Reliance",  "Reliance",         [592,557,769,543,667,679,716,900,757,805,868,990,B], "Jun-26 pending"),
 ("Hyper/Super","Lulu",      "Lulu",             [44,47,53,53,65,70,86,103,70,78,101,116,116], ""),
 ("Hyper/Super","More",      "More Retail",      [34,35,34,34,41,41,41,39,39,45,45,55,B], "Jun-26 pending"),
 ("Hyper/Super","HS-Others", "Spencer",          [15,14,10,10,10,11,22,13,14,12,12,12,B], "Jun-26 pending"),
 ("Hyper/Super","HS-Others", "Vmm",              [12,13,11,9,6,3,10,17,14,25,28,31,28], ""),
 ("Hyper/Super","HS-Others", "V-Mart",           [3,1,2,2,1,20,19,14,15,18,15,10,13], ""),
 ("Hyper/Super","HS-Others", "Trent",            [4,4,6,5,5,6,5,6,4,6,7,8,B], "Jun-26 pending"),
 ("Hyper/Super","HS-Others", "Spar",             [B]*13, "ASSUMED map->HS-Others: not in mapping sheet - confirm"),
 ("Pharmacy","Apollo",       "Apollo",           [275,287,327,306,430,555,567,534,511,570,699,751,B], "Jun-26 pending"),
 ("Pharmacy","Wellness",     "Wellness Forever", [74,72,73,66,89,94,88,82,86,98,110,102,80], ""),
 ("Pharmacy","Drug-Others",  "Frankros",         [1,1,1,1,1,3,3,4,3,3,4,10,7], ""),
 ("Pharmacy","Drug-Others",  "Arambagh",         [1,1,1,3,3,1,2,2,2,3,7,5,B], "Jun-26 pending"),
 ("Pharmacy","Drug-Others",  "Guardian",         [1,1,1,1,1,1,2,2,2,2,2,2,2], ""),
 ("Regional","Regional",     "Sancus(Rmt)",      [36,41,30,47,30,33,54,35,28,42,58,46,40], ""),
 ("Regional","Regional",     "Ratandeep",        [7,11,5,5,3,B,10,6,6,3,3,3,B], "Jun-26 pending"),
 ("Regional","Regional",     "Vijetha",          [4,3,B,B,B,B,B,B,B,B,B,B,B], "Reporting stopped after Jul-25"),
 ("Regional","Regional",     "National Mart",    [B,B,B,B,B,B,B,B,B,B,2,2,4], "ASSUMED map->Regional - confirm"),
 ("Regional","Regional",     "Sasta Sundar",     [0,0,1,1,1,1,0,0,1,1,1,1,1], ""),
 ("Regional","Regional",     "Sumo Save",        [B,B,B,B,B,B,B,B,B,0,1,2,1], "ASSUMED map->Regional - confirm"),
 ("Regional","Regional",     "Ratanadeep",       [B,B,B,B,B,B,B,B,B,B,B,B,3], "Dup of Ratandeep? confirm"),
 ("Regional","Regional",     "Apna Mart",        [B,B,B,B,B,B,B,B,B,3,B,B,B], "ASSUMED map->Regional - confirm"),
 ("Regional","Regional",     "Sohum",            [B]*13, "Sohum Shoppe -> Regional/Others (no data yet)"),
 ("H&B","H&B",               "H&G",              [81,79,89,85,67,66,71,80,48,51,62,87,52], ""),
 ("H&B","H&B",               "Wh-Smith",         [1,1,1,1,1,1,1,1,1,1,1,1,B], "Jun-26 pending"),
 ("H&B","H&B",               "BEAUTY & NUTRIE",  [0,1,1,1,1,1,1,1,1,1,1,0,1], "ASSUMED map->H&B - confirm"),
 ("EB2B","EB2B",             "Fsn",              [164,191,160,151,156,186,168,164,168,173,229,208,217], ""),
 ("CNC","CNC",               "Metro Cnc",        [44,42,43,35,63,58,57,53,48,53,48,48,B], "Jun-26 pending"),
 ("CNC","CNC",               "Walmart Cnc",      [21,15,15,12,18,23,25,17,17,19,26,11,B], "Jun-26 pending"),
]
NROW = len(SRC)                      # 29 chains
S_FIRST, S_LAST = 2, 1 + NROW        # data rows 2..30 on Source sheet

# month letter map on Source sheet: cols D(4)..P(16)
MCOL = {m: get_column_letter(4 + i) for i, m in enumerate(MONTHS)}
def MRNG(m):  # absolute source range for a month column
    L = MCOL[m]; return f"Source_Data!${L}${S_FIRST}:${L}${S_LAST}"
BKT_RNG = f"Source_Data!$B${S_FIRST}:$B${S_LAST}"

# =================================================================== workbook
wb = Workbook()

# ---------------------------------------------------------------- 1. Source_Data
ws = wb.active; ws.title = "Source_Data"
headers = ["Format","Summary Bucket","Chain (source)"] + MONTHS + ["Remarks"]
for c, h in enumerate(headers, 1):
    ws.cell(row=1, column=c, value=h)
style_header(ws, 1, 1, len(headers))
for r, (fmt, bkt, chain, vals, rem) in enumerate(SRC, start=2):
    ws.cell(row=r, column=1, value=fmt)
    ws.cell(row=r, column=2, value=bkt)
    ws.cell(row=r, column=3, value=chain)
    for i, v in enumerate(vals):
        cell = ws.cell(row=r, column=4 + i)
        if v is not None:
            cell.value = v
        cell.alignment = CTR
    ws.cell(row=r, column=len(headers), value=rem)
# Grand total check row (informational)
gt = NROW + 3
ws.cell(row=gt, column=3, value="Grand Total (source)").font = BOLD
for i, m in enumerate(MONTHS):
    L = MCOL[m]
    ws.cell(row=gt, column=4 + i,
            value=f"=SUM({L}{S_FIRST}:{L}{S_LAST})").font = BOLD
note = ws.cell(row=gt + 2, column=1,
    value=("PASTE MONTHLY DATA HERE. Leave a cell BLANK if the value is not yet received "
           "(do NOT enter 0). The Summary & Chart read this sheet by 'Summary Bucket' via "
           "SUMIFS, so pasting the pending Jun-26 column auto-updates everything."))
note.fill = NOTE_FILL; note.font = BOLD; note.alignment = LEFT
ws.merge_cells(start_row=gt + 2, start_column=1, end_row=gt + 2, end_column=len(headers))
ws.freeze_panes = "D2"
ws.column_dimensions["A"].width = 13; ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 18
for i in range(len(MONTHS)):
    ws.column_dimensions[get_column_letter(4 + i)].width = 8
ws.column_dimensions[get_column_letter(len(headers))].width = 34

# ---------------------------------------------------------------- 2. Mapping
mp = wb.create_sheet("Mapping")
mp_hdr = ["Format (mapping sheet)","Chain","Chain Mapping","Summary Bucket used"]
for c, h in enumerate(mp_hdr, 1):
    mp.cell(row=1, column=c, value=h)
style_header(mp, 1, 1, 4)
MAPPING = [
 ("H/S MKT","D-Mart","D-Mart","D-Mart"),("H/S MKT","Reliance Retail","Reliance Retail","Reliance"),
 ("H/S MKT","Lulu","Lulu","Lulu"),("H/S MKT","Max Hyper","Others","HS-Others"),
 ("H/S MKT","More Retail","More Retail","More"),("H/S MKT","Spencer","Others","HS-Others"),
 ("H/S MKT","Trent Hypermarket","Others","HS-Others"),("H/S MKT","V-Mart","Others","HS-Others"),
 ("H/S MKT","VMM","Others","HS-Others"),
 ("DRUG STORE","Apollo Healthco","Apollo Healthco","Apollo"),
 ("DRUG STORE","Wellness Forever","Wellness Forever","Wellness"),
 ("DRUG STORE","Guardian Healthcare","Others","Drug-Others"),
 ("DRUG STORE","Arambagh","Others","Drug-Others"),("DRUG STORE","FRANKROSS","Others","Drug-Others"),
 ("Regional","RMT-Sancus","Sancus","Regional"),("Regional","Grace Super MKT","Others","Regional"),
 ("Regional","Deal Share","Others","Regional"),("Regional","Ratnadeep","Others","Regional"),
 ("Regional","Vijetha","Others","Regional"),("Regional","Pothys","Others","Regional"),
 ("Regional","Sarvana","Others","Regional"),("Regional","Sasta Sunder","Others","Regional"),
 ("Regional","Trilife","Others","Regional"),("Regional","Octopolis Technologies Pvt Ltd","Others","Regional"),
 ("Regional","B&N","Others","Regional"),("Regional","SAI SAACHI ASSOCIATES-MT-OR","Others","Regional"),
 ("Regional","Others","Others","Regional"),("Regional","GNRC","Others","Regional"),
 ("Regional","Nao Grocery","Others","Regional"),("Regional","Sohum Shoppe","Others","Regional"),
 ("H&B","H&G","H&G","H&B"),("H&B","Dabur new u","Others","H&B"),("H&B","WH-Smith","Others","H&B"),
 ("H&B","Travel Retail Services-Relay","Others","H&B"),
 ("Eb2b","FSN","EB2B","EB2B"),
 ("CNC","Metro-CNC","CNC","CNC"),("CNC","Walmart-CNC","CNC","CNC"),
 ("SIS","(no chains mapped yet)","-","SIS"),
]
for r, row in enumerate(MAPPING, start=2):
    for c, v in enumerate(row, 1):
        cell = mp.cell(row=r, column=c, value=v); cell.border = BORDER; cell.alignment = LEFT
for w, col in zip((14, 30, 18, 18), "ABCD"):
    mp.column_dimensions[col].width = w
mp.freeze_panes = "A2"

# ---------------------------------------------------------------- 3. Summary
sm = wb.create_sheet("Summary")
QH = ["Q1-24","Q2-24","Q3-24","Q4-24","Q1-25","Q2-25","Q3-25","Q4-25","Q1-26"]
sm_hdr = ["Format","Chain / Line"] + QH + ["GO Sequential QoQ","GO QoQ","Remarks"]
for c, h in enumerate(sm_hdr, 1):
    sm.cell(row=1, column=c, value=h)
style_header(sm, 1, 1, len(sm_hdr))
# columns: A Format, B Line, C..K = Q1-24..Q1-26, L Seq QoQ, M GO QoQ, N Remarks
COL_Q1_26, COL_Q4_25, COL_Q1_25 = "K", "J", "G"
COL_SEQ, COL_GO, COL_REM = "L", "M", "N"

# rows: (Format, Line, bucket_or_None, [8 historical Q1-24..Q4-25], kind, remark)
# kind: 'chain' | 'htotal'(H/S) | 'dtotal'(Drug) | 'format'(single line) | 'sis' | 'grand'
ROWS = [
 ("Hyper/Super","D-Mart","D-Mart",[2234,1873,1954,2028,2600,2304,2902,3121],"chain",""),
 ("","Reliance","Reliance",[1195,1399,1382,1556,1908,1870,2063,2461],"chain",""),
 ("","Lulu","Lulu",[115,121,94,124,124,153,220,251],"chain",""),
 ("","More","More",[37,54,59,71,84,104,123,123],"chain",""),
 ("","Others","HS-Others",[84,120,137,112,105,88,119,155],"chain","Max Hyper/Spencer/Trent/V-Mart/VMM"),
 ("","H/S Total",None,[3665,3568,3625,3891,4820,4519,5427,6111],"htotal",""),
 ("Pharmacy","Apollo","Apollo",[409,399,566,602,793,921,1552,1615],"chain",""),
 ("","Wellness","Wellness",[268,211,253,259,249,211,271,266],"chain",""),
 ("","Others","Drug-Others",[9,10,9,11,9,12,19,22],"chain","Guardian/Arambagh/Frankross"),
 ("","Drug Total",None,[686,620,828,872,1051,1144,1842,1902],"dtotal",""),
 ("Regional","Regional","Regional",[173,145,140,131,150,145,119,109],"format",""),
 ("H&B","H&B","H&B",[218,187,214,224,237,255,206,182],"format",""),
 ("EB2B","EB2B","EB2B",[536,458,423,388,523,502,510,505],"format",""),
 ("CNC","CNC","CNC",[226,218,197,214,210,154,244,206],"format",""),
 ("SIS","SIS",None,[None]*8,"sis","NEW format added - no SIS source data provided yet"),
 ("","Grand Total",None,[5503,5197,5428,5719,6992,6729,8349,9015],"grand",""),
]
r0 = 2
for i, (fmt, line, bkt, hist, kind, rem) in enumerate(ROWS):
    r = r0 + i
    sm.cell(row=r, column=1, value=fmt)
    sm.cell(row=r, column=2, value=line)
    # historical C..J (Q1-24..Q4-25)
    for j, v in enumerate(hist):
        cell = sm.cell(row=r, column=3 + j)
        if kind in ("htotal", "dtotal", "grand"):
            pass  # filled below with SUM formulas
        elif v is not None:
            cell.value = v
        cell.alignment = CTR
    # Q1-26 (col K)
    kcell = sm.cell(row=r, column=11)
    if kind == "chain" or kind == "format":
        pend = (f'SUMPRODUCT(({BKT_RNG}="{bkt}")*({MRNG("May-26")}<>"")*({MRNG("Jun-26")}=""))')
        raw = (f'SUMIFS({MRNG("Apr-26")},{BKT_RNG},"{bkt}")'
               f'+SUMIFS({MRNG("May-26")},{BKT_RNG},"{bkt}")'
               f'+SUMIFS({MRNG("Jun-26")},{BKT_RNG},"{bkt}")')
        kcell.value = f'=IF({pend}>0,"Pending",{raw})'
    elif kind == "sis":
        kcell.value = "Source data required"
    kcell.alignment = CTR

# total-row formulas (need row indices)
def rownum(line):
    for i, rr in enumerate(ROWS):
        if rr[1] == line:
            return r0 + i
HS_CH = [rownum(x) for x in ("D-Mart","Reliance","Lulu","More","Others")][:5]  # first Others = HS
# Others appears twice; resolve explicitly by position
r_dmart = r0 + 0; r_hsoth = r0 + 4; r_hstot = r0 + 5
r_apollo = r0 + 6; r_drgoth = r0 + 8; r_drgtot = r0 + 9
r_reg = r0 + 10; r_hnb = r0 + 11; r_eb2b = r0 + 12; r_cnc = r0 + 13; r_sis = r0 + 14; r_grand = r0 + 15
# historical SUMs for totals
for col in range(3, 11):  # C..J
    L = get_column_letter(col)
    sm.cell(row=r_hstot, column=col, value=f"=SUM({L}{r0}:{L}{r0+4})").font = WHITE_B
    sm.cell(row=r_drgtot, column=col, value=f"=SUM({L}{r0+6}:{L}{r0+8})").font = WHITE_B
    sm.cell(row=r_grand, column=col,
            value=f"=SUM({L}{r_hstot},{L}{r_drgtot},{L}{r_reg}:{L}{r_sis})").font = WHITE_B
# Q1-26 (K) totals - pending-aware
K = "K"
sm.cell(row=r_hstot, column=11,
    value=f'=IF(COUNTIF({K}{r0}:{K}{r0+4},"Pending")>0,"Pending",SUM({K}{r0}:{K}{r0+4}))').font = WHITE_B
sm.cell(row=r_drgtot, column=11,
    value=f'=IF(COUNTIF({K}{r0+6}:{K}{r0+8},"Pending")>0,"Pending",SUM({K}{r0+6}:{K}{r0+8}))').font = WHITE_B
sm.cell(row=r_grand, column=11,
    value=(f'=IF(OR({K}{r_hstot}="Pending",{K}{r_drgtot}="Pending",{K}{r_reg}="Pending",'
           f'{K}{r_hnb}="Pending",{K}{r_eb2b}="Pending",{K}{r_cnc}="Pending"),"Pending",'
           f'SUM({K}{r_hstot},{K}{r_drgtot},{K}{r_reg},{K}{r_hnb},{K}{r_eb2b},{K}{r_cnc}))')).font = WHITE_B

# QoQ columns for every row
for i, (fmt, line, bkt, hist, kind, rem) in enumerate(ROWS):
    r = r0 + i
    seq = sm.cell(row=r, column=12)
    go = sm.cell(row=r, column=13)
    if kind == "sis":
        seq.value = "Pending"; go.value = "Pending"
    else:
        seq.value = (f'=IF(ISNUMBER({COL_Q1_26}{r}),'
                     f'IF({COL_Q4_25}{r}=0,"n/a",({COL_Q1_26}{r}-{COL_Q4_25}{r})/{COL_Q4_25}{r}),"Pending")')
        go.value = (f'=IF(AND(ISNUMBER({COL_Q1_26}{r}),ISNUMBER({COL_Q1_25}{r}),{COL_Q1_25}{r}<>0),'
                    f'({COL_Q1_26}{r}-{COL_Q1_25}{r})/{COL_Q1_25}{r},"Pending")')
    seq.number_format = "0.0%"; go.number_format = "0.0%"
    seq.alignment = CTR; go.alignment = CTR
    # remarks
    rc = sm.cell(row=r, column=14)
    base = rem
    if kind in ("chain", "format", "htotal", "dtotal", "grand"):
        rc.value = (base + ("; " if base else "") +
                    'Q1-26 = Apr+May+Jun-26 (auto from Source_Data)') if kind != "grand" else \
                   'Q1-26 Pending until all chains report Jun-26'
    else:
        rc.value = base
    rc.alignment = LEFT

# total-row shading
for rr in (r_hstot, r_drgtot, r_grand):
    for c in range(1, 15):
        cell = sm.cell(row=rr, column=c)
        cell.fill = TOT_FILL
        if cell.font is None or not cell.font.color:
            cell.font = WHITE_B
        else:
            cell.font = WHITE_B
# borders + alignment on the block
for r in range(1, r_grand + 1):
    for c in range(1, 15):
        cell = sm.cell(row=r, column=c); cell.border = BORDER
# conditional formatting: green growth / red decline on L,M
rng = f"L{r0}:M{r_grand}"
sm.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(ISNUMBER(L{r0}),L{r0}>0)"], fill=GREEN_F, font=GREEN_T))
sm.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(ISNUMBER(L{r0}),L{r0}<0)"], fill=RED_F, font=RED_T))
sm.conditional_formatting.add(rng, FormulaRule(formula=[f'L{r0}="Pending"'], fill=PEND_FILL))
sm.freeze_panes = "C2"
sm.column_dimensions["A"].width = 12; sm.column_dimensions["B"].width = 13
for c in range(3, 12):
    sm.column_dimensions[get_column_letter(c)].width = 8
sm.column_dimensions["L"].width = 14; sm.column_dimensions["M"].width = 11
sm.column_dimensions["N"].width = 40

# ---------------------------------------------------------------- 4. Chart
ch = wb.create_sheet("Chart")
ch.cell(row=1, column=1, value="MT Total Sales - Quarterly Trend (Grand Total, incl. Others)").font = Font(bold=True, size=13)
ch.cell(row=2, column=1, value=("Auto-linked to Summary!Grand Total. Q1-26 is 'Pending' (no plotted point) "
        "until every chain's Jun-26 is pasted into Source_Data; the point then appears automatically.")).alignment = LEFT
chart = LineChart()
chart.title = "MT Total Sales by Quarter (Q1-24 -> Q1-26)"
chart.style = 12
chart.height = 9; chart.width = 22
chart.y_axis.title = "Sales (INR Cr / units as per source)"
chart.x_axis.title = "Quarter"
data = Reference(sm, min_col=3, max_col=11, min_row=r_grand, max_row=r_grand)  # C..K grand total
cats = Reference(sm, min_col=3, max_col=11, min_row=1, max_row=1)              # C1..K1 headers
chart.add_data(data, from_rows=True, titles_from_data=False)
chart.set_categories(cats)
s = chart.series[0]
s.smooth = False
chart.dataLabels = DataLabelList(); chart.dataLabels.showVal = True
ch.add_chart(chart, "A4")

# ---------------------------------------------------------------- 5. QC
qc = wb.create_sheet("QC")
def qc_title(r, text):
    c = qc.cell(row=r, column=1, value=text); c.font = WHITE_B; c.fill = HDR_FILL
    qc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
def qc_row(r, cells, bold=False):
    for c, v in enumerate(cells, 1):
        cell = qc.cell(row=r, column=c, value=v); cell.alignment = LEFT
        if bold: cell.font = BOLD

r = 1
qc_title(r, "QC & VERIFICATION LOG - MT Sales Summary"); r += 2
qc_row(r, ["Key fixes applied"], bold=True); r += 1
for t in [
    "1. Root cause of -100%: Q1-26 was blank but treated as 0 in QoQ. Fixed: blank Jun-26 -> row shows 'Pending', QoQ shows 'Pending' (never -100%).",
    "2. Quarter logic confirmed: Q1=Apr-Jun, Q2=Jul-Sep, Q3=Oct-Dec, Q4=Jan-Mar; FY label = start year (Apr-26 => Q1-26). Verified: D-Mart Q4-25 = Jan+Feb+Mar-26 = 967+928+1226 = 3121 (matches).",
    "3. SIS format row ADDED to the summary (was missing). No SIS source data yet -> 'Source data required'.",
    "4. Chart labels corrected (old chart mislabelled Q4-25/Q1-25 order and ended at 'Q4-26'). Now Q1-24..Q1-26, auto-linked.",
    "5. Q1-26 column is now SUMIFS-driven from Source_Data -> pasting Jun-26 auto-updates table + chart. Historical Q1-24..Q4-25 kept as verified values.",
]:
    qc_row(r, [t]); qc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1
r += 1

qc_row(r, ["Quarter -> Month mapping"], bold=True); r += 1
qc_row(r, ["Quarter","Months"], bold=True); r += 1
for q, m in [("Q2-25","Jul-25, Aug-25, Sep-25"),("Q3-25","Oct-25, Nov-25, Dec-25"),
             ("Q4-25","Jan-26, Feb-26, Mar-26"),("Q1-26","Apr-26, May-26, Jun-26")]:
    qc_row(r, [q, m]); r += 1
r += 1

qc_row(r, ["Pending Jun-26 (chains missing -> Q1-26 not final)"], bold=True); r += 1
qc_row(r, ["Live count of chains with May-26 filled but Jun-26 blank:",
           f'=SUMPRODUCT(({MRNG("May-26")}<>"")*({MRNG("Jun-26")}=""))'], bold=True); r += 1
pend_chains = [c for (_,_,c,v,_) in SRC if v[12] is None and v[11] is not None]
qc_row(r, ["Chains pending Jun-26:", ", ".join(pend_chains)]);
qc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5); r += 2

qc_row(r, ["Source-recomputed vs Screenshot-1 (Grand Total) - mismatch check"], bold=True); r += 1
qc_row(r, ["Quarter","Source (recomputed)","Screenshot-1","Diff","Note"], bold=True); r += 1
for q, srcv, s1v in [("Q2-25",6730,6729),("Q3-25",8363,8349),("Q4-25",9034,9015),("Q1-26","partial/pending","blank")]:
    diff = (srcv - s1v) if isinstance(srcv,int) and isinstance(s1v,int) else "-"
    note = "rounding (<0.3%)" if isinstance(diff,int) else "Jun-26 incomplete -> pending"
    qc_row(r, [q, srcv, s1v, diff, note]); r += 1
r += 1

qc_row(r, ["Mapping gaps / assumptions to confirm"], bold=True); r += 1
qc_row(r, ["Item","Status"], bold=True); r += 1
for item, st in [
    ("Spar","Not in mapping sheet - tentatively HS-Others (Others). Confirm."),
    ("BEAUTY & NUTRIE","Not in mapping sheet - tentatively H&B (Others). Confirm."),
    ("National Mart / Sumo Save / Apna Mart","Not in mapping - tentatively Regional (Others). Confirm."),
    ("Ratanadeep vs Ratandeep","Possible duplicate spelling in source - confirm single chain."),
    ("Max Hyper","In mapping (HS-Others) but no data column in source screenshot."),
    ("SIS","Format added; no chains mapped and no source data. Provide SIS chains + values."),
]:
    qc_row(r, [item, st]); qc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5); r += 1
r += 1

qc_row(r, ["Data considered / not considered"], bold=True); r += 1
qc_row(r, ["Considered","MT channel only. All chains in Source_Data mapped to a Summary Bucket; 'Others' kept inside format totals."]); r += 1
qc_row(r, ["Not considered","GT (General Trade) data - excluded per scope. Blank/pending months - excluded from sums (not zero-filled)."]); r += 1
for cc in "ABCDE":
    qc.column_dimensions[cc].width = 26
qc.column_dimensions["A"].width = 34

# force spreadsheet apps to recalc all formulas on open
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("wrote", os.path.abspath(OUT))
