#!/usr/bin/env python3
"""
Milestone 2: Zone & Category Contribution Excel Scorecards
Produces:
  exports/Zone_Scorecard_FY27_YTD.xlsx
  exports/Category_Brand_Contribution_FY27.xlsx
"""
import json
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import os; os.system("pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

em = json.loads(Path('dashboard/enriched_metrics.json').read_text())

NAVY="10254A"; WHITE="FFFFFF"; LIGHT="EAF0FB"; GREEN="1E8449"; RED="C0392B"; AMBER="E67E22"; GRAY="95A5A6"
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, r, ncols, col=NAVY):
    for c in range(1, ncols+1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def cs(ws, r, c, val, bold=False, align="left", color=None, bg=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(bold=bold, color=color or "000000", name='Calibri', size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = border
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    if fmt: cell.number_format = fmt
    return cell

# ── Zone Scorecard ─────────────────────────────────────────────────────────────
wb1 = Workbook()
ws = wb1.active
ws.title = "Zone Scorecard"
ws.freeze_panes = 'B3'

ws.merge_cells('A1:J1')
tc = ws.cell(row=1, column=1,
    value=f"Zone Scorecard — FY27 YTD ({em['fy27_period']}) | NSV in ₹ Lakh | Growth TGT: +20% on FY26")
tc.font = Font(bold=True, color=WHITE, size=12, name='Calibri')
tc.fill = PatternFill("solid", fgColor=NAVY)
tc.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

COLS = ['Zone', 'FY26 NSV (₹L)', 'FY27 TGT (₹L)', 'FY27 YTD ACH (₹L)', 'ACH %',
        'Gap vs TGT (₹L)', 'Run-Rate (Ann.)', 'Pipeline Ratio', 'Status', 'MoM Trend (FY27)']
hdr(ws, 2, len(COLS))
for i, h in enumerate(COLS, 1):
    ws.cell(row=2, column=i).value = h
ws.row_dimensions[2].height = 40

zones = em.get('by_zone', [])
total_fy26 = sum(z['fy26_nsv'] for z in zones)
total_ytd  = sum(z['fy27_ytd'] for z in zones)
total_tgt  = sum(z['fy27_tgt'] for z in zones)

for ri, z in enumerate(zones, 3):
    bg = LIGHT if ri % 2 == 0 else WHITE
    status = z.get('status', '')
    sc = GREEN if status == 'on_track' else (AMBER if status == 'monitor' else RED)
    cs(ws, ri, 1, z['name'], bold=True, bg=bg)
    cs(ws, ri, 2, z['fy26_nsv'], align="right", bg=bg, fmt='#,##0.00')
    cs(ws, ri, 3, z['fy27_tgt'], align="right", bg=bg, fmt='#,##0.00')
    cs(ws, ri, 4, z['fy27_ytd'], align="right", bg=bg, fmt='#,##0.00')
    ach = z.get('fy27_ach_pct')
    ac = cs(ws, ri, 5, f"{ach}%" if ach is not None else '–', align="center", bold=True, bg=bg)
    if ach is not None:
        ac.font = Font(bold=True, color=GREEN if ach>=100 else (AMBER if ach>=80 else RED), name='Calibri', size=9)
    gap = z.get('gap_vs_tgt', 0)
    gc = cs(ws, ri, 6, gap, align="right", bg=bg, fmt='#,##0.00')
    if gap < 0: gc.font = Font(color=RED, bold=True, name='Calibri', size=9)
    cs(ws, ri, 7, z['fy27_run_rate'], align="right", bg=bg, fmt='#,##0.00')
    p27 = z.get('pipeline_ratio_fy27')
    pc = cs(ws, ri, 8, f"{p27:.2f}" if p27 else '–', align="center", bold=True, bg=bg)
    if p27:
        pc.font = Font(bold=True, color=RED if p27>1.4 else (AMBER if p27<0.75 else GREEN), name='Calibri', size=9)
    stc = cs(ws, ri, 9, status.replace('_',' ').title(), align="center", bold=True, bg=bg)
    stc.font = Font(bold=True, color=sc, name='Calibri', size=9)
    monthly = z.get('monthly_fy27', {})
    trend = '  '.join(f"{m[:3]}:₹{v:.0f}L" for m, v in sorted(monthly.items()))
    cs(ws, ri, 10, trend or '–', bg=bg)

# Totals
tr = len(zones) + 3
for c in range(1, len(COLS)+1):
    ws.cell(row=tr, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=tr, column=c).font = Font(bold=True, color=WHITE, name='Calibri', size=9)
    ws.cell(row=tr, column=c).border = border
ws.cell(row=tr, column=1).value = 'TOTAL'
for col, val in [(2, total_fy26), (3, total_tgt), (4, total_ytd)]:
    ws.cell(row=tr, column=col).value = round(val, 2)
    ws.cell(row=tr, column=col).number_format = '#,##0.00'
    ws.cell(row=tr, column=col).alignment = Alignment(horizontal="right")
ov_ach = round(total_ytd/total_tgt*100, 1) if total_tgt else None
ws.cell(row=tr, column=5).value = f"{ov_ach}%" if ov_ach else '–'
ws.cell(row=tr, column=5).alignment = Alignment(horizontal="center")

for i, w in enumerate([20, 14, 14, 16, 9, 14, 14, 14, 14, 45], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

Path('exports').mkdir(exist_ok=True)
out1 = f"exports/Zone_Scorecard_FY27_YTD_{date.today().strftime('%Y%m%d')}.xlsx"
wb1.save(out1)
print(f"✅ Zone scorecard saved: {out1}")
print(f"   Zones: {len(zones)} | Overall ACH: {ov_ach}% | FY26: ₹{total_fy26:,.2f}L | FY27 YTD: ₹{total_ytd:,.2f}L")

# ── Category + Brand Contribution ─────────────────────────────────────────────
wb2 = Workbook()
wc = wb2.active
wc.title = "Category Contribution"
wc.merge_cells('A1:G1')
ct = wc.cell(row=1, column=1,
    value=f"Category Contribution — FY27 YTD ({em['fy27_period']}) vs FY26 | ₹ Lakh")
ct.font = Font(bold=True, color=WHITE, size=12, name='Calibri')
ct.fill = PatternFill("solid", fgColor=NAVY)
ct.alignment = Alignment(horizontal="center", vertical="center")
wc.row_dimensions[1].height = 28

CCOLS = ['Category', 'FY26 NSV (₹L)', 'FY27 YTD (₹L)', 'YoY Trend', 'FY26 Mix %', 'FY27 Mix %', 'Run-Rate (Ann.)']
hdr(wc, 2, len(CCOLS))
for i, h in enumerate(CCOLS, 1):
    wc.cell(row=2, column=i).value = h
wc.row_dimensions[2].height = 36

cats = em.get('by_category', [])
c26_tot = sum(c['fy26_nsv'] for c in cats)
c27_tot = sum(c['fy27_ytd'] for c in cats)
for ri, cat in enumerate(cats, 3):
    bg = LIGHT if ri % 2 == 0 else WHITE
    cs(wc, ri, 1, cat['name'], bold=True, bg=bg)
    cs(wc, ri, 2, cat['fy26_nsv'], align="right", bg=bg, fmt='#,##0.00')
    cs(wc, ri, 3, cat['fy27_ytd'], align="right", bg=bg, fmt='#,##0.00')
    trend = '▲' if cat['fy27_ytd'] > cat['fy26_nsv']/3 else '▼'  # rough ytd vs monthly avg
    cs(wc, ri, 4, trend, align="center", bg=bg, bold=True,
       color=GREEN if trend=='▲' else RED)
    mix26 = round(cat['fy26_nsv']/c26_tot*100, 1) if c26_tot else 0
    mix27 = round(cat['fy27_ytd']/c27_tot*100, 1) if c27_tot else 0
    cs(wc, ri, 5, f"{mix26}%", align="center", bg=bg)
    cs(wc, ri, 6, f"{mix27}%", align="center", bg=bg, bold=True,
       color=GREEN if mix27 > mix26 else (RED if mix27 < mix26-1 else "000000"))
    cs(wc, ri, 7, round(cat['fy27_ytd']*12/em['months_ytd'], 2), align="right", bg=bg, fmt='#,##0.00')
for i, w in enumerate([28, 14, 14, 10, 10, 10, 14], 1):
    wc.column_dimensions[get_column_letter(i)].width = w

# Brand tab
wb2_brands = wb2.create_sheet("Brand Contribution")
brands = em.get('by_brand', [])
b26_tot = sum(b['fy26_nsv'] for b in brands)
b27_tot = sum(b['fy27_ytd'] for b in brands)
BCOLS = ['Brand', 'FY26 NSV (₹L)', 'FY27 YTD (₹L)', 'FY26 Mix %', 'FY27 Mix %', 'Mix Shift', 'Run-Rate (Ann.)']
hdr(wb2_brands, 1, len(BCOLS))
for i, h in enumerate(BCOLS, 1):
    wb2_brands.cell(row=1, column=i).value = h
wb2_brands.row_dimensions[1].height = 36

for ri, b in enumerate(sorted(brands, key=lambda x: -x['fy26_nsv']), 2):
    bg = LIGHT if ri % 2 == 0 else WHITE
    cs(wb2_brands, ri, 1, b['name'], bold=True, bg=bg)
    cs(wb2_brands, ri, 2, b['fy26_nsv'], align="right", bg=bg, fmt='#,##0.00')
    cs(wb2_brands, ri, 3, b['fy27_ytd'], align="right", bg=bg, fmt='#,##0.00')
    m26 = round(b['fy26_nsv']/b26_tot*100,1) if b26_tot else 0
    m27 = round(b['fy27_ytd']/b27_tot*100,1) if b27_tot else 0
    cs(wb2_brands, ri, 4, f"{m26}%", align="center", bg=bg)
    cs(wb2_brands, ri, 5, f"{m27}%", align="center", bg=bg, bold=True)
    shift = round(m27-m26, 1)
    sc = cs(wb2_brands, ri, 6, f"+{shift}%" if shift>=0 else f"{shift}%", align="center", bold=True, bg=bg)
    sc.font = Font(bold=True, color=GREEN if shift>0 else (RED if shift<-1 else GRAY), name='Calibri', size=9)
    cs(wb2_brands, ri, 7, round(b['fy27_ytd']*12/em['months_ytd'], 2), align="right", bg=bg, fmt='#,##0.00')

for i, w in enumerate([24, 14, 14, 11, 11, 10, 14], 1):
    wb2_brands.column_dimensions[get_column_letter(i)].width = w

out2 = f"exports/Category_Brand_Contribution_FY27_{date.today().strftime('%Y%m%d')}.xlsx"
wb2.save(out2)
print(f"✅ Category + Brand contribution saved: {out2}")
print(f"   Categories: {len(cats)} | Brands: {len(brands)}")
