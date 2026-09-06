#!/usr/bin/env python3
"""
Forecast Excel schema validator.

Checks the structure and internal consistency of a Dynamic_Multi_Brand_Forecast
Excel workbook BEFORE it is fed to load_forecast_detail.py.  Can be called as
a standalone script or imported as a library (validate_file() returns the result
list; the caller decides whether to abort).

Exit codes (standalone):
  0  — all checks PASS or WARN only
  1  — at least one FAIL

Usage:
  python scripts/validate_forecast_excel.py path/to/forecast.xlsx
  python scripts/validate_forecast_excel.py path/to/forecast.xlsx --strict   # WARN → FAIL
"""
from __future__ import annotations
import argparse, math, re, sys
from pathlib import Path
from typing import Any

import openpyxl

# ── constants ──────────────────────────────────────────────────────────────────

REQUIRED_SHEETS = [
    "Forecast_Database",
    "Target_Master",
    "Control_Panel",
    "Forecast_Review",
    "Article_Master",
]

# Minimum required columns in Forecast_Database (case-sensitive, as they appear
# in the header row).  Additional columns are allowed.
REQUIRED_FDB_COLS = {
    "Forecast_Month",
    "Financial_Year",
    "Brand",
    "Chain",
    "Zone",
    "Forecast_Value",
    "Target_Value",
    "NSV_Rate",
    "Final_Forecast_Qty",
    "System_Forecast_Qty",
}

REQUIRED_TM_COLS = {"Month", "Target NSV Value"}

# Month label regex: "Sep'26", "Oct'26", etc.
_MONTH_RE = re.compile(r"^[A-Za-z]{3}'\d{2}$")

# Reconciliation tolerance: monthly Forecast_Value sum vs Target NSV Value
_RECON_TOL_PCT = 1.0   # percent

# ── result dataclass ───────────────────────────────────────────────────────────

PASS = "PASS"
WARN = "WARN"
FAIL = "FAIL"


class Result:
    __slots__ = ("level", "check", "detail")

    def __init__(self, level: str, check: str, detail: str = ""):
        self.level = level
        self.check = check
        self.detail = detail

    def __repr__(self) -> str:
        icon = {"PASS": "✓", "WARN": "⚠", "FAIL": "✗"}[self.level]
        line = f"  {icon} {self.level:<4}  {self.check}"
        if self.detail:
            line += f"\n          {self.detail}"
        return line


# ── helpers ────────────────────────────────────────────────────────────────────

def _sheet_rows(wb: openpyxl.Workbook, sheet: str) -> list[tuple]:
    ws = wb[sheet]
    return [tuple(c.value for c in row) for row in ws.iter_rows()]


def _find_header_row(rows: list[tuple], sentinel: str) -> int | None:
    """Return index of the first row whose first non-None cell equals sentinel."""
    for i, row in enumerate(rows):
        first = next((v for v in row if v is not None), None)
        if first is not None and str(first).strip() == sentinel:
            return i
    return None


def _col_index(headers: tuple, name: str) -> int | None:
    for i, h in enumerate(headers):
        if h is not None and str(h).strip() == name:
            return i
    return None


def _is_numeric(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return not math.isnan(float(v)) and not math.isinf(float(v))
    try:
        float(str(v).replace(",", ""))
        return True
    except ValueError:
        return False


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


# ── individual checks ──────────────────────────────────────────────────────────

def _chk_sheets(wb: openpyxl.Workbook) -> list[Result]:
    results = []
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        results.append(Result(FAIL, "Required sheets present",
                               f"Missing: {missing}"))
    else:
        results.append(Result(PASS, "Required sheets present",
                               f"Found all {len(REQUIRED_SHEETS)} required sheets"))
    return results


def _chk_fdb_columns(wb: openpyxl.Workbook) -> tuple[list[Result], dict | None]:
    """Check Forecast_Database columns; return (results, meta) where meta has index map."""
    results = []
    if "Forecast_Database" not in wb.sheetnames:
        return results, None

    rows = _sheet_rows(wb, "Forecast_Database")
    hi = _find_header_row(rows, "Forecast_Month")
    if hi is None:
        results.append(Result(FAIL, "Forecast_Database header row",
                               "No row with 'Forecast_Month' found"))
        return results, None

    headers = rows[hi]
    present = {str(h).strip() for h in headers if h is not None}
    missing = REQUIRED_FDB_COLS - present
    if missing:
        results.append(Result(FAIL, "Forecast_Database required columns",
                               f"Missing columns: {sorted(missing)}"))
    else:
        results.append(Result(PASS, "Forecast_Database required columns",
                               f"All {len(REQUIRED_FDB_COLS)} required columns present "
                               f"({len(present)} total)"))

    # Build index map for subsequent checks
    meta = {
        "header_idx": hi,
        "data_start": hi + 1,
        "rows": rows,
        "col": {str(h).strip(): i for i, h in enumerate(headers) if h is not None},
    }
    return results, meta


def _chk_fdb_data(meta: dict) -> list[Result]:
    """Check that Forecast_Database has non-empty data rows."""
    results = []
    data_rows = [r for r in meta["rows"][meta["data_start"]:]
                 if any(v is not None for v in r)]
    if not data_rows:
        results.append(Result(FAIL, "Forecast_Database data rows",
                               "No data rows found below the header"))
    else:
        results.append(Result(PASS, "Forecast_Database data rows",
                               f"{len(data_rows):,} rows"))
    return results


def _chk_month_format(meta: dict) -> list[Result]:
    """Validate that all Forecast_Month values match the expected format."""
    results = []
    col = meta["col"].get("Forecast_Month")
    if col is None:
        return results

    bad: list[str] = []
    seen: set[str] = set()
    for row in meta["rows"][meta["data_start"]:]:
        if all(v is None for v in row):
            continue
        v = row[col] if col < len(row) else None
        if v is None:
            continue
        s = str(v).strip()
        if not _MONTH_RE.match(s):
            bad.append(repr(s))
        else:
            seen.add(s)

    if bad:
        sample = bad[:5]
        results.append(Result(FAIL, "Forecast_Month format",
                               f"Invalid values (sample): {sample}; "
                               f"expected Mon'YY e.g. Sep'26"))
    else:
        results.append(Result(PASS, "Forecast_Month format",
                               f"All values match Mon'YY; months present: "
                               f"{sorted(seen)}"))
    return results


def _chk_value_columns(meta: dict) -> list[Result]:
    """Check Forecast_Value and Target_Value are numeric and non-negative."""
    results = []
    for col_name in ("Forecast_Value", "Target_Value", "NSV_Rate"):
        ci = meta["col"].get(col_name)
        if ci is None:
            continue
        bad_rows = 0
        neg_rows = 0
        total_rows = 0
        for row in meta["rows"][meta["data_start"]:]:
            if all(v is None for v in row):
                continue
            v = row[ci] if ci < len(row) else None
            if v is None:
                continue
            total_rows += 1
            f = _to_float(v)
            if f is None:
                bad_rows += 1
            elif f < 0:
                neg_rows += 1

        if total_rows == 0:
            results.append(Result(WARN, f"{col_name} values", "No non-null values found"))
        elif bad_rows > 0:
            results.append(Result(FAIL, f"{col_name} values",
                                   f"{bad_rows:,}/{total_rows:,} rows have non-numeric values"))
        elif neg_rows > 0:
            results.append(Result(WARN, f"{col_name} values",
                                   f"{neg_rows:,}/{total_rows:,} rows have negative values"))
        else:
            results.append(Result(PASS, f"{col_name} values",
                                   f"All {total_rows:,} non-null values are numeric and ≥ 0"))
    return results


def _chk_brands_chains(meta: dict) -> list[Result]:
    """Check at least one brand and one chain are present."""
    results = []
    for dim, col_name in [("Brand", "Brand"), ("Chain", "Chain")]:
        ci = meta["col"].get(col_name)
        if ci is None:
            results.append(Result(WARN, f"{dim} coverage", "Column not found"))
            continue
        vals = {str(row[ci]).strip()
                for row in meta["rows"][meta["data_start"]:]
                if any(v is not None for v in row)
                and ci < len(row) and row[ci] is not None}
        if not vals:
            results.append(Result(FAIL, f"{dim} coverage", "No non-null values found"))
        else:
            results.append(Result(PASS, f"{dim} coverage",
                                   f"{len(vals)} unique {dim.lower()}s: "
                                   f"{sorted(vals)[:8]}{'…' if len(vals)>8 else ''}"))
    return results


def _chk_target_master(wb: openpyxl.Workbook) -> tuple[list[Result], dict[str, float]]:
    """Validate Target_Master and return {month: target_nsv} mapping."""
    results = []
    targets: dict[str, float] = {}

    if "Target_Master" not in wb.sheetnames:
        return results, targets

    rows = _sheet_rows(wb, "Target_Master")
    # Header row starts with "MonthStart"; "Month" is the second column.
    hi = _find_header_row(rows, "MonthStart")
    if hi is None:
        results.append(Result(FAIL, "Target_Master header row",
                               "No row with 'MonthStart' found in first column"))
        return results, targets

    headers = rows[hi]
    month_ci = _col_index(headers, "Month")
    tgt_ci   = _col_index(headers, "Target NSV Value")

    if month_ci is None or tgt_ci is None:
        missing = [c for c, ci in [("Month", month_ci), ("Target NSV Value", tgt_ci)]
                   if ci is None]
        results.append(Result(FAIL, "Target_Master required columns",
                               f"Missing: {missing}"))
        return results, targets

    results.append(Result(PASS, "Target_Master required columns", ""))
    bad_months = []
    for row in rows[hi + 1:]:
        if all(v is None for v in row):
            continue
        mo = row[month_ci] if month_ci < len(row) else None
        tv = row[tgt_ci]   if tgt_ci  < len(row) else None
        if mo is None:
            continue
        mo_s = str(mo).strip()
        if not _MONTH_RE.match(mo_s):
            bad_months.append(repr(mo_s))
            continue
        f = _to_float(tv)
        if f is None:
            results.append(Result(WARN, f"Target_Master [{mo_s}] Target NSV Value",
                                   f"Non-numeric: {tv!r}"))
        else:
            targets[mo_s] = f

    if bad_months:
        results.append(Result(FAIL, "Target_Master Month format",
                               f"Invalid values: {bad_months}"))
    elif targets:
        results.append(Result(PASS, "Target_Master months",
                               f"{len(targets)} months: {sorted(targets)} "
                               f"(values in ₹)"))
    return results, targets


def _chk_reconciliation(meta: dict, targets: dict[str, float]) -> list[Result]:
    """Cross-check: sum of Forecast_Value by month ≈ Target_Master NSV Value."""
    results = []
    if not targets or "Forecast_Value" not in meta["col"]:
        results.append(Result(WARN, "Monthly reconciliation",
                               "Skipped — Target_Master or Forecast_Value not available"))
        return results

    fv_ci = meta["col"]["Forecast_Value"]
    fm_ci = meta["col"].get("Forecast_Month")
    if fm_ci is None:
        return results

    sums: dict[str, float] = {}
    for row in meta["rows"][meta["data_start"]:]:
        if all(v is None for v in row):
            continue
        mo = row[fm_ci] if fm_ci < len(row) else None
        fv = row[fv_ci] if fv_ci < len(row) else None
        if mo is None:
            continue
        mo_s = str(mo).strip()
        f = _to_float(fv)
        if f is not None:
            sums[mo_s] = sums.get(mo_s, 0.0) + f

    all_pass = True
    detail_lines = []
    for mo, tgt in sorted(targets.items()):
        actual = sums.get(mo, 0.0)
        if tgt == 0:
            pct_diff = 0.0 if actual == 0 else 100.0
        else:
            pct_diff = abs(actual - tgt) / abs(tgt) * 100
        status = PASS if pct_diff <= _RECON_TOL_PCT else FAIL
        if status == FAIL:
            all_pass = False
        detail_lines.append(
            f"{mo}: db_sum={actual/1e5:.2f}L tgt={tgt/1e5:.2f}L diff={pct_diff:.2f}% [{status}]"
        )
        results.append(Result(status, f"Reconciliation [{mo}]",
                               detail_lines[-1]))

    if all_pass and detail_lines:
        # Replace per-month rows with one summary
        results = results[-len(detail_lines):]   # keep them for detail output
    return results


def _chk_fdb_month_coverage(meta: dict, targets: dict[str, float]) -> list[Result]:
    """Every Target_Master month must have rows in Forecast_Database."""
    results = []
    if not targets:
        return results

    fm_ci = meta["col"].get("Forecast_Month")
    if fm_ci is None:
        return results

    db_months = {str(row[fm_ci]).strip()
                 for row in meta["rows"][meta["data_start"]:]
                 if any(v is not None for v in row)
                 and fm_ci < len(row) and row[fm_ci] is not None}

    missing = set(targets.keys()) - db_months
    extra   = db_months - set(targets.keys())

    if missing:
        results.append(Result(FAIL, "Forecast_Database month coverage",
                               f"Target months missing from DB: {sorted(missing)}"))
    else:
        results.append(Result(PASS, "Forecast_Database month coverage",
                               f"All {len(targets)} target months present in DB; "
                               f"extra DB months: {sorted(extra) or 'none'}"))
    return results


# ── public API ─────────────────────────────────────────────────────────────────

def validate_file(path: str | Path) -> list[Result]:
    """Run all checks against *path*.  Returns list of Result objects."""
    path = Path(path)
    results: list[Result] = []

    if not path.exists():
        return [Result(FAIL, "File exists", f"Not found: {path}")]

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        return [Result(FAIL, "File readable", str(exc))]

    results += _chk_sheets(wb)

    fdb_results, meta = _chk_fdb_columns(wb)
    results += fdb_results

    if meta:
        results += _chk_fdb_data(meta)
        results += _chk_month_format(meta)
        results += _chk_value_columns(meta)
        results += _chk_brands_chains(meta)

    tm_results, targets = _chk_target_master(wb)
    results += tm_results

    if meta and targets:
        results += _chk_fdb_month_coverage(meta, targets)
        results += _chk_reconciliation(meta, targets)

    wb.close()
    return results


def print_report(path: str | Path, results: list[Result]) -> None:
    width = 56
    print("=" * width)
    print(" Forecast Excel Schema Validator")
    print(f" File: {Path(path).name}")
    print("=" * width)
    for r in results:
        print(repr(r))
    counts = {PASS: 0, WARN: 0, FAIL: 0}
    for r in results:
        counts[r.level] = counts.get(r.level, 0) + 1
    print("-" * width)
    print(f" SUMMARY: {counts[PASS]} PASS  {counts[WARN]} WARN  {counts[FAIL]} FAIL")
    print("=" * width)


# ── CLI ────────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("xlsx", help="Path to the forecast Excel workbook")
    ap.add_argument("--strict", action="store_true",
                    help="Treat WARN items as FAIL (non-zero exit)")
    args = ap.parse_args()

    results = validate_file(args.xlsx)
    print_report(args.xlsx, results)

    has_fail = any(r.level == FAIL for r in results)
    has_warn = any(r.level == WARN for r in results)
    if has_fail or (args.strict and has_warn):
        sys.exit(1)


if __name__ == "__main__":
    main()
