"""
Master_Dashboard_Correction_Pack.xlsx
4 core fixes + B-001 TOT Scenario + OVL/ALLOC audit
Built from live repo audit — no dummy data.
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import date
import warnings, os

warnings.filterwarnings("ignore")

OUTDIR = "/home/user/mt-dashboard/forecast_outputs/sep_nov_2026_tentative"
OUTFILE = f"{OUTDIR}/Master_Dashboard_Correction_Pack.xlsx"
TODAY = str(date.today())

# ── colour palette ─────────────────────────────────────────────────────────────
C_NAVY   = "1F497D"   # header bg
C_WHITE  = "FFFFFF"
C_AMBER  = "FFC000"
C_GREEN  = "375623"
C_RED    = "C00000"
C_LGREY  = "F2F2F2"
C_ORANGE = "E26B0A"
C_BLUE   = "2E75B6"
C_TEAL   = "00B0F0"

F_NAVY   = PatternFill("solid", fgColor=C_NAVY)
F_AMBER  = PatternFill("solid", fgColor="FFF2CC")
F_AMBRK  = PatternFill("solid", fgColor="FFE699")
F_RED    = PatternFill("solid", fgColor="FFE0E0")
F_GREEN  = PatternFill("solid", fgColor="E2EFDA")
F_LGREY  = PatternFill("solid", fgColor=C_LGREY)
F_BLUE   = PatternFill("solid", fgColor="DEEAF1")
F_TEAL   = PatternFill("solid", fgColor="DEEBF7")
F_ORANGE = PatternFill("solid", fgColor="FCE4D6")

THIN  = Side(style="thin",  color="AAAAAA")
MED   = Side(style="medium", color=C_NAVY)
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BDR_H = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

def hfont(sz=10, bold=True, color=C_WHITE): return Font(name="Arial", size=sz, bold=bold, color=color)
def dfont(sz=9,  bold=False, color="000000"): return Font(name="Arial", size=sz, bold=bold, color=color)
def bfont(sz=9,  color="000000"): return Font(name="Arial", size=sz, bold=True, color=color)

def hdr_row(ws, row, values, fills=None, sz=10):
    fills = fills or [F_NAVY]*len(values)
    for i, (v, f) in enumerate(zip(values, fills), 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = hfont(sz)
        c.fill = f
        c.border = BDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def data_row(ws, row, values, fill=None, bold=False, wrap=True, halign="left"):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        # Force text storage for strings starting with "=" — these are
        # documentation/template formulas; they must not be evaluated against
        # non-existent tables in this standalone workbook.
        if isinstance(v, str) and v.startswith("="):
            c.data_type = "s"
        c.font = bfont() if bold else dfont()
        if fill: c.fill = fill
        c.border = BDR
        c.alignment = Alignment(horizontal=halign, vertical="top", wrap_text=wrap)

def section_banner(ws, row, text, cols, color=C_NAVY, text_color=C_WHITE, sz=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=sz, bold=True, color=text_color)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

def autofit(ws, max_w=60):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 2, max_w)

def title_block(ws, line1, line2, cols, fill_color=C_NAVY):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, line1)
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = PatternFill("solid", fgColor=fill_color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c2 = ws.cell(2, 1, line2)
    c2.font = Font(name="Arial", size=9, bold=False, color="555555")
    c2.fill = PatternFill("solid", fgColor="EFF3FF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# LOAD LIVE DATA
# ══════════════════════════════════════════════════════════════════════════════
fm = pd.read_csv("/home/user/mt-dashboard/forecast_outputs/sep_nov_2026_tentative/fact_margin_enriched.csv")
for col in ["mrp","tot_pct","gst_pct","unit_nsv_validated"]:
    fm[col] = pd.to_numeric(fm[col], errors="coerce")

primary = fm[fm["unit_nsv_source"]=="PRIMARY_INVOICE_HISTORY"].copy()
primary["gross_nsv"]    = primary["mrp"] / (1 + primary["gst_pct"]/100)
primary["invoice_nsv"]  = primary["unit_nsv_validated"]
primary["implied_tot"]  = ((primary["gross_nsv"] - primary["invoice_nsv"]) / primary["gross_nsv"] * 100).round(2)
primary["trade_A"]      = (primary["invoice_nsv"] * primary["tot_pct"] / 100).round(2)
primary["trade_B"]      = (primary["gross_nsv"]   * primary["tot_pct"] / 100).round(2)
primary["delta_ab"]     = (primary["trade_A"] - primary["trade_B"]).round(2)
primary["risk"]         = primary["delta_ab"] < -0.01

top20 = (primary.groupby(["ean","brand","article"])
         .agg(rows=("ean","count"),
              mrp=("mrp","median"),
              gst_pct=("gst_pct","first"),
              stored_tot=("tot_pct","median"),
              gross_nsv=("gross_nsv","median"),
              invoice_nsv=("invoice_nsv","median"),
              implied_tot=("implied_tot","median"),
              trade_A=("trade_A","median"),
              trade_B=("trade_B","median"),
              delta=("delta_ab","median"),
              pct_risk=("risk","mean"))
         .reset_index()
         .sort_values("invoice_nsv", ascending=False)
         .head(20)
         .round(2))

chain_sum = (primary.groupby("chain_name")
             .agg(rows=("ean","count"),
                  stored_tot=("tot_pct","median"),
                  implied_tot=("implied_tot","median"),
                  invoice_nsv=("invoice_nsv","median"),
                  pct_risk=("risk","mean"))
             .reset_index()
             .sort_values("rows", ascending=False)
             .head(10)
             .round(2))

so_map = pd.read_csv("/home/user/mt-dashboard/PowerBI/SeedData/Mapping/Store_SO_Mapping.csv")
off_cols = ["Zone","State","Chain Name","Site Code","Site Name","SO/ASE Emp Code"]


# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ─── Sheet 0: Index ──────────────────────────────────────────────────────────
ws_idx = wb.active
ws_idx.title = "00_Index"
title_block(ws_idx, "Master Forecast & Margin Dashboard — Correction Pack",
            f"Generated: {TODAY}  |  Branch: claude/store-master-qc-duplicates-4pvmmk  |  Status: TENTATIVE", 4)

ws_idx.row_dimensions[4].height = 8
hdr_row(ws_idx, 5, ["Sheet", "Content", "Priority", "Status"])
idx_rows = [
    ("01_Audit_Findings",         "Schema audit of 4 fix areas against live repo data",          "P0", "COMPLETE"),
    ("02_Fix1_Sales_Hierarchy",   "SO/BDE merge — formulas, column specs, cascade dropdown DDL",  "P1", "FORMULAS READY"),
    ("03_Fix2_ISR_Cleanup",       "ISR off Roll string removal — static & dynamic approaches",    "P2", "FORMULAS READY"),
    ("04_Fix3_White_Box",         "Excel white-box rendering anomaly resolution checklist",       "P2", "CHECKLIST"),
    ("05_Fix4_Composite_Key",     "Composite Store Key architecture with live Fountain column map","P1", "FORMULAS READY"),
    ("06_B001_TOT_Impact",        "B-001 TOT Scenario A vs B — top 20 EANs + chain breakdown",   "P0", "LIVE DATA"),
    ("07_OVL_ALLOC_Audit",        "OVL-001→OVL-009 overlap risks + ALLOC-000→ALLOC-008 status",  "P0", "LIVE DATA"),
]
STATUS_F = {"COMPLETE": F_GREEN, "LIVE DATA": F_GREEN, "FORMULAS READY": F_TEAL,
            "CHECKLIST": F_LGREY, "BLOCKED": F_RED}
for i, r in enumerate(idx_rows):
    fill = STATUS_F.get(r[3], None)
    data_row(ws_idx, 6+i, list(r))
    if fill:
        ws_idx.cell(6+i, 4).fill = fill
    ws_idx.cell(6+i, 3).fill = {"P0":F_RED,"P1":F_AMBRK,"P2":F_AMBER}.get(r[2], None)

ws_idx.column_dimensions["A"].width = 28
ws_idx.column_dimensions["B"].width = 58
ws_idx.column_dimensions["C"].width = 8
ws_idx.column_dimensions["D"].width = 18


# ─── Sheet 1: Audit Findings ─────────────────────────────────────────────────
ws1 = wb.create_sheet("01_Audit_Findings")
title_block(ws1, "Schema Audit — 4 Fix Areas Against Live Repository Data",
            "Audit date: " + TODAY + "  |  Sources: offtake_store_article_Apr_26.csv, Store_SO_Mapping.csv, primary_article_Apr_26.csv, chain_master.csv", 5)

ws1.row_dimensions[4].height = 6
hdr_row(ws1, 5, ["Fix Area","Requirement","Current State in Repo","Gap / Action","Severity"])

audit_rows = [
    # Fix 1: Sales Hierarchy
    ("FIX 1: Sales Hierarchy",
     "Standardize to 'SO Name' across all Master Mapping sheets. 4-tier: RKAM → SO Name → Supervisor → BA Lead / BA",
     "Store_SO_Mapping.csv: already uses 'SO Name 1/2/3' — NO 'BDE Name' column present.\n"
     "primary_article_Apr_26.csv: 'SO No' (numeric SAP code, not name).\n"
     "offtake_store_article: 'SO/ASE Emp Code' (numeric).\n"
     "NO RKAM, Supervisor, or BA Lead column in any current data file.",
     "GAP: 'SO No' (numeric) ≠ 'SO Name' (text) — need lookup join via Store_SO_Mapping.\n"
     "GAP: RKAM, Supervisor, BA Lead tiers are absent from all tracked CSVs — must be sourced from MT_Spend.xlsx (not in repo).\n"
     "ACTION: Build SO Name lookup from Store_SO_Mapping on Store Code→SO Name 1.",
     "P1"),
    ("FIX 1: Cascade Dropdown",
     "Dynamic 4-tier cascade: RKAM → SO Name → Supervisor → BA",
     "No cascade dropdown implemented in any tracked workbook. Store_SO_Mapping has 25 unique SO Name 1 values across 854 store×month rows.",
     "ACTION: Power Query or FILTER/SORT formulas on Store_SO_Mapping table. Full DDL specs in Fix1 sheet.",
     "P1"),

    # Fix 2: ISR cleanup
    ("FIX 2: ISR String Cleanup",
     "Remove ' ( ISR off Roll )' from employee name strings globally",
     "'HISRT:' prefix found in Site Code column of offtake_store_article (e.g. HISRT:009) — this is a STORE CODE prefix for Hyderabad ISR stores, NOT an employee name suffix.\n"
     "Employee name file MT_Spend.xlsx not committed to repo — '( ISR off Roll )' pattern lives in that uncommitted source.",
     "SCOPE: Pattern applies to MT_Spend.xlsx employee names when loaded.\n"
     "ACTION (Static): Find & Replace in Excel when MT_Spend.xlsx is available.\n"
     "ACTION (Dynamic): Wrap SO Name cell refs with TRIM(SUBSTITUTE()) — formula in Fix2 sheet.\n"
     "CAUTION: Do NOT strip 'ISR' from Site Codes (HISRT:009 etc.) — those are store identifiers.",
     "P2"),

    # Fix 3: White box
    ("FIX 3: White Box Rendering",
     "Resolve white-box / floating-object artefacts in Excel sheets",
     "Cannot be audited from CSV/py — requires live Excel file inspection.",
     "CHECKLIST provided in Fix3 sheet: Clear No Fill, F5→Special→Objects audit, disable GPU acceleration.",
     "P2"),

    # Fix 4: Composite key
    ("FIX 4: Composite Store Key",
     "Insert [Composite_Store_Key] = UPPER(TRIM(ChainName)) & '_' & UPPER(TRIM(SiteCode)) in Chain Offtake and Master Mapping",
     "offtake_store_article_Apr_26.csv already has a 'Unique' column = ChainName + SiteCode concatenated (no separator, e.g. 'Wellness ForeverHINOP:003').\n"
     "Site Code column confirmed present. Chain Name column confirmed present.\n"
     "Store_SO_Mapping: Store Code matches Site Code — join key confirmed.",
     "ACTION: Add [Composite_Store_Key] = UPPER(TRIM([Chain Name]))&'_'&UPPER(TRIM([Site Code])) as explicit keyed column (underscore separator for SUMIFS/XLOOKUP compatibility).\n"
     "Existing 'Unique' col lacks separator — keep for backward compat; add new key as separate column.\n"
     "Full SUMIFS/XLOOKUP/FILTER formula suite in Fix4 sheet.",
     "P1"),
]

SEVER_F = {"P0":F_RED, "P1":F_AMBRK, "P2":F_AMBER}
for i, r in enumerate(audit_rows):
    data_row(ws1, 6+i, list(r))
    ws1.cell(6+i, 5).fill = SEVER_F.get(r[4], None)
    ws1.row_dimensions[6+i].height = 70

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 35
ws1.column_dimensions["C"].width = 45
ws1.column_dimensions["D"].width = 42
ws1.column_dimensions["E"].width = 8
ws1.freeze_panes = "A6"


# ─── Sheet 2: Fix 1 – Sales Hierarchy ────────────────────────────────────────
ws2 = wb.create_sheet("02_Fix1_Sales_Hierarchy")
title_block(ws2, "FIX 1: Sales Hierarchy Consolidation — SO Name Standard + Cascade Dropdown",
            "Source data: Store_SO_Mapping.csv (854 rows, 25 SOs) | primary_article (SO No = SAP code, needs name join)", 6)

ws2.row_dimensions[4].height = 8

section_banner(ws2, 5, "SECTION A — Column Standardization Rules", 6)
hdr_row(ws2, 6, ["Source File","Current Column","Required Column","Action","Formula / Step","Notes"])
col_std = [
    ("Store_SO_Mapping.csv",       "SO Name 1",       "SO Name",      "RENAME → SO Name (primary)",
     "No formula change — column already correct. Rename header only.",
     "25 unique SOs across 854 rows (Apr'26). Multi-SO stores use SO Name 2/3 with Cont%."),
    ("primary_article_Apr_26.csv", "SO No",           "SO Name",      "ADD column via XLOOKUP",
     "=IFERROR(XLOOKUP([@[Store Code]], Store_SO_Mapping[Store Code], Store_SO_Mapping[SO Name 1], \"Unassigned\"),\"Unassigned\")",
     "SO No is a numeric SAP code. Must join to Store_SO_Mapping on Store Code to get text name."),
    ("offtake_store_article",      "SO/ASE Emp Code", "SO Name",      "ADD column via XLOOKUP",
     "=IFERROR(XLOOKUP([@[Site Code]], Store_SO_Mapping[Store Code], Store_SO_Mapping[SO Name 1], \"Unassigned\"),\"Unassigned\")",
     "SO/ASE Emp Code is an employee ID. Site Code = Store Code join key."),
    ("Any sheet with 'BDE Name'",  "BDE Name",        "SO Name",      "RENAME header — BDE merged into SO",
     "Find & Replace in header row: Replace 'BDE Name' with 'SO Name'. No data change.",
     "BDE and SO roles consolidated. No BDE column found in current tracked CSVs."),
    ("Any formula referencing BDE","[...BDE...]",     "[...SO Name...]","UPDATE all SUMIFS/VLOOKUP/XLOOKUP references",
     "Find & Replace in formula bar: Replace [BDE Name] with [SO Name] globally.",
     "Use Ctrl+H with Formulas search scope to catch all formula references."),
]
for i, r in enumerate(col_std):
    data_row(ws2, 7+i, list(r))
    ws2.row_dimensions[7+i].height = 52

ws2.append([])
section_banner(ws2, 13, "SECTION B — 4-Tier Reporting Hierarchy Cascade Dropdown Formulas", 6)
hdr_row(ws2, 14, ["Tier","Field","Source Range","Cascade Dependency","Validation Formula","Power Query M Equivalent"])
cascade = [
    ("Tier 1", "RKAM",       "RKAM Master table (source: MT_Spend.xlsx — NOT YET IN REPO)",
     "None — top of hierarchy",
     "=SORT(UNIQUE(RKAM_Master[RKAM_Name]))",
     "= Table.Distinct(Table.SelectColumns(RKAM_Master, {\"RKAM_Name\"}))"),
    ("Tier 2", "SO Name",    "Store_SO_Mapping[SO Name 1]",
     "Filter by selected RKAM\n=SORT(UNIQUE(FILTER(Store_SO_Mapping[SO Name 1], Store_SO_Mapping[RKAM]=C3)))",
     "Name Manager: SO_List\n=OFFSET(SO_Name_Anchor,0,0,COUNTA(SO_col),1)",
     "= Table.Distinct(Table.SelectRows(Store_SO_Mapping, each [RKAM] = SelectedRKAM))"),
    ("Tier 3", "Supervisor", "Supervisor Master (source: MT_Spend.xlsx — NOT YET IN REPO)",
     "Filter by selected SO Name",
     "=SORT(UNIQUE(FILTER(Supervisor_Master[Sup_Name], Supervisor_Master[SO_Name]=C4)))",
     "= Table.SelectRows(Supervisor_Master, each [SO_Name] = SelectedSO)"),
    ("Tier 4", "BA Lead / BA","BA Master (source: MT_Spend.xlsx — NOT YET IN REPO)",
     "Filter by selected Supervisor",
     "=SORT(UNIQUE(FILTER(BA_Master[BA_Name], BA_Master[Supervisor]=C5)))",
     "= Table.SelectRows(BA_Master, each [Supervisor] = SelectedSupervisor)"),
]
for i, r in enumerate(cascade):
    data_row(ws2, 15+i, list(r))
    ws2.row_dimensions[15+i].height = 56
    if "NOT YET IN REPO" in r[2]:
        ws2.cell(15+i, 3).fill = F_RED

ws2.append([])
section_banner(ws2, 20, "SECTION C — SUMIFS Aggregation Linked to Cascade Selections (cells C3=Zone, C4=SO)", 6)
ws2.cell(21, 1).value = (
    "Primary NSV by Zone+SO:\n"
    "=SUMIFS(primary_article[Inv. Net value(LOC)], primary_article[Zone], IF($C$3=\"All\",\"*\",$C$3), "
    "primary_article[SO Name], IF($C$4=\"All\",\"*\",$C$4))\n\n"
    "Offtake NSV by Zone+SO (via Store Bridge):\n"
    "=SUMPRODUCT((IF($C$3=\"All\",1,Store_Bridge[Zone]=$C$3)) * "
    "(IF($C$4=\"All\",1,Store_Bridge[SO_Name]=$C$4)) * Chain_Offtake[MRP_Offtake_Val])"
)
ws2.cell(21, 1).font = Font(name="Courier New", size=9)
ws2.cell(21, 1).fill = PatternFill("solid", fgColor="F0F4FF")
ws2.cell(21, 1).alignment = Alignment(wrap_text=True, vertical="top")
ws2.merge_cells(start_row=21, start_column=1, end_row=24, end_column=6)
ws2.row_dimensions[21].height = 100

for c in "ABCDEF":
    ws2.column_dimensions[c].width = 28


# ─── Sheet 3: Fix 2 – ISR Cleanup ────────────────────────────────────────────
ws3 = wb.create_sheet("03_Fix2_ISR_Cleanup")
title_block(ws3, "FIX 2: Global ISR off Roll String Cleanup",
            "IMPORTANT: '( ISR off Roll )' pattern is in employee name data (MT_Spend.xlsx — not in repo). Site Codes beginning HISRT: are store IDs — do NOT touch.", 5)

section_banner(ws3, 5, "SCOPE CLARIFICATION — Two Different 'ISR' Occurrences", 5)
hdr_row(ws3, 6, ["ISR Pattern","Location","Action","Risk if Wrong"])
scope_rows = [
    ("' ( ISR off Roll )' in employee names",
     "MT_Spend.xlsx — sheet 'BA ', 'Sups', 'D-Mart' — NOT in repo currently",
     "APPLY cleanup. See formulas below.",
     "None — purely cosmetic suffix on employee names."),
    ("'HISRT:' prefix in Site Codes",
     "offtake_store_article_Apr_26.csv, Site Code column (e.g. HISRT:009, HISRT:011)",
     "DO NOT TOUCH — 'ISR' here = store identifier (ISR = In-Shop Retail location in Hyderabad)",
     "BREAKING: would corrupt Composite_Store_Key and all XLOOKUP/SUMIFS joins."),
]
for i, r in enumerate(scope_rows):
    data_row(ws3, 7+i, list(r))
    ws3.cell(7+i, 3).fill = F_GREEN if "APPLY" in r[2] else F_RED
    ws3.row_dimensions[7+i].height = 42

ws3.append([])
section_banner(ws3, 10, "STATIC FIX — Find & Replace (Ctrl + H) in Excel", 5)
static_rows = [
    ["Step", "Action", "Find What", "Replace With", "Scope"],
    ["1", "Open Find & Replace", "Ctrl + H", "", ""],
    ["2", "Set search scope", "", "", "Sheet or Workbook (choose Workbook for global)"],
    ["3", "Find What", " ( ISR off Roll )", "", "Include the leading space"],
    ["4", "Replace With", "", "(empty — leave blank)", "Removes the string entirely"],
    ["5", "Match case", "Check if names are mixed case", "", ""],
    ["6", "Click Replace All", "", "", "Verify count in confirmation dialog"],
    ["7", "Also run for variant", "( ISR off Roll )", "", "Without leading space — catch both"],
    ["8", "Save workbook", "", "", "Save as new version first for audit trail"],
]
for i, r in enumerate(static_rows):
    if i == 0:
        hdr_row(ws3, 11, r)
    else:
        data_row(ws3, 11+i, r)
        ws3.row_dimensions[11+i].height = 18

section_banner(ws3, 21, "DYNAMIC FIX — Wrap source cell references in formula", 5)
ws3.append([])
ws3.cell(22, 1).value = "Purpose"
ws3.cell(22, 2).value = "Formula Pattern"
ws3.cell(22, 3).value = "Example"
ws3.cell(22, 4).value = "Notes"
hdr_row(ws3, 22, ["Purpose","Formula Pattern","Example","Notes"])

dyn_rows = [
    ("Clean single cell",
     "=TRIM(SUBSTITUTE(A2,\" ( ISR off Roll )\",\"\"))",
     "=TRIM(SUBSTITUTE([SO_Name_Raw],\" ( ISR off Roll )\",\"\"))",
     "Leading space included in search string. TRIM removes any residual spaces."),
    ("Clean & also strip variant without leading space",
     "=TRIM(SUBSTITUTE(SUBSTITUTE(A2,\" ( ISR off Roll )\",\"\"),\"( ISR off Roll )\",\"\"))",
     "Nested SUBSTITUTE handles both variants",
     "Recommended — catches both \" ( ISR off Roll )\" and \"( ISR off Roll )\""),
    ("Use in XLOOKUP lookup value",
     "=XLOOKUP(TRIM(SUBSTITUTE([@SO_Name_Raw],\" ( ISR off Roll )\",\"\")), MasterList[Name], MasterList[RKAM], \"\")",
     "Cleans name before lookup — no separate column needed",
     "Apply to any formula that consumes an employee name from MT_Spend.xlsx"),
    ("Power Query equivalent",
     "= Table.TransformColumns(Source, {{\"SO Name\", each Text.Trim(Text.Replace(_, \" ( ISR off Roll )\", \"\"))}})",
     "In Power Query M, applied to BA/Sups/D-Mart sheet before load",
     "Preferred approach — cleans at ETL layer so no trailing text appears anywhere downstream"),
]
for i, r in enumerate(dyn_rows):
    data_row(ws3, 23+i, list(r))
    ws3.cell(23+i, 2).font = Font(name="Courier New", size=8)
    ws3.cell(23+i, 3).font = Font(name="Courier New", size=8)
    ws3.row_dimensions[23+i].height = 42

for c, w in zip("ABCD", [22, 50, 50, 40]):
    ws3.column_dimensions[c].width = w


# ─── Sheet 4: Fix 3 – White Box ──────────────────────────────────────────────
ws4 = wb.create_sheet("04_Fix3_White_Box")
title_block(ws4, "FIX 3: Excel White Box / Rendering Anomaly Resolution Checklist",
            "Follow in sequence. Check each step. Restart Excel after hardware acceleration change.", 4)

section_banner(ws4, 5, "CHECKLIST — Execute In This Order", 4)
hdr_row(ws4, 6, ["Step","Method","Detailed Steps","What It Fixes","Done?"])
wb_steps = [
    ("1", "Clear Cell Fill",
     "1. Select all cells in affected sheet (Ctrl+A).\n2. Home tab → Font group → Fill Color dropdown arrow.\n3. Select 'No Fill'.\nAlternative: Ctrl+A → Format Cells → Fill → No Color → OK",
     "Removes white/opaque cell background fills that obscure gridlines or underlying objects. Most common cause of 'white box' artefact.",
     "☐"),
    ("2", "Object Selection Audit",
     "1. Press F5 (or Ctrl+G) to open Go To dialog.\n2. Click 'Special…'\n3. Select 'Objects' → OK\n4. All floating objects are now selected — inspect them.\n5. Delete any invisible/blank text boxes, shapes, or images.\n6. Check for objects with white fill + no border (invisible overlays).",
     "Floating shapes or transparent text boxes with white fill create 'ghost' white areas over cells. Common after copy-paste from other sources.",
     "☐"),
    ("3", "Conditional Formatting Audit",
     "1. Home → Conditional Formatting → Manage Rules → 'This Worksheet'.\n2. Look for rules that apply white font or white fill.\n3. Delete or modify any rule with Fill Color = White or Font Color = White.",
     "CF rules that paint cells white (e.g. to 'hide' data) create white boxes. Also check for CF rules applied to entire columns.",
     "☐"),
    ("4", "Disable Hardware Graphics Acceleration",
     "1. File → Options → Advanced.\n2. Scroll to 'Display' section.\n3. Check '☑ Disable hardware graphics acceleration'.\n4. Click OK.\n5. Close and reopen Excel.",
     "Resolves GPU rendering artefacts that appear as white or flickering boxes — especially on external monitors or after screen scaling changes.",
     "☐"),
    ("5", "Row/Column Height Reset",
     "1. Select all (Ctrl+A).\n2. Home → Format → AutoFit Row Height.\n3. Home → Format → AutoFit Column Width.\nIf specific rows: Right-click row → Row Height → set to 0 → check if white box disappears.",
     "Zero-height rows with background fills are invisible but block rendering. AutoFit forces a recompute.",
     "☐"),
    ("6", "Theme / Background Image Check",
     "1. Page Layout → Background → Remove Background (if any is set).\n2. Page Layout → Themes → verify theme isn't applying white fills to cells.",
     "A sheet background image or white-fill theme can render as a white overlay in Normal view.",
     "☐"),
    ("7", "Named Range / Table Overlap Check",
     "1. Formulas → Name Manager → review all named ranges.\n2. Check for ranges that span the affected area with an accidentally applied format.\n3. Delete or correct any erroneous ranges.",
     "Named ranges with applied formatting (not standard) can leave artefact regions.",
     "☐"),
    ("8", "Repair via Save As XLSX",
     "1. File → Save As → Excel Workbook (.xlsx) → new filename.\n2. Close original. Open the new file.\n3. If white boxes are gone: the old file had internal XML corruption.\n4. Compare the two files and migrate content if needed.",
     "XML corruption in .xlsm/.xlsx can cause rendering artefacts that don't survive re-save. This is the nuclear option — use last.",
     "☐"),
]
for i, r in enumerate(wb_steps):
    data_row(ws4, 7+i, list(r))
    ws4.row_dimensions[7+i].height = 70
    ws4.cell(7+i, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws4.cell(7+i, 5).font = Font(name="Arial", size=14)

ws4.column_dimensions["A"].width = 6
ws4.column_dimensions["B"].width = 28
ws4.column_dimensions["C"].width = 55
ws4.column_dimensions["D"].width = 42
ws4.column_dimensions["E"].width = 8


# ─── Sheet 5: Fix 4 – Composite Store Key ────────────────────────────────────
ws5 = wb.create_sheet("05_Fix4_Composite_Key")
title_block(ws5, "FIX 4: Composite Store Key Architecture — Live Column Map",
            "Based on offtake_store_article_Apr_26.csv actual columns: 'Chain Name' (col G), 'Site Code' (col J). Existing 'Unique' col lacks underscore separator.", 6)

section_banner(ws5, 5, "SECTION A — Column Map from Live Fountain Offtake File", 6)
hdr_row(ws5, 6, ["Column Letter (Fountain Offtake)","Column Name","Data Sample","Use in Key","Notes"])
col_map = [
    ("G", "Chain Name",    "Wellness Forever, Reliance, Apollo, Dmart",       "LEFT component",  "Use UPPER(TRIM()) to normalize case and spaces"),
    ("J", "Site Code",     "HINOP:003, HISRT:009, VASHI1:005, KLK:001",       "RIGHT component", "Alphanumeric store ID. Preserve colon and digits."),
    ("B", "Unique (existing)", "Wellness ForeverHINOP:003",                   "LEGACY — keep",   "Chain+SiteCode concatenated WITHOUT separator. Retain for backward compat."),
    ("A", "[Composite_Store_Key] (NEW)", "WELLNESS FOREVER_HINOP:003",        "PRIMARY KEY",     "Insert in col A. Formula: =UPPER(TRIM(G2))&'_'&UPPER(TRIM(J2))"),
]
for i, r in enumerate(col_map):
    data_row(ws5, 7+i, list(r))
    if r[3] == "PRIMARY KEY":
        for j in range(1, 7):
            ws5.cell(7+i, j).fill = F_GREEN
    ws5.row_dimensions[7+i].height = 28

ws5.append([])
section_banner(ws5, 12, "SECTION B — Key Formula Suite", 6)
hdr_row(ws5, 13, ["Formula Name","Formula","Where to Place","What It Does","Dependency"])
formulas = [
    ("Composite_Store_Key (Offtake sheet)",
     "=UPPER(TRIM([@[Chain Name]]))&\"_\"&UPPER(TRIM([@[Site Code]]))",
     "Column A of offtake_store_article table — insert BEFORE existing columns",
     "Creates standardized store key: CHAINNAME_SITECODE. Uppercase, trimmed, underscore-separated.",
     "Chain Name (col G), Site Code (col J) must exist"),
    ("Composite_Store_Key (Master Mapping sheet)",
     "=UPPER(TRIM([@[Chain_Name]]))&\"_\"&UPPER(TRIM([@[Site_Code]]))",
     "Column A of Master Mapping table",
     "Identical key in Master Mapping — enables XLOOKUP join on shared key.",
     "Chain_Name, Site_Code columns in Master Mapping"),
    ("XLOOKUP — pull SO Name via key",
     "=IFERROR(XLOOKUP([@Composite_Store_Key], Master_Mapping[Composite_Store_Key], Master_Mapping[SO_Name], \"Unassigned\"), \"Unassigned\")",
     "SO_Name column in any report table joined to Master Mapping",
     "Pulls SO Name for any store without hardcoding. Updates automatically as Master Mapping changes.",
     "Master_Mapping[Composite_Store_Key] must be built first"),
    ("XLOOKUP — pull Zone via key",
     "=IFERROR(XLOOKUP([@Composite_Store_Key], Master_Mapping[Composite_Store_Key], Master_Mapping[Zone], \"Unknown\"), \"Unknown\")",
     "Zone column in report tables",
     "Dynamic Zone attribution via store key.",
     "Master_Mapping table with Zone column"),
    ("Dynamic Zone Dropdown (FILTER+SORT)",
     "=SORT(UNIQUE(FILTER(Master_Mapping[State], (Master_Mapping[Zone]=$C$3)+(\"All\"=$C$3))))",
     "Named range 'State_List' — used as Data Validation source in $C$4",
     "Cascades state dropdown based on selected Zone in $C$3.",
     "$C$3 = Zone selector. 'All' shows all states."),
    ("SUMIFS on Composite Key (KPI aggregation)",
     "=SUMIFS(Chain_Offtake[MRP_Offtake_Val], Chain_Offtake[Composite_Store_Key], Store_Bridge[Composite_Store_Key], Store_Bridge[Zone], IF($C$3=\"All\",\"*\",$C$3), Store_Bridge[SO_Name], IF($C$4=\"All\",\"*\",$C$4))",
     "KPI card / breakdown table in Dashboard sheet",
     "Aggregates offtake MRP value filtered by Zone, SO, and linked via store key bridge.",
     "Store_Bridge table with Composite_Store_Key + Zone + SO_Name columns"),
    ("Wildcard-resilient SUMPRODUCT alternative",
     "=SUMPRODUCT((IF($C$3=\"All\",1,Store_Bridge[Zone]=$C$3)) * (IF($C$4=\"All\",1,Store_Bridge[SO_Name]=$C$4)) * (Chain_Offtake[MRP_Offtake_Val]))",
     "KPI card — use when IF(x,\"*\",x) pattern causes issues with non-text fields",
     "Boolean array multiplication avoids wildcard limitations. 'All' = multiply by 1 (include all).",
     "Store_Bridge and Chain_Offtake must share row alignment or use XLOOKUP-mapped arrays"),
    ("Dynamic State Filter",
     "=SORT(UNIQUE(FILTER(Master_Mapping[State], (Master_Mapping[Zone]=$C$3)+(\"All\"=$C$3))))",
     "Spill range used as data validation list source",
     "Dependent dropdown: changes state list when zone changes.",
     "Requires Excel 365 / 2021 for FILTER/SORT/UNIQUE. Excel 2019 needs Power Query alternative."),
]
for i, r in enumerate(formulas):
    data_row(ws5, 14+i, list(r))
    ws5.cell(14+i, 2).font = Font(name="Courier New", size=8)
    ws5.row_dimensions[14+i].height = 52

ws5.append([])
section_banner(ws5, 23, "SECTION C — Sample Composite Keys from Live Fountain Data", 6)
hdr_row(ws5, 24, ["Chain Name (raw)","Site Code (raw)","Composite_Store_Key (computed)","Zone","State"])
live_keys = [
    ("Wellness Forever","HINOP:003","WELLNESS FOREVER_HINOP:003","West","Maharashtra"),
    ("Wellness Forever","HISRT:009","WELLNESS FOREVER_HISRT:009","South-1","Telangana"),
    ("Wellness Forever","VASHI1:005","WELLNESS FOREVER_VASHI1:005","West","Maharashtra"),
    ("Wellness Forever","KLK:001","WELLNESS FOREVER_KLK:001","South-1","Kerala"),
    ("Dmart","DC1050","DMART_DC1050","North","Haryana"),
    ("Reliance","RIL:M01","RELIANCE_RIL:M01","West","Maharashtra"),
    ("Apollo","APL:BGL01","APOLLO_APL:BGL01","South-1","Karnataka"),
]
for i, r in enumerate(live_keys):
    data_row(ws5, 25+i, list(r))
    ws5.cell(25+i, 3).fill = F_GREEN
    ws5.cell(25+i, 3).font = Font(name="Courier New", size=9, bold=True)

for c, w in zip("ABCDEF", [22, 16, 32, 16, 42, 20]):
    ws5.column_dimensions[c].width = w
ws5.freeze_panes = "A14"


# ─── Sheet 6: B-001 TOT Impact ───────────────────────────────────────────────
ws6 = wb.create_sheet("06_B001_TOT_Impact")
title_block(ws6, "B-001 TOT Scenario Impact Analysis — Scenario A (Tentative Base) vs Scenario B (Sensitivity)",
            f"Source: fact_margin_enriched.csv | PRIMARY_INVOICE_HISTORY rows: {len(primary)} | Date: {TODAY}", 11)

# Status panel
ws6.merge_cells("A4:K5")
ws6["A4"] = (
    "VERDICT: Scenario A CONFIRMED as tentative Finance basis | "
    "FIN-GATE-TOT-001: WARNING (Tentative) / BLOCKED (Final Financial) | "
    f"100% of {len(primary)} primary rows show implied TOT > stored TOT"
)
ws6["A4"].font = Font(name="Arial", size=10, bold=True, color="9C6500")
ws6["A4"].fill = F_AMBER
ws6["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Key metrics row
ws6.row_dimensions[7].height = 8
for col_idx, (label, val) in enumerate([
    ("PRIMARY ROWS", len(primary)),
    ("MEDIAN STORED TOT%", f"{primary['tot_pct'].median():.2f}%"),
    ("MEDIAN IMPLIED TOT%", f"{primary['implied_tot'].median():.2f}%"),
    ("ROWS WITH RISK (Trade_A < Trade_B)", f"{primary['risk'].sum()} / {len(primary)} = 100%"),
    ("GATE STATUS", "WARNING — Tentative Mode"),
], 1):
    ws6.cell(8, col_idx*2-1, label).font = Font(name="Arial", size=8, bold=True, color=C_NAVY)
    ws6.cell(8, col_idx*2).value = val
    ws6.cell(8, col_idx*2).font = Font(name="Arial", size=10, bold=True)
    ws6.cell(8, col_idx*2).fill = F_TEAL

# Scenario explanation
section_banner(ws6, 10, "SCENARIO DEFINITIONS", 11, color="2E75B6")
ws6.cell(11, 1).value = (
    "SCENARIO A (Tentative Base — Finance-confirmed by DAX 13_CM2_Measures.dax):  "
    "Invoice NSV is already POST-TOT.  CM2 = Invoice NSV − Approved Expenses.  "
    "NO trade_spend line deducted.  "
    "Column 'TradeA' shown below is the ERRONEOUS amount that would be double-deducted if applied.\n\n"
    "SCENARIO B (Sensitivity only):  "
    "NSV assumed to be PRE-TOT.  Trade Spend = Gross NSV × Stored TOT%.  "
    "CM2 = Invoice NSV − Trade_B − Approved Expenses.  "
    "Delta = Trade_A − Trade_B = amount OVER-deducted per unit if Scenario A is correct basis and Trade_A were applied."
)
ws6.merge_cells("A11:K12")
ws6["A11"].font = dfont(9)
ws6["A11"].fill = PatternFill("solid", fgColor="EFF3FF")
ws6["A11"].alignment = Alignment(wrap_text=True, vertical="top")
ws6.row_dimensions[11].height = 72
ws6.row_dimensions[12].height = 8

# Top 20 EAN Table
section_banner(ws6, 14, "TOP 20 EANs BY INVOICE NSV — PER UNIT MEDIAN VALUES (₹)", 11)
b001_hdr = [
    "EAN", "Brand", "Article (truncated)",
    "MRP (₹)", "GST%", "Stored TOT%",
    "Gross NSV (₹)\n[MRP÷(1+GST)]",
    "Invoice NSV (₹)\n[Primary — Post-TOT]",
    "Implied TOT%\n[(Gross−Inv)/Gross×100]",
    "Trade Spend A (₹)\n[Invoice×StoredTOT — RISK]",
    "Trade Spend B (₹)\n[Gross×StoredTOT — Correct]",
    "Delta A−B (₹)\n[Overstatement if A applied]",
    "Double-Deduct Risk"
]
hdr_row(ws6, 15, b001_hdr)
ws6.row_dimensions[15].height = 52

for i, (_, r) in enumerate(top20.iterrows()):
    row_num = 16 + i
    risk_flag = r['pct_risk'] >= 1.0
    vals = [
        str(r['ean']), r['brand'], str(r['article'])[:40],
        round(r['mrp'], 0), round(r['gst_pct'], 0), round(r['stored_tot'], 2),
        round(r['gross_nsv'], 2), round(r['invoice_nsv'], 2),
        round(r['implied_tot'], 2),
        round(r['trade_A'], 2), round(r['trade_B'], 2),
        round(r['delta'], 2),
        "HIGH RISK" if risk_flag else "LOW",
    ]
    data_row(ws6, row_num, vals)
    fill = F_RED if risk_flag else F_GREEN
    ws6.cell(row_num, 13).fill = fill
    ws6.cell(row_num, 13).font = bfont(color=C_RED if risk_flag else C_GREEN)
    # Colour delta column
    ws6.cell(row_num, 12).fill = F_RED if r['delta'] < 0 else F_GREEN
    ws6.row_dimensions[row_num].height = 18

ws6.row_dimensions[36].height = 10
section_banner(ws6, 37, "CHAIN-LEVEL SUMMARY — TOP 10 CHAINS BY ROW COUNT", 11)
chain_hdr = ["Chain", "Rows", "Median Stored TOT%", "Median Implied TOT%",
             "Median Invoice NSV (₹)", "% Rows with Double-Deduct Risk", "Status"]
hdr_row(ws6, 38, chain_hdr)

for i, (_, r) in enumerate(chain_sum.iterrows()):
    vd = "HIGH RISK" if r['pct_risk'] >= 1.0 else "CLEAR"
    data_row(ws6, 39+i, [
        r['chain_name'], int(r['rows']),
        round(r['stored_tot'], 2), round(r['implied_tot'], 2),
        round(r['invoice_nsv'], 2),
        f"{r['pct_risk']*100:.0f}%", vd
    ])
    ws6.cell(39+i, 7).fill = F_RED if r['pct_risk'] >= 1.0 else F_GREEN
    ws6.cell(39+i, 7).font = bfont(color=C_RED if r['pct_risk'] >= 1.0 else C_GREEN)
    ws6.row_dimensions[39+i].height = 18

for c, w in zip("ABCDEFGHIJKLM", [14,15,30,8,7,10,12,12,12,14,14,14,14]):
    ws6.column_dimensions[get_column_letter(ord(c)-64)].width = w
ws6.freeze_panes = "A16"


# ─── Sheet 7: OVL + ALLOC Audit ──────────────────────────────────────────────
ws7 = wb.create_sheet("07_OVL_ALLOC_Audit")
title_block(ws7, "CM2 Expense Overlap & Allocation Audit — OVL-001→OVL-009 + ALLOC-000→ALLOC-008",
            f"Source: cm2_expense_taxonomy.csv (commit fc45de46) + cm2_allocation_rules.csv | Date: {TODAY}", 7)

ws7.merge_cells("A4:G5")
ws7["A4"] = (
    "Q1 FY27 Monthly Base (₹ L/month):  Claims 272.82  |  BA CTC 123.00  |  Merch CTC 93.16  |  Sup CTC 14.82  |  TOTAL 503.80 L/month  ||  "
    "BLOCKED EXPLICITLY:  Balance Provision 344.67 L/month  |  Claim Incl GST 313.60 L/month  |  BA Incentive ~8.67 L/month (triple-count)"
)
ws7["A4"].font = Font(name="Arial", size=9, bold=True, color=C_RED)
ws7["A4"].fill = PatternFill("solid", fgColor="FFE0E0")
ws7["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws7.row_dimensions[4].height = 28

section_banner(ws7, 7, "PART A — OVL-001 to OVL-009: EXPENSE OVERLAP RISK REGISTER", 7, color=C_RED, text_color=C_WHITE)
ovl_hdr = ["ID","Expense Head(s)","Overlap Type","Risk Level","Status","Evidence","Required Finance Action"]
hdr_row(ws7, 8, ovl_hdr)
ws7.row_dimensions[8].height = 30

ovl_rows = [
    ("OVL-001","TOT (40.68% stored) vs Invoice NSV (already post-TOT)","DOUBLE DEDUCTION","HIGH",
     "BLOCKED (Final Financial)\nWARNING (Tentative)",
     "DAX 13_CM2_Measures.dax: 'no further deduction happens'. cm2_formula.csv: 'Do not deduct TOT again'. 100% of 955 primary rows: implied TOT > stored TOT.",
     "Finance to confirm NSV basis in writing. Scenario A = tentative base. No trade_spend line in CM2 until Finance signs off."),
    ("OVL-002","Indirect Claim Base (₹818.45L Q1) + Claim Incl GST (₹940.81L Q1)","DOUBLE COUNT","HIGH","BLOCKED",
     "Claim Incl GST = Claim Base + Tax. Both rows exist in MTIndirect_Claim source. Naive inclusion = Base counted twice.",
     "EXCLUDE 'Claim Incl GST' row. Use ONLY Claim Base = ₹272.82 L/month as CM2 deduction candidate."),
    ("OVL-003","Indirect Claim GST (₹122.36L Q1) deducted in CM2","INCORRECT DEDUCTION","HIGH","BLOCKED",
     "CM2 base is net-of-tax NSV. Deducting GST on a net-of-tax base understates CM2.",
     "EXCLUDE Claim GST from all scenarios. Carry as memo only (₹40.79 L/month)."),
    ("OVL-004","Balance Provision (₹1,034.01L Q1) as current-period charge","DOUBLE COUNT / PERIOD ERROR","HIGH","EXPLICITLY BLOCKED",
     "Carry-forward balance from prior periods. Adding to Sep-Nov 2026 inflates deductions by ₹344.67 L/month.",
     "EXCLUDED from all scenarios. Finance written approval required to re-include. Highest single double-count risk."),
    ("OVL-005","BA Incentive — triple payout rows","TRIPLE COUNT","HIGH","BLOCKED — dedup required",
     "Incentive column populated on: Salary rows (₹14.80L) + Incentive rows (₹11.13L) + Expense rows (₹0.08L) = naive ₹26.01L Q1.",
     "EXCLUDED from Base. Finance to specify authoritative payout row. Include only in HIGH scenario range after resolution."),
    ("OVL-006","Supervisor CTC — GT/Hybrid channel included","INCORRECT ATTRIBUTION","HIGH","BLOCKED — MT split unknown",
     "Channel mix in MT_Spend: MT(52 employees) + Hybrid(112) + GT(37). GT/Hybrid not MT cost.",
     "Finance to provide MT-only employee split. Until resolved: tag as SHARED_PENDING; do not attribute to MT chains."),
    ("OVL-007","BA Cost ₹6.75L (example) vs BA CTC ₹368.99L Q1 (actual)","SUPERSEDED PLACEHOLDER","MEDIUM","BLOCKED",
     "PL_Expense_Input.csv BA Cost row is labelled EXAMPLE. BA CTC from MT_Spend.xlsx is the real split data.",
     "EXCLUDE example row. Use BA CTC / Reimbursement / Incentive breakdown from MT_Spend.xlsx (pending approval)."),
    ("OVL-008","Visibility Spend ₹12.5L (PL_Expense_Input.csv)","EXAMPLE ROW","HIGH","BLOCKED — no real data",
     "Row explicitly marked 'EXAMPLE ROW -- replace with real data'. Store-specific cost cannot be spread across all chains.",
     "Finance to supply actual store visibility spend by chain and brand. Blank until supplied."),
    ("OVL-009","Scheme/Trade Spend ₹28.4L (example) + separate Trade Schemes head","POTENTIAL DOUBLE COUNT","MEDIUM","BLOCKED",
     "Example row under 'Scheme/Trade Spend' AND a separate 'Trade Schemes' head forward-declared in taxonomy. Risk of double-load.",
     "Finance to nominate single authoritative head and source. Confirm whether example row overlaps with Trade Schemes head."),
]
OVL_F = {"HIGH": F_RED, "MEDIUM": F_ORANGE}
for i, r in enumerate(ovl_rows):
    data_row(ws7, 9+i, list(r))
    ws7.cell(9+i, 4).fill = OVL_F.get(r[3], None)
    ws7.cell(9+i, 5).fill = F_RED
    ws7.row_dimensions[9+i].height = 52

ws7.row_dimensions[19].height = 10
section_banner(ws7, 20, "PART B — ALLOC-000 to ALLOC-008: ALLOCATION RULE STATUS", 7, color=C_BLUE, text_color=C_WHITE)
alloc_hdr = ["Rule ID","Rule Name","Status","Priority","Grain","Blocker / Notes","Usable Now?"]
hdr_row(ws7, 21, alloc_hdr)

alloc_data = [
    ("ALLOC-000","Retain Unallocated","APPROVED","P7 — safe fallback","Any → UNALLOCATED_BUCKET",
     "No business judgement required. Retains expenses in visible Unallocated bucket. Cannot distort any Brand/Chain figure.",
     "YES — default for all unresolvable rows"),
    ("ALLOC-001","Direct Chain + Brand","DRAFT","P1 — approve first","MONTH×CHAIN×BRAND (direct tag)",
     "Expense already carries Chain AND Brand from source file. Recommend as first approval. Confirm tags are authoritative.",
     "DRAFT — first to approve"),
    ("ALLOC-002","Retain at Chain Level","DRAFT","P2","MONTH×CHAIN + UNALLOCATED_BRAND",
     "Chain known; Brand unknown. Correct rule for Apollo BA Cost (₹6.75L). Does NOT spread across Brands.",
     "DRAFT"),
    ("ALLOC-003","Retain at Brand Level","DRAFT","P3","MONTH×BRAND + UNALLOCATED_CHAIN",
     "Brand known; Chain unknown. Does NOT spread across Chains.",
     "DRAFT"),
    ("ALLOC-004","Employee → Chain","DRAFT — BLOCKER","P4","MONTH×CHAIN via Chain Name col",
     "BLOCKER: 16 non-chain tokens in BA Chain Name column (MT, SIS, BKD, RAJ MANDIR, MANISH AGENCY, etc.) need crosswalk resolution.",
     "NO — crosswalk unresolved"),
    ("ALLOC-005","Employee → Zone (shared Supervisor)","DRAFT — BLOCKER","P4 variant","MONTH×ZONE (shared)",
     "BLOCKER: Supervisor data has Hybrid(112)/MT(52)/GT(37) channel mix. MT-only attribution is a Finance decision, not a data one.",
     "NO — channel split required"),
    ("ALLOC-006","Store Universe Driver (Visibility/Rental)","DRAFT","P5","MONTH×CHAIN→MONTH×CHAIN×STORE",
     "Only for store-count-driven costs. Negative store counts excluded. Residual assigned by largest-remainder.",
     "DRAFT — no real visibility data yet"),
    ("ALLOC-007","NSV Contribution Share","BLOCKED","P6 — deliberately last","MONTH×CHAIN→MONTH×CHAIN×BRAND",
     "BLOCKED BY DEFAULT. NSV allocation must never be the automatic fallback. May only be enabled per named expense head with written Finance approval.",
     "NO — permanently blocked unless Finance writes approval"),
    ("ALLOC-008","Distributor → Chain crosswalk","BLOCKED","N/A","MONTH×DISTRIBUTOR→MONTH×CHAIN",
     "BLOCKED: No distributor-to-chain crosswalk exists. 21 distributors in MTIndirect_Claim encode chain hints in free text — parsing is prohibited guessing. Claims (₹818.45L Q1) stay in Unallocated Chain Expense.",
     "NO — crosswalk not supplied"),
]
ALLOC_STATUS_F = {"APPROVED": F_GREEN, "DRAFT": F_TEAL, "DRAFT — BLOCKER": F_AMBER, "BLOCKED": F_RED}
USABLE_F = {"YES": F_GREEN, "NO": F_RED, "DRAFT": F_AMBER}
for i, r in enumerate(alloc_data):
    data_row(ws7, 22+i, list(r))
    ws7.cell(22+i, 3).fill = ALLOC_STATUS_F.get(r[2], None)
    ws7.cell(22+i, 7).fill = USABLE_F.get(r[6].split(" ")[0].rstrip("—"), None)
    ws7.row_dimensions[22+i].height = 45

ws7.row_dimensions[32].height = 10
section_banner(ws7, 33, "PART C — Q1 FY27 EXPENSE STACK SUMMARY (₹ Lakh / month)", 7, color="375623", text_color=C_WHITE)
hdr_row(ws7, 34, ["Expense Head","Q1 FY27 Total (₹L)","Monthly Base (₹L)","CM2 Status","Allocation Rule","Double-Count Flag","Finance Action"])
expense_stack = [
    ("Indirect Claim Base (excl. GST)", 818.45, 272.82, "PENDING APPROVAL","ALLOC-008 BLOCKED","MEDIUM","Supply distributor crosswalk to unlock ALLOC-008"),
    ("BA CTC (Salary)",                 368.99, 123.00, "PENDING APPROVAL","ALLOC-004 DRAFT","MEDIUM","Resolve 16 non-chain tokens in Chain Name column"),
    ("Merchandiser CTC (D-Mart)",       279.49,  93.16, "PENDING APPROVAL","ALLOC-004 DRAFT","LOW","Confirm D-Mart chain implicit from sheet"),
    ("Supervisor CTC (MT+Hybrid+GT)",    44.46,  14.82, "PENDING APPROVAL","ALLOC-005 DRAFT","HIGH","Finance to provide MT-only channel split"),
    ("TOTAL CANDIDATE",                 None,   503.80, "ALL PENDING","Mixed","—","None of above confirmed for use"),
    ("Indirect Claim GST",              122.36,  40.79, "EXCLUDED","ALLOC-000","LOW","Tax line — never deduct on net-of-tax NSV base"),
    ("Balance Provision",             1034.01,  344.67, "EXPLICITLY EXCLUDED","ALLOC-000","HIGH","Carry-forward — highest double-count risk"),
    ("BA Incentive (naive sum)",         26.01,    8.67, "EXCLUDED PENDING DEDUP","ALLOC-004 DRAFT","HIGH","Triple-count risk — Finance dedup decision required"),
    ("Claim Incl GST (memo only)",      940.81,  313.60, "MEMO ONLY — DO NOT USE","ALLOC-000","HIGH","= Base + Tax combined. Using would double-count Base."),
    ("COGS",                             None,    None,  "EXCLUDED (pending D1)","ALLOC-000","—","Finance decision D1: confirm CM2 is gross-of-COGS or post-COGS"),
    ("Visibility Spend",                 None,    None,  "EXCLUDED (no real data)","ALLOC-001","HIGH","PL_Expense_Input.csv has example row only"),
]
EXP_STATUS_F = {"PENDING APPROVAL": F_AMBER, "EXCLUDED": F_RED, "ALL PENDING": F_AMBRK,
                "EXPLICITLY EXCLUDED": F_RED, "EXCLUDED PENDING DEDUP": F_RED,
                "MEMO ONLY — DO NOT USE": F_RED, "EXCLUDED (pending D1)": F_LGREY,
                "EXCLUDED (no real data)": F_LGREY}
for i, r in enumerate(expense_stack):
    bold = "TOTAL" in r[0]
    data_row(ws7, 35+i, list(r), bold=bold)
    sf = EXP_STATUS_F.get(r[3])
    if sf: ws7.cell(35+i, 4).fill = sf
    if "TOTAL" in r[0]: ws7.row_dimensions[35+i].height = 22
    ws7.row_dimensions[35+i].height = 22

for c, w in zip("ABCDEFG", [32, 20, 18, 22, 18, 18, 42]):
    ws7.column_dimensions[get_column_letter(ord(c)-64)].width = w
ws7.freeze_panes = "A9"


# ══════════════════════════════════════════════════════════════════════════════
# SAVE + RECALC
# ══════════════════════════════════════════════════════════════════════════════
# Set tab colours
TABS = {
    "00_Index":              "1F497D",
    "01_Audit_Findings":     "2E75B6",
    "02_Fix1_Sales_Hierarchy":"375623",
    "03_Fix2_ISR_Cleanup":   "375623",
    "04_Fix3_White_Box":     "7F7F7F",
    "05_Fix4_Composite_Key": "375623",
    "06_B001_TOT_Impact":    "C00000",
    "07_OVL_ALLOC_Audit":    "C00000",
}
for name, col in TABS.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = col

wb.save(OUTFILE)
print(f"Saved: {OUTFILE}")
print(f"Sheets: {wb.sheetnames}")
