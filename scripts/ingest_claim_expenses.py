#!/usr/bin/env python3
"""
Ingest real chain-level claim expenses from the Apr-Jun 2026 claim master into:
  1. dashboard/claim_data.json  — standalone claim audit file
  2. dashboard/data.js          — injects D.claims into DASH and patches D.cm2.by_chain

Usage:
    python scripts/ingest_claim_expenses.py \\
        --src <path-to-xlsx> \\
        --out-claim dashboard/claim_data.json \\
        --out-datajs dashboard/data.js
"""

import argparse, json, re, sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Only map Excel chain names → canonical names that ACTUALLY exist in D.cm2
# ---------------------------------------------------------------------------
CHAIN_MAP = {
    "D-Mart":                 "DMart",
    "Avenue Supermarts Ltd.": "DMart",
    "Apollo":                 "Apollo",
    "Reliance Retail":        "Reliance Retail",
    "Health & Glow":          "Health & Glow",
}


def inr_to_lakh(v: float) -> float:
    return round(v / 1e5, 2) if v else 0.0


def load_chain_summary(xlsx_path: Path) -> list[dict]:
    """Chain Summary sheet → per-chain totals in INR and Lakhs."""
    ws = openpyxl.load_workbook(xlsx_path, data_only=True)["Chain Summary"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not row[0]:
            continue
        total_inr = row[7] or 0.0
        rows.append({
            "chain_raw":  str(row[0]).strip(),
            "canonical":  CHAIN_MAP.get(str(row[0]).strip()),
            "total_inr":  round(total_inr, 2),
            "total_lakh": inr_to_lakh(total_inr),
            "total_claim_lakh": inr_to_lakh(total_inr),   # canonical field name for D.claims
            "by_category_inr": {
                "Chain Promo (On Invoice)":       round(row[1] or 0, 2),
                "Extra Margin / Rate Difference": round(row[2] or 0, 2),
                "Freight / Transportation":       round(row[3] or 0, 2),
                "Incentive":                      round(row[4] or 0, 2),
                "Off Invoice / Debit Note Promo": round(row[5] or 0, 2),
                "Visibility":                     round(row[6] or 0, 2),
            },
        })
    return rows


def load_claim_master(xlsx_path: Path) -> list[dict]:
    """Claim Master sheet → transaction rows."""
    ws = openpyxl.load_workbook(xlsx_path, data_only=True)["Claim Master"]
    rows = []
    header_found = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not header_found:
            if row[0] == "Month":
                header_found = True
            continue
        if not row[0]:
            continue
        rows.append({
            "month":      str(row[0]),
            "distributor": str(row[1]),
            "chain":       str(row[3]),
            "category":    str(row[4]),
            "actual_expense": row[9] or 0.0,
            "review_status": str(row[13]) if row[13] else "Included",
        })
    return rows


def build_by_chain_claims(chain_rows: list) -> dict:
    """
    Build D.claims.by_chain object keyed by display chain name.
    Uses total_claim_lakh (Lakhs) with by_category breakdown.
    Chains not in CHAIN_MAP appear under their raw name so the UI can still
    show them even if they don't cross-reference D.primary.
    """
    by_chain: dict[str, dict] = {}
    for r in chain_rows:
        name = r["canonical"] or r["chain_raw"]  # use canonical if available
        if name not in by_chain:
            by_chain[name] = {
                "total_claim_lakh": 0.0,
                "source_chains": [],
                "by_category": {},
            }
        by_chain[name]["total_claim_lakh"] = round(
            by_chain[name]["total_claim_lakh"] + r["total_claim_lakh"], 2)
        by_chain[name]["source_chains"].append(r["chain_raw"])
        for cat, amt_inr in r["by_category_inr"].items():
            by_chain[name]["by_category"][cat] = round(
                by_chain[name]["by_category"].get(cat, 0) + inr_to_lakh(amt_inr), 2)
    return by_chain


def build_by_dist_claims(txn_rows: list) -> dict:
    """Build D.claims.by_distributor from transaction rows."""
    agg: dict[str, dict] = {}
    for r in txn_rows:
        d = r["distributor"]
        m = r["month"]
        amt = r["actual_expense"]
        if d not in agg:
            agg[d] = {"total_inr": 0.0, "by_month": {}}
        agg[d]["total_inr"] = round(agg[d]["total_inr"] + amt, 2)
        agg[d]["by_month"][m] = round(agg[d]["by_month"].get(m, 0) + amt, 2)

    result = {}
    for d, data in agg.items():
        months = data["by_month"]
        n_months = len(months) or 1
        total_lakh = inr_to_lakh(data["total_inr"])
        result[d] = {
            "total_claim_lakh": total_lakh,
            "avg_monthly": round(total_lakh / n_months, 2),
            "apr_lakh": inr_to_lakh(months.get("Apr-2026", 0)),
            "may_lakh": inr_to_lakh(months.get("May-2026", 0)),
            "jun_lakh": inr_to_lakh(months.get("Jun-2026", 0)),
        }
    return result


def build_claim_data(chain_rows, txn_rows, xlsx_path) -> dict:
    """Full claim_data.json payload."""
    by_chain  = build_by_chain_claims(chain_rows)
    by_dist   = build_by_dist_claims(txn_rows)
    grand_inr = sum(r["total_inr"] for r in chain_rows)
    grand_lakh = inr_to_lakh(grand_inr)
    exc_count = sum(1 for r in txn_rows if r.get("review_status") in ["Exception", "Pending"])
    return {
        "metadata": {
            "source": xlsx_path.name,
            "months": ["Apr-2026", "May-2026", "Jun-2026"],
            "period": "Q1 FY27",
            "generated_at": datetime.now().isoformat(),
        },
        "claims": {
            "grand_total_inr":  round(grand_inr, 2),
            "grand_total_lakh": grand_lakh,
            "by_chain":         by_chain,
            "by_distributor":   by_dist,
            "quality_summary": {
                "total_records":         len(txn_rows),
                "exceptions_count":      exc_count,
                "chains_with_claims":    len(by_chain),
                "distributors_with_claims": len(by_dist),
            },
        },
    }


def inject_claims_into_datajs(datajs_path: Path, by_chain: dict, by_dist: dict,
                               grand_lakh: float) -> None:
    """
    Inject D.claims key into data.js DASH object, and patch D.cm2.by_chain
    expense values for the 4 matched canonical chains.
    """
    text = datajs_path.read_text(encoding="utf-8")

    # ── 1. Inject / replace D.claims in DASH ──────────────────────────────
    claims_payload = {"by_chain": by_chain, "by_distributor": by_dist}
    claims_json = json.dumps(claims_payload, ensure_ascii=False, separators=(",", ":"))

    if '"claims":' in text:
        # Replace existing claims block
        idx = text.find('"claims":')
        start = text.index("{", idx)
        depth, end = 0, start
        for i, ch in enumerate(text[start:], start):
            if ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0: end = i + 1; break
        text = text[:idx] + f'"claims":{claims_json}' + text[end:]
        print(f"  ✅ Replaced D.claims in data.js  ({len(by_chain)} chains, {len(by_dist)} distributors)")
    else:
        # Insert after opening brace of DASH object
        dash_idx = text.find("window.DASH = {")
        if dash_idx == -1:
            dash_idx = text.find("window.DASH={")
        insert_after = text.index("{", dash_idx) + 1
        text = text[:insert_after] + f'"claims":{claims_json},' + text[insert_after:]
        print(f"  ✅ Injected D.claims into data.js  ({len(by_chain)} chains, {len(by_dist)} distributors)")

    # ── 2. Patch D.cm2.by_chain expense for matched chains ────────────────
    cm2_idx = text.find('"cm2":')
    if cm2_idx != -1:
        start = text.index("{", cm2_idx)
        depth, end = 0, start
        for i, ch in enumerate(text[start:], start):
            if ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0: end = i + 1; break

        cm2 = json.loads(text[start:end])

        # Build canonical → claim lakh from chains that have matching names
        matched = {name: data["total_claim_lakh"] for name, data in by_chain.items()
                   if name in CHAIN_MAP.values()}
        # All unmatched chains roll into Unmapped Chain
        unmatched_lakh = sum(data["total_claim_lakh"] for name, data in by_chain.items()
                             if name not in CHAIN_MAP.values())

        total_exp = 0.0
        for c in cm2["by_chain"]:
            name = c["name"]
            claim = matched.get(name, 0.0)
            if name == "Unmapped Chain":
                claim = round(claim + unmatched_lakh, 2)
            if claim > 0:
                c["expense"] = claim
                if c.get("nsv", 0):
                    c["cm2_value"] = round(c["nsv"] - claim, 2)
                    c["cm2_pct"] = round(c["cm2_value"] / c["nsv"] * 100, 1)
            total_exp += c.get("expense", 0)

        total_nsv = cm2["total_nsv"]
        cm2["total_expense"] = round(total_exp, 2)
        cm2["expense_pct_of_nsv"] = round(total_exp / total_nsv * 100, 1) if total_nsv else 0
        cm2["cm2_value"] = round(total_nsv - total_exp, 2)
        cm2["cm2_pct"] = round(cm2["cm2_value"] / total_nsv * 100, 1) if total_nsv else 0
        cm2["has_expense_data"] = True
        cm2["provisional"] = True
        cm2["provisional_label"] = "CM2 PROVISIONAL — Q1 FY27 Claim Master (Finance sign-off pending)"
        cm2["provisional_reasons"] = [
            "Source: Distributor_Chain_Claim_Master_AprJun_2026.xlsx (Apr–Jun 2026 actual claims)",
            "51 exceptions pending owner decision (chain mapping + legacy workbooks)",
            "NSV spans FY25–FY27; claim expense covers Q1 FY27 only",
        ]

        cm2_json = json.dumps(cm2, ensure_ascii=False, separators=(",", ":"))
        text = text[:start] + cm2_json + text[end:]
        print(f"  ✅ Patched D.cm2: total_expense={total_exp:.2f}L  cm2_pct={cm2['cm2_pct']}%")

    datajs_path.write_text(text, encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src",        required=True, help="Path to claim master .xlsx")
    ap.add_argument("--out-claim",  default="dashboard/claim_data.json")
    ap.add_argument("--out-datajs", default="dashboard/data.js")
    args = ap.parse_args()

    xlsx_path   = Path(args.src)
    claim_path  = Path(args.out_claim)
    datajs_path = Path(args.out_datajs)

    if not xlsx_path.exists():
        print(f"❌ Source not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    print("=" * 70)
    print("CLAIM EXPENSE INGESTION — Real Apr-Jun 2026 Data")
    print("=" * 70)

    print("\n1. Loading Excel source...")
    chain_rows = load_chain_summary(xlsx_path)
    txn_rows   = load_claim_master(xlsx_path)
    print(f"   Chain Summary: {len(chain_rows)} chains")
    print(f"   Claim Master:  {len(txn_rows)} transactions")

    print("\n2. Building data structures...")
    by_chain  = build_by_chain_claims(chain_rows)
    by_dist   = build_by_dist_claims(txn_rows)
    grand_inr = sum(r["total_inr"] for r in chain_rows)
    grand_lakh = inr_to_lakh(grand_inr)
    print(f"   by_chain:      {len(by_chain)} entries")
    print(f"   by_distributor:{len(by_dist)} entries")
    print(f"   Grand total:   INR {grand_inr:,.0f}  ({grand_lakh:.2f} L)")

    print("\n3. Writing claim_data.json...")
    claim_data = build_claim_data(chain_rows, txn_rows, xlsx_path)
    claim_path.write_text(json.dumps(claim_data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"   ✅ Written: {claim_path}")

    print("\n4. Patching data.js...")
    inject_claims_into_datajs(datajs_path, by_chain, by_dist, grand_lakh)

    print()
    print("=" * 70)
    print("✅ INGESTION COMPLETE")
    print(f"   {claim_path}")
    print(f"   {datajs_path} (D.claims injected + D.cm2 patched)")
    print("=" * 70)


if __name__ == "__main__":
    main()
