"""
Build NPI chain-wise output with Central zone mapping for Maharashtra (Vidarbha).
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re, os

SRC = "/root/.claude/uploads/f72862e2-ac8f-514a-a251-d4833c7268e5/74ad52cf-MT_Chain_Wise__Article_Wise_NPI_for_TY__Central_mapping_for_maharastra.xlsx"
OUT = "/home/user/mt-dashboard/PowerBI/SeedData/NPI_ChainWise_WithVidarbha.xlsx"

# ── Load Sheet 1 (header on row index 1, row 0 is blank) ──────────────────────
npi = pd.read_excel(SRC, sheet_name=0, header=1)
print(f"Sheet1 rows: {len(npi)}, cols: {list(npi.columns)}")

# ── Load Sheet 2 (Central mapping) ────────────────────────────────────────────
cm = pd.read_excel(SRC, sheet_name=1, header=0)
print(f"Sheet2 rows: {len(cm)}, cols: {list(cm.columns)}")
print(cm.head(3).to_string())

# Normalise chain names for matching
def norm(s):
    if not isinstance(s, str): return ""
    return re.sub(r'\s+', ' ', s.strip().lower())

cm["_chain_norm"] = cm["Chain Name"].apply(norm)
central_chains_raw = set(cm["_chain_norm"].unique())
# Add short aliases: "reliance retail ltd." → also match "reliance" in NPI
CHAIN_ALIASES = {
    "reliance retail ltd.": "reliance",
    "reliance retail": "reliance",
}
central_chains = set()
for c in central_chains_raw:
    central_chains.add(c)
    if c in CHAIN_ALIASES:
        central_chains.add(CHAIN_ALIASES[c])
print(f"\nCentral mapping chains (normalised + aliases): {central_chains}")

# ── Apply Central zone re-mapping ─────────────────────────────────────────────
# Logic: Maharashtra rows whose chain (normalised) is in Central mapping
# → reclassify Zone to "Central" and add Sub_Zone="Vidarbha"

npi["_chain_norm"] = npi["Chain Name"].apply(norm)
npi_state_col = None
for c in npi.columns:
    if "state" in str(c).lower():
        npi_state_col = c
        break

print(f"\nState column in NPI: {npi_state_col}")
print("Zone values:", npi["Zone"].value_counts().to_string())

maha_mask = npi[npi_state_col].apply(lambda x: isinstance(x, str) and "maharashtra" in x.lower())
central_mask = maha_mask & npi["_chain_norm"].isin(central_chains)

print(f"\nMaharashtra rows total: {maha_mask.sum()}")
print(f"Maharashtra rows → Central (Vidarbha) reclassified: {central_mask.sum()}")

npi["Zone_Revised"] = npi["Zone"].copy()
npi["Sub_Zone"] = ""
npi.loc[central_mask, "Zone_Revised"] = "Central"
npi.loc[central_mask, "Sub_Zone"] = "Vidarbha"

# Chain breakdown of reclassified rows
print("\nReclassified rows by chain:")
print(npi[central_mask].groupby("Chain Name").size().to_string())

# ── Chain-wise NPI status pivot (REVISED zone) ────────────────────────────────
STATUS_COLS = ["Disc", "EPD", "NPD", "NPD not to show", "epd", "Freebie"]

def npi_pivot(df, zone_col="Zone_Revised"):
    grp = df.groupby([zone_col, npi_state_col, "Chain Name", "NPI Status"]).size().reset_index(name="Count")
    piv = grp.pivot_table(index=[zone_col, npi_state_col, "Chain Name"],
                          columns="NPI Status", values="Count", fill_value=0).reset_index()
    for s in STATUS_COLS:
        if s not in piv.columns:
            piv[s] = 0
    # Reorder status cols
    status_present = [s for s in STATUS_COLS if s in piv.columns]
    extra = [c for c in piv.columns if c not in [zone_col, npi_state_col, "Chain Name"] + status_present]
    piv = piv[[zone_col, npi_state_col, "Chain Name"] + status_present + extra]
    piv["Total"] = piv[status_present + extra].sum(axis=1)
    piv = piv.sort_values([zone_col, npi_state_col, "Chain Name"]).reset_index(drop=True)
    return piv

pivot_revised  = npi_pivot(npi, "Zone_Revised")
pivot_original = npi_pivot(npi, "Zone")

print(f"\nRevised pivot rows: {len(pivot_revised)}")
print(pivot_revised[pivot_revised["Zone_Revised"]=="Central"].to_string())

# ── Maharashtra detail (before vs after) ──────────────────────────────────────
maha_detail = npi[maha_mask].copy()
maha_detail = maha_detail.drop(columns=["_chain_norm"], errors="ignore")

# ── Build output Excel ────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# Colour scheme
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "DEEAF1"
GREEN      = "E2EFDA"
YELLOW     = "FFF2CC"
WHITE      = "FFFFFF"

def hdr_font(bold=True, color="FFFFFF", sz=11):
    return Font(name="Arial", bold=bold, color=color, size=sz)

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def write_df_to_ws(ws, df, start_row=1, header_fill=DARK_BLUE, alt_fill=LIGHT_BLUE):
    cols = list(df.columns)
    # Header row
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=start_row, column=ci, value=col)
        cell.font = hdr_font(bold=True, color="FFFFFF", sz=10)
        cell.fill = hdr_fill(header_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), start_row+1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border()
            if ri % 2 == 0:
                cell.fill = hdr_fill(alt_fill)
            # Right-align numbers
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
    return ri  # last data row

def auto_col_width(ws, df, start_row=1):
    for ci, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 10), 40)

wb = openpyxl.Workbook()

# ── Sheet 1: Chain-Wise NPI Summary (Revised zones) ──────────────────────────
ws1 = wb.active
ws1.title = "Chain-Wise NPI (Revised)"
ws1.freeze_panes = "D2"

# Title
ws1.merge_cells("A1:L1")
t = ws1["A1"]
t.value = "NPI Chain-Wise Summary — Revised Zones (Maharashtra Vidarbha → Central)"
t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t.fill = hdr_fill(DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 22

last = write_df_to_ws(ws1, pivot_revised, start_row=2)
auto_col_width(ws1, pivot_revised, start_row=2)

# Highlight Central zone rows
zone_col_name = pivot_revised.columns[0]  # "Zone_Revised"
for ri in range(3, last+1):
    if ws1.cell(row=ri, column=1).value == "Central":
        for ci in range(1, len(pivot_revised.columns)+1):
            ws1.cell(row=ri, column=ci).fill = hdr_fill(YELLOW)

# ── Sheet 2: Chain-Wise NPI Summary (Original zones) ─────────────────────────
ws2 = wb.create_sheet("Chain-Wise NPI (Original)")
ws2.freeze_panes = "D2"

ws2.merge_cells("A1:L1")
t2 = ws2["A1"]
t2.value = "NPI Chain-Wise Summary — Original Zones (Before Vidarbha Remapping)"
t2.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t2.fill = hdr_fill(MID_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 22

write_df_to_ws(ws2, pivot_original, start_row=2, header_fill=MID_BLUE)
auto_col_width(ws2, pivot_original, start_row=2)

# ── Sheet 3: Maharashtra Vidarbha Remapped Detail ─────────────────────────────
ws3 = wb.create_sheet("Vidarbha Remapped Detail")
ws3.freeze_panes = "A2"

ws3.merge_cells("A1:P1")
t3 = ws3["A1"]
t3.value = f"Maharashtra → Central (Vidarbha) Remapped Rows  [{central_mask.sum()} articles]"
t3.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
t3.fill = hdr_fill("375623")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 20

vid_df = npi[central_mask].copy()
vid_df = vid_df.drop(columns=["_chain_norm", "Unnamed: 0"], errors="ignore")
write_df_to_ws(ws3, vid_df, start_row=2, header_fill="375623", alt_fill=GREEN)
auto_col_width(ws3, vid_df, start_row=2)

# ── Sheet 4: Central Mapping Reference ───────────────────────────────────────
ws4 = wb.create_sheet("Central Mapping (101 Stores)")
ws4.freeze_panes = "A2"

ws4.merge_cells("A1:G1")
t4 = ws4["A1"]
t4.value = "Central Zone Mapping — 101 Vidarbha Stores (Maharashtra)"
t4.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
t4.fill = hdr_fill("7030A0")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 20

cm_out = cm.drop(columns=["_chain_norm"], errors="ignore")
write_df_to_ws(ws4, cm_out, start_row=2, header_fill="7030A0", alt_fill="EAD1DC")
auto_col_width(ws4, cm_out, start_row=2)

# ── Sheet 5: Zone-wise NPI Totals ────────────────────────────────────────────
ws5 = wb.create_sheet("Zone Totals")

ws5.merge_cells("A1:H1")
t5 = ws5["A1"]
t5.value = "Zone-Wise NPI Article Count Summary (Revised)"
t5.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
t5.fill = hdr_fill(DARK_BLUE)
t5.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 20

status_cols = [c for c in pivot_revised.columns if c not in ["Zone_Revised", npi_state_col, "Chain Name", "Total"]]
zone_sum = pivot_revised.groupby("Zone_Revised")[status_cols + ["Total"]].sum().reset_index()
zone_sum.columns = ["Zone"] + status_cols + ["Total"]
zone_sum = zone_sum.sort_values("Total", ascending=False).reset_index(drop=True)
write_df_to_ws(ws5, zone_sum, start_row=2)
auto_col_width(ws5, zone_sum, start_row=2)

wb.save(OUT)
print(f"\nSaved → {OUT}")
print("\nZone totals (revised):")
print(zone_sum.to_string())
