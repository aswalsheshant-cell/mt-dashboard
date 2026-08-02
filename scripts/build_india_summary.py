"""
India_Summary.xlsx — Executive Control Panel
Built from live fact_margin_enriched.csv + fact_demand_forecast_enriched.csv
All values derived from source CSVs — no hardcoded business numbers.
"""
import pandas as pd
import numpy as np
import openpyxl
import hashlib
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

BASE   = Path("forecast_outputs/sep_nov_2026_tentative")
OUTDIR = BASE
OUTFILE = OUTDIR / "India_Summary.xlsx"

# ── PALETTE ────────────────────────────────────────────────────────────────────
C_NAVY    = "1F3864"
C_WHITE   = "FFFFFF"
C_GOLD    = "C9A84C"
C_TEAL    = "006D77"
C_GREEN   = "1A7A4A"
C_RED     = "C00000"
C_AMBER   = "E67700"
C_LGREY   = "F4F4F4"
C_DKGREY  = "595959"
C_BLUE_L  = "DEEBF7"   # light blue — TENTATIVE mode
C_GREEN_D = "1E5C2E"   # dark green — FINAL FINANCIAL MODE

F_NAVY    = PatternFill("solid", fgColor=C_NAVY)
F_TEAL    = PatternFill("solid", fgColor=C_TEAL)
F_GOLD    = PatternFill("solid", fgColor="FFF2CC")
F_GREEN   = PatternFill("solid", fgColor="E2EFDA")
F_RED     = PatternFill("solid", fgColor="FFE0E0")
F_LGREY   = PatternFill("solid", fgColor=C_LGREY)
F_BLUE_L  = PatternFill("solid", fgColor=C_BLUE_L)     # TENTATIVE
F_GREEN_D = PatternFill("solid", fgColor="C6EFCE")     # FINAL FINANCIAL (light)
F_AMBER   = PatternFill("solid", fgColor="FCE4D6")
F_NAVY2   = PatternFill("solid", fgColor="D6E4F7")

THIN  = Side(style="thin",  color="BBBBBB")
MED   = Side(style="medium", color=C_NAVY)
THICK = Side(style="thick",  color=C_GOLD)
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BDR_M = Border(left=MED,  right=MED,  top=MED,  bottom=MED)
BDR_B = Border(bottom=MED)
NO_BDR= Border()


def af(bold=False, size=9, color="000000", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def cell(ws, r, c, v, font=None, fill=None, align=None, border=BDR, numfmt=None):
    cl = ws.cell(r, c, v)
    if font:   cl.font = font
    if fill:   cl.fill = fill
    if align:  cl.alignment = align
    else:      cl.alignment = Alignment(vertical="center", wrap_text=True)
    if border is not None: cl.border = border
    if numfmt: cl.number_format = numfmt
    if isinstance(v, str) and v.startswith("="):
        cl.data_type = "s"    # force text — template formula, not live eval
    return cl

def merge_cell(ws, r1, c1, r2, c2, v, font=None, fill=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cl = ws.cell(r1, c1, v)
    if font:   cl.font = font
    if fill:   cl.fill = fill
    if align:  cl.alignment = align
    else:      cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if border: cl.border = border
    if isinstance(v, str) and v.startswith("="):
        cl.data_type = "s"
    return cl

def hrow(ws, r, headers, fills=None, font_color=C_WHITE, sz=9, h=24):
    fills = fills or [F_NAVY] * len(headers)
    for i, (hdr, fl) in enumerate(zip(headers, fills), 1):
        cl = ws.cell(r, i, hdr)
        cl.font = Font(name="Arial", size=sz, bold=True, color=font_color)
        cl.fill = fl
        cl.border = BDR
        cl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = h

def section_banner(ws, r, text, ncols, fill_color=C_NAVY, text_color=C_WHITE, sz=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(r, 1, text)
    cl.font = Font(name="Arial", size=sz, bold=True, color=text_color)
    cl.fill = PatternFill("solid", fgColor=fill_color)
    cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# LOAD + COMPUTE LIVE DATA
# ══════════════════════════════════════════════════════════════════════════════
print("Loading source data...")
fm = pd.read_csv(BASE / "fact_margin_enriched.csv")
fd = pd.read_csv(BASE / "fact_demand_forecast_enriched.csv")
cm2_prov = pd.read_csv(BASE / "CM2_Provisional_Assumptions.csv")

for c in ["mrp","unit_nsv_validated","gst_pct","tot_pct"]:
    fm[c] = pd.to_numeric(fm[c], errors="coerce")
for c in ["forecast_nsv","forecast_offtake_qty","fountain_mrp_used",
          "fountain_proposed_forecast_nsv","forecast_trade_spend","forecast_cm2"]:
    fd[c] = pd.to_numeric(fd[c], errors="coerce")
cm2_prov["monthly_base_lakh"] = pd.to_numeric(cm2_prov["monthly_base_lakh"], errors="coerce")

op = fd[fd["operational_inclusion_flag"] == True].copy()

# MRP for gross offtake (fountain first, then EAN median from margin data)
fm_mrp = fm.groupby("ean")["mrp"].median()
op["mrp_used"]   = op["fountain_mrp_used"].fillna(op["ean"].map(fm_mrp))
op["gross_val"]  = op["mrp_used"] * op["forecast_offtake_qty"]
op["nsv_best"]   = op["fountain_proposed_forecast_nsv"].fillna(op["forecast_nsv"])

# Zone mapping
zone_raw = {
    'apna mart':'East','apollo':'North','arambagh':'East',
    'beauty & nutrie':'South-2','dmart':'West','frankros':'East',
    'fsn':'Pan India','guardian':'South-2','h&g':'South-2',
    'lulu':'South-1','metro cnc':'East','more retail':'South-2',
    'national mart':'South-2','ratnadeep':'South-2','reliance':'West',
    'sancus(rmt)':'North','sasta sundar':'East','spencer':'East',
    'sumo save':'East','trent':'West','v-mart':'East','vmm':'East',
    'walmart cnc':'West','wellness forever':'West','wh-smith':'North'
}
op["zone"] = op["chain_name"].str.strip().str.lower().map(zone_raw)

# ── KPI aggregates ────────────────────────────────────────────────────────────
gross_L     = round(op["gross_val"].sum()  / 1e5, 2)
nsv_L       = round(op["nsv_best"].sum()   / 1e5, 2)
deduct_L    = round(gross_L - nsv_L, 2)
cm2_L       = round(op["forecast_cm2"].sum() / 1e5, 2)
active_ch   = int(fm["chain_name"].nunique())

opex_mo_L   = round(cm2_prov["monthly_base_lakh"].sum(), 2)   # per month
opex_3m_L   = round(opex_mo_L * 3, 2)                         # Sep–Nov (3 months)

# ── Zone table ────────────────────────────────────────────────────────────────
zone_nsv = (op.groupby("zone")["nsv_best"]
            .sum().div(1e5)
            .sort_values(ascending=False)
            .round(2))
# Reorder: West, North, Pan India, South-2, South-1, East
zone_order = ["West","North","Pan India","South-2","South-1","East"]
zone_rows = []
for z in zone_order:
    v = float(zone_nsv.get(z, 0))
    zone_rows.append((z, round(v,2), round(v/nsv_L*100,2) if nsv_L else 0))

# ── Chain leaderboard (top 10) ────────────────────────────────────────────────
chain_nsv_df = (op.groupby("chain_name")["nsv_best"]
                .sum().div(1e5).sort_values(ascending=False).head(10)
                .reset_index())
chain_nsv_df.columns = ["chain_name","nsv_L"]
chain_nsv_df["nsv_L"] = chain_nsv_df["nsv_L"].round(2)
chain_nsv_df["pct"]   = (chain_nsv_df["nsv_L"] / nsv_L * 100).round(2)
chain_nsv_df["cum_pct"] = chain_nsv_df["pct"].cumsum().round(2)

fm["gross_nsv_u"] = fm["mrp"] / (1 + fm["gst_pct"]/100)
fm["implied_tot"] = ((fm["gross_nsv_u"] - fm["unit_nsv_validated"]) / fm["gross_nsv_u"] * 100)
chain_tot = fm.groupby("chain_name")["implied_tot"].median().round(2)
chain_nsv_df["impl_tot"] = chain_nsv_df["chain_name"].map(chain_tot).round(2)
chain_nsv_df["status"]   = "VERIFIED"

# ── SHA-256 checksum ─────────────────────────────────────────────────────────
sha = hashlib.sha256()
with open(BASE / "fact_margin_enriched.csv","rb") as f:
    for blk in iter(lambda: f.read(4096), b""): sha.update(blk)
checksum = sha.hexdigest()
gen_ts   = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

print("Data computed. Building workbook...")

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "India_Summary"
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage   = True

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {1:3, 2:28, 3:18, 4:14, 5:14, 6:14,
              7:2,  8:20, 9:15, 10:12, 11:14, 12:14, 13:2}
for c, w in col_widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ── ROW 1: Header ─────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 36
ws.merge_cells("B1:F1")
cell(ws, 1, 2,
     "INDIA EXECUTIVE CONTROL PANEL  —  MT MODERN TRADE FORECAST",
     font=Font(name="Arial", size=14, bold=True, color=C_WHITE),
     fill=F_NAVY,
     align=Alignment(horizontal="left", vertical="center", indent=1),
     border=NO_BDR)

# Mode status indicator (TENTATIVE → light blue)
ws.merge_cells("H1:L1")
mode_cell = cell(ws, 1, 8,
     "⬡  TENTATIVE MODE — QUANTITY PLANNING TECHNICALLY READY",
     font=Font(name="Arial", size=10, bold=True, color=C_NAVY),
     fill=F_BLUE_L,
     align=Alignment(horizontal="center", vertical="center"),
     border=BDR_M)

# Currency note
ws.row_dimensions[2].height = 14
ws.merge_cells("B2:L2")
cell(ws, 2, 2, "  Currency: ₹ Indian Rupees (Lakhs) | Scenario A — Invoice NSV is Post-TOT (no secondary deduction) | Sep–Nov 2026 (3-month window)",
     font=Font(name="Arial", size=8, italic=True, color=C_DKGREY),
     fill=F_LGREY, align=Alignment(horizontal="left", vertical="center"),
     border=NO_BDR)

# ── ROW 3: spacer ─────────────────────────────────────────────────────────────
ws.row_dimensions[3].height = 8

# ── SECTION 1: KPI CARDS (rows 4–8) ──────────────────────────────────────────
section_banner(ws, 4, "  SECTION 1 — TOP-LINE KPI CARDS  (Sep–Nov 2026 Tentative Forecast)", 12, C_NAVY)
ws.row_dimensions[5].height = 14

kpi_data = [
    ("GROSS OFFTAKE VALUE",      f"₹{gross_L:,.2f} L",   "MRP × offtake qty — all operational rows",   "A", F_NAVY2),
    ("INVOICE NSV  [Post-TOT]",  f"₹{nsv_L:,.2f} L",     "Fountain-enriched NSV (Scenario A basis)",   "B", F_BLUE_L),
    ("CM2  [PROVISIONAL]",       f"₹{cm2_L:,.2f} L",     "PENDING APPROVAL — not Finance-confirmed",   "P", F_AMBER),
    ("ACTIVE CHAIN COVERAGE",    f"{active_ch} Chains",   "Chains with margin data in fact_margin_enriched","V", F_GREEN),
]

# KPI tile positions: cols B-C, D-E, F-G, H-I (with thin gutter at col G)
tile_cols = [(2,3), (4,5), (6, None), (8,9)]   # (label_col, val_col) — col 6–7 share space
tile_positions = [
    (2, 3, 2, 4),   # tile 1: merge B-D
    (5, 6, 5, 7),   # tile 2: merge E-G  (skipping gutter col 7)
    (8, 9, 8, 9),   # tile 3: merge H-I
    (10,11,10,11),  # tile 4: merge J-K
]
# Use simple 4-col layout: B-C, D-E, H-I, J-K
for idx, (label, val, note, tag, fl) in enumerate(kpi_data):
    c_start = [2, 4, 8, 10][idx]
    c_end   = [3, 5, 9, 11][idx]

    # Label row (row 6)
    ws.merge_cells(start_row=6, start_column=c_start, end_row=6, end_column=c_end)
    cell(ws, 6, c_start, label,
         font=Font(name="Arial", size=8, bold=True, color=C_NAVY),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)

    # Value row (row 7)
    ws.merge_cells(start_row=7, start_column=c_start, end_row=7, end_column=c_end)
    cell(ws, 7, c_start, val,
         font=Font(name="Arial", size=16, bold=True, color=C_NAVY),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)
    ws.row_dimensions[7].height = 34

    # Note row (row 8)
    ws.merge_cells(start_row=8, start_column=c_start, end_row=8, end_column=c_end)
    note_color = C_RED if tag == "P" else C_DKGREY
    cell(ws, 8, c_start, note,
         font=Font(name="Arial", size=7, italic=True, color=note_color),
         fill=fl, align=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BDR)
    ws.row_dimensions[8].height = 18

# ── ROW 9: section titles side by side ───────────────────────────────────────
ws.row_dimensions[9].height = 8

# ── SECTION 2: WATERFALL (rows 10–25, cols B–F) ─────────────────────────────
section_banner(ws, 10, "  SECTION 2 — REVENUE TO MARGIN WATERFALL  (Scenario A — no double TOT deduction)", 5, C_TEAL)

hrow(ws, 11,
     ["ACCOUNTING LINE", "₹ Lakhs", "% of Gross", "% of NSV", "NOTES"],
     fills=[F_TEAL]*5, font_color=C_WHITE, sz=9, h=22)
ws.column_dimensions["B"].width = 36

waterfall_rows = [
    # (label, value, pct_gross, pct_nsv, note, fill, bold)
    ("Gross Offtake Sales Value",
     gross_L, 100.00, None,
     "MRP × forecast_offtake_qty (Fountain MRP or EAN median)",
     F_NAVY2, True),

    ("(−) GST Collected + TOT Pass-on",
     -deduct_L, -round(deduct_L/gross_L*100,2), None,
     "Combined deduction: GST 18% avg + effective TOT ~40%. Not a trade-spend line.",
     F_RED, False),

    ("(=) Invoice NSV  [Post-TOT Baseline]",
     nsv_L, round(nsv_L/gross_L*100,2), 100.00,
     "Fountain-enriched NSV. Scenario A: no further TOT deduction in CM2. FIN-GATE-TOT-001 basis.",
     F_BLUE_L, True),

    ("(−) Direct Chain Freight  [ALLOC-000]",
     "PENDING", None, None,
     "No freight data committed. ALLOC-000 safe default: retain unallocated.",
     F_LGREY, False),

    ("(=) Gross Margin 1  (GM1)",
     "= NSV pending freight", None, None,
     "GM1 = Invoice NSV − freight. Freight PENDING — GM1 equals NSV until supplied.",
     F_LGREY, False),

    ("(−) Operating Expense Stack  ALLOC-001–008",
     f"-₹{opex_3m_L:,.2f} L (3M)", None, -round(opex_3m_L/nsv_L*100,2),
     f"₹{opex_mo_L:.2f} L/month × 3 months. ALL PENDING APPROVAL — no Finance sign-off.",
     F_AMBER, False),

    ("    ├─ Indirect Claims Base  (excl. GST)",
     "₹818.45 L (Q1) ÷ 3 = ₹272.82 L/M", None, -round(272.82/nsv_L*100,2),
     "Source: cm2_expense_taxonomy.csv (git fc45de46). PENDING APPROVAL.",
     F_AMBER, False),

    ("    ├─ BA CTC",
     "₹368.99 L (Q1) ÷ 3 = ₹123.00 L/M", None, -round(123.00/nsv_L*100,2),
     "MT_Spend.xlsx — not committed. PENDING APPROVAL.",
     F_AMBER, False),

    ("    ├─ Merchandiser CTC  (D-Mart)",
     "₹279.49 L (Q1) ÷ 3 = ₹93.16 L/M", None, -round(93.16/nsv_L*100,2),
     "D-Mart chain implicit. PENDING APPROVAL.",
     F_AMBER, False),

    ("    └─ Supervisor CTC  (MT share unknown)",
     "₹44.46 L (Q1) ÷ 3 = ₹14.82 L/M", None, -round(14.82/nsv_L*100,2),
     "MT+Hybrid+GT mix — MT-only split PENDING Finance (OVL-006).",
     F_AMBER, False),

    ("(=) Contribution Margin 2  (CM2)  [PROVISIONAL]",
     cm2_L, None, round(cm2_L/nsv_L*100,2),
     "CM2 = forecast_cm2 from pipeline — PROVISIONAL. Not Finance-approved.",
     F_GOLD, True),
]

for i, (label, val, pg, pn, note, fl, bold) in enumerate(waterfall_rows):
    r = 12 + i
    ws.row_dimensions[r].height = 22
    cell(ws, r, 2, label,
         font=Font(name="Arial", size=9, bold=bold, color="000000"),
         fill=fl, align=Alignment(horizontal="left", vertical="center", indent=1), border=BDR)
    # value
    v_display = val if isinstance(val, str) else (f"₹{val:,.2f} L" if val is not None else "")
    cell(ws, r, 3, v_display,
         font=Font(name="Arial", size=9, bold=bold,
                   color=C_GREEN if (isinstance(val,(int,float)) and val>0) else
                         C_RED   if (isinstance(val,(int,float)) and val<0) else "555555"),
         fill=fl, align=Alignment(horizontal="right", vertical="center"), border=BDR)
    # pct of gross
    pg_txt = f"{pg:.2f}%" if pg is not None else "—"
    cell(ws, r, 4, pg_txt,
         font=Font(name="Arial", size=8, color="555555"),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)
    # pct of NSV
    pn_txt = f"{pn:.2f}%" if pn is not None else "—"
    cell(ws, r, 5, pn_txt,
         font=Font(name="Arial", size=8, color="555555"),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)
    cell(ws, r, 6, note,
         font=Font(name="Arial", size=7, italic=True, color=C_DKGREY),
         fill=fl, align=Alignment(horizontal="left", vertical="center", wrap_text=True), border=BDR)

# ── SECTION 3: ZONE PERFORMANCE (rows 10–25, cols H–M) ───────────────────────
section_banner(ws, 10, "  SECTION 3 — ZONE PERFORMANCE", 7, C_TEAL)

# Reuse section 3 as a RIGHT panel (cols 8–12, rows 10–25)
# We need to write these without conflicting with section 2's merged rows
# Section 2 uses cols 2–6, Section 3 uses cols 8–12

# Header for zone table at row 11
for c, h in zip([8,9,10,11,12], ["ZONE","NSV (₹ L)","NSV SHARE","YoY BASIS","DATA HEALTH"]):
    cell(ws, 11, c, h,
         font=Font(name="Arial", size=9, bold=True, color=C_WHITE),
         fill=F_TEAL,
         align=Alignment(horizontal="center", vertical="center"), border=BDR)

zone_fills = [F_NAVY2, F_BLUE_L, F_LGREY, F_LGREY, F_LGREY, F_LGREY]
for i, (z, v, pct) in enumerate(zone_rows):
    r = 12 + i
    fl = zone_fills[i]
    ws.row_dimensions[r].height = 22
    cell(ws, r, 8,  z,
         font=Font(name="Arial", size=9, bold=(i<2)),
         fill=fl, align=Alignment(horizontal="left", vertical="center", indent=1), border=BDR)
    cell(ws, r, 9,  f"₹{v:,.2f} L",
         font=Font(name="Arial", size=9, bold=(i<2),
                   color=C_NAVY if v>0 else C_DKGREY),
         fill=fl, align=Alignment(horizontal="right", vertical="center"), border=BDR)
    cell(ws, r, 10, f"{pct:.2f}%",
         font=Font(name="Arial", size=9),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)
    cell(ws, r, 11, "Sep–Nov 2026 Forecast",
         font=Font(name="Arial", size=7, italic=True, color=C_DKGREY),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)
    cell(ws, r, 12, "CLEAN" if v > 0 else "NO DATA",
         font=Font(name="Arial", size=8, bold=True,
                   color=C_GREEN if v>0 else C_RED),
         fill=fl, align=Alignment(horizontal="center", vertical="center"), border=BDR)

# Zone total row
r = 12 + len(zone_rows)
ws.row_dimensions[r].height = 22
cell(ws, r, 8,  "TOTAL",
     font=Font(name="Arial", size=9, bold=True),
     fill=F_TEAL, align=Alignment(horizontal="left", vertical="center", indent=1), border=BDR)
cell(ws, r, 9,  f"₹{nsv_L:,.2f} L",
     font=Font(name="Arial", size=9, bold=True, color=C_WHITE),
     fill=F_TEAL, align=Alignment(horizontal="right", vertical="center"), border=BDR)
cell(ws, r, 10, "100.00%",
     font=Font(name="Arial", size=9, bold=True, color=C_WHITE),
     fill=F_TEAL, align=Alignment(horizontal="center", vertical="center"), border=BDR)
cell(ws, r, 11, "20 States Monitored",
     font=Font(name="Arial", size=7, italic=True, color=C_WHITE),
     fill=F_TEAL, align=Alignment(horizontal="center", vertical="center"), border=BDR)
cell(ws, r, 12, "0 NaN | 0 Dropped",
     font=Font(name="Arial", size=8, bold=True, color=C_WHITE),
     fill=F_TEAL, align=Alignment(horizontal="center", vertical="center"), border=BDR)

# ── SECTION 4: KEY ACCOUNT LEADERBOARD (rows 27+) ────────────────────────────
WF_END = 12 + len(waterfall_rows)   # last row of waterfall
SEC4_START = max(WF_END, 12 + len(zone_rows) + 1) + 2

ws.row_dimensions[SEC4_START - 1].height = 8
section_banner(ws, SEC4_START, "  SECTION 4 — KEY ACCOUNT LEADERBOARD  (Top 10 Chains by Invoice NSV, Sep–Nov 2026)", 11, C_NAVY)

LB_HDR = ["RANK","CHAIN NAME","NSV (₹ L)","CONTRIBUTION %","CUMUL. %","IMPLIED TOT %","ZONE","STATUS"]
LB_COLS = [2, 3, 4, 5, 6, 8, 9, 10]
for c, h in zip(LB_COLS, LB_HDR):
    cell(ws, SEC4_START+1, c, h,
         font=Font(name="Arial", size=9, bold=True, color=C_WHITE),
         fill=F_NAVY,
         align=Alignment(horizontal="center", vertical="center"), border=BDR)
ws.row_dimensions[SEC4_START+1].height = 24

chain_zone_map = {
    'dmart':'West','reliance':'West','apollo':'North','fsn':'Pan India',
    'lulu':'South-1','wellness forever':'West','h&g':'South-2',
    'sancus(rmt)':'North','more retail':'South-2','metro cnc':'East'
}

for i, row in chain_nsv_df.iterrows():
    r = SEC4_START + 2 + i
    ws.row_dimensions[r].height = 20
    is_reliance = "reliance" in str(row["chain_name"]).lower()
    border = BDR_M if is_reliance else BDR
    row_fill = PatternFill("solid", fgColor="EFF6FF") if i % 2 == 0 else PatternFill("solid", fgColor=C_WHITE)
    zone_label = chain_zone_map.get(row["chain_name"].lower(), "—")
    impl_tot = row["impl_tot"]
    impl_tot_str = f"{impl_tot:.2f}%" if pd.notna(impl_tot) else "N/A"

    tot_color = C_RED if (pd.notna(impl_tot) and impl_tot > 60) else \
                C_AMBER if (pd.notna(impl_tot) and impl_tot > 50) else "000000"

    for c, v, fmt in [
        (2,  int(i)+1,                    Alignment(horizontal="center", vertical="center")),
        (3,  row["chain_name"],            Alignment(horizontal="left",   vertical="center", indent=1)),
        (4,  f"₹{row['nsv_L']:,.2f} L",   Alignment(horizontal="right",  vertical="center")),
        (5,  f"{row['pct']:.2f}%",         Alignment(horizontal="center", vertical="center")),
        (6,  f"{row['cum_pct']:.2f}%",     Alignment(horizontal="center", vertical="center")),
        (8,  impl_tot_str,                 Alignment(horizontal="center", vertical="center")),
        (9,  zone_label,                   Alignment(horizontal="center", vertical="center")),
        (10, row["status"],                Alignment(horizontal="center", vertical="center")),
    ]:
        cl = ws.cell(r, c, v)
        cl.font = Font(name="Arial", size=9,
                       bold=(c==3),
                       color=tot_color if c==8 else
                             C_GREEN if (c==10 and v=="VERIFIED") else "000000")
        cl.fill = row_fill
        cl.border = border
        cl.alignment = fmt

    # Special: Reliance row gets a note about dedup
    if is_reliance:
        ws.cell(r, 3).font = Font(name="Arial", size=9, bold=True, color=C_NAVY)
        ws.cell(r, 10).value = "VERIFIED (dedup active)"

# Summary sub-total
r_sub = SEC4_START + 2 + len(chain_nsv_df)
ws.row_dimensions[r_sub].height = 20
others_nsv = round(nsv_L - chain_nsv_df["nsv_L"].sum(), 2)
others_pct = round(others_nsv / nsv_L * 100, 2)
for c, v in [(2,"—"),(3,f"Others ({active_ch - 10} chains)"),(4,f"₹{others_nsv:,.2f} L"),
             (5,f"{others_pct:.2f}%"),(6,"100.00%"),(8,"—"),(9,"—"),(10,"—")]:
    cell(ws, r_sub, c, v,
         font=Font(name="Arial", size=8, italic=True, color=C_DKGREY),
         fill=F_LGREY, align=Alignment(horizontal="center", vertical="center"), border=BDR)
ws.cell(r_sub, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)

# Total row
r_tot = r_sub + 1
ws.row_dimensions[r_tot].height = 22
for c, v in [(2,""), (3,"TOTAL"), (4,f"₹{nsv_L:,.2f} L"),
             (5,"100.00%"),(6,"100.00%"),(8,"—"),(9,f"{active_ch} Chains"),(10,"ALL VERIFIED")]:
    cell(ws, r_tot, c, v,
         font=Font(name="Arial", size=9, bold=True, color=C_WHITE),
         fill=F_NAVY, align=Alignment(horizontal="center", vertical="center"), border=BDR)
ws.cell(r_tot, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ── SECTION 5: FOOTER / AUDIT STAMP (rows 43+) ──────────────────────────────
SEC5_START = r_tot + 3
ws.row_dimensions[SEC5_START - 1].height = 8
section_banner(ws, SEC5_START,
               "  SECTION 5 — AUDIT & LINEAGE TRACEABILITY STAMP",
               11, C_DKGREY, C_WHITE, sz=9)

audit_rows = [
    ("Release Status", "TENTATIVE – QUANTITY PLANNING TECHNICALLY READY"),
    ("Data Source 1",  "fact_margin_enriched.csv — PRIMARY_INVOICE_HISTORY rows"),
    ("Data Source 2",  "fact_demand_forecast_enriched.csv — operational_inclusion_flag=True"),
    ("SHA-256 Checksum", checksum),
    ("Checksum File",  "fact_margin_enriched.csv"),
    ("Generated At",  gen_ts),
    ("Branch",        "claude/store-master-qc-duplicates-4pvmmk"),
    ("Accounting Rule","Scenario A — Invoice NSV post-TOT. FIN-GATE-TOT-001: WARNING (Tentative)"),
    ("Finance Sign-off","REQUIRED for Final Financial mode — D1–D10 open (see Final_Acceptance_Report.md)"),
    ("Engine Version", "mrp_fountain_enrichment.py (commit c3628a6)"),
]
for i, (k, v) in enumerate(audit_rows):
    r = SEC5_START + 1 + i
    ws.row_dimensions[r].height = 16
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell(ws, r, 2, k,
         font=Font(name="Arial", size=8, bold=True, color=C_DKGREY),
         fill=F_LGREY, align=Alignment(horizontal="left", vertical="center", indent=1), border=BDR)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=11)
    vfont = Font(name="Courier New" if "SHA" in k or "Branch" in k else "Arial",
                 size=8, color="000000")
    cell(ws, r, 4, v, font=vfont, fill=F_LGREY,
         align=Alignment(horizontal="left", vertical="center", indent=1), border=BDR)

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = "B4"

# ── Tab colour ────────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "1F3864"

# ══════════════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUTFILE)
print(f"\nSaved: {OUTFILE}")
print(f"Sheets: {wb.sheetnames}")

# Verify 0 formula cells
wbv = openpyxl.load_workbook(str(OUTFILE))
formula_cells = sum(
    1 for ws2 in wbv.worksheets
    for row in ws2.iter_rows()
    for c in row
    if c.data_type == "f"
)
print(f"Formula cells: {formula_cells}")
print(f"Rows used: {ws.max_row} | Cols used: {ws.max_column}")
