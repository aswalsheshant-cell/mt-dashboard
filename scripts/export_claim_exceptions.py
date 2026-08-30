#!/usr/bin/env python3
"""
Automated Claim Audit Formatter — Export unresolved claims for Finance review.

Reads claim_reconciliation_matrix.json (output from reconcile_claim_exceptions.py)
and generates a clean Excel/CSV export for the Finance and Commercial teams.

Formats the 22% "Under Review" and "Unresolved" claim records with:
- Raw claim chain name
- Suggested canonical chain (with confidence score)
- Claim value in lakhs
- Exception type (unmapped, legacy, missing evidence, etc.)
- Rule applied
- Recommended action

No modifications to existing files. Read-only export only.

Usage:
    python scripts/export_claim_exceptions.py [--format csv|xlsx] [--output <path>]

Examples:
    python scripts/export_claim_exceptions.py                          # → claim_exceptions_review.csv
    python scripts/export_claim_exceptions.py --format xlsx             # → claim_exceptions_review.xlsx
    python scripts/export_claim_exceptions.py --output /path/to/file   # Custom output path
"""

import json
import sys
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

RECONCILIATION_MATRIX_PATH = Path("dashboard/claim_reconciliation_matrix.json")
DEFAULT_OUTPUT_CSV = Path("dashboard/claim_exceptions_review.csv")
DEFAULT_OUTPUT_XLSX = Path("dashboard/claim_exceptions_review.xlsx")

# Confidence score thresholds for recommendation levels
CONFIDENCE_THRESHOLDS = {
    0.95: "HIGH — Auto-approve recommended",
    0.80: "MEDIUM — Manual review recommended",
    0.60: "LOW — Requires commercial input",
    0.0: "CRITICAL — Manual research required",
}


class ClaimExceptionExporter:
    """Export unresolved claim exceptions for Finance review."""

    def __init__(self):
        self.reconciliation_data = None
        self.exceptions = []
        self.statistics = {}

    def load_reconciliation_matrix(self) -> bool:
        """Load claim reconciliation matrix from JSON."""
        if not RECONCILIATION_MATRIX_PATH.exists():
            print(f"❌ Reconciliation matrix not found: {RECONCILIATION_MATRIX_PATH}")
            return False

        try:
            with open(RECONCILIATION_MATRIX_PATH) as f:
                self.reconciliation_data = json.load(f)
            print(f"✓ Loaded reconciliation matrix ({len(self.reconciliation_data.get('reconciliation_matrix', []))} records)")
            return True
        except json.JSONDecodeError as e:
            print(f"❌ Invalid JSON in reconciliation matrix: {e}")
            return False

    def extract_review_records(self):
        """Extract records that require review (status = REVIEW or UNRESOLVED)."""
        matrix = self.reconciliation_data.get("reconciliation_matrix", [])

        # Filter for records needing review
        self.exceptions = [
            record
            for record in matrix
            if record.get("status") in ["REVIEW", "UNRESOLVED"]
        ]

        # Calculate statistics
        total_value = sum(r.get("claim_value_lakh", 0) for r in self.exceptions)
        review_count = sum(1 for r in self.exceptions if r.get("status") == "REVIEW")
        unresolved_count = sum(
            1 for r in self.exceptions if r.get("status") == "UNRESOLVED"
        )

        self.statistics = {
            "total_records": len(self.exceptions),
            "review_records": review_count,
            "unresolved_records": unresolved_count,
            "total_value_lakh": round(total_value, 2),
            "review_pct": (review_count / len(self.exceptions) * 100) if self.exceptions else 0,
            "generated_at": datetime.now().isoformat(),
        }

        print(
            f"✓ Extracted {len(self.exceptions)} exceptions for review (₹{total_value:,.2f}L)"
        )

    def get_confidence_recommendation(self, confidence_score: float) -> str:
        """Get action recommendation based on confidence score."""
        for threshold, recommendation in sorted(
            CONFIDENCE_THRESHOLDS.items(), reverse=True
        ):
            if confidence_score >= threshold:
                return recommendation
        return CONFIDENCE_THRESHOLDS[0.0]

    def format_csv(self) -> str:
        """Generate CSV format export."""
        lines = []

        # Header
        lines.append(
            "Raw Chain Name,Suggested Chain,Confidence Score,Claim Value (₹L),Exception Type,Status,Rule Applied,Recommendation,Notes"
        )

        # Data rows
        for exc in self.exceptions:
            raw_chain = exc.get("raw_claim_chain", "—")
            canonical = exc.get("canonical_chain", "—")
            confidence = exc.get("confidence_score", 0)
            value = exc.get("claim_value_lakh", 0)
            exc_type = exc.get("exception_type", "unknown")
            status = exc.get("status", "UNKNOWN")
            rule = exc.get("rule_applied", "—")
            recommendation = self.get_confidence_recommendation(confidence)

            # Escape quotes and handle commas
            raw_chain_safe = f'"{raw_chain}"' if "," in raw_chain else raw_chain
            canonical_safe = f'"{canonical}"' if "," in canonical else canonical
            rule_safe = f'"{rule}"' if "," in rule else rule

            lines.append(
                f'{raw_chain_safe},{canonical_safe},{confidence:.3f},{value:.2f},{exc_type},{status},{rule_safe},"{recommendation}","Review required"'
            )

        # Summary footer
        lines.append("")
        lines.append("SUMMARY")
        lines.append(
            f'Total Records,{self.statistics["total_records"]},,,,'
        )
        lines.append(
            f'Total Value (₹L),{self.statistics["total_value_lakh"]:,.2f},,,,'
        )
        lines.append(
            f'Review Status,{self.statistics["review_records"]} REVIEW + {self.statistics["unresolved_records"]} UNRESOLVED,,,,'
        )
        lines.append(
            f'Generated At,{self.statistics["generated_at"]},,,,'
        )

        return "\n".join(lines)

    def format_xlsx(self) -> None:
        """Generate XLSX format export (requires openpyxl)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("⚠️  openpyxl not installed. Falling back to CSV.")
            return False

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exceptions Review"

        # Header row
        headers = [
            "Raw Chain Name",
            "Suggested Chain",
            "Confidence Score",
            "Claim Value (₹L)",
            "Exception Type",
            "Status",
            "Rule Applied",
            "Recommendation",
            "Notes",
        ]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for exc in self.exceptions:
            raw_chain = exc.get("raw_claim_chain", "—")
            canonical = exc.get("canonical_chain", "—")
            confidence = exc.get("confidence_score", 0)
            value = exc.get("claim_value_lakh", 0)
            exc_type = exc.get("exception_type", "unknown")
            status = exc.get("status", "UNKNOWN")
            rule = exc.get("rule_applied", "—")
            recommendation = self.get_confidence_recommendation(confidence)

            # Determine status color
            status_fill = None
            if status == "REVIEW":
                status_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif status == "UNRESOLVED":
                status_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            row = [
                raw_chain,
                canonical,
                confidence,
                value,
                exc_type,
                status,
                rule,
                recommendation,
                "Review required",
            ]
            ws.append(row)

            # Apply color to status cell
            if status_fill:
                ws.cell(row=ws.max_row, column=6).fill = status_fill

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 35
        ws.column_dimensions["I"].width = 20

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Claim Exceptions Review — Finance Approval Gate"])
        ws_summary.append([])
        ws_summary.append(["Total Records", self.statistics["total_records"]])
        ws_summary.append(["Under Review", self.statistics["review_records"]])
        ws_summary.append(["Unresolved", self.statistics["unresolved_records"]])
        ws_summary.append(["Total Value (₹L)", self.statistics["total_value_lakh"]])
        ws_summary.append(["Review %", f"{self.statistics['review_pct']:.1f}%"])
        ws_summary.append(["Generated", self.statistics["generated_at"]])

        # Save workbook
        wb.save(DEFAULT_OUTPUT_XLSX)
        print(f"✓ XLSX export saved: {DEFAULT_OUTPUT_XLSX}")
        return True

    def export(self, output_path: Path = None, fmt: str = "csv") -> bool:
        """Generate and save export file."""
        print()
        print("=" * 70)
        print("CLAIM EXCEPTION EXPORTER — Finance Review")
        print("=" * 70)
        print()

        # Load data
        if not self.load_reconciliation_matrix():
            return False

        # Extract exceptions
        self.extract_review_records()

        # Export
        print("\nExporting...")

        if fmt == "xlsx":
            success = self.format_xlsx()
            if not success:
                fmt = "csv"  # Fallback

        if fmt == "csv":
            csv_content = self.format_csv()
            output_path = output_path or DEFAULT_OUTPUT_CSV
            with open(output_path, "w") as f:
                f.write(csv_content)
            print(f"✓ CSV export saved: {output_path}")

        # Print statistics
        print()
        print("=" * 70)
        print("EXPORT SUMMARY")
        print("=" * 70)
        print(f"Records exported: {self.statistics['total_records']}")
        print(f"  - Under Review: {self.statistics['review_records']}")
        print(f"  - Unresolved: {self.statistics['unresolved_records']}")
        print(f"Total value: ₹{self.statistics['total_value_lakh']:,.2f}L")
        print(f"Generated: {self.statistics['generated_at']}")
        print()
        print(f"📊 File: {output_path or DEFAULT_OUTPUT_CSV}")
        print()
        print("Next Steps:")
        print("  1. Share with Finance Controller for approval")
        print("  2. Review Recommendation column for action priority")
        print("  3. Assign canonical chain names to unmapped records")
        print("  4. Document source evidence for missing claims")
        print("  5. Re-run reconciliation_matrix with approved mappings")
        print("=" * 70)

        return True


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Automated Claim Audit Formatter for Finance Review"
    )
    parser.add_argument(
        "--format",
        choices=["csv", "xlsx"],
        default="csv",
        help="Export format (default: csv)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output file path (default: dashboard/claim_exceptions_review.[csv|xlsx])",
    )

    args = parser.parse_args()

    exporter = ClaimExceptionExporter()
    success = exporter.export(output_path=args.output, fmt=args.format)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
