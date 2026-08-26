#!/usr/bin/env python3
"""
Tier 3b: Batch Promo & Trade Spend Ingestion Pipeline

Ingests multi-month promotional data from Excel files into data_master.json
and regenerates dashboard/data.js with 5-month (Apr-Aug'26) promo metrics.

USAGE:
  python scripts/batch_load_promos.py --src <promo_excel_dir> --master <data_master.json> --output <dashboard/data.js>

  Default: python scripts/batch_load_promos.py
    (reads: PowerBI/RawDataFolders/Promo_Monthly/*.xlsx, updates: data_master.json + dashboard/data.js)

This pipeline is ADDITIVE: it merges into the existing promo.* block rather
than replacing it, so a previously ingested month (e.g. Aug-26, loaded via
load_promo_detail.py into promo.detail) is preserved. Newly parsed months are
written into promo.monthly[<Month-YY>] using the same per-chain shape as
promo.detail, and promo.mom_trend/summary fields are (re)computed across all
months present so the dashboard's MoM trendline always reflects everything
ingested so far — this run's months plus any already on file.

Pipeline:
  1. Multi-file parsing: single-month and multi-month Excel workbooks
  2. Chain canonicalization: aliases -> ChainMaster.csv canonical names
  3. Brand canonicalization: casing/spacing variants -> canonical names
  4. Offer depth parser: "%", "BxGy"/BOGO, and Rs-off / selling-price-vs-MRP
     forms all normalized to a discount-depth percentage
  5. Per-month aggregation: SKU count, avg depth, by-chain/brand/category
  6. Merge into data_master.json promo.monthly (additive) + regenerate data.js
"""
from __future__ import annotations
import json
import re
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =====================================================================
# CHAIN CANONICALIZATION (aligned with PowerBI/SeedData/Masters/ChainMaster.csv)
# =====================================================================
CHAIN_ALIASES = {
    "Dmart": "D-Mart", "D-Mart": "D-Mart",
    "Apna Klub": "Apna Klub",
    "Apna Mart": "Apna Mart",
    "Apollo": "Apollo",
    "Arambagh": "Arambagh",
    "Ascent": "Ascent Wellness", "Ascent Wellness": "Ascent Wellness",
    "Azorte": "Azorte",
    "B&N": "B&N", "Beauty & Nutire": "B&N",
    "Broadway": "Broadway",
    "Deal Share": "Deal Share", "Deal share": "Deal Share",
    "Eremedium": "Eremedium",
    "Frank Ross": "Frank Ross", "Frankross": "Frank Ross",
    "GNRC": "G N R C Medishop Pvt Ltd",
    "Grace": "Grace Super MKT",
    "Guardian": "Guardian", "Gaurdian": "Guardian",
    "Health & Glow": "Health & Glow", "H&G": "Health & Glow",
    "Lifestyle": "Lifestyle",
    "Lulu": "Lulu",
    "Max Hyper": "Max Hyper",
    "Medanta": "Medanta",
    "Metro": "Metro CNC", "Metro CNC": "Metro CNC", "Metro C&C": "Metro CNC",
    "More Retail": "More Retail", "More": "More Retail",
    "National Mart": "National Mart",
    "Nykaa": "Nykaa SS", "Nykaa SS": "Nykaa SS",
    "Pothys": "Pothys",
    "RMT-Sancus": "RMT-Sancus", "Sancus": "RMT-Sancus", "Sancus RMT-Delhi": "RMT-Sancus",
    "Ratnadeep": "Ratnadeep",
    "Relay": "Relay",
    "Reliance": "Reliance Retail", "Reliance Retail": "Reliance Retail",
    "Saravana": "Sarvana", "Sarvana": "Sarvana",
    "SastaSundar": "SastaSundar", "Sasta Sundar": "SastaSundar",
    "Shoppers": "Shoppers Stop", "Shoppers Stop": "Shoppers Stop",
    "Sohum": "Sohum Shoppe", "Sohum Shoppe": "Sohum Shoppe",
    "Spencers": "Spencers", "Spencer": "Spencers",
    "Sumo Save": "Sumo Save", "Sumosave": "Sumo Save",
    "Today's Basket": "Today's Basket",
    "Trent": "Trent",
    "V-Mart": "V-Mart", "V Mart": "V-Mart",
    "Vijetha": "Vijetha",
    "Vishal": "Vishal Mega Mart", "Vishal Mega Mart": "Vishal Mega Mart", "VMM": "Vishal Mega Mart",
    "WH-Smith": "WH-Smith",
    "Walmart": "Walmart CNC", "Walmart CNC": "Walmart CNC", "Wal-mart": "Walmart CNC",
    "Wellness Forever": "Wellness Forever",
}


def canonicalize_chain(raw_chain: str) -> str | None:
    """Map a raw chain name (any known alias/casing/spacing variant) to its canonical name."""
    if not raw_chain:
        return None
    return CHAIN_ALIASES.get(str(raw_chain).strip())


BRAND_ALIASES = {
    "mamaearth": "Mamaearth",
    "the derma co.": "The Derma Co.", "the derma co": "The Derma Co.",
    "aqualogica": "Aqualogica",
    "bblunt": "BBLUNT",
    "dr. sheth's": "Dr. Sheth's", "dr.sheth's": "Dr. Sheth's", "dr sheth's": "Dr. Sheth's",
}


def canonicalize_brand(raw_brand: str) -> str:
    """Normalize brand name casing/spacing to a canonical form."""
    if not raw_brand:
        return "Unknown"
    cleaned = str(raw_brand).strip()
    if not cleaned or cleaned.upper() == "#N/A":
        return "Unknown"
    return BRAND_ALIASES.get(cleaned.lower(), cleaned)


# =====================================================================
# OFFER DEPTH PARSER
# Normalizes the free-form "Offer to consumer" cell into a discount-depth %.
# =====================================================================
def parse_offer_depth(offer_val, mrp_val=None) -> float | None:
    """Parse a raw consumer offer into a normalized discount depth (0-100%)."""
    if offer_val is None:
        return None

    # Numeric fraction/percent already in the sheet (0.1 = 10%)
    if isinstance(offer_val, (int, float)):
        v = float(offer_val)
        if 0 < v <= 1:
            return round(v * 100, 1)
        if 1 < v <= 100:
            return round(v, 1)
        return None

    text = str(offer_val).strip()
    if not text:
        return None
    low = text.lower()

    # BxGy patterns ("B1G1", "B2G1", "b2g1 on invoice") and bare BOGO
    bxgy = re.search(r'b\s*(\d)\s*g\s*(\d)', low)
    if bxgy:
        x, y = int(bxgy.group(1)), int(bxgy.group(2))
        return round(y / (x + y) * 100, 1)
    if 'bogo' in low:
        return 50.0

    # Explicit percent: "25% Off", "Flat 20% Off", "10% discount to consumer"
    pct = re.search(r'(\d+(?:\.\d+)?)\s*%', text)
    if pct:
        return round(float(pct.group(1)), 1)

    # Flat rupee amount off MRP: "Flat Rs 50 off MRP 250" / "Rs.50 off"
    flat_off = re.search(r'rs\.?\s*(\d+(?:\.\d+)?)\s*off', low)
    if flat_off and mrp_val:
        try:
            amt, mrp = float(flat_off.group(1)), float(mrp_val)
            if mrp > 0:
                return round(amt / mrp * 100, 1)
        except (ValueError, TypeError):
            pass

    # Selling-price-style offers vs MRP: "@Rs.149", "At Rs 199", "SP Rs 169",
    # "Selling price 269", "49 Zone", "Power of 49"
    sp = re.search(r'(?:rs\.?\s*|selling price\.?\s*|power of\s*|^)(\d+(?:\.\d+)?)', low)
    if sp and mrp_val:
        try:
            price, mrp = float(sp.group(1)), float(mrp_val)
            if 0 < price < mrp:
                return round((mrp - price) / mrp * 100, 1)
        except (ValueError, TypeError):
            pass

    return None  # Unparseable free text (e.g. "Kitted pack no claim")


def extract_month_year(date_val) -> tuple[int, int] | None:
    """Extract (month, year) from a datetime cell or common date string formats."""
    if not date_val:
        return None
    try:
        if hasattr(date_val, 'month') and hasattr(date_val, 'year'):
            return (date_val.month, date_val.year)
        if isinstance(date_val, str):
            if '/' in date_val:
                parts = date_val.split('/')
                if len(parts) == 3:
                    return (int(parts[0]), int(parts[2]))
            elif '-' in date_val:
                parts = date_val.split('-')
                if len(parts) == 3:
                    return (int(parts[1]), int(parts[0]))
    except (ValueError, AttributeError):
        pass
    return None


# Filenames encode the covered month unambiguously; the "Period"/"To" cells in
# some source files have day/month swapped by Excel on entry (confirmed: the
# July file's Period reads Jan-07 instead of Jul-01), so filename is the
# trustworthy fallback, not those cells.
FILENAME_MONTH_HINTS = {"apr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8}
MONTH_ABBR = {4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug"}


def month_year_from_filename(filename: str) -> tuple[int, int] | None:
    """Infer (month, year) from a filename containing a month token, e.g. 'Promo_July_2026.xlsx'."""
    name_lower = filename.lower()
    for token, month in FILENAME_MONTH_HINTS.items():
        if token in name_lower:
            return (month, 2026)
    return None


def load_kam_sheet(wb) -> dict:
    """Read an optional 'Chain Name | KAM | Remarks' sheet into {canonical_chain: (kam, received)}."""
    kam_map = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if header[:2] != ["Chain Name", "KAM"]:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            canonical = canonicalize_chain(row[0])
            if not canonical:
                continue
            kam = str(row[1]).strip() if len(row) > 1 and row[1] else None
            remarks = str(row[2]).strip().lower() if len(row) > 2 and row[2] else ""
            received = "received" in remarks if remarks else None
            kam_map[canonical] = (kam, received)
    return kam_map


def load_promo_files(src_dir: str) -> tuple[list[dict], dict]:
    """Load all Excel promo files from directory. Returns (records, kam_map)."""
    promo_records = []
    kam_map = {}
    src_path = Path(src_dir)

    if not src_path.exists():
        print(f"  ⚠ Promo directory not found: {src_dir}")
        return promo_records, kam_map

    xlsx_files = sorted(set(src_path.glob("*.xlsx")))

    if not xlsx_files:
        print(f"  ⚠ No Excel files found in {src_dir}")
        return promo_records, kam_map

    print(f"\n[1] Loading {len(xlsx_files)} promo Excel file(s)...")

    for xlsx_file in xlsx_files:
        print(f"  📄 {xlsx_file.name}")
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            kam_map.update(load_kam_sheet(wb))

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row = [cell.value for cell in ws[1]]
                if not header_row or header_row[0] != "Chain Name" or "Brand" not in header_row:
                    continue  # skip KAM/pivot/summary sheets

                chain_idx = header_row.index("Chain Name")
                brand_idx = header_row.index("Brand")
                offer_idx = header_row.index("Offer to consumer")
                mrp_idx = header_row.index("MRP") if "MRP" in header_row else None
                cat_idx = header_row.index("Category") if "Category" in header_row else None
                month_idx = header_row.index("Month") if "Month" in header_row else None

                # Fallback for files with no per-row "Month" column: the
                # filename itself encodes the single month the file covers.
                fallback_month_year = month_year_from_filename(xlsx_file.name)

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[chain_idx]:
                        continue

                    canonical_chain = canonicalize_chain(row[chain_idx])
                    if not canonical_chain:
                        print(f"    ✗ Unmapped chain: '{row[chain_idx]}' (row {row_num})")
                        continue

                    # Prefer the explicit per-row "Month" column; fall back to
                    # the filename-derived month for single-month files.
                    month_year = None
                    if month_idx is not None:
                        month_year = extract_month_year(row[month_idx])
                    if not month_year:
                        month_year = fallback_month_year
                    if not month_year:
                        print(f"    ⚠ Missing month/year for {canonical_chain} (row {row_num})")
                        continue
                    month, year = month_year

                    mrp_val = row[mrp_idx] if mrp_idx is not None else None
                    depth = parse_offer_depth(row[offer_idx], mrp_val)

                    promo_records.append({
                        "chain": canonical_chain,
                        "brand": canonicalize_brand(row[brand_idx]),
                        "category": str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] else "Unknown",
                        "month": month,
                        "year": year,
                        "discount_depth": depth,
                        "source": xlsx_file.name,
                    })
        except Exception as e:
            print(f"    ✗ Error reading {xlsx_file.name}: {e}")

    print(f"  ✓ Loaded {len(promo_records)} promo records")
    return promo_records, kam_map


def build_month_detail(month_key: str, records: list[dict], kam_map: dict, source: str) -> dict:
    """Build a promo.detail-shaped block (as used by dashboard/index.html) for one month."""
    by_chain_recs = defaultdict(list)
    for r in records:
        by_chain_recs[r["chain"]].append(r)

    by_chain = []
    for chain, recs in sorted(by_chain_recs.items(), key=lambda kv: -len(kv[1])):
        depths = [r["discount_depth"] for r in recs if r["discount_depth"] is not None]
        kam, received = kam_map.get(chain, (None, None))
        by_chain.append({
            "name": chain,
            "skus": len(recs),
            "brands": len(set(r["brand"] for r in recs)),
            "categories": len(set(r["category"] for r in recs)),
            "avg_offer_pct": round(sum(depths) / len(depths), 1) if depths else None,
            "offer_parseable_pct": round(len(depths) / len(recs) * 100) if recs else 0,
            "kam": kam,
            "received": received,
        })

    by_brand_recs = defaultdict(list)
    for r in records:
        by_brand_recs[r["brand"]].append(r)
    by_brand = [
        {"name": b, "skus": len(recs)}
        for b, recs in sorted(by_brand_recs.items(), key=lambda kv: -len(kv[1]))
    ]

    by_cat_recs = defaultdict(list)
    for r in records:
        by_cat_recs[r["category"]].append(r)
    by_category = [
        {"name": c, "skus": len(recs)}
        for c, recs in sorted(by_cat_recs.items(), key=lambda kv: -len(kv[1]))
    ]

    received_flags = [kam_map.get(c, (None, None))[1] for c in by_chain_recs]
    chains_received = sum(1 for f in received_flags if f is True)
    chains_pending = sum(1 for f in received_flags if f is False)

    all_depths = [r["discount_depth"] for r in records if r["discount_depth"] is not None]

    return {
        "month": month_key,
        "source": source,
        "total_skus": len(records),
        "chains_in_promo": len(by_chain_recs),
        "brands_in_promo": len(by_brand_recs),
        "chains_received": chains_received,
        "chains_pending": chains_pending,
        "avg_depth": round(sum(all_depths) / len(all_depths), 1) if all_depths else None,
        "by_chain": by_chain,
        "by_brand": by_brand,
        "by_category": by_category,
        "kam_status": {"pending": [c for c in by_chain_recs if kam_map.get(c, (None, None))[1] is False]},
    }


def build_top_level_summary(all_months: dict, records_by_month: dict) -> dict:
    """Aggregate every month in `promo.monthly` into the top-level promo KPI fields."""
    chain_totals = defaultdict(lambda: {"promos": 0, "depths": [], "brands": set()})
    brand_totals = defaultdict(int)
    cat_totals = defaultdict(int)
    n_promos = 0
    all_depths = []

    for month_key, month_block in all_months.items():
        for c in month_block.get("by_chain", []):
            chain_totals[c["name"]]["promos"] += c["skus"]
            if c.get("avg_offer_pct") is not None:
                chain_totals[c["name"]]["depths"].append(c["avg_offer_pct"])
        for b in month_block.get("by_brand", []):
            brand_totals[b["name"]] += b["skus"]
        for cat in month_block.get("by_category", []):
            cat_totals[cat["name"]] += cat["skus"]
        n_promos += month_block.get("total_skus", 0)
        depths = [c["avg_offer_pct"] for c in month_block.get("by_chain", []) if c.get("avg_offer_pct") is not None]
        month_avg = month_block.get("avg_depth")
        if month_avg is None and depths:
            month_avg = round(sum(depths) / len(depths), 1)
        if month_avg is not None:
            all_depths.append(month_avg)
        # Per-chain distinct brand tracking, from this run's raw records only
        # (the pre-existing Aug-26 detail block has no raw records to draw from).
        for r in records_by_month.get(month_key, []):
            chain_totals[r["chain"]]["brands"].add(r["brand"])

    by_chain = [
        {
            "name": name,
            "promos": vals["promos"],
            "avg_depth": round(sum(vals["depths"]) / len(vals["depths"]), 1) if vals["depths"] else None,
            "brands": len(vals["brands"]) if vals["brands"] else None,
        }
        for name, vals in sorted(chain_totals.items(), key=lambda kv: -kv[1]["promos"])
    ]

    return {
        "n_promos": n_promos,
        "avg_depth": round(sum(all_depths) / len(all_depths), 1) if all_depths else 0,
        "by_chain": by_chain,
        "by_brand": [{"name": n, "promos": p} for n, p in sorted(brand_totals.items(), key=lambda kv: -kv[1])],
        "by_category": [{"name": n, "promos": p} for n, p in sorted(cat_totals.items(), key=lambda kv: -kv[1])],
    }


def load_master(master_path: str) -> dict:
    with open(master_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(
        description="Batch-ingest promo Excel files into data_master.json (additive) and regenerate dashboard/data.js"
    )
    parser.add_argument("--src", default="PowerBI/RawDataFolders/Promo_Monthly")
    parser.add_argument("--master", default="data_master.json")
    parser.add_argument("--output", default="dashboard/data.js")
    args = parser.parse_args()

    print("=" * 80)
    print("TIER 3b: BATCH PROMO INGESTION PIPELINE (additive)")
    print("=" * 80)

    promo_records, kam_map = load_promo_files(args.src)
    if not promo_records:
        print("\n✗ No promo records loaded. Aborting.")
        return 1

    print(f"\n[2] Loading data_master.json...")
    try:
        master = load_master(args.master)
        print(f"  ✓ Loaded (metadata: {master['metadata'].get('status', 'unknown')})")
    except FileNotFoundError:
        print(f"  ✗ File not found: {args.master}")
        return 1
    except json.JSONDecodeError as e:
        print(f"  ✗ JSON parse error: {e}")
        return 1

    existing_promo = master.get("promo") or {}
    existing_monthly = dict(existing_promo.get("monthly") or {})

    # Seed from the pre-existing single-month `detail` block (e.g. Aug-26,
    # ingested previously via load_promo_detail.py) so it survives as one of
    # the months in the new multi-month collection instead of being dropped.
    prev_detail = existing_promo.get("detail")
    if prev_detail and prev_detail.get("month") and prev_detail["month"] not in existing_monthly:
        existing_monthly[prev_detail["month"]] = prev_detail

    print(f"\n[3] Building per-month detail blocks from {len(promo_records)} new records...")
    recs_by_month = defaultdict(list)
    month_sources = {}
    for r in promo_records:
        key = f"{MONTH_ABBR.get(r['month'], r['month'])}-{str(r['year'])[-2:]}"
        recs_by_month[key].append(r)
        month_sources[key] = r["source"]

    for month_key, recs in recs_by_month.items():
        existing_monthly[month_key] = build_month_detail(month_key, recs, kam_map, month_sources[month_key])
        print(f"  ✓ {month_key}: {len(recs)} SKU-promo lines, {existing_monthly[month_key]['chains_in_promo']} chains")

    print(f"  ✓ Total months on file: {len(existing_monthly)}")

    # Latest month (by calendar order among what's covered) stays `detail` for
    # backward compatibility with the existing single-month dashboard section.
    order = {v: k for k, v in MONTH_ABBR.items()}
    def month_sort_key(mk):
        abbr, yy = mk.split('-')
        return (int(yy), order.get(abbr, 99))
    latest_month = sorted(existing_monthly.keys(), key=month_sort_key)[-1]

    top_level = build_top_level_summary(existing_monthly, recs_by_month)
    def month_avg_depth(block):
        if block.get("avg_depth") is not None:
            return block["avg_depth"]
        depths = [c["avg_offer_pct"] for c in block.get("by_chain", []) if c.get("avg_offer_pct") is not None]
        return round(sum(depths) / len(depths), 1) if depths else None

    mom_trend = [
        {
            "month": mk,
            "skus": existing_monthly[mk]["total_skus"],
            "avg_depth": month_avg_depth(existing_monthly[mk]),
            "chains": existing_monthly[mk]["chains_in_promo"],
        }
        for mk in sorted(existing_monthly.keys(), key=month_sort_key)
    ]

    promo_block = {
        **top_level,
        "lines": existing_promo.get("lines", []),
        "detail": existing_monthly[latest_month],
        "monthly": existing_monthly,
        "months_available": [mk for mk, _ in sorted(existing_monthly.items(), key=lambda kv: month_sort_key(kv[0]))],
        "mom_trend": mom_trend,
    }

    print(f"\n[4] Merging additive promo block into data_master.json...")
    master["promo"] = promo_block
    print(f"  ✓ Merged — {len(existing_monthly)} months: {promo_block['months_available']}")

    print(f"\n[5] Saving updated data_master.json...")
    with open(args.master, 'w', encoding='utf-8') as f:
        json.dump(master, f, indent=2, ensure_ascii=False)
    print(f"  ✓ Saved")

    print(f"\n[6] Regenerating dashboard/data.js via sync_data_js.py...")
    import subprocess
    result = subprocess.run(
        ["python", "scripts/sync_data_js.py", "--source", args.master, "--output", args.output],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        print(f"  ✓ Dashboard data.js regenerated ({args.output})")
    else:
        print(f"  ✗ Failed to regenerate data.js\n    {result.stderr}")
        return 1

    print(f"\n" + "=" * 80)
    print("PROMO BATCH INGESTION COMPLETE")
    print("=" * 80)
    print(f"""
Source:           {args.src}
Master:           {args.master} (UPDATED, additive)
Dashboard:        {args.output} (REGENERATED)
Months on file:   {promo_block['months_available']}
Total SKU-lines:  {top_level['n_promos']}

Status:           ✓ READY FOR DEPLOYMENT
""")
    return 0


if __name__ == "__main__":
    exit(main())
