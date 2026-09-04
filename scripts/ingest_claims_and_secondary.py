"""
Ingests:
  1. Updated claim Excel (f14f5602) → claim_data.json + D.claims patch in data.js
  2. Secondary sales repository (All_Sancus_Months) → D.secondary_sales patch in data.js

Usage:
  python scripts/ingest_claims_and_secondary.py \
      --claim   <path/to/Claim_Invoice_Mapped.xlsx> \
      --secsale <path/to/Secondary_Sales_All_Sancus_Months.xlsx> \
      --datajs  dashboard/data.js \
      --claimjson dashboard/claim_data.json
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def inr_to_lakh(v):
    return round(float(v or 0) / 1e5, 4)

def _read_sheet(wb, name):
    ws = wb[name]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if headers is None:
            headers = [str(c) if c is not None else f"_col{i}" for c, i in zip(row, range(len(row)))]
            continue
        if not any(v is not None for v in row):
            break
        rows.append({h: v for h, v in zip(headers, row)})
    return headers, rows


# ---------------------------------------------------------------------------
# 1. Claim ingestion
# ---------------------------------------------------------------------------

CHAIN_CANON = {
    "D-Mart": "DMart",
    "Dmart": "DMart",
    "DMART": "DMart",
    "Avenue Supermarts Ltd.": "DMart",
    "Reliance Retail Ltd.": "Reliance Retail",
    "RELIANCE": "Reliance Retail",
    "Reliance": "Reliance Retail",
    "Lulu Hyper": "Lulu",
    "APOLLO": "Apollo",
    "Nykaa (FSN)": "Nykaa (FSN)",
    "Nykaa": "Nykaa (FSN)",
    "Spencer's": "Spencer's",
    "Spencer": "Spencer's",
    "SUMOSAVE": "Sumo Save",
    "RATNADEEP RETAIL PVT LTD": "Ratnadeep",
    "EMAMI FRANKROSS": "Frankross",
    "Frankross": "Frankross",
    "GHL Pharma & Diagnostic Pvt Ltd": "GHL Pharma",
    "MERABO LABS PRIVATE LIMITED": "Merabo Labs",
    "Arambagh": "Arambagh",
    "Spar": "Spar",
    "Trilife": "Trilife",
    "Sastasundar": "Sastasundar",
    "Sasta Sundar": "Sastasundar",
    "Wellness Forever": "Wellness Forever",
    "Health & Glow": "Health & Glow",
    "More Retail": "More Retail",
    "More": "More Retail",
    "Unspecified / Review": "Unspecified",
    "AirPlaza": "AirPlaza",
    "Combined Charge": "Combined Charge",
    "FLEET LABS": "Fleet Labs",
    "Transportation": "Transportation",
    "Beauty & Nutrition": "Beauty & Nutrition",
    "Sancus Network": "Sancus Network",
    "Vishal Mega Mart": "Vishal Mega Mart",
}

BRAND_CANON = {
    "ME": "Mamaearth",
    "MAMAEARTH": "Mamaearth",
    "TDC": "The Derma Co.",
    "THE DERMA CO.": "The Derma Co.",
    "The Derma Co": "The Derma Co.",
    "AQ": "Aqualogica",
    "AQUALOGICA": "Aqualogica",
    "BB": "BBLUNT",
    "Bblunt": "BBLUNT",
    "DR Sheths": "Dr. Sheth's",
    "DR. SHETH's": "Dr. Sheth's",
    "Dr. Sheths": "Dr. Sheth's",
    "DRS": "Dr. Sheth's",
    "Unmapped": None,
    "Unknown": None,
    "Mixed / Not Split": None,
    "Other / Mixed": None,
}

DIST_CANON = {
    "A Z Enterprises": "AZ Enterprises",
    "Sri Vijaya Durga Agencies": "SVDA",
    "Venkateshwara Agencies": "VA",
    "D.L. Sales": "DL Sales",
    "Balaji Associates": "Balaji",
    "Real Time Logistics": "RealTime",
    "Sai Saachi Associates": "Sai Saachi",
    "Sancus Network": "Sancus Networks Pvt Ltd.",
    "Chhabra Traders": "Chhabra Traders",
    "CHOUDHARY ENTERPRISES": "Choudhary Enterprises",
    "Just Mark": "Just Mark",
    "Kiran Trading Co.": "Kiran Trading",
    "Kottaram": "Kottaram",
    "Mark Enterprises": "Mark Enterprises",
    "Pragati Sales": "Pragati Sales",
    "R R Traders": "RR Traders",
    "Srijan Enterprises": "Srijan",
    "United Marketing": "United Marketing",
    "Sehaj Enterprises": "Sehaj Enterprises",
}

def canon_chain(name):
    return CHAIN_CANON.get(name, name) if name else "Unknown"

def canon_brand(name):
    if not name: return None
    return BRAND_CANON.get(name, name)

def canon_dist(name):
    if not name: return "Unknown"
    return DIST_CANON.get(name, name)


def build_claim_data(wb):
    """Build claim_data dict from openpyxl workbook (f14f5602 version)."""
    import openpyxl

    # ---- Chain Summary ----
    ws_chain = wb["Chain Summary"]
    chain_rows = list(ws_chain.iter_rows(values_only=True))
    # row[0] = Chain, row[1..6] = categories, row[7] = Total
    chain_header = chain_rows[2]  # ('Chain', 'Chain Promo...', ... , 'Total', 'Review Note')
    by_chain = {}
    for row in chain_rows[3:]:
        if not row[0] or not isinstance(row[0], str): continue
        chain_raw = row[0]
        chain_name = canon_chain(chain_raw)
        total_inr = float(row[7] or 0)
        total_lakh = inr_to_lakh(total_inr)
        cats = {}
        for j, cat in enumerate(chain_header[1:7], start=1):
            cats[str(cat)] = inr_to_lakh(row[j] or 0)
        by_chain[chain_name] = {
            "total_claim_lakh": total_lakh,
            "source_chain": chain_raw,
            "by_category": cats,
        }

    # ---- Brand Summary ----
    ws_brand = wb["Brand Summary"]
    brand_rows = list(ws_brand.iter_rows(values_only=True))
    # row[2] = headers ('Brand', 'Apr-2026', 'May-2026', 'Jun-2026', '3-Month Total', ...)
    by_brand = {}
    for row in brand_rows[3:]:
        if not row[0] or not isinstance(row[0], str): continue
        brand_name = canon_brand(row[0])
        if brand_name is None: continue
        apr = inr_to_lakh(row[1] or 0)
        may = inr_to_lakh(row[2] or 0)
        jun = inr_to_lakh(row[3] or 0)
        total = inr_to_lakh(row[4] or 0)
        if brand_name in by_brand:
            by_brand[brand_name]["apr"] += apr
            by_brand[brand_name]["may"] += may
            by_brand[brand_name]["jun"] += jun
            by_brand[brand_name]["total_lakh"] += total
        else:
            by_brand[brand_name] = {
                "apr_lakh": apr, "may_lakh": may, "jun_lakh": jun,
                "total_lakh": total,
            }

    # ---- Distributor Summary ----
    ws_dist = wb["Distributor Summary"]
    dist_rows = list(ws_dist.iter_rows(values_only=True))
    by_distributor = {}
    for row in dist_rows[3:]:
        if not row[0] or not isinstance(row[0], str): continue
        dist_name = canon_dist(row[0])
        total_inr = float(row[4] or 0)
        by_distributor[dist_name] = {
            "total_claim_lakh": inr_to_lakh(total_inr),
            "source_name": row[0],
            "apr_lakh": inr_to_lakh(row[1] or 0),
            "may_lakh": inr_to_lakh(row[2] or 0),
            "jun_lakh": inr_to_lakh(row[3] or 0),
        }

    # ---- Exceptions ----
    ws_exc = wb["Exceptions"]
    exc_rows = list(ws_exc.iter_rows(values_only=True))
    exceptions = []
    for row in exc_rows[3:]:
        if not row[0]: continue
        exceptions.append({
            "month": str(row[0]) if row[0] else "",
            "distributor": str(row[1]) if row[1] else "",
            "issue": str(row[2]) if row[2] else "",
            "source_file": str(row[3]) if row[3] else "",
            "details": str(row[4]) if row[4] else "",
        })

    # ---- KPIs ----
    grand_total_lakh = sum(v["total_claim_lakh"] for v in by_chain.values())
    included_lakh = sum(
        v["total_claim_lakh"] for k, v in by_chain.items()
        if k not in ("AirPlaza", "Avenue Supermarts Ltd.", "Beauty & Nutrition",
                     "Combined Charge", "Fleet Labs", "Transportation", "Unspecified")
        and v["total_claim_lakh"] > 0
    )

    return {
        "metadata": {
            "source_file": "Distributor_Chain_Claim_Master_AprJun_2026_Invoice_Mapped.xlsx",
            "period": "Apr–Jun 2026",
            "currency": "INR (values in Lakh)",
            "provisional": True,
            "generated_by": "scripts/ingest_claims_and_secondary.py",
        },
        "claims": {
            "grand_total_lakh": round(grand_total_lakh, 2),
            "included_mapped_lakh": round(included_lakh, 2),
            "by_chain": by_chain,
            "by_brand": by_brand,
            "by_distributor": by_distributor,
            "exceptions": exceptions,
            "quality_summary": {
                "n_chains": len(by_chain),
                "n_brands": len(by_brand),
                "n_distributors": len(by_distributor),
                "n_exceptions": len(exceptions),
            },
        },
    }


# ---------------------------------------------------------------------------
# 2. Secondary sales ingestion
# ---------------------------------------------------------------------------

def build_secondary_sales(wb):
    """Build secondary_sales dict from All_Sancus_Months workbook."""
    ws = wb["Sales Repository"]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        if not row[0]:
            break
        rows.append(row)

    ci = {h: i for i, h in enumerate(headers)}

    MONTHS = ["2026-04", "2026-05", "2026-06"]
    MONTH_LABELS = {"2026-04": "Apr-2026", "2026-05": "May-2026", "2026-06": "Jun-2026"}

    def zero_mom():
        return {m: 0.0 for m in MONTHS}

    mom_dist = defaultdict(zero_mom)
    mom_chain = defaultdict(zero_mom)
    mom_brand = defaultdict(zero_mom)
    mom_region = defaultdict(zero_mom)
    mom_dist_qty = defaultdict(zero_mom)
    total_nsv = 0.0

    for r in rows:
        m = str(r[ci["Source_Month"]]) if r[ci["Source_Month"]] else ""
        d = str(r[ci["Distributor"]]) if r[ci["Distributor"]] else "Unknown"
        c = str(r[ci["Chain"]]) if r[ci.get("Chain", 99)] is not None else "Unknown"
        b_raw = str(r[ci["Brand"]]) if "Brand" in ci and r[ci["Brand"]] else "Unknown"
        rg = str(r[ci["Region"]]) if r[ci["Region"]] else "Unknown"
        nsv = float(r[ci["NSV_Value"]] or 0) if "NSV_Value" in ci else 0.0
        qty = float(r[ci["Quantity"]] or 0) if "Quantity" in ci else 0.0

        c = canon_chain(c)
        b = canon_brand(b_raw) or "Unknown/Unmapped"

        if m in MONTHS:
            mom_dist[d][m] += nsv
            mom_chain[c][m] += nsv
            mom_brand[b][m] += nsv
            mom_region[rg][m] += nsv
            mom_dist_qty[d][m] += qty
            total_nsv += nsv

    def to_lakh_list(d):
        return {
            "apr_lakh": round(d["2026-04"] / 1e5, 2),
            "may_lakh": round(d["2026-05"] / 1e5, 2),
            "jun_lakh": round(d["2026-06"] / 1e5, 2),
            "total_lakh": round(sum(d.values()) / 1e5, 2),
        }

    by_distributor = [
        {"name": d, **to_lakh_list(mv)}
        for d, mv in sorted(mom_dist.items(), key=lambda x: -sum(x[1].values()))
    ]
    by_chain = [
        {"name": c, **to_lakh_list(mv)}
        for c, mv in sorted(mom_chain.items(), key=lambda x: -sum(x[1].values()))
    ]
    by_brand = [
        {"name": b, **to_lakh_list(mv)}
        for b, mv in sorted(mom_brand.items(), key=lambda x: -sum(x[1].values()))
        if b not in ("Unknown/Unmapped",)
    ]
    by_region = [
        {"name": r, **to_lakh_list(mv)}
        for r, mv in sorted(mom_region.items(), key=lambda x: -sum(x[1].values()))
    ]

    return {
        "metadata": {
            "source_file": "Secondary_Sales_Master_Repository_All_Sancus_Months.xlsx",
            "period": "Apr–Jun 2026",
            "currency": "INR (values in Lakh)",
            "provisional": True,
            "note": "NSV as reported by distributor registers; North registers not available.",
        },
        "total_lakh": round(total_nsv / 1e5, 2),
        "total_cr": round(total_nsv / 1e7, 2),
        "n_rows": len(rows),
        "n_distributors": len(mom_dist),
        "months": ["2026-04", "2026-05", "2026-06"],
        "month_labels": ["Apr-2026", "May-2026", "Jun-2026"],
        "by_distributor": by_distributor,
        "by_chain": by_chain,
        "by_brand": by_brand,
        "by_region": by_region,
    }


# ---------------------------------------------------------------------------
# 3. Patch data.js
# ---------------------------------------------------------------------------

def patch_datajs_block(content, block_key, json_val, after_key=None):
    """Replace or insert a top-level key in the DASH object within data.js."""
    json_str = json.dumps(json_val, ensure_ascii=False, separators=(",", ":"))
    key_pattern = rf'"{re.escape(block_key)}":\s*'

    existing = re.search(key_pattern, content)
    if existing:
        # Find start of value
        val_start = existing.end()
        # Walk braces/brackets to find end
        ch = content[val_start]
        if ch in ('{', '['):
            opener = ch
            closer = '}' if ch == '{' else ']'
            depth = 0
            i = val_start
            while i < len(content):
                if content[i] == opener:
                    depth += 1
                elif content[i] == closer:
                    depth -= 1
                    if depth == 0:
                        val_end = i + 1
                        break
                i += 1
            return content[:val_start] + json_str + content[val_end:]
        else:
            # scalar — find comma or closing brace
            m = re.search(r'[,}]', content[val_start:])
            if m:
                val_end = val_start + m.start()
                return content[:val_start] + json_str + content[val_end:]
    else:
        # Insert after opening brace of DASH object
        dash_open = content.find('"data":') if '"data":' in content else content.find('{')
        insert_after = content.find('{', dash_open) + 1
        insert_str = f'"{block_key}":{json_str},'
        return content[:insert_after] + insert_str + content[insert_after:]

    return content


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--claim", required=True)
    ap.add_argument("--secsale", required=True)
    ap.add_argument("--datajs", default="dashboard/data.js")
    ap.add_argument("--claimjson", default="dashboard/claim_data.json")
    args = ap.parse_args()

    import openpyxl

    print("=== 1. Loading claim Excel …")
    wb_claim = openpyxl.load_workbook(args.claim, read_only=True, data_only=True)
    claim_data = build_claim_data(wb_claim)
    wb_claim.close()

    print(f"   Chains: {claim_data['claims']['quality_summary']['n_chains']}")
    print(f"   Brands: {claim_data['claims']['quality_summary']['n_brands']}")
    print(f"   Distributors: {claim_data['claims']['quality_summary']['n_distributors']}")
    print(f"   Grand total: {claim_data['claims']['grand_total_lakh']:.2f} L")
    print(f"   Included: {claim_data['claims']['included_mapped_lakh']:.2f} L")

    print("\n=== 2. Loading secondary sales Excel …")
    wb_sec = openpyxl.load_workbook(args.secsale, read_only=True, data_only=True)
    sec_data = build_secondary_sales(wb_sec)
    wb_sec.close()

    print(f"   Rows: {sec_data['n_rows']}")
    print(f"   Distributors: {sec_data['n_distributors']}")
    print(f"   Total NSV: {sec_data['total_lakh']:.2f} L / {sec_data['total_cr']:.2f} Cr")

    print("\n=== 3. Writing claim_data.json …")
    Path(args.claimjson).write_text(
        json.dumps(claim_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"   Written: {args.claimjson}")

    print("\n=== 4. Patching data.js …")
    datajs_path = Path(args.datajs)
    content = datajs_path.read_text(encoding="utf-8")
    orig_len = len(content)

    # Inject claims
    claims_payload = claim_data["claims"]
    content = patch_datajs_block(content, "claims", claims_payload)
    print(f"   Patched D.claims  (chains:{claims_payload['quality_summary']['n_chains']}, brands:{claims_payload['quality_summary']['n_brands']})")

    # Inject secondary_sales
    content = patch_datajs_block(content, "secondary_sales", sec_data)
    print(f"   Patched D.secondary_sales  ({sec_data['n_rows']} rows, {sec_data['total_cr']} Cr)")

    # Also patch cm2 expense for top chains
    # Build cm2 chain patch from claim data
    top_chains_expense = {
        k: round(v["total_claim_lakh"], 2)
        for k, v in claims_payload["by_chain"].items()
        if v["total_claim_lakh"] > 0
    }
    print(f"   Top chain expense keys: {list(top_chains_expense.keys())[:5]}")

    datajs_path.write_text(content, encoding="utf-8")
    new_len = len(content)
    print(f"   data.js: {orig_len:,} → {new_len:,} bytes (Δ {new_len-orig_len:+,})")

    print("\n=== 5. Writing exceptions CSV …")
    csv_path = Path("docs/claim_exceptions_review.csv")
    csv_path.parent.mkdir(exist_ok=True)
    lines = ["Month,Distributor,Issue,Source File,Details"]
    for exc in claims_payload.get("exceptions", []):
        def esc(s):
            s = str(s or "")
            return f'"{s.replace(chr(34), chr(34)*2)}"'
        lines.append(",".join(esc(exc[f]) for f in ("month","distributor","issue","source_file","details")))
    csv_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"   Written: {csv_path} ({len(lines)-1} exceptions)")

    print("\nDone.")


if __name__ == "__main__":
    main()
