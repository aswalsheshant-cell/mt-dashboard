#!/usr/bin/env python3
"""
Derive TOT% and MRP Corrected Rate from Primary data.

Input: MTEB2BMTDPrimaryAugXX._3.xlsx (or similar)
Output: Same file with columns 15 (MRP corrected Rate) and 21 (TOT%) populated

Formulas:
  TOT% = (MRP Rate × Inv Qty - Inv. Net value) / (MRP Rate × Inv Qty) × 100
  MRP_Corrected = MRP Rate × (1 - TOT% / 100)

Expected columns (by position):
  14: MRP Rate (source)
  15: MRP corrected Rate (output - currently empty)
  17: Inv Qty (source)
  18: Inv. Net value(LOC) (source)
  21: TOT% (output - currently empty)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, numbers
import sys
from pathlib import Path

def derive_tot_and_mrp(input_file, output_file=None):
    """
    Load Primary data, derive TOT% and MRP Corrected Rate, write output.

    Args:
        input_file: Path to MTEB2BMTDPrimaryAugXX._3.xlsx
        output_file: Path to output file (default: input_file with "_UPDATED" suffix)

    Returns:
        Dict with processing stats
    """
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    if output_file is None:
        output_file = input_path.parent / (input_path.stem + "_UPDATED" + input_path.suffix)

    output_path = Path(output_file)

    # Load workbook
    print(f"Loading {input_file}...")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    print(f"Active sheet: {ws.title}")

    # Column mapping (1-indexed in openpyxl)
    # Column 14 = Column N (MRP Rate)
    # Column 15 = Column O (MRP corrected Rate) - output
    # Column 17 = Column Q (Inv Qty)
    # Column 18 = Column R (Inv. Net value(LOC))
    # Column 21 = Column U (TOT%)  - output

    col_mrp_rate = 14         # N
    col_mrp_corrected = 15    # O
    col_inv_qty = 17          # Q
    col_inv_net_value = 18    # R
    col_tot = 21              # U

    stats = {
        'total_rows': 0,
        'rows_calculated': 0,
        'rows_skipped': 0,
        'errors': []
    }

    # Iterate through data rows (assuming row 1 is header)
    for row_idx in range(2, ws.max_row + 1):
        stats['total_rows'] += 1

        try:
            # Get source values
            mrp_rate = ws.cell(row=row_idx, column=col_mrp_rate).value
            inv_qty = ws.cell(row=row_idx, column=col_inv_qty).value
            inv_net_value = ws.cell(row=row_idx, column=col_inv_net_value).value

            # Skip if critical values are missing
            if any(v is None for v in [mrp_rate, inv_qty, inv_net_value]):
                stats['rows_skipped'] += 1
                continue

            # Convert to float
            mrp_rate = float(mrp_rate)
            inv_qty = float(inv_qty)
            inv_net_value = float(inv_net_value)

            # Calculate TOT%
            # TOT% = (MRP Rate × Inv Qty - Inv. Net value) / (MRP Rate × Inv Qty) × 100
            mrp_extended = mrp_rate * inv_qty
            if mrp_extended == 0:
                tot_pct = 0
            else:
                tot_pct = ((mrp_extended - inv_net_value) / mrp_extended) * 100

            # Calculate MRP Corrected
            # MRP_Corrected = MRP Rate × (1 - TOT% / 100)
            mrp_corrected = mrp_rate * (1 - tot_pct / 100)

            # Write to sheet
            ws.cell(row=row_idx, column=col_tot).value = tot_pct
            ws.cell(row=row_idx, column=col_mrp_corrected).value = mrp_corrected

            # Apply number formatting if desired
            ws.cell(row=row_idx, column=col_tot).number_format = '0.00'
            ws.cell(row=row_idx, column=col_mrp_corrected).number_format = '0.00'

            stats['rows_calculated'] += 1

            if row_idx % 1000 == 0:
                print(f"  Processed {row_idx}/{ws.max_row} rows...")

        except Exception as e:
            stats['errors'].append({
                'row': row_idx,
                'error': str(e)
            })

    # Save output
    print(f"\nSaving to {output_file}...")
    wb.save(output_file)
    wb.close()

    # Print summary
    print("\n" + "="*60)
    print("DERIVATION SUMMARY")
    print("="*60)
    print(f"Total rows processed:    {stats['total_rows']}")
    print(f"Rows calculated:         {stats['rows_calculated']}")
    print(f"Rows skipped (missing):  {stats['rows_skipped']}")
    if stats['errors']:
        print(f"Rows with errors:        {len(stats['errors'])}")
        print("\nFirst 5 errors:")
        for err in stats['errors'][:5]:
            print(f"  Row {err['row']}: {err['error']}")
    print("="*60)
    print(f"\n✓ Output saved: {output_file}")

    return stats

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python derive_tot_and_mrp.py <input_xlsx> [output_xlsx]")
        print("\nExample:")
        print("  python derive_tot_and_mrp.py MTEB2BMTDPrimaryAug26._3.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        stats = derive_tot_and_mrp(input_file, output_file)
        sys.exit(0 if not stats['errors'] else 1)
    except Exception as e:
        print(f"❌ Fatal error: {e}", file=sys.stderr)
        sys.exit(1)
