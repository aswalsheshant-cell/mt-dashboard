"""Output Trust and Release Control layer (agent/policies/AI_LEVERAGE_AND_JUDGMENT.md
enforcement §A-§F, plus the Output Release Gate extension).

A file being generated successfully is not the same claim as a file being
safe to share. This module is the single place that decides whether an
output is DRAFT, VALIDATED, or APPROVED_FOR_SHARING, and separates every
key conclusion into Fact / Inference / Recommendation with an explicit
confidence level and cited evidence -- never a bare assertion.

Nothing here requires DuckDB/Ollama. `redaction_scan()` uses `openpyxl`
(already an optional dependency in agent/requirements.txt) and degrades to
an explicit "cannot scan" result, never a silent pass, if it isn't
installed.
"""
from __future__ import annotations

from dataclasses import dataclass, field

# --------------------------------------------------------------------- #
# 1. Output Release Gate
# --------------------------------------------------------------------- #
DRAFT = "DRAFT"
VALIDATED = "VALIDATED"
APPROVED_FOR_SHARING = "APPROVED_FOR_SHARING"

_CHECKLIST_ITEMS = (
    "source_validated",
    "business_rules_applied",
    "totals_reconciled",
    "period_status_confirmed",
    "mappings_approved",
    "exceptions_disclosed",
    "output_visually_checked",
    "version_and_timestamp_added",
    "confidentiality_level_confirmed",
)


@dataclass
class ReleaseChecklist:
    source_validated: bool = False
    business_rules_applied: bool = False
    totals_reconciled: bool = False
    period_status_confirmed: bool = False
    mappings_approved: bool = False
    exceptions_disclosed: bool = False
    output_visually_checked: bool = False
    version_and_timestamp_added: bool = False
    confidentiality_level_confirmed: bool = False

    def failing_items(self) -> list:
        return [item for item in _CHECKLIST_ITEMS if not getattr(self, item)]


def evaluate_release(checklist: ReleaseChecklist, human_approved: bool = False) -> tuple:
    """Returns (status, blocking_reasons). A checklist with ANY failing
    item can never reach APPROVED_FOR_SHARING, even if `human_approved` is
    True -- human approval authorizes sharing, it does not substitute for
    an unreconciled total or an unreviewed hidden sheet.
    """
    failing = checklist.failing_items()
    if failing:
        return DRAFT, failing
    if not human_approved:
        return VALIDATED, ["awaiting explicit human sharing approval"]
    return APPROVED_FOR_SHARING, []


# --------------------------------------------------------------------- #
# 2. Data lineage summary
# --------------------------------------------------------------------- #
def format_lineage(output_name: str, source_files: list, transformations: list, validation_results: dict) -> str:
    lines = ["Output:", output_name, "", "Source files:"]
    lines += [f"- {s}" for s in source_files] or ["- (none)"]
    lines += ["", "Transformation:"]
    lines += [f"- {t}" for t in transformations] or ["- (none)"]
    lines += ["", "Validation:"]
    lines += [f"- {k}: {v}" for k, v in validation_results.items()] or ["- (none)"]
    return "\n".join(lines)


# --------------------------------------------------------------------- #
# 3/4/6. Fact / Inference / Recommendation, confidence, claim-to-evidence
# --------------------------------------------------------------------- #
FACT = "FACT"
INFERENCE = "INFERENCE"
RECOMMENDATION = "RECOMMENDATION"
_CLAIM_KINDS = (FACT, INFERENCE, RECOMMENDATION)

HIGH = "HIGH"
MEDIUM = "MEDIUM"
LOW = "LOW"
_CONFIDENCE_LEVELS = (HIGH, MEDIUM, LOW)


@dataclass
class Claim:
    statement: str
    kind: str                              # FACT | INFERENCE | RECOMMENDATION
    evidence: list = field(default_factory=list)
    confidence: str = MEDIUM               # HIGH | MEDIUM | LOW


def validate_claim(claim: Claim) -> tuple:
    """A FACT or INFERENCE with no cited evidence is rejected outright --
    it must not reach a final response as if it were a confirmed number.
    RECOMMENDATIONs are proposals, not measurements, so they aren't held
    to the same evidence bar, but an unknown kind/confidence is still
    rejected rather than silently accepted.
    """
    if claim.kind not in _CLAIM_KINDS:
        return False, f"unknown claim kind '{claim.kind}' -- must be one of {_CLAIM_KINDS}"
    if claim.confidence not in _CONFIDENCE_LEVELS:
        return False, f"unknown confidence level '{claim.confidence}' -- must be one of {_CONFIDENCE_LEVELS}"
    if claim.kind in (FACT, INFERENCE) and not claim.evidence:
        return False, f"{claim.kind} claim '{claim.statement}' has no supporting evidence -- rejected, not stated as fact"
    return True, "ok"


def format_claim(claim: Claim) -> str:
    lines = [f"{claim.kind.title()} ({claim.confidence.title()} confidence): {claim.statement}"]
    if claim.evidence:
        lines.append("Evidence:")
        lines += [f"- {e}" for e in claim.evidence]
    return "\n".join(lines)


# --------------------------------------------------------------------- #
# 5. Materiality classification (Critical / Material / Informational)
# --------------------------------------------------------------------- #
CRITICAL = "Critical"
MATERIAL = "Material"
INFORMATIONAL = "Informational"


def classify_materiality(pct_change: float | None = None, abs_impact: float | None = None,
                          contribution_pct: float | None = None,
                          mapping_exception_value: float | None = None,
                          financial_threshold: float = 1_000_000,       # ₹10L
                          growth_threshold: float = 0.10,                # ±10%
                          contribution_threshold: float = 0.80,          # Top 80%
                          mapping_exception_threshold: float = 500_000,  # ₹5L
                          critical_multiple: float = 5.0) -> str:
    """`critical_multiple` is a documented assumption, not a business rule
    handed down by the user: a financial/mapping impact at or above
    `critical_multiple` x its own materiality threshold is escalated from
    Material to Critical. Adjust the multiple per call if leadership wants
    a different split -- it is a parameter, not a hardcoded constant.
    """
    if mapping_exception_value is not None and abs(mapping_exception_value) > mapping_exception_threshold:
        return CRITICAL if abs(mapping_exception_value) >= mapping_exception_threshold * critical_multiple else MATERIAL
    if abs_impact is not None and abs(abs_impact) >= financial_threshold:
        return CRITICAL if abs(abs_impact) >= financial_threshold * critical_multiple else MATERIAL
    if pct_change is not None and abs(pct_change) >= growth_threshold:
        return MATERIAL
    if contribution_pct is not None and contribution_pct >= contribution_threshold:
        return MATERIAL
    return INFORMATIONAL


# --------------------------------------------------------------------- #
# 7. Scope and period labels
# --------------------------------------------------------------------- #
@dataclass
class ScopeLabel:
    reporting_period: str
    period_status: str              # "Complete" | "Partial"
    data_type: str                  # "Primary" | "Secondary" | "Both"
    tax_basis: str                  # e.g. "excluding tax"
    unit_basis: str                 # e.g. "NSV Rs Cr"
    data_refresh_date: str
    scope_exclusions: list = field(default_factory=list)

    def format(self) -> str:
        lines = [
            f"Reporting period: {self.reporting_period} ({self.period_status})",
            f"Data type: {self.data_type}",
            f"Basis: {self.unit_basis}, {self.tax_basis}",
            f"Data refresh date: {self.data_refresh_date}",
        ]
        if self.scope_exclusions:
            lines.append(f"Scope exclusions: {', '.join(self.scope_exclusions)}")
        return "\n".join(lines)


# --------------------------------------------------------------------- #
# 8. Audience-based output modes
# --------------------------------------------------------------------- #
AUDIENCE_LEADERSHIP = "leadership"
AUDIENCE_ANALYST = "analyst"
AUDIENCE_OPERATIONS = "operations"

AUDIENCE_SECTIONS = {
    AUDIENCE_LEADERSHIP: ["key_insights", "business_impact", "risk", "required_action"],
    AUDIENCE_ANALYST: ["reconciliation", "mapping_exceptions", "assumptions", "detailed_tables", "qc_evidence"],
    AUDIENCE_OPERATIONS: ["account_exceptions", "owner", "action", "deadline", "source_issue"],
}


def select_sections(audience: str) -> list:
    if audience not in AUDIENCE_SECTIONS:
        raise ValueError(f"unknown audience '{audience}' -- must be one of {sorted(AUDIENCE_SECTIONS)}")
    return list(AUDIENCE_SECTIONS[audience])


# --------------------------------------------------------------------- #
# 9. Pre-share redaction check (real, working, degrades gracefully)
# --------------------------------------------------------------------- #
_SUSPICIOUS_KEYWORDS = (
    "password", "token", "api key", "apikey", "secret", "confidential",
    "internal only", "do not share", "cm2", "cost price", "margin assumption",
)


def redaction_scan(xlsx_path) -> tuple:
    """Scans an xlsx for hidden sheets, suspicious keywords in sheet names
    or cell text, and cell comments (often reviewer notes not meant for
    external eyes). Returns (clean, issues). If openpyxl isn't installed,
    returns (False, ["openpyxl not installed -- cannot scan, do not treat
    as clean"]) -- never silently reports a clean scan it didn't do.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, ["openpyxl not installed -- cannot scan this file for redaction issues; "
                        "do not mark APPROVED_FOR_SHARING without a manual check"]

    issues = []
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            issues.append(f"hidden sheet '{ws.title}' present -- must be reviewed before sharing")
        title_lower = ws.title.lower()
        for kw in _SUSPICIOUS_KEYWORDS:
            if kw in title_lower:
                issues.append(f"sheet name '{ws.title}' matches suspicious keyword '{kw}'")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    val_lower = cell.value.lower()
                    for kw in _SUSPICIOUS_KEYWORDS:
                        if kw in val_lower:
                            issues.append(f"{ws.title}!{cell.coordinate}: cell text matches suspicious keyword '{kw}'")
                if cell.comment is not None:
                    issues.append(f"{ws.title}!{cell.coordinate}: has a cell comment -- review before external sharing")
    wb.close()
    return (len(issues) == 0), issues


# --------------------------------------------------------------------- #
# 10. Visual QC -- honest scope: only the automatable subset
# --------------------------------------------------------------------- #
def formula_error_scan(xlsx_path) -> tuple:
    """The only part of 'visual QC' this environment can actually check
    without rendering the file: cells whose CACHED value is an Excel
    error literal (#REF!, #N/A, #VALUE!, ...). Chart placement, fonts,
    colors, truncated labels, and axis scale are NOT checked here -- they
    require a human look or a render step, and `output_visually_checked`
    on ReleaseChecklist must reflect that, not be set from this function
    alone.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, ["openpyxl not installed -- cannot scan for formula errors"]

    error_literals = ("#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!")
    issues = []
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in error_literals:
                    issues.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
    wb.close()
    return (len(issues) == 0), issues


# --------------------------------------------------------------------- #
# 11. Reproducibility / version filename
# --------------------------------------------------------------------- #
_STATUS_SLUG = {DRAFT: "Draft", VALIDATED: "Validated", APPROVED_FOR_SHARING: "Approved"}


def build_version_filename(base: str, status: str, version: int, date: str, ext: str = "xlsx") -> str:
    slug = _STATUS_SLUG.get(status, status)
    return f"{base}_{slug}_v{version}_{date}.{ext}"


# --------------------------------------------------------------------- #
# 12. Human approval matrix
# --------------------------------------------------------------------- #
APPROVAL_MATRIX = {
    "read_and_analyse_data": False,
    "generate_draft_report": False,
    "apply_approved_alias": False,
    "create_new_mapping_assumption": True,
    "change_grand_total_logic": True,
    "overwrite_approved_file": True,
    "commit_or_push": True,
    "publish_or_share": True,
}


def requires_approval(action_key: str) -> bool:
    if action_key not in APPROVAL_MATRIX:
        raise KeyError(f"unknown action '{action_key}' -- add it to APPROVAL_MATRIX explicitly, never assume")
    return APPROVAL_MATRIX[action_key]


# --------------------------------------------------------------------- #
# The final-output structure (AI_LEVERAGE_AND_JUDGMENT.md rule 15)
# --------------------------------------------------------------------- #
def format_final_response(*, run_status: str, business_outcome: str, outcome_achieved: str,
                           key_results: list, validation_evidence: dict, assumptions: list,
                           exceptions_and_risks: list, confidence: list, files: list,
                           sharing_status: str, decision_required: str | None) -> str:
    lines = [f"Run Status: {run_status}", "", "Business Outcome:", business_outcome, "",
             f"Outcome Achieved: {outcome_achieved}", "", "Key Results:"]
    lines += [f"- {k}" for k in key_results] if key_results else ["- (none)"]
    lines += ["", "Validation Evidence:"]
    lines += [f"- {k}: {v}" for k, v in validation_evidence.items()] if validation_evidence else ["- (none)"]
    lines += ["", "Assumptions:"]
    lines += [f"- {a}" for a in assumptions] if assumptions else ["- (none)"]
    lines += ["", "Exceptions and Risks:"]
    lines += [f"- {r}" for r in exceptions_and_risks] if exceptions_and_risks else ["- (none)"]
    lines += ["", "Confidence:"]
    lines += [f"- {c}" for c in confidence] if confidence else ["- (none)"]
    lines += ["", "Files Created or Changed:"]
    lines += [f"- {f}" for f in files] if files else ["- (none)"]
    lines += ["", f"Sharing Status: {sharing_status}", "", f"Decision Required: {decision_required or 'None'}"]
    return "\n".join(lines)
