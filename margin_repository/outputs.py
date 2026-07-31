# -*- coding: utf-8 -*-
"""Generate the 10 charter outputs + dashboard dataset + templates.

All outputs are DERIVED from the append-only repository history; the history
itself is never mutated here.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from schema import (REPO_COLS, ARTICLE_COLS, COMMERCIAL_COLS, CONDITION_COLS,
                    DATE_COLS, COMMERCIAL_PCT_COLS)
from validation import qc_report

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SEV_FILL = {"PASS": "C6E9C6", "WARNING": "FFF2A8", "FAIL": "FCD5A5", "BLOCKED": "F4B7B7"}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write(ws, df, sev_col=None, wrap=()):
    cols = list(df.columns)
    for c, col in enumerate(cols, 1):
        cell = ws.cell(1, c, col)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    text_cols = {"EAN", "SKU Code", "Site Code", "Pack Size", "Version Number"}
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        fill = None
        if sev_col and sev_col in df.columns:
            fill = SEV_FILL.get(str(row[sev_col]).upper())
        for c, col in enumerate(cols, 1):
            v = row[col]
            if pd.isna(v):
                v = ""
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9); cell.border = BORDER
            if col in text_cols:
                cell.number_format = "@"; cell.value = "" if v == "" else str(v)
            cell.alignment = Alignment(wrap_text=col in wrap, vertical="top")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
    for c, col in enumerate(cols, 1):
        w = 34 if col in wrap else max(11, min(26, len(str(col)) + 3))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    if len(df):
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(df) + 1)


def _num(df, col):
    return pd.to_numeric(df.get(col), errors="coerce")


def build_outputs(repo, out_path, changelog=None, removed=None, impact=None,
                  as_of=None):
    """repo: MarginRepository. Writes the full outputs workbook."""
    as_of = as_of or pd.Timestamp.today().normalize().date().isoformat()
    hist = repo.history.copy()
    cur = repo.current(include_held=False)
    cur_all = repo.current(include_held=True)
    wb = Workbook(); wb.remove(wb.active)

    # 1. Chain Article Margin Master (current published) + full history
    master_cols = [c for c in REPO_COLS if c in cur_all.columns]
    _write(wb.create_sheet("1_Chain_Article_Master"),
           cur_all[master_cols] if not cur_all.empty else pd.DataFrame(columns=master_cols),
           sev_col="QC_Severity", wrap=("Validation_Flags",))
    _write(wb.create_sheet("1b_Full_Version_History"),
           hist[[c for c in REPO_COLS if c in hist.columns]] if not hist.empty
           else pd.DataFrame(columns=REPO_COLS), sev_col="QC_Severity")

    # 2. Margin Change Log
    cl = changelog if changelog is not None else pd.DataFrame()
    _write(wb.create_sheet("2_Margin_Change_Log"),
           cl if not cl.empty else pd.DataFrame(columns=[
               "Chain", "Brand", "Article", "EAN", "Field", "Old Value",
               "New Value", "Difference", "Reason", "Effective From", "Version"]),
           wrap=("Reason",))

    # 3. Commercial Difference Report (field diffs, pivoted per article)
    if not cl.empty:
        diff = cl.pivot_table(index=["Chain", "Brand", "Article", "EAN"],
                              columns="Field", values="Difference",
                              aggfunc="first").reset_index()
    else:
        diff = pd.DataFrame(columns=["Chain", "Brand", "Article", "EAN"])
    _write(wb.create_sheet("3_Commercial_Difference"), diff)

    # 4. Missing Margin Report
    if not cur_all.empty:
        fem = _num(cur_all, "Final Effective Margin %")
        tm = _num(cur_all, "Trade Margin %")
        miss = cur_all[(tm.isna()) | (tm == 0) | (fem.isna()) |
                       cur_all["Validation_Flags"].str.contains("MISSING_COMMERCIALS|BLANK_TRADE_MARGIN", na=False)]
        miss = miss[["Chain", "Brand", "Category", "Article", "EAN", "Pack Size",
                     "Trade Margin %", "Final Effective Margin %", "Validation_Flags"]]
    else:
        miss = pd.DataFrame(columns=["Chain", "Brand", "Article", "EAN"])
    _write(wb.create_sheet("4_Missing_Margin"), miss, wrap=("Validation_Flags",))

    # 5. Duplicate Article Report
    if not cur_all.empty:
        dup = cur_all[cur_all["Validation_Flags"].str.contains(
            "DUPLICATE_EAN|DUPLICATE_CHAIN_ARTICLE|DUPLICATE_EFFECTIVE_DATE", na=False)]
        dup = dup[["Chain", "Brand", "Article", "EAN", "Pack Size", "MRP",
                   "Effective From", "Version Number", "Validation_Flags"]]
    else:
        dup = pd.DataFrame(columns=["Chain", "Article", "EAN"])
    _write(wb.create_sheet("5_Duplicate_Article"), dup, wrap=("Validation_Flags",))

    # 6. Margin QC Report
    qc = qc_report(cur_all) if not cur_all.empty else pd.DataFrame(
        columns=["QC Metric", "Value"])
    _write(wb.create_sheet("6_Margin_QC"), qc)

    # 7. Forecast Ready Margin Table (latest approved margin per Chain x EAN)
    fr_cols = ["Chain", "Brand", "Category", "Sub Category", "Article", "EAN",
               "Pack Size", "MRP", "Trade Margin %", "Final Effective Margin %",
               "GST %", "Effective From", "Effective To", "Version Number"]
    fr = cur[[c for c in fr_cols if c in cur.columns]] if not cur.empty \
        else pd.DataFrame(columns=fr_cols)
    _write(wb.create_sheet("7_Forecast_Ready_Margin"), fr)

    # 8. CM2 Ready Margin Table (component breakdown for CM2 modelling)
    cm2_cols = ["Chain", "Brand", "Category", "Article", "EAN", "Pack Size", "MRP",
                "GST %"] + COMMERCIAL_PCT_COLS + ["Final Effective Margin %",
                                                  "Effective From", "Version Number"]
    cm2 = cur[[c for c in cm2_cols if c in cur.columns]] if not cur.empty \
        else pd.DataFrame(columns=cm2_cols)
    _write(wb.create_sheet("8_CM2_Ready_Margin"), cm2)

    # 9. Dashboard Dataset (aggregations)
    _build_dashboard(wb, cur, cl, removed)

    # 10. Executive Summary
    _build_exec_summary(wb, repo, cur, cur_all, cl, removed, qc, as_of)

    wb.save(out_path)
    return out_path


def _agg_margin(cur, by):
    if cur.empty:
        return pd.DataFrame(columns=[by, "Articles", "Avg Final Margin %", "Avg Trade Margin %"])
    g = cur.copy()
    g["_fem"] = _num(g, "Final Effective Margin %")
    g["_tm"] = _num(g, "Trade Margin %")
    out = g.groupby(by).agg(Articles=(by, "size"),
                            **{"Avg Final Margin %": ("_fem", "mean"),
                               "Avg Trade Margin %": ("_tm", "mean")}).reset_index()
    out["Avg Final Margin %"] = out["Avg Final Margin %"].round(2)
    out["Avg Trade Margin %"] = out["Avg Trade Margin %"].round(2)
    return out.sort_values("Avg Final Margin %", ascending=False)


def _build_dashboard(wb, cur, cl, removed):
    ws = wb.create_sheet("9_Dashboard_Dataset")
    blocks = [
        ("Chain-wise Margin", _agg_margin(cur, "Chain")),
        ("Brand-wise Margin", _agg_margin(cur, "Brand")),
        ("Category Margin", _agg_margin(cur, "Category")),
    ]
    # commercial expiry calendar
    if not cur.empty and "Effective To" in cur.columns:
        exp = cur.copy()
        exp["_to"] = pd.to_datetime(exp["Effective To"], errors="coerce")
        exp = exp[exp["_to"].notna()].sort_values("_to")
        expiry = exp[["Chain", "Brand", "Article", "EAN", "Effective To"]].head(200)
    else:
        expiry = pd.DataFrame(columns=["Chain", "Article", "EAN", "Effective To"])
    # articles without commercials
    if not cur.empty:
        tm = _num(cur, "Trade Margin %")
        no_comm = cur[(tm.isna()) | (tm == 0)][["Chain", "Brand", "Article", "EAN"]]
    else:
        no_comm = pd.DataFrame(columns=["Chain", "Article", "EAN"])
    # top margin changes
    if cl is not None and not cl.empty:
        top = cl.copy()
        top["_d"] = pd.to_numeric(top["Difference"], errors="coerce").abs()
        top = top.sort_values("_d", ascending=False).head(25)[
            ["Chain", "Brand", "Article", "EAN", "Field", "Old Value", "New Value", "Difference"]]
    else:
        top = pd.DataFrame(columns=["Chain", "Article", "Field", "Difference"])
    blocks += [("Commercial Expiry Calendar", expiry),
               ("Articles Without Commercials", no_comm),
               ("Top Margin Changes", top)]

    r = 1
    for title, tbl in blocks:
        ws.cell(r, 1, title).font = Font(bold=True, size=11, color="1F4E78")
        r += 1
        if tbl.empty:
            ws.cell(r, 1, "(none)"); r += 2; continue
        for c, col in enumerate(tbl.columns, 1):
            cell = ws.cell(r, c, col); cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        r += 1
        for _, row in tbl.iterrows():
            for c, col in enumerate(tbl.columns, 1):
                ws.cell(r, c, "" if pd.isna(row[col]) else row[col])
            r += 1
        r += 2
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 20


def _build_exec_summary(wb, repo, cur, cur_all, cl, removed, qc, as_of):
    ws = wb.create_sheet("10_Executive_Summary")
    n_records = len(repo.history)
    n_current = len(cur)
    chains = int(cur_all["Chain"].nunique()) if not cur_all.empty else 0
    new_ct = int((repo.history.get("Change_Type") == "NEW").sum()) if n_records else 0
    chg_ct = int((repo.history.get("Change_Type") == "CHANGED").sum()) if n_records else 0
    health = qc.set_index("QC Metric")["Value"].get("Repository Health %", "n/a") if not qc.empty else "n/a"
    conf = qc.set_index("QC Metric")["Value"].get("Confidence Score %", "n/a") if not qc.empty else "n/a"
    blocked = qc.set_index("QC Metric")["Value"].get("BLOCKED", 0) if not qc.empty else 0
    lines = [
        ("Chain x Article Margin Repository - Executive Summary", ""),
        ("As of", as_of),
        ("", ""),
        ("Total versioned records (append-only history)", n_records),
        ("Current published articles (forecast-ready)", n_current),
        ("Chains covered", chains),
        ("New articles ingested (lifetime)", new_ct),
        ("Changed-margin versions (lifetime)", chg_ct),
        ("Change-log entries this run", 0 if cl is None else len(cl)),
        ("Articles removed from latest file (reported, not deleted)",
         0 if removed is None else len(removed)),
        ("", ""),
        ("Repository Health %", health),
        ("Confidence Score %", conf),
        ("BLOCKED records (held from publish)", blocked),
        ("", ""),
        ("Non-negotiables enforced", "Append-only - never overwrites; EAN-priority key; "
         "full audit trail; validate-before-publish; all inconsistencies flagged."),
        ("Downstream feeds", "Sheets 7 (Forecast-Ready) and 8 (CM2-Ready) expose the "
         "latest APPROVED chain-specific margin for demand/primary/offtake/CM2/planning."),
        ("Integration points (stubbed)", "Forecast/CM2/planning model refresh and "
         "on-file-arrival automation are wired as callable hooks - connect to live "
         "systems when available. No margins are fabricated."),
    ]
    for i, (k, v) in enumerate(lines, 1):
        a = ws.cell(i, 1, k); b = ws.cell(i, 2, v)
        if i == 1:
            a.font = Font(bold=True, size=13, color="1F4E78")
        elif v == "" and k:
            a.font = Font(bold=True, size=10)
        else:
            a.font = Font(size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 70


# ---------------------------------------------------------------------------
# Blank templates
# ---------------------------------------------------------------------------
def build_repository_template(path):
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Repository_Master")
    groups = [("ARTICLE INFORMATION", ARTICLE_COLS),
              ("COMMERCIAL INFORMATION", COMMERCIAL_COLS),
              ("COMMERCIAL CONDITIONS", CONDITION_COLS),
              ("DATE / VERSION", DATE_COLS)]
    c = 1
    group_fill = PatternFill("solid", fgColor="2E75B6")
    for gname, gcols in groups:
        ws.merge_cells(start_row=1, start_column=c, end_row=1, end_column=c + len(gcols) - 1)
        gc = ws.cell(1, c, gname); gc.fill = group_fill
        gc.font = Font(bold=True, color="FFFFFF"); gc.alignment = Alignment(horizontal="center")
        for j, col in enumerate(gcols):
            hc = ws.cell(2, c + j, col); hc.fill = HEADER_FILL; hc.font = HEADER_FONT
            hc.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(c + j)].width = 16
        c += len(gcols)
    ws.freeze_panes = "A3"
    # data-dictionary sheet
    dd = wb.create_sheet("Data_Dictionary")
    rows = [("PRIMARY KEY", "Chain + Brand + Category + Sub Category + Article + EAN + "
             "Pack Size + MRP + Effective From. EAN takes priority; never Article alone."),
            ("Final Effective Margin %", "Provided value respected; else derived = sum of "
             "earned components minus Consumer Offer % and Cash Discount %."),
            ("Version Number", "Auto-incremented by the engine on any tracked-field change. "
             "Never overwrite; every version is retained."),
            ("Approval Status", "APPROVED records feed Forecast/CM2. PENDING/blank are held."),
            ("GST %", "Must be one of 0, 5, 12, 18, 28."),
            ("Status", "ACTIVE / INACTIVE / DELISTED. Inactive articles are flagged.")]
    dd.cell(1, 1, "Field").font = HEADER_FONT; dd.cell(1, 1).fill = HEADER_FILL
    dd.cell(1, 2, "Rule").font = HEADER_FONT; dd.cell(1, 2).fill = HEADER_FILL
    for i, (k, v) in enumerate(rows, 2):
        dd.cell(i, 1, k).font = Font(bold=True, size=9)
        vc = dd.cell(i, 2, v); vc.alignment = Alignment(wrap_text=True, vertical="top")
    dd.column_dimensions["A"].width = 26; dd.column_dimensions["B"].width = 90
    wb.save(path); return path


def build_import_template(path):
    """Blank input template for source commercial files."""
    wb = Workbook(); ws = wb.active; ws.title = "Commercial_Import"
    cols = ARTICLE_COLS + COMMERCIAL_COLS + CONDITION_COLS + DATE_COLS
    for c, col in enumerate(cols, 1):
        hc = ws.cell(1, c, col); hc.fill = HEADER_FILL; hc.font = HEADER_FONT
        hc.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    wb.save(path); return path
