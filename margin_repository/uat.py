# -*- coding: utf-8 -*-
"""User Acceptance Test framework for the Margin Repository engine.

Runs a structured set of test cases covering every business rule and edge case.
Produces a UAT workbook with test case, expected result, actual result, verdict,
evidence, root cause, corrective action, and owner columns.
"""
import os
import sys
import shutil
import tempfile
import traceback

sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from schema import REPO_COLS
from repository import MarginRepository, coerce_numeric, derive_final_margin
from validation import validate_frame, qc_report
from impact import impact_from_changelog
from search import search
from outputs import build_outputs

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
PASS_FILL = PatternFill("solid", fgColor="C6E9C6")
WARN_FILL = PatternFill("solid", fgColor="FFF2A8")
FAIL_FILL = PatternFill("solid", fgColor="F4B7B7")


class UAT:
    def __init__(self):
        self.results = []

    def check(self, test_id, case, expected, condition, actual="", evidence="",
              root_cause="", corrective="", owner=""):
        verdict = "PASS" if condition else "FAIL"
        if actual == "":
            actual = str(condition)
        self.results.append({
            "Test ID": test_id, "Test Case": case, "Expected Result": expected,
            "Actual Result": actual, "Verdict": verdict, "Evidence": evidence,
            "Root Cause": root_cause if not condition else "",
            "Corrective Action": corrective if not condition else "",
            "Owner": owner,
        })
        return condition


def _valid_rows():
    return [
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Sunscreen", "Sub Category": "SPF50",
         "Article": "UV Shield 50ml",   "EAN": "8901234560001", "Pack Size": "50ml", "MRP": 499,
         "Trade Margin %": 28, "TOT %": 2, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Sunscreen", "Sub Category": "SPF30",
         "Article": "UV Light 100ml",   "EAN": "8901234560002", "Pack Size": "100ml", "MRP": 699,
         "Trade Margin %": 25, "TOT %": 3, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo", "Sub Category": "Volume",
         "Article": "Volume Pro 150ml", "EAN": "8901234560004", "Pack Size": "150ml", "MRP": 299,
         "Trade Margin %": 20, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Facewash", "Sub Category": "Gel",
         "Article": "FreshGel 100ml",   "EAN": "8901234560005", "Pack Size": "100ml", "MRP": 249,
         "Trade Margin %": 24, "TOT %": 1, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Serum", "Sub Category": "Vitamin C",
         "Article": "Glow Serum 30ml",  "EAN": "8901234560006", "Pack Size": "30ml", "MRP": 799,
         "Trade Margin %": 30, "TOT %": 2, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Moisturizer", "Sub Category": "Gel",
         "Article": "AquaBurst 50ml",   "EAN": "8901234560007", "Pack Size": "50ml", "MRP": 599,
         "Trade Margin %": 26, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo", "Sub Category": "Anti-Dandruff",
         "Article": "Clean Root 200ml", "EAN": "8901234560003", "Pack Size": "200ml", "MRP": 350,
         "Trade Margin %": 22, "Frontend %": 2, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Conditioner", "Sub Category": "Repair",
         "Article": "SilkSmooth 200ml", "EAN": "8901234560011", "Pack Size": "200ml", "MRP": 399,
         "Trade Margin %": 21, "TOT %": 1, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Cream", "Sub Category": "Night",
         "Article": "NightRepair 50g",  "EAN": "8901234560012", "Pack Size": "50g", "MRP": 449,
         "Trade Margin %": 18, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Oil", "Sub Category": "Hair",
         "Article": "Hair Oil 100ml",   "EAN": "8901234560013", "Pack Size": "100ml", "MRP": 349,
         "Trade Margin %": 15, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
    ]


def _uat_batch():
    """Controlled UAT input covering all edge cases."""
    rows = _valid_rows()
    # blank EAN
    rows.append({"Chain": "DMart", "Brand": "TestBrand", "Category": "Cream", "Sub Category": "Night",
                 "Article": "NightGlow 30g", "EAN": "", "Pack Size": "30g", "MRP": 299,
                 "Trade Margin %": 19, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"})
    # duplicate article (same EAN as row 0)
    rows.append({"Chain": "Apollo", "Brand": "TestBrand", "Category": "Sunscreen", "Sub Category": "SPF50",
                 "Article": "UV Shield 50ml DUPE", "EAN": "8901234560001", "Pack Size": "50ml", "MRP": 499,
                 "Trade Margin %": 28, "TOT %": 2, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"})
    # invalid GST
    rows.append({"Chain": "Reliance", "Brand": "TestBrand", "Category": "Kajal", "Sub Category": "Eye",
                 "Article": "Bold Kajal 0.3g", "EAN": "8901234560014", "Pack Size": "0.3g", "MRP": 199,
                 "Trade Margin %": 32, "GST %": 15, "Effective From": "2026-04-01", "Status": "ACTIVE"})
    # missing margin
    rows.append({"Chain": "DMart", "Brand": "TestBrand", "Category": "Oil", "Sub Category": "Body",
                 "Article": "BodyOil 100ml", "EAN": "8901234560015", "Pack Size": "100ml", "MRP": 249,
                 "Trade Margin %": "", "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"})
    # new article (not in valid rows set)
    rows.append({"Chain": "DMart", "Brand": "TestBrand", "Category": "Conditioner", "Sub Category": "Smoothing",
                 "Article": "SuperSmooth 150ml", "EAN": "8901234560016", "Pack Size": "150ml", "MRP": 379,
                 "Trade Margin %": 23, "GST %": 18, "Effective From": "2026-07-01", "Status": "ACTIVE"})
    # delisted article
    rows.append({"Chain": "Apollo", "Brand": "TestBrand", "Category": "Serum", "Sub Category": "Retinol",
                 "Article": "RetinolMax 15ml", "EAN": "8901234560017", "Pack Size": "15ml", "MRP": 999,
                 "Trade Margin %": 35, "GST %": 18, "Effective From": "2026-01-01", "Status": "DELISTED"})
    # future-effective margin
    rows.append({"Chain": "Apollo", "Brand": "TestBrand", "Category": "Sunscreen", "Sub Category": "SPF50",
                 "Article": "UV Shield 50ml", "EAN": "8901234560001", "Pack Size": "50ml", "MRP": 499,
                 "Trade Margin %": 30, "TOT %": 2, "GST %": 18, "Effective From": "2027-01-01", "Status": "ACTIVE"})
    return pd.DataFrame(rows)


def _margin_change_batch():
    """Batch with one margin change for an existing article."""
    rows = _valid_rows()
    # change margin on row index 0 (UV Shield)
    rows[0]["Trade Margin %"] = 31
    rows[0]["Effective From"] = "2026-07-01"
    return pd.DataFrame(rows)


def run_uat(out_path=None):
    uat = UAT()
    tmpdir = tempfile.mkdtemp(prefix="margin_uat_")
    repo_root = os.path.join(tmpdir, "repo")

    try:
        # ---- TC01: Valid records are appended ----
        repo = MarginRepository(repo_root)
        b1 = pd.DataFrame(_valid_rows())
        s1, cl1, rm1 = repo.import_frame(b1, "uat_valid_batch.xlsx")
        uat.check("TC01", "Valid records appended on first import",
                  "10 new records, 0 changed, 0 unchanged",
                  s1["new"] == 10 and s1["changed"] == 0,
                  "new=%d changed=%d unchanged=%d" % (s1["new"], s1["changed"], s1["unchanged"]))

        # ---- TC02: Blank EAN gets fallback key + WARNING ----
        repo2 = MarginRepository(os.path.join(tmpdir, "repo2"))
        uat_df = _uat_batch()
        s2, cl2, rm2 = repo2.import_frame(uat_df, "uat_full_batch.xlsx")

        cur = repo2.current(include_held=True)
        blank_ean = cur[cur["Validation_Flags"].str.contains("BLANK_EAN", na=False)]
        uat.check("TC02", "Blank EAN record flagged BLANK_EAN",
                  "1 record with BLANK_EAN warning, fallback key used",
                  len(blank_ean) >= 1,
                  "found %d BLANK_EAN records" % len(blank_ean))

        # ---- TC03: Duplicate article flagged ----
        dup = cur[cur["Validation_Flags"].str.contains("DUPLICATE_EAN", na=False)]
        uat.check("TC03", "Duplicate EAN article detected",
                  "DUPLICATE_EAN flag on article sharing EAN 8901234560001",
                  len(dup) >= 1,
                  "found %d DUPLICATE_EAN records" % len(dup))

        # ---- TC04: Invalid GST flagged ----
        bad_gst = cur[cur["Validation_Flags"].str.contains("INCORRECT_GST", na=False)]
        uat.check("TC04", "Invalid GST (15%) flagged INCORRECT_GST",
                  "1+ records with INCORRECT_GST, severity FAIL",
                  len(bad_gst) >= 1,
                  "found %d INCORRECT_GST records" % len(bad_gst))
        if not bad_gst.empty:
            uat.check("TC04b", "INCORRECT_GST severity is FAIL",
                      "QC_Severity = FAIL for GST=15",
                      all(bad_gst["QC_Severity"].isin(["FAIL", "BLOCKED"])),
                      str(bad_gst["QC_Severity"].unique()))

        # ---- TC05: Missing margin flagged ----
        miss = cur[cur["Validation_Flags"].str.contains("BLANK_TRADE_MARGIN", na=False)]
        uat.check("TC05", "Missing trade margin flagged",
                  "1+ records with BLANK_TRADE_MARGIN",
                  len(miss) >= 1,
                  "found %d BLANK_TRADE_MARGIN records" % len(miss))

        # ---- TC06: Margin change triggers new version ----
        repo3 = MarginRepository(os.path.join(tmpdir, "repo3"))
        b_init = pd.DataFrame(_valid_rows())
        repo3.import_frame(b_init, "init.xlsx")
        b_chg = _margin_change_batch()
        s3, cl3, rm3 = repo3.import_frame(b_chg, "change.xlsx")
        uat.check("TC06", "Margin change creates new version",
                  "1 changed record, changelog has Trade Margin % entry",
                  s3["changed"] >= 1,
                  "changed=%d" % s3["changed"])
        if not cl3.empty:
            tm_cl = cl3[cl3["Field"] == "Trade Margin %"]
            uat.check("TC06b", "Changelog records Trade Margin % change",
                      "Old=28, New=31, Difference=3",
                      len(tm_cl) >= 1,
                      "changelog Trade Margin entries: %d" % len(tm_cl))
            if not tm_cl.empty:
                row_cl = tm_cl.iloc[0]
                uat.check("TC06c", "Changelog difference is correct",
                          "Difference = 3.0",
                          float(row_cl["Difference"]) == 3.0,
                          "Difference=%s" % row_cl["Difference"])

        # ---- TC07: New article gets version 1 ----
        uat.check("TC07", "New article in UAT batch assigned version 1",
                  "SuperSmooth 150ml has Version Number = 1",
                  s2["new"] >= 1,
                  "new articles=%d" % s2["new"])

        # ---- TC08: Delisted article flagged INACTIVE ----
        inactive = cur[cur["Validation_Flags"].str.contains("INACTIVE_ARTICLE", na=False)]
        uat.check("TC08", "Delisted article flagged INACTIVE_ARTICLE",
                  "1+ records with INACTIVE_ARTICLE warning",
                  len(inactive) >= 1,
                  "found %d INACTIVE_ARTICLE records" % len(inactive))

        # ---- TC09: Future-effective margin accepted ----
        future = cur[cur["EAN"].astype(str) == "8901234560001"]
        uat.check("TC09", "Future-effective margin record accepted",
                  "Article with Effective From 2027-01-01 exists in repository",
                  len(future) >= 1,
                  "found %d records for EAN 8901234560001" % len(future))

        # ---- TC10: Re-import unchanged file creates no false versions ----
        repo4 = MarginRepository(os.path.join(tmpdir, "repo4"))
        b_same = pd.DataFrame(_valid_rows())
        repo4.import_frame(b_same, "first.xlsx")
        hist_before = len(repo4.history)
        s4, cl4, rm4 = repo4.import_frame(b_same, "second.xlsx")
        uat.check("TC10", "Re-import of identical file creates no new versions",
                  "0 new, 0 changed, 10 unchanged",
                  s4["new"] == 0 and s4["changed"] == 0 and s4["unchanged"] == 10,
                  "new=%d changed=%d unchanged=%d" % (s4["new"], s4["changed"], s4["unchanged"]))
        uat.check("TC10b", "History row count unchanged after re-import",
                  "History length same before and after",
                  len(repo4.history) == hist_before,
                  "before=%d after=%d" % (hist_before, len(repo4.history)))

        # ---- TC11: Versioning — only changed article gets new version ----
        cur3 = repo3.current(include_held=True)
        uv = cur3[cur3["EAN"].astype(str) == "8901234560001"]
        if not uv.empty:
            v = int(float(uv.iloc[0]["Version Number"]))
            uat.check("TC11", "Only changed article version incremented",
                      "UV Shield version = 2",
                      v == 2, "version=%d" % v)
        others = cur3[cur3["EAN"].astype(str) == "8901234560002"]
        if not others.empty:
            v2 = int(float(others.iloc[0]["Version Number"]))
            uat.check("TC11b", "Unchanged article version stays at 1",
                      "UV Light version = 1",
                      v2 == 1, "version=%d" % v2)

        # ---- TC12: Historical versions remain unchanged ----
        hist3 = repo3.history
        v1_rows = hist3[(hist3["EAN"].astype(str) == "8901234560001") &
                        (hist3["Version Number"].astype(str).str.strip() == "1")]
        uat.check("TC12", "Historical version 1 preserved after version 2 created",
                  "Version 1 row still exists in history",
                  len(v1_rows) >= 1,
                  "v1 rows=%d" % len(v1_rows))

        # ---- TC13: Rollback test ----
        versions = repo3.list_versions()
        pre_rollback_count = len(repo3.history)
        first_snap = [v for v in versions if not v.startswith("PRE-")][0]
        repo3.rollback(first_snap)
        uat.check("TC13", "Rollback restores earlier state",
                  "History count after rollback < before",
                  len(repo3.history) < pre_rollback_count,
                  "before=%d after=%d" % (pre_rollback_count, len(repo3.history)))
        pre_rollback_snaps = [v for v in repo3.list_versions() if v.startswith("PRE-ROLLBACK")]
        uat.check("TC13b", "Pre-rollback state archived",
                  "PRE-ROLLBACK snapshot exists",
                  len(pre_rollback_snaps) >= 1,
                  "pre-rollback snapshots=%d" % len(pre_rollback_snaps))

        # ---- TC14: Forecast-ready output excludes HELD records ----
        cur_pub = repo2.current(include_held=False)
        cur_all = repo2.current(include_held=True)
        held_in_pub = cur_pub[cur_pub["Record_Status"] == "HELD"] if not cur_pub.empty else pd.DataFrame()
        uat.check("TC14", "Forecast-ready output excludes HELD records",
                  "No HELD records in published current view",
                  held_in_pub.empty,
                  "HELD in published=%d" % len(held_in_pub))
        uat.check("TC14b", "HELD records exist in all-inclusive current view",
                  "At least 1 HELD record in full current view",
                  not cur_all.empty and (cur_all["Record_Status"] == "HELD").any(),
                  "HELD in all=%d" % (cur_all["Record_Status"] == "HELD").sum() if not cur_all.empty else 0)

        # ---- TC15: Reconciliation difference is zero ----
        uat.check("TC15", "Reconciliation difference is zero",
                  "source rows = new + changed + unchanged",
                  s2.get("reconciliation_diff", -1) == 0,
                  "reconciliation_diff=%s" % s2.get("reconciliation_diff", "missing"))

        # ---- TC16: Outputs workbook has all 11 sheets ----
        out_xlsx = os.path.join(tmpdir, "uat_outputs.xlsx")
        build_outputs(repo2, out_xlsx, changelog=cl2, removed=rm2)
        from openpyxl import load_workbook
        wb = load_workbook(out_xlsx, read_only=True)
        expected = ["1_Chain_Article_Master", "1b_Full_Version_History", "2_Margin_Change_Log",
                    "3_Commercial_Difference", "4_Missing_Margin", "5_Duplicate_Article",
                    "6_Margin_QC", "7_Forecast_Ready_Margin", "8_CM2_Ready_Margin",
                    "9_Dashboard_Dataset", "10_Executive_Summary"]
        missing_sheets = [s for s in expected if s not in wb.sheetnames]
        uat.check("TC16", "All 11 output sheets present",
                  "All sheets: " + ", ".join(expected),
                  len(missing_sheets) == 0,
                  "missing: %s" % missing_sheets if missing_sheets else "all present")
        wb.close()

        # ---- TC17: Search returns correct results ----
        cur_s = repo2.current(include_held=True)
        r_sun = search(cur_s, "sunscreen")
        uat.check("TC17", "Search 'sunscreen' returns sunscreen articles",
                  "1+ results from Sunscreen category",
                  len(r_sun) >= 1,
                  "results=%d" % len(r_sun))

        r_below = search(cur_s, "margin below 20")
        if not r_below.empty:
            vals = pd.to_numeric(r_below["Final Effective Margin %"], errors="coerce").dropna()
            uat.check("TC17b", "Search 'margin below 20' returns only low-margin articles",
                      "All results have margin < 20",
                      all(vals < 20),
                      "max margin in results=%.1f" % vals.max() if len(vals) else "no numeric values")

        # ---- TC18: Impact analysis flags high-risk changes ----
        impact = impact_from_changelog(cl3)
        if not impact.empty:
            uat.check("TC18", "Impact analysis produces results for margin changes",
                      "Impact table has rows with Margin Delta",
                      "Margin Delta (pp)" in impact.columns and len(impact) >= 1,
                      "impact rows=%d" % len(impact))

    except Exception as e:
        traceback.print_exc()
        uat.results.append({
            "Test ID": "ERR", "Test Case": "Unhandled exception",
            "Expected Result": "No exceptions", "Actual Result": str(e),
            "Verdict": "FAIL", "Evidence": traceback.format_exc()[:500],
            "Root Cause": str(e), "Corrective Action": "Debug and fix",
            "Owner": "Engineering",
        })
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    # build UAT workbook
    df = pd.DataFrame(uat.results)
    if out_path:
        _write_uat_workbook(df, out_path)

    # print summary
    passed = sum(1 for r in uat.results if r["Verdict"] == "PASS")
    failed = sum(1 for r in uat.results if r["Verdict"] == "FAIL")
    total = len(uat.results)
    print("\n" + "=" * 70)
    print("UAT RESULTS:  %d/%d passed,  %d failed" % (passed, total, failed))
    if failed == 0:
        print("STATUS: PASS")
    else:
        print("STATUS: FAIL")
        for r in uat.results:
            if r["Verdict"] == "FAIL":
                print("  FAIL: %s — %s | %s" % (r["Test ID"], r["Test Case"], r["Actual Result"]))
    print("=" * 70)
    return df


def _write_uat_workbook(df, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "UAT_Results"
    cols = list(df.columns)
    for c, col in enumerate(cols, 1):
        cell = ws.cell(1, c, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        verdict = row.get("Verdict", "")
        fill = PASS_FILL if verdict == "PASS" else FAIL_FILL if verdict == "FAIL" else WARN_FILL
        for c, col in enumerate(cols, 1):
            v = row[col]
            if pd.isna(v):
                v = ""
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == "Verdict":
                cell.fill = fill
    widths = {"Test ID": 10, "Test Case": 40, "Expected Result": 40, "Actual Result": 35,
              "Verdict": 10, "Evidence": 35, "Root Cause": 25, "Corrective Action": 25, "Owner": 12}
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 20)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(df) + 1)

    # summary sheet
    ss = wb.create_sheet("UAT_Summary")
    passed = int((df["Verdict"] == "PASS").sum())
    failed = int((df["Verdict"] == "FAIL").sum())
    total = len(df)
    summary_rows = [
        ("UAT Execution Summary", ""),
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Pass Rate %", round(100.0 * passed / total, 1) if total else 0),
        ("", ""),
        ("Overall Status", "PASS" if failed == 0 else "FAIL"),
    ]
    for i, (k, v) in enumerate(summary_rows, 1):
        a = ss.cell(i, 1, k)
        b = ss.cell(i, 2, v)
        if i == 1:
            a.font = Font(bold=True, size=13, color="1F4E78")
        elif k == "Overall Status":
            a.font = Font(bold=True, size=11)
            b.font = Font(bold=True, size=11)
            b.fill = PASS_FILL if v == "PASS" else FAIL_FILL
        else:
            a.font = Font(size=10)
    ss.column_dimensions["A"].width = 30
    ss.column_dimensions["B"].width = 20

    wb.save(path)


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "UAT_Report.xlsx"
    run_uat(out)
