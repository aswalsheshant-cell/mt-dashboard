# -*- coding: utf-8 -*-
"""Real-data pilot runner and validation framework.

Creates an isolated pilot environment, imports a real chain margin file,
produces a full pilot report package including profiling, reconciliation,
exception report, and validation sample workbook.
"""
import os
import sys
import json
import shutil

sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from schema import REPO_COLS
from config import load_config, save_config, DEFAULT_CONFIG
from repository import MarginRepository, file_checksum, coerce_numeric
from ingest import normalize_file
from impact import impact_from_changelog
from validation import qc_report
from outputs import build_outputs
from release import (build_reconciliation_report, build_release_checklist,
                     build_business_rule_register, build_known_limitations,
                     HEADER_FILL, HEADER_FONT, BORDER, PASS_FILL, FAIL_FILL,
                     WARN_FILL, PENDING_FILL, _styled_header)

PILOT_DIRS = [
    "input/pending", "input/processed", "input/rejected",
    "repository/test", "repository/production", "repository/archive",
    "output/snapshots", "output/exceptions", "output/impact_reports",
    "config", "logs", "manifests",
]


def create_pilot_environment(base_path):
    """Create the isolated pilot folder structure."""
    for d in PILOT_DIRS:
        os.makedirs(os.path.join(base_path, d), exist_ok=True)
    cfg_path = os.path.join(base_path, "config", "validation_config.json")
    if not os.path.exists(cfg_path):
        save_config(DEFAULT_CONFIG, cfg_path)
    return base_path


def profile_source_file(path, sheet=0):
    """Pre-import profiling report for a source file."""
    from ingest import _read_raw, _detect_header_row, SOURCE_COLS
    from schema import canon_header

    ext = os.path.splitext(path)[1].lower()
    raw = _read_raw(path, sheet)
    hrow, score = _detect_header_row(raw)
    headers_raw = raw.iloc[hrow].tolist()
    headers_canon = [canon_header(v) for v in headers_raw]
    mapped = [h for h in headers_canon if h in SOURCE_COLS]
    unmapped = [h for h in headers_canon if h not in SOURCE_COLS and h is not None]

    # detect duplicate headers
    seen = {}
    dup_headers = []
    for h in headers_canon:
        if h is not None:
            seen[h] = seen.get(h, 0) + 1
    dup_headers = [h for h, c in seen.items() if c > 1]

    data_rows = len(raw) - hrow - 1
    profile = {
        "file_path": path,
        "file_type": ext,
        "total_raw_rows": len(raw),
        "header_row": hrow,
        "header_match_score": score,
        "data_rows": data_rows,
        "total_columns": len(headers_raw),
        "mapped_columns": len(mapped),
        "unmapped_columns": len(unmapped),
        "mapped_list": mapped,
        "unmapped_list": unmapped,
        "duplicate_headers": dup_headers,
        "raw_headers": [str(h) for h in headers_raw],
        "canonical_headers": [str(h) for h in headers_canon],
    }
    return profile


def run_pilot(source_path, pilot_base, sheet=0, config_path=None):
    """Run a complete pilot import and produce all reports.

    Returns (summary, report_paths) where report_paths is a dict of
    artifact name -> file path.
    """
    pilot_base = create_pilot_environment(pilot_base)
    cfg = load_config(config_path or os.path.join(pilot_base, "config", "validation_config.json"))

    reports = {}

    # Step 1: Profile the source file
    print("=" * 70)
    print("PILOT IMPORT — PRE-IMPORT PROFILING")
    print("=" * 70)
    profile = profile_source_file(source_path, sheet)
    print("  File:            %s" % profile["file_path"])
    print("  Type:            %s" % profile["file_type"])
    print("  Data rows:       %d" % profile["data_rows"])
    print("  Mapped columns:  %d / %d" % (profile["mapped_columns"], profile["total_columns"]))
    print("  Unmapped:        %s" % (", ".join(profile["unmapped_list"]) or "(none)"))
    print("  Duplicate hdrs:  %s" % (", ".join(profile["duplicate_headers"]) or "(none)"))

    profile_path = os.path.join(pilot_base, "output", "Pilot_Profile_Report.json")
    with open(profile_path, "w") as f:
        json.dump(profile, f, indent=2)
    reports["profile"] = profile_path

    # Step 2: Check for duplicate import
    repo_root = os.path.join(pilot_base, "repository", "test")
    repo = MarginRepository(repo_root)
    prev, cs = repo.check_duplicate_import(source_path)
    if prev:
        print("\n  WARNING: This file was previously imported as batch '%s' on %s" %
              (prev["batch"], prev["ts"]))
        print("  Proceeding anyway for pilot validation.\n")

    # Step 3: Normalize and import
    print("\n" + "=" * 70)
    print("PILOT IMPORT — INGESTION")
    print("=" * 70)
    df, meta = normalize_file(source_path, sheet)
    print("  Normalized %d rows (header row %d, %d columns mapped)" %
          (meta["rows"], meta["header_row"], meta["mapped"]))

    source_basename = os.path.basename(source_path)
    summary, changelog, removed = repo.import_frame(
        df, source_file=source_basename, source_path=source_path,
        submitted_by="pilot_runner", business_owner="PILOT_TEST")

    # Step 4: Print reconciliation
    print("\n" + "=" * 70)
    print("PILOT IMPORT — RECONCILIATION")
    print("=" * 70)
    for k in ["rows_in_file", "new", "changed", "unchanged", "pass", "warning",
              "fail", "blocked", "published_forecast_ready", "held",
              "fallback_article_keys", "removed_from_file", "changelog_entries",
              "reconciliation_diff"]:
        print("  %-35s %s" % (k, summary.get(k, "N/A")))

    recon_ok = summary["reconciliation_diff"] == 0
    print("\n  RECONCILIATION: %s" % ("PASS" if recon_ok else "FAIL — DIFFERENCE = %d" % summary["reconciliation_diff"]))

    # Step 5: Generate reports
    print("\n" + "=" * 70)
    print("PILOT — GENERATING REPORTS")
    print("=" * 70)

    # Reconciliation report
    recon_path = os.path.join(pilot_base, "output", "Pilot_Reconciliation.xlsx")
    build_reconciliation_report(summary, recon_path)
    reports["reconciliation"] = recon_path
    print("  Reconciliation report: %s" % recon_path)

    # Full outputs workbook
    impact = impact_from_changelog(changelog, cfg=cfg)
    out_path = os.path.join(pilot_base, "output", "Pilot_Outputs.xlsx")
    build_outputs(repo, out_path, changelog=changelog, removed=removed, impact=impact)
    reports["outputs"] = out_path
    print("  Outputs workbook:     %s" % out_path)

    # Exception report
    exc_path = os.path.join(pilot_base, "output", "exceptions", "Pilot_Exceptions.xlsx")
    _build_exception_report(repo, exc_path)
    reports["exceptions"] = exc_path
    print("  Exception report:     %s" % exc_path)

    # Validation sample
    sample_path = os.path.join(pilot_base, "output", "Pilot_Validation_Sample.xlsx")
    _build_validation_sample(repo, summary, sample_path)
    reports["validation_sample"] = sample_path
    print("  Validation sample:    %s" % sample_path)

    # Move source to processed
    processed_path = os.path.join(pilot_base, "input", "processed", source_basename)
    shutil.copy2(source_path, processed_path)
    print("  Source copied to:     %s" % processed_path)

    # Final verdict
    print("\n" + "=" * 70)
    blocked_count = summary.get("blocked", 0)
    fail_count = summary.get("fail", 0)
    if not recon_ok:
        verdict = "FAIL"
        reason = "Reconciliation difference is not zero"
    elif blocked_count > 0 or fail_count > 0:
        verdict = "PASS WITH WARNINGS"
        reason = "%d BLOCKED, %d FAIL records detected — review exceptions" % (blocked_count, fail_count)
    else:
        verdict = "PASS"
        reason = "All checks passed"
    print("PILOT VERDICT: %s" % verdict)
    print("  %s" % reason)
    print("=" * 70)

    summary["pilot_verdict"] = verdict
    summary["pilot_reason"] = reason
    reports["summary"] = summary

    return summary, reports


def _build_exception_report(repo, path):
    """Produce an exception workbook with all non-PASS records."""
    cur = repo.current(include_held=True)
    if cur.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Exceptions"
        ws.cell(1, 1, "No records in repository")
        wb.save(path)
        return

    wb = Workbook()
    wb.remove(wb.active)

    # sheet 1: all FAIL/BLOCKED records
    held = cur[cur["Record_Status"] == "HELD"]
    _write_exception_sheet(wb, "FAIL_BLOCKED", held,
                           ["Chain", "Brand", "Article", "EAN", "Pack Size", "MRP",
                            "Trade Margin %", "Final Effective Margin %", "GST %",
                            "QC_Severity", "Validation_Flags", "Record_Status"])

    # sheet 2: WARNING records
    warn = cur[cur["QC_Severity"] == "WARNING"]
    _write_exception_sheet(wb, "WARNINGS", warn,
                           ["Chain", "Brand", "Article", "EAN", "Pack Size",
                            "Trade Margin %", "Final Effective Margin %",
                            "QC_Severity", "Validation_Flags"])

    # sheet 3: duplicates
    dups = cur[cur["Validation_Flags"].str.contains("DUPLICATE", na=False)]
    _write_exception_sheet(wb, "DUPLICATES", dups,
                           ["Chain", "Brand", "Article", "EAN", "Pack Size", "MRP",
                            "Effective From", "Version Number", "Validation_Flags"])

    # sheet 4: fallback keys
    fallback = cur[cur["Article_Key"].str.startswith("ALT|", na=False)]
    _write_exception_sheet(wb, "FALLBACK_KEYS", fallback,
                           ["Chain", "Brand", "Article", "EAN", "Pack Size", "MRP",
                            "Article_Key", "Validation_Flags"])

    # sheet 5: summary
    ss = wb.create_sheet("SUMMARY")
    rows = [
        ("Exception Summary", ""),
        ("", ""),
        ("Total Records", len(cur)),
        ("PASS", int((cur["QC_Severity"] == "PASS").sum())),
        ("WARNING", int((cur["QC_Severity"] == "WARNING").sum())),
        ("FAIL", int((cur["QC_Severity"] == "FAIL").sum())),
        ("BLOCKED", int((cur["QC_Severity"] == "BLOCKED").sum())),
        ("Duplicate records", len(dups)),
        ("Fallback-key records", len(fallback)),
    ]
    for i, (k, v) in enumerate(rows, 1):
        a = ss.cell(i, 1, k)
        ss.cell(i, 2, v)
        if i == 1:
            a.font = Font(bold=True, size=13, color="1F4E78")
        else:
            a.font = Font(size=10)
    ss.column_dimensions["A"].width = 30
    ss.column_dimensions["B"].width = 15

    wb.save(path)


def _write_exception_sheet(wb, name, df, cols):
    ws = wb.create_sheet(name)
    use_cols = [c for c in cols if c in df.columns]
    if df.empty:
        ws.cell(1, 1, "(no records)")
        return
    _styled_header(ws, use_cols)
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        sev = str(row.get("QC_Severity", ""))
        fill = None
        if sev == "BLOCKED":
            fill = PatternFill("solid", fgColor="F4B7B7")
        elif sev == "FAIL":
            fill = PatternFill("solid", fgColor="FCD5A5")
        elif sev == "WARNING":
            fill = PatternFill("solid", fgColor="FFF2A8")
        for c, col in enumerate(use_cols, 1):
            v = row.get(col, "")
            if pd.isna(v):
                v = ""
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(col == "Validation_Flags"), vertical="top")
            if fill:
                cell.fill = fill
    for c, col in enumerate(use_cols, 1):
        w = 35 if col == "Validation_Flags" else max(12, min(22, len(col) + 3))
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def _build_validation_sample(repo, summary, path):
    """Build a structured validation sample workbook for manual review."""
    cur = repo.current(include_held=True)
    hist = repo.history
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation_Sample"

    cols = ["Sample Category", "Chain", "Brand", "Article", "EAN", "Pack Size", "MRP",
            "Trade Margin %", "Final Effective Margin %", "GST %", "Version Number",
            "QC_Severity", "Validation_Flags", "Change_Type",
            "Source Value", "Repository Value", "Expected Result", "Actual Result",
            "Status", "Reviewer Remarks"]
    _styled_header(ws, cols)

    samples = []

    if not cur.empty:
        # unchanged articles (up to 10)
        unchanged = hist[hist["Change_Type"] == "NEW"].head(10) if not hist.empty else pd.DataFrame()
        for _, r in unchanged.iterrows():
            samples.append(_sample_row("Unchanged Article", r, "Record preserved as imported"))

        # changed articles (up to 5)
        changed = hist[hist["Change_Type"] == "CHANGED"].head(5) if not hist.empty else pd.DataFrame()
        for _, r in changed.iterrows():
            samples.append(_sample_row("Changed Article", r, "New version created"))

        # new articles (up to 5)
        new = hist[hist["Change_Type"] == "NEW"].tail(5) if not hist.empty else pd.DataFrame()
        for _, r in new.iterrows():
            samples.append(_sample_row("New Article", r, "Ingested as version 1"))

        # all duplicates
        dups = cur[cur["Validation_Flags"].str.contains("DUPLICATE", na=False)]
        for _, r in dups.iterrows():
            samples.append(_sample_row("Duplicate Record", r, "Flagged for review"))

        # all FAIL records
        fails = cur[cur["QC_Severity"] == "FAIL"]
        for _, r in fails.iterrows():
            samples.append(_sample_row("FAIL Record", r, "Held from forecast"))

        # all BLOCKED records
        blocked = cur[cur["QC_Severity"] == "BLOCKED"]
        for _, r in blocked.iterrows():
            samples.append(_sample_row("BLOCKED Record", r, "Blocked from publish"))

        # version history examples (up to 3 articles with >1 version)
        if not hist.empty:
            vn = pd.to_numeric(hist["Version Number"], errors="coerce")
            multi = hist[vn > 1]["Article_Key"].unique()[:3]
            for ak in multi:
                versions = hist[hist["Article_Key"] == ak].sort_values("Version Number")
                for _, r in versions.iterrows():
                    samples.append(_sample_row("Version History", r,
                                               "Version %s preserved" % r.get("Version Number", "")))

    for ri, row in enumerate(samples, 2):
        for c, col in enumerate(cols, 1):
            v = row.get(col, "")
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(col in ("Validation_Flags", "Reviewer Remarks")),
                                       vertical="top")

    widths = {"Sample Category": 18, "Chain": 14, "Brand": 14, "Article": 22, "EAN": 16,
              "Pack Size": 10, "MRP": 8, "Trade Margin %": 14, "Final Effective Margin %": 18,
              "GST %": 8, "Version Number": 12, "QC_Severity": 12, "Validation_Flags": 35,
              "Change_Type": 12, "Source Value": 14, "Repository Value": 14,
              "Expected Result": 20, "Actual Result": 20, "Status": 10, "Reviewer Remarks": 25}
    for c, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(col, 14)
    ws.freeze_panes = "A2"

    wb.save(path)


def _sample_row(category, r, expected):
    return {
        "Sample Category": category,
        "Chain": r.get("Chain", ""),
        "Brand": r.get("Brand", ""),
        "Article": r.get("Article", ""),
        "EAN": r.get("EAN", ""),
        "Pack Size": r.get("Pack Size", ""),
        "MRP": r.get("MRP", ""),
        "Trade Margin %": r.get("Trade Margin %", ""),
        "Final Effective Margin %": r.get("Final Effective Margin %", ""),
        "GST %": r.get("GST %", ""),
        "Version Number": r.get("Version Number", ""),
        "QC_Severity": r.get("QC_Severity", ""),
        "Validation_Flags": r.get("Validation_Flags", ""),
        "Change_Type": r.get("Change_Type", ""),
        "Source Value": "",
        "Repository Value": "",
        "Expected Result": expected,
        "Actual Result": "",
        "Status": "",
        "Reviewer Remarks": "",
    }


def build_approval_authority_matrix(path):
    """Generate Approval_Authority_Matrix.xlsx — blank for business to fill."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Approval_Matrix"

    cols = ["Role", "Scope", "Authority Level", "Approval Limit",
            "Escalation Path", "Backup Approver", "Name", "Department", "Remarks"]
    _styled_header(ws, cols)

    roles = [
        ("Submitter (Maker)", "Submit new or changed margin records",
         "Submit only — cannot approve own submissions",
         "All articles within assigned chain(s)", "Reviewer", "", "", "KAM / Channel Analyst", ""),
        ("Reviewer (Checker)", "Review and validate margin submissions",
         "Verify data accuracy, flag issues, recommend approval or rejection",
         "All submissions within assigned chains/categories", "Approver", "", "", "Commercial Excellence / Sales Ops", ""),
        ("Approver", "Approve margin records for production use",
         "Final approval for standard margin changes",
         "Margin changes up to ±3 pp", "Finance Approver", "", "", "Sales Head / Commercial Owner", ""),
        ("Finance Approver", "Approve margin/GST-sensitive changes",
         "Required for high-risk or GST changes",
         "All margin changes, GST changes, blocked records",
         "Commercial Head", "", "", "Finance Head / Tax", ""),
        ("Emergency Approver", "Approve urgent margin changes outside normal process",
         "Bypass normal review cycle for time-critical changes",
         "All — must be documented and ratified within 48 hours",
         "Commercial Head + Finance Head", "", "", "Sales Head or delegated authority", ""),
        ("Retrospective Date Approver", "Approve backdated effective dates",
         "Required when Effective From is in the past",
         "All backdated changes", "Finance Approver + Commercial Head", "", "",
         "Sales Head + Finance Head (joint approval)", ""),
        ("High-Risk Escalation", "Margin changes exceeding ±5 pp or below floor",
         "Dual approval required — commercial + finance",
         "Changes blocked by risk threshold configuration",
         "Executive Sponsor", "", "", "Commercial Head + Finance Head", ""),
    ]

    for ri, row_data in enumerate(roles, 2):
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 7 and v == "":
                cell.fill = PENDING_FILL

    widths = [22, 35, 35, 30, 25, 18, 18, 25, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # dual approval thresholds sheet
    ts = wb.create_sheet("Risk_Thresholds")
    ts.cell(1, 1, "Margin Risk Thresholds").font = Font(bold=True, size=13, color="1F4E78")
    ts.cell(2, 1, "These thresholds are configurable in config/validation_config.json").font = Font(size=9, italic=True)

    th_cols = ["Margin Change Range", "Risk Tier", "Required Approval", "Status"]
    _styled_header(ts, th_cols, row=4)
    thresholds = [
        ("Up to ±1.0 pp", "NORMAL", "Standard review (Reviewer)", "PENDING_APPROVAL"),
        ("±1.0 to ±3.0 pp", "WARNING", "Checker + Approver", "PENDING_APPROVAL"),
        ("±3.0 to ±5.0 pp", "HIGH_RISK", "Commercial Approval required", "PENDING_APPROVAL"),
        ("Above ±5.0 pp", "BLOCKED", "Finance + Commercial dual approval", "PENDING_APPROVAL"),
        ("Margin below 0% or above ceiling", "BLOCKED", "Blocked until reviewed", "APPROVED_BY_DEFAULT"),
    ]
    for ri, row_data in enumerate(thresholds, 5):
        for c, v in enumerate(row_data, 1):
            cell = ts.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 4:
                if "APPROVED" in v:
                    cell.fill = PASS_FILL
                else:
                    cell.fill = PENDING_FILL
    for c, w in enumerate([25, 14, 35, 20], 1):
        ts.column_dimensions[get_column_letter(c)].width = w

    # GST controls sheet
    gs = wb.create_sheet("GST_Controls")
    gs.cell(1, 1, "GST Validation Controls").font = Font(bold=True, size=13, color="1F4E78")
    gs.cell(2, 1, "Configurable in config/validation_config.json").font = Font(size=9, italic=True)

    g_cols = ["GST Scenario", "Configured Severity", "Proposed Behavior", "Status"]
    _styled_header(gs, g_cols, row=4)
    gst_rules = [
        ("Blank GST (not provided)", "WARNING",
         "Warn if GST is not required for calculation; investigate if calculation depends on it",
         "PENDING_APPROVAL"),
        ("Invalid GST value (not in {0,5,12,18,28})", "FAIL",
         "Hold record from forecast until corrected", "PENDING_APPROVAL"),
        ("GST inconsistent with approved article master", "BLOCKED",
         "Block until reconciled with master data", "PENDING_APPROVAL"),
        ("GST changed from previous approved version", "WARNING",
         "Flag for Finance/Tax review before approval", "PENDING_APPROVAL"),
        ("Unsupported GST rate (negative or >28)", "BLOCKED",
         "Block — clearly invalid", "APPROVED_BY_DEFAULT"),
    ]
    for ri, row_data in enumerate(gst_rules, 5):
        for c, v in enumerate(row_data, 1):
            cell = gs.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 4:
                if "APPROVED" in v:
                    cell.fill = PASS_FILL
                else:
                    cell.fill = PENDING_FILL
    for c, w in enumerate([40, 18, 45, 20], 1):
        gs.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)
    return path


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Margin Repository Pilot Runner")
    sub = p.add_subparsers(dest="cmd")

    run = sub.add_parser("run", help="Run pilot import")
    run.add_argument("source", help="Path to source margin file")
    run.add_argument("--pilot-dir", default="./pilot_environment")
    run.add_argument("--sheet", default=0, type=int)
    run.add_argument("--config", default=None)

    env = sub.add_parser("setup", help="Create pilot environment only")
    env.add_argument("--pilot-dir", default="./pilot_environment")

    matrix = sub.add_parser("approval-matrix", help="Generate approval authority matrix")
    matrix.add_argument("--out", default="Approval_Authority_Matrix.xlsx")

    args = p.parse_args()
    if args.cmd == "run":
        run_pilot(args.source, args.pilot_dir, args.sheet, args.config)
    elif args.cmd == "setup":
        create_pilot_environment(args.pilot_dir)
        print("Pilot environment created at: %s" % args.pilot_dir)
    elif args.cmd == "approval-matrix":
        build_approval_authority_matrix(args.out)
        print("Approval authority matrix: %s" % args.out)
    else:
        p.print_help()
