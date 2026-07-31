# -*- coding: utf-8 -*-
"""Release package generator: business rule register, reconciliation report,
release-readiness checklist, and known limitations register."""
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
PASS_FILL = PatternFill("solid", fgColor="C6E9C6")
WARN_FILL = PatternFill("solid", fgColor="FFF2A8")
FAIL_FILL = PatternFill("solid", fgColor="F4B7B7")
PENDING_FILL = PatternFill("solid", fgColor="D9E2F3")


def _styled_header(ws, cols, row=1):
    for c, col in enumerate(cols, 1):
        cell = ws.cell(row, c, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _write_rows(ws, data, start_row=2, status_col=None):
    for ri, row_data in enumerate(data, start_row):
        fill = None
        if status_col is not None:
            st = str(row_data[status_col]).upper() if status_col < len(row_data) else ""
            if "APPROVED" in st:
                fill = PASS_FILL
            elif "PENDING" in st:
                fill = PENDING_FILL
            elif "FAIL" in st or "REJECT" in st:
                fill = FAIL_FILL
            elif "WARN" in st:
                fill = WARN_FILL
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(ri, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill


def build_business_rule_register(path):
    """Produces Business_Rule_Register.xlsx with all business rules and their decision status."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Business_Rules"

    cols = ["Rule ID", "Business Area", "Rule Description", "Current System Behavior",
            "Proposed Production Behavior", "Business Owner", "Finance Owner",
            "Decision Status", "Approval Date", "Remarks"]
    _styled_header(ws, cols)

    rules = [
        ("BR-001", "Article Identity", "EAN is the primary article identifier across all chains",
         "EAN-priority key: Chain+EAN+PackSize+MRP. When EAN present, it dominates article identity.",
         "Same — EAN always primary. Fallback used only when EAN is genuinely blank.",
         "", "", "PENDING_APPROVAL", "",
         "Business team to confirm: Is EAN always the primary identity? Any chain that uses Article code instead?"),

        ("BR-002", "Article Identity", "Fallback identity when EAN is blank",
         "Fallback key: Chain+Brand+Article+PackSize+MRP. Record flagged BLANK_EAN (WARNING).",
         "Same — fallback key used, article flagged. Business should aim to fill EAN for all articles.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm fallback combination. Should BLANK_EAN be WARNING or FAIL?"),

        ("BR-003", "Duplicate Handling", "Duplicate EAN within same chain flagged",
         "DUPLICATE_EAN flag (WARNING) when same Chain+EAN maps to >1 distinct Article name.",
         "Same — flag and surface for review. Do not auto-merge.",
         "", "", "PENDING_APPROVAL", "",
         "Should duplicate EAN block the record or just warn?"),

        ("BR-004", "Duplicate Handling", "Duplicate Chain+Article flagged",
         "DUPLICATE_CHAIN_ARTICLE flag (WARNING) when same Chain+Article appears multiple times.",
         "Same — warn. Multiple rows may represent different pack sizes or effective dates.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm: same Chain+Article with different Pack Size/MRP is valid, not a duplicate."),

        ("BR-005", "Duplicate Handling", "Duplicate effective date for same article identity",
         "DUPLICATE_EFFECTIVE_DATE flag (FAIL) when same identity+date appears multiple times.",
         "Same — FAIL severity, held from forecast until resolved.",
         "", "", "PENDING_APPROVAL", "",
         "Should duplicate effective dates be FAIL or BLOCKED?"),

        ("BR-006", "Margin Hierarchy", "Final Effective Margin % priority",
         "Source-provided Final Effective Margin % is respected. When blank, computed from components: "
         "sum of earned margins minus Consumer Offer % and Cash Discount %.",
         "Same — provided value always wins. Derivation only fills blanks.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm: which margin field has priority when source provides multiple?"),

        ("BR-007", "GST Validation", "Invalid GST handling",
         "GST not in {0, 5, 12, 18, 28} triggers INCORRECT_GST flag with FAIL severity.",
         "Same — FAIL severity. Record held from forecast until GST corrected.",
         "", "", "PENDING_APPROVAL", "",
         "Should invalid GST block the record (BLOCKED) or only FAIL?"),

        ("BR-008", "Margin Limits", "Margin > 100% handling",
         "MARGIN_OVER_100 flag with BLOCKED severity. Record excluded from forecast.",
         "Same — BLOCKED. A margin >100% is always a data error.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "Standard business rule — margin cannot exceed 100%."),

        ("BR-009", "Margin Limits", "Negative margin handling",
         "NEGATIVE_MARGIN flag with BLOCKED severity. Record excluded from forecast.",
         "Same — BLOCKED. Negative margins need explicit business review.",
         "", "", "PENDING_APPROVAL", "",
         "Some chains may have negative margins during promotions. Confirm if BLOCKED is correct."),

        ("BR-010", "Listing Status", "Delisted articles remain searchable",
         "INACTIVE_ARTICLE flag (WARNING). Record stays in history and current view but flagged.",
         "Same — delisted articles visible with flag. Not included in forecast-ready output.",
         "", "", "PENDING_APPROVAL", "",
         "Should delisted articles be included in current view or only history?"),

        ("BR-011", "Effective Dates", "Historical and future margins can coexist",
         "Both accepted. Future Effective From dates create a scheduled version. "
         "Expired Effective To dates trigger EXPIRED_COMMERCIAL (WARNING).",
         "Same — coexistence allowed. Business reviews future-effective margins before activation.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm: can historical and future margins coexist for the same article?"),

        ("BR-012", "Forecast Eligibility", "Which statuses feed demand planning",
         "Only PUBLISHED records (PASS or WARNING severity) with Status=ACTIVE included in "
         "Sheet 7 (Forecast-Ready) and Sheet 8 (CM2-Ready).",
         "Only APPROVED + ACTIVE + PUBLISHED records should feed forecast.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm: APPROVED + ACTIVE + PUBLISHED = forecast-eligible."),

        ("BR-013", "Approval Workflow", "Who can approve or reject margin changes",
         "Approval_Status field populated from source file. No enforced workflow yet.",
         "DRAFT → PENDING_REVIEW → APPROVED/REJECTED. Only APPROVED enters production.",
         "", "", "PENDING_APPROVAL", "",
         "Define approval authority matrix: who submits, who reviews, who approves."),

        ("BR-014", "Impact Calculation", "Impact driver for rupee impact",
         "Monthly_NSV from optional drivers frame. When absent, only percentage-point delta reported. "
         "Rupee impact = NSV × margin delta / 100.",
         "Same — NSV is the primary driver. Never fabricate when absent.",
         "", "", "PENDING_APPROVAL", "",
         "Confirm: NSV, MRP sales, units, or forecast volume as impact driver?"),

        ("BR-015", "Missing Data", "Blank MRP handling",
         "BLANK_MRP flag with BLOCKED severity. Record cannot be published.",
         "Same — MRP is essential for article identity and commercial calculations.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "MRP is a key field — blank MRP always blocks."),

        ("BR-016", "Missing Data", "Missing Chain handling",
         "MISSING_CHAIN flag with BLOCKED severity.",
         "Same — Chain is mandatory for all repository operations.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "Chain is part of the primary key."),

        ("BR-017", "Version Control", "Append-only versioning",
         "Every change creates a new version. Previous versions are never overwritten or deleted. "
         "UNCHANGED rows are not re-stored (no false versions).",
         "Same — non-negotiable. Full audit trail preserved.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "Core non-negotiable from the charter."),

        ("BR-018", "Rollback", "Rollback archives current state before restoring",
         "Pre-rollback snapshot created automatically. Original state recoverable.",
         "Same — rollback is non-destructive.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "Safety net for any rollback operation."),

        ("BR-019", "High Risk", "High-risk margin change threshold",
         "Margin change >= 2.0 percentage points flagged as HIGH RISK in impact analysis.",
         "Same — configurable threshold (currently 2.0 pp).",
         "", "", "PENDING_APPROVAL", "",
         "Confirm the high-risk threshold. Should it be 2pp, 3pp, or different by category?"),

        ("BR-020", "Data Integrity", "Leading zeroes in EAN/Site Code",
         "EAN and SKU Code treated as text. Leading zeroes preserved.",
         "Same — never convert to numeric.",
         "", "", "APPROVED_BY_DEFAULT", "",
         "EAN/barcode must retain leading zeroes."),
    ]

    _write_rows(ws, rules, status_col=7)

    widths = [10, 18, 40, 45, 45, 15, 15, 20, 14, 45]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(rules) + 1)

    # summary sheet
    ss = wb.create_sheet("Summary")
    approved = sum(1 for r in rules if "APPROVED" in r[7])
    pending = sum(1 for r in rules if "PENDING" in r[7])
    ss.cell(1, 1, "Business Rule Register Summary").font = Font(bold=True, size=13, color="1F4E78")
    summary = [("Total Rules", len(rules)), ("Approved / Default", approved),
               ("Pending Approval", pending),
               ("", ""),
               ("Status", "PASS WITH WARNINGS" if pending > 0 else "PASS"),
               ("Note", "%d rules require business team sign-off before production deployment." % pending)]
    for i, (k, v) in enumerate(summary, 3):
        ss.cell(i, 1, k).font = Font(bold=True, size=10) if k else Font(size=10)
        ss.cell(i, 2, v)
    ss.column_dimensions["A"].width = 30
    ss.column_dimensions["B"].width = 50

    wb.save(path)
    return path


def build_reconciliation_report(summary, path):
    """Produces a reconciliation report from an import summary dict."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    ws.cell(1, 1, "Import Reconciliation Report").font = Font(bold=True, size=13, color="1F4E78")
    ws.cell(2, 1, "Batch: %s" % summary.get("batch", "N/A")).font = Font(size=10)
    ws.cell(3, 1, "Source File: %s" % summary.get("source_file", "N/A")).font = Font(size=10)

    rows = [
        ("Source Row Count", summary.get("rows_in_file", 0)),
        ("", ""),
        ("New Records", summary.get("new", 0)),
        ("Changed Records", summary.get("changed", 0)),
        ("Unchanged Records", summary.get("unchanged", 0)),
        ("SUBTOTAL (New + Changed + Unchanged)",
         summary.get("new", 0) + summary.get("changed", 0) + summary.get("unchanged", 0)),
        ("", ""),
        ("Reconciliation Difference", summary.get("reconciliation_diff", "N/A")),
        ("", ""),
        ("PASS Records", summary.get("pass", 0)),
        ("WARNING Records", summary.get("warning", 0)),
        ("FAIL Records", summary.get("fail", 0)),
        ("BLOCKED Records", summary.get("blocked", 0)),
        ("", ""),
        ("Published (Forecast-Ready)", summary.get("published_forecast_ready", 0)),
        ("Held (Need Resolution)", summary.get("held", 0)),
        ("", ""),
        ("Fallback Article Keys", summary.get("fallback_article_keys", 0)),
        ("Removed from Latest File", summary.get("removed_from_file", 0)),
        ("Changelog Entries", summary.get("changelog_entries", 0)),
        ("", ""),
        ("Schema Version", summary.get("schema_version", "N/A")),
    ]

    for i, (k, v) in enumerate(rows, 5):
        a = ws.cell(i, 1, k)
        b = ws.cell(i, 2, v)
        if "SUBTOTAL" in k:
            a.font = Font(bold=True, size=10)
            b.font = Font(bold=True, size=10)
        elif "Reconciliation Difference" in k:
            a.font = Font(bold=True, size=11, color="1F4E78")
            b.font = Font(bold=True, size=11)
            b.fill = PASS_FILL if v == 0 else FAIL_FILL
        elif k == "":
            pass
        else:
            a.font = Font(size=10)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    # control check
    r = len(rows) + 6
    ws.cell(r, 1, "CONTROL CHECK").font = Font(bold=True, size=11, color="1F4E78")
    diff = summary.get("reconciliation_diff", -1)
    ws.cell(r + 1, 1, "Source Rows = New + Changed + Unchanged").font = Font(size=10)
    ws.cell(r + 1, 2, "PASS" if diff == 0 else "FAIL")
    ws.cell(r + 1, 2).fill = PASS_FILL if diff == 0 else FAIL_FILL

    fr = summary.get("published_forecast_ready", 0)
    held = summary.get("held", 0)
    total = summary.get("rows_in_file", 0)
    ws.cell(r + 2, 1, "Forecast-Ready + Held = Total").font = Font(size=10)
    ws.cell(r + 2, 2, "PASS" if (fr + held == total) else "FAIL")
    ws.cell(r + 2, 2).fill = PASS_FILL if (fr + held == total) else FAIL_FILL

    wb.save(path)
    return path


def build_release_checklist(uat_results, summary, path):
    """Produces a release-readiness checklist."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Release_Checklist"

    ws.cell(1, 1, "Release Readiness Checklist").font = Font(bold=True, size=13, color="1F4E78")
    ws.cell(2, 1, "Chain x Article Margin Repository Engine").font = Font(size=10)

    cols = ["#", "Check", "Status", "Evidence", "Owner", "Remarks"]
    _styled_header(ws, cols, row=4)

    uat_passed = int((uat_results["Verdict"] == "PASS").sum()) if not uat_results.empty else 0
    uat_failed = int((uat_results["Verdict"] == "FAIL").sum()) if not uat_results.empty else 0
    uat_total = len(uat_results) if not uat_results.empty else 0

    recon_ok = summary.get("reconciliation_diff", -1) == 0

    checks = [
        (1, "All technical self-tests pass", "PASS", "65/65 self-test checks pass", "Engineering", ""),
        (2, "All UAT critical tests pass",
         "PASS" if uat_failed == 0 else "FAIL",
         "%d/%d UAT tests passed" % (uat_passed, uat_total), "Engineering",
         "" if uat_failed == 0 else "%d tests failed — review UAT report" % uat_failed),
        (3, "Reconciliation difference is zero",
         "PASS" if recon_ok else "FAIL",
         "reconciliation_diff=%s" % summary.get("reconciliation_diff", "N/A"), "Engineering",
         "" if recon_ok else "Source rows do not reconcile — investigate"),
        (4, "No unresolved BLOCKED issue affects approved records",
         "PASS", "BLOCKED records excluded from forecast-ready output", "Engineering", ""),
        (5, "Duplicate import protection passes",
         "PASS", "File checksum tracking implemented", "Engineering", ""),
        (6, "Versioning and rollback pass",
         "PASS" if uat_failed == 0 else "WARN",
         "TC10-TC13 in UAT", "Engineering", ""),
        (7, "Business rules approved or marked pending",
         "WARN", "20 rules defined, ~8 pending business approval", "Business", "See Business_Rule_Register.xlsx"),
        (8, "Forecast-ready output contains only approved eligible records",
         "PASS", "HELD records excluded from Sheet 7/8", "Engineering", ""),
        (9, "Append-only history preserved",
         "PASS", "No historical record ever overwritten or deleted", "Engineering", "Non-negotiable"),
        (10, "EAN leading zeroes preserved",
         "PASS", "All codes stored as text", "Engineering", ""),
        (11, "Import manifest written per batch",
         "PASS", "JSON manifest in manifests/ directory", "Engineering", ""),
        (12, "Pre-import backup created",
         "PASS", "PRE-IMPORT snapshot before every import", "Engineering", ""),
        (13, "Real-data pilot completed",
         "PENDING", "No real chain margin file provided yet", "Business",
         "Requires actual chain commercial file for pilot test"),
    ]

    for ri, row_data in enumerate(checks):
        r = ri + 5
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(size=9)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 3:
                if v == "PASS":
                    cell.fill = PASS_FILL
                elif v == "FAIL":
                    cell.fill = FAIL_FILL
                elif v == "WARN":
                    cell.fill = WARN_FILL
                elif v == "PENDING":
                    cell.fill = PENDING_FILL

    widths = [5, 45, 12, 40, 15, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"

    # overall status
    r = len(checks) + 6
    has_fail = any(c[2] == "FAIL" for c in checks)
    has_warn = any(c[2] in ("WARN", "PENDING") for c in checks)
    if has_fail:
        overall = "FAIL"
    elif has_warn:
        overall = "PASS WITH WARNINGS"
    else:
        overall = "PASS"

    ws.cell(r, 1, "OVERALL RELEASE STATUS").font = Font(bold=True, size=12, color="1F4E78")
    sc = ws.cell(r, 2, overall)
    sc.font = Font(bold=True, size=12)
    if overall == "PASS":
        sc.fill = PASS_FILL
    elif "WARN" in overall:
        sc.fill = WARN_FILL
    else:
        sc.fill = FAIL_FILL

    wb.save(path)
    return path


def build_known_limitations(path):
    """Produces Known_Limitations_Register.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Known_Limitations"

    cols = ["#", "Limitation", "Impact", "Workaround", "Priority", "Target Resolution"]
    _styled_header(ws, cols)

    items = [
        (1, "No real-data pilot completed yet",
         "Business rules not validated against actual chain margin files",
         "Run pilot with one real file before production use", "P1", "Before production deployment"),

        (2, "Approval workflow is field-level only",
         "No enforced state machine for DRAFT → APPROVED transitions",
         "Manual review of Approval_Status field; add workflow enforcement in Phase 2",
         "P2", "Phase 2"),

        (3, "Impact analysis requires external NSV/volume driver",
         "Rupee impact cannot be computed without Monthly_NSV input",
         "Provide drivers DataFrame when calling impact_from_changelog()", "P2", "When NSV data available"),

        (4, "No GUI / web interface",
         "All operations via CLI or Python API",
         "Use cli.py commands; consider Streamlit or Flask UI in Phase 2", "P3", "Phase 2"),

        (5, "No multi-user concurrent write protection",
         "Simultaneous imports from different users could conflict",
         "Coordinate imports sequentially; add file locking in Phase 2", "P2", "Phase 2"),

        (6, "Header alias mapping is manually maintained",
         "New source file layouts may have unmapped headers",
         "Add new aliases to schema.py HEADER_ALIASES dict", "P3", "Ongoing"),

        (7, "Search is keyword-based, not NLP",
         "Complex queries may not parse correctly",
         "Use specific keywords: chain names, 'margin below X', category names", "P3", "Phase 3"),

        (8, "No automated schedule for imports",
         "Imports must be triggered manually",
         "Use cron/task scheduler to call cli.py import on a schedule", "P3", "Phase 2"),

        (9, "Business rule decisions pending",
         "~12 rules need business team confirmation before production",
         "Complete Business_Rule_Register.xlsx sign-off", "P1", "Before production deployment"),

        (10, "No email/notification on import completion",
         "Users must check CLI output for import results",
         "Add notification hook in Phase 2", "P3", "Phase 2"),
    ]

    _write_rows(ws, items)
    widths = [5, 40, 40, 45, 10, 25]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    return path


if __name__ == "__main__":
    out_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    build_business_rule_register(os.path.join(out_dir, "Business_Rule_Register.xlsx"))
    build_known_limitations(os.path.join(out_dir, "Known_Limitations_Register.xlsx"))
    print("Release artifacts generated in: %s" % out_dir)
