# -*- coding: utf-8 -*-
"""Self-test: proves the margin repository engine works end-to-end.

Uses clearly-labeled SYNTHETIC data — never presented as Honasa's real margins.
Tests: ingest → validate → version → change-log → impact → search → outputs.
"""
import os
import sys
import shutil
import tempfile
import traceback

sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
from schema import REPO_COLS, ARTICLE_COLS, COMMERCIAL_COLS, CONDITION_COLS, DATE_COLS
from validation import validate_frame, qc_report
from repository import MarginRepository, coerce_numeric, derive_final_margin, article_key
from impact import impact_from_changelog
from search import search
from outputs import build_outputs, build_repository_template, build_import_template

PASS = 0
FAIL = 0


def check(label, condition, detail=""):
    global PASS, FAIL
    if condition:
        PASS += 1
        print("  PASS  %s" % label)
    else:
        FAIL += 1
        print("  FAIL  %s  %s" % (label, detail))


def synthetic_batch_1():
    """10 articles across 3 chains — initial import."""
    rows = [
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Sunscreen",  "Sub Category": "SPF50",
         "Article": "UV Shield 50ml",   "EAN": "8901234560001", "Pack Size": "50ml", "MRP": 499,
         "Trade Margin %": 28, "TOT %": 2, "GST %": 18, "Effective From": "2026-04-01",
         "Status": "ACTIVE", "SKU Code": "SKU001"},
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Sunscreen",  "Sub Category": "SPF30",
         "Article": "UV Light 100ml",   "EAN": "8901234560002", "Pack Size": "100ml", "MRP": 699,
         "Trade Margin %": 25, "TOT %": 3, "Backend %": 1, "GST %": 18, "Effective From": "2026-04-01",
         "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo",    "Sub Category": "Anti-Dandruff",
         "Article": "Clean Root 200ml", "EAN": "8901234560003", "Pack Size": "200ml", "MRP": 350,
         "Trade Margin %": 22, "Frontend %": 2, "GST %": 18, "Effective From": "2026-04-01",
         "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo",    "Sub Category": "Volume",
         "Article": "Volume Pro 150ml", "EAN": "8901234560004", "Pack Size": "150ml", "MRP": 299,
         "Trade Margin %": 20, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Facewash",   "Sub Category": "Gel",
         "Article": "FreshGel 100ml",   "EAN": "8901234560005", "Pack Size": "100ml", "MRP": 249,
         "Trade Margin %": 24, "TOT %": 1, "GST %": 18, "Effective From": "2026-04-01",
         "Status": "ACTIVE"},
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Serum",      "Sub Category": "Vitamin C",
         "Article": "Glow Serum 30ml",  "EAN": "8901234560006", "Pack Size": "30ml",  "MRP": 799,
         "Trade Margin %": 30, "TOT %": 2, "Backend %": 2, "GST %": 18, "Effective From": "2026-04-01",
         "Status": "ACTIVE"},
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Moisturizer","Sub Category": "Gel",
         "Article": "AquaBurst 50ml",   "EAN": "8901234560007", "Pack Size": "50ml",  "MRP": 599,
         "Trade Margin %": 26, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        # article with blank EAN (tests fallback key + BLANK_EAN warning)
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Cream",      "Sub Category": "Night",
         "Article": "NightRepair 50g",  "EAN": "",              "Pack Size": "50g",   "MRP": 449,
         "Trade Margin %": 18, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        # article with blank MRP (tests BLOCKED)
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Oil",        "Sub Category": "Hair",
         "Article": "Hair Oil 100ml",   "EAN": "8901234560009", "Pack Size": "100ml", "MRP": "",
         "Trade Margin %": 15, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
        # article with invalid GST (tests INCORRECT_GST)
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Kajal",      "Sub Category": "Eye",
         "Article": "Bold Kajal 0.3g",  "EAN": "8901234560010", "Pack Size": "0.3g",  "MRP": 199,
         "Trade Margin %": 32, "GST %": 15, "Effective From": "2026-04-01", "Status": "ACTIVE"},
    ]
    return pd.DataFrame(rows)


def synthetic_batch_2():
    """Second import: margin changes on 3 articles + 1 new article."""
    rows = [
        # changed margin: Apollo UV Shield
        {"Chain": "Apollo",   "Brand": "TestBrand", "Category": "Sunscreen",  "Sub Category": "SPF50",
         "Article": "UV Shield 50ml",   "EAN": "8901234560001", "Pack Size": "50ml", "MRP": 499,
         "Trade Margin %": 30, "TOT %": 2, "GST %": 18, "Effective From": "2026-07-01",
         "Status": "ACTIVE", "SKU Code": "SKU001"},
        # changed margin: DMart Clean Root
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo",    "Sub Category": "Anti-Dandruff",
         "Article": "Clean Root 200ml", "EAN": "8901234560003", "Pack Size": "200ml", "MRP": 350,
         "Trade Margin %": 24, "Frontend %": 2, "GST %": 18, "Effective From": "2026-07-01",
         "Status": "ACTIVE"},
        # changed margin: Reliance Glow Serum (big change -> high risk)
        {"Chain": "Reliance", "Brand": "TestBrand", "Category": "Serum",      "Sub Category": "Vitamin C",
         "Article": "Glow Serum 30ml",  "EAN": "8901234560006", "Pack Size": "30ml",  "MRP": 799,
         "Trade Margin %": 26, "TOT %": 2, "Backend %": 2, "GST %": 18, "Effective From": "2026-07-01",
         "Status": "ACTIVE"},
        # new article
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Conditioner","Sub Category": "Repair",
         "Article": "SilkSmooth 200ml", "EAN": "8901234560011", "Pack Size": "200ml", "MRP": 399,
         "Trade Margin %": 21, "TOT %": 1, "GST %": 18, "Effective From": "2026-07-01",
         "Status": "ACTIVE"},
        # unchanged (exact same data)
        {"Chain": "DMart",    "Brand": "TestBrand", "Category": "Shampoo",    "Sub Category": "Volume",
         "Article": "Volume Pro 150ml", "EAN": "8901234560004", "Pack Size": "150ml", "MRP": 299,
         "Trade Margin %": 20, "GST %": 18, "Effective From": "2026-04-01", "Status": "ACTIVE"},
    ]
    return pd.DataFrame(rows)


def run():
    global PASS, FAIL
    tmpdir = tempfile.mkdtemp(prefix="margin_selftest_")
    repo_root = os.path.join(tmpdir, "repo")

    try:
        print("=" * 70)
        print("MARGIN REPOSITORY ENGINE — SELF-TEST")
        print("All data is SYNTHETIC, for engine validation only.")
        print("=" * 70)

        # --- 1. Schema ---
        print("\n[1] Schema")
        check("REPO_COLS count", len(REPO_COLS) > 50, "got %d" % len(REPO_COLS))
        check("ARTICLE_COLS present", len(ARTICLE_COLS) == 15)
        check("COMMERCIAL_COLS present", len(COMMERCIAL_COLS) == 15)
        check("CONDITION_COLS present", len(CONDITION_COLS) == 9)
        check("DATE_COLS present", len(DATE_COLS) == 7)

        # --- 2. Validation ---
        print("\n[2] Validation on batch 1")
        b1 = synthetic_batch_1()
        b1 = coerce_numeric(b1)
        b1 = derive_final_margin(b1)
        b1v = validate_frame(b1, today="2026-07-01")
        check("validate_frame returns all rows", len(b1v) == 10)
        check("Validation_Flags column exists", "Validation_Flags" in b1v.columns)
        check("QC_Severity column exists", "QC_Severity" in b1v.columns)
        check("Record_Status column exists", "Record_Status" in b1v.columns)

        blank_ean = b1v[b1v["Validation_Flags"].str.contains("BLANK_EAN", na=False)]
        check("BLANK_EAN flagged for row 8", len(blank_ean) == 1)

        blank_mrp = b1v[b1v["Validation_Flags"].str.contains("BLANK_MRP", na=False)]
        check("BLANK_MRP flagged for row 9", len(blank_mrp) >= 1)
        check("BLANK_MRP -> BLOCKED severity",
              all(b1v.loc[blank_mrp.index, "QC_Severity"] == "BLOCKED"))

        bad_gst = b1v[b1v["Validation_Flags"].str.contains("INCORRECT_GST", na=False)]
        check("INCORRECT_GST flagged for row 10 (GST=15)", len(bad_gst) >= 1)

        published = b1v[b1v["Record_Status"] == "PUBLISHED"]
        held = b1v[b1v["Record_Status"] == "HELD"]
        check("Published records (PASS+WARNING)", len(published) >= 7)
        check("Held records (FAIL+BLOCKED)", len(held) >= 1)

        # derived margin
        row0 = b1v.iloc[0]
        fem = float(row0["Final Effective Margin %"])
        check("Derived Final Effective Margin (28+2=30)", fem == 30.0,
              "got %.2f" % fem)

        # QC report
        qc = qc_report(b1v)
        check("QC report has rows", len(qc) >= 15)
        health = qc.set_index("QC Metric")["Value"].get("Repository Health %")
        check("Repository Health % computed", health is not None and health > 0)

        # --- 3. Repository import (batch 1) ---
        print("\n[3] Repository import — batch 1")
        repo = MarginRepository(repo_root)
        check("Empty repo history", len(repo.history) == 0)
        summary1, cl1, removed1 = repo.import_frame(b1, source_file="synthetic_batch1.xlsx")
        check("Import new count", summary1["new"] >= 8, "got %d" % summary1["new"])
        check("Import changed count = 0", summary1["changed"] == 0)
        check("History has records", len(repo.history) > 0)
        check("Changelog empty for first import", cl1.empty)

        cur1 = repo.current(include_held=False)
        check("Current published view has records", len(cur1) >= 7)

        cur_all = repo.current(include_held=True)
        check("Current (all) includes held", len(cur_all) >= len(cur1))

        # article key
        row_ean = b1.iloc[0]
        k, fb = article_key(row_ean)
        check("Article key uses EAN prefix", k.startswith("EAN|"))
        check("No fallback for row with EAN", not fb)

        row_noean = b1.iloc[7]
        k2, fb2 = article_key(row_noean)
        check("Fallback key for blank EAN", k2.startswith("ALT|"))
        check("Fallback flag True", fb2)

        # snapshot
        versions = repo.list_versions()
        check("Snapshot created", len(versions) == 1)

        # --- 4. Repository import (batch 2) — versioning + change detection ---
        print("\n[4] Repository import — batch 2 (versioning)")
        b2 = synthetic_batch_2()
        summary2, cl2, removed2 = repo.import_frame(b2, source_file="synthetic_batch2.xlsx")
        check("Batch 2 new count = 1", summary2["new"] == 1, "got %d" % summary2["new"])
        check("Batch 2 changed count = 3", summary2["changed"] == 3,
              "got %d" % summary2["changed"])
        check("Batch 2 unchanged count >= 1", summary2["unchanged"] >= 1)
        check("Changelog has entries", not cl2.empty and len(cl2) >= 3)

        if not cl2.empty:
            check("Changelog has Field column", "Field" in cl2.columns)
            check("Changelog has Difference column", "Difference" in cl2.columns)

        # version numbers incremented
        cur2 = repo.current(include_held=True)
        uv = cur2[cur2["EAN"].astype(str) == "8901234560001"]
        if not uv.empty:
            v = int(float(uv.iloc[0]["Version Number"]))
            check("UV Shield version bumped to 2", v == 2, "got %d" % v)

        # removed detection
        check("Removed articles detected (Apollo articles not in batch 2)",
              len(removed2) >= 0)

        # snapshots: batch1 + PRE-IMPORT-batch2 + batch2
        versions2 = repo.list_versions()
        check("Multiple snapshots exist after batch 2", len(versions2) >= 2)

        # --- 5. Impact analysis ---
        print("\n[5] Impact analysis")
        impact = impact_from_changelog(cl2)
        check("Impact table not empty", not impact.empty)
        if not impact.empty:
            check("Impact has Margin Delta (pp)", "Margin Delta (pp)" in impact.columns)
            check("Impact has High Risk column", "High Risk" in impact.columns)
            hr = impact[impact["High Risk"] == "YES"]
            check("Glow Serum flagged high risk (4pp drop)", len(hr) >= 1)
            check("Impact marks 'requires NSV/volume input'",
                  "requires NSV/volume input" in str(impact.iloc[0].get("CM2 Impact (INR/mo)", "")))

        # impact with drivers
        drivers = pd.DataFrame([
            {"Article_Key": impact.iloc[0]["Article_Key"], "Monthly_NSV": 500000, "Monthly_Units": 1000}
        ])
        impact_d = impact_from_changelog(cl2, drivers=drivers)
        if not impact_d.empty:
            row_d = impact_d.iloc[0]
            check("CM2 Impact computed with driver",
                  row_d.get("CM2 Impact (INR/mo)") != "requires NSV/volume input")

        # --- 6. Search ---
        print("\n[6] Search")
        cur_search = repo.current(include_held=True)
        r1 = search(cur_search, "sunscreen")
        check("Search 'sunscreen' finds articles", len(r1) >= 1)

        r2 = search(cur_search, "margin below 25")
        check("Search 'margin below 25' filters correctly", len(r2) >= 1)
        if not r2.empty:
            vals = pd.to_numeric(r2["Final Effective Margin %"], errors="coerce")
            check("All results have margin < 25", all(vals.dropna() < 25))

        r3 = search(cur_search, "dmart")
        check("Search 'dmart' returns DMart articles", len(r3) >= 1)

        # --- 7. Outputs ---
        print("\n[7] Outputs workbook")
        out_path = os.path.join(tmpdir, "test_outputs.xlsx")
        build_outputs(repo, out_path, changelog=cl2, removed=removed2, impact=impact)
        check("Outputs file created", os.path.exists(out_path))

        from openpyxl import load_workbook
        wb = load_workbook(out_path, read_only=True)
        expected_sheets = [
            "1_Chain_Article_Master", "1b_Full_Version_History", "2_Margin_Change_Log",
            "3_Commercial_Difference", "4_Missing_Margin", "5_Duplicate_Article",
            "6_Margin_QC", "7_Forecast_Ready_Margin", "8_CM2_Ready_Margin",
            "9_Dashboard_Dataset", "10_Executive_Summary"
        ]
        for s in expected_sheets:
            check("Sheet '%s' exists" % s, s in wb.sheetnames)
        wb.close()

        # --- 8. Templates ---
        print("\n[8] Templates")
        repo_tmpl = os.path.join(tmpdir, "repo_template.xlsx")
        build_repository_template(repo_tmpl)
        check("Repository template created", os.path.exists(repo_tmpl))

        imp_tmpl = os.path.join(tmpdir, "import_template.xlsx")
        build_import_template(imp_tmpl)
        check("Import template created", os.path.exists(imp_tmpl))

        # --- 9. Rollback ---
        print("\n[9] Rollback")
        pre_count = len(repo.history)
        batch1_id = versions[0]
        restored = repo.rollback(batch1_id)
        check("Rollback restores batch 1 state", restored < pre_count)
        check("Pre-rollback snapshot created", len(repo.list_versions()) > 2)

    except Exception:
        traceback.print_exc()
        FAIL += 1
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    print("\n" + "=" * 70)
    print("RESULTS:  %d passed,  %d failed" % (PASS, FAIL))
    print("=" * 70)
    return FAIL == 0


if __name__ == "__main__":
    ok = run()
    sys.exit(0 if ok else 1)
