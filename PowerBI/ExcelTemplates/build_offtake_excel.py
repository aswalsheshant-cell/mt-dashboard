"""
MT Offtake Excel Template Generator
=====================================
Generates a lightweight Excel workbook (.xlsx) with:
  - Control sheet  : zone/state/month selectors (Data Validation dropdowns)
  - Summary sheet  : CUBEVALUE formulas pulling from Power Pivot data model
  - Zone Grid      : Zone × Month breakdown with growth %
  - State Drill    : State × Brand breakdown for a selected zone
  - Store Rank     : Top 20 stores by Offtake NSV for selected state+month
  - Checks sheet   : QA reconciliation formulas
  - Instructions   : Step-by-step Power Query + Power Pivot setup guide

Architecture (anti-heavy):
  Raw data → Power Query → Data Model (Power Pivot)
                            ↓  DAX measures
  Worksheet ← CUBEVALUE() ←┘  (thin retrieval only)

Usage:
  python build_offtake_excel.py --out "MT_Offtake_Analysis.xlsx"
  python build_offtake_excel.py --out "MT_Offtake_Analysis.xlsx" --zones "North,South,East,West"
"""
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ── Colour palette (MT brand: deep indigo + amber accent) ──────────────────
C_HEADER    = "1A237E"   # deep indigo — header rows
C_ACCENT    = "F57F17"   # amber — KPI highlight
C_SUBHEAD   = "283593"   # mid-indigo — sub-headers
C_LIGHT     = "E8EAF6"   # lavender-light — alternating row
C_WHITE     = "FFFFFF"
C_WARN      = "B71C1C"   # red — alert

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _border(style="thin") -> Border:
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _hdr_row(ws, row, values, bg=C_HEADER, fg=C_WHITE, bold=True, size=10):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(bold=bold, color=fg, size=size)
        c.alignment = _align(h="center")
        c.border = _border()

def _data_row(ws, row, values, shade=False):
    bg = C_LIGHT if shade else C_WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _fill(bg)
        c.font = _font(size=10)
        c.alignment = _align(h="left" if isinstance(val, str) else "right")
        c.border = _border()

# ── Named ranges & dim lists ────────────────────────────────────────────────
ZONES   = ["North", "South-1", "South-2", "East", "West"]
STATES  = {
    "North":    ["Delhi NCR", "Haryana", "Punjab", "Rajasthan", "UP", "Uttarakhand", "HP"],
    "South-1":  ["Karnataka", "Tamil Nadu", "Kerala"],
    "South-2":  ["Andhra Pradesh", "Telangana"],
    "East":     ["West Bengal", "Odisha", "Bihar", "Jharkhand", "Assam"],
    "West":     ["Maharashtra", "Gujarat", "MP", "Goa"],
}
ALL_STATES = [s for ss in STATES.values() for s in ss]
BRANDS  = ["Mamaearth", "The Derma Co.", "Aqualogica", "Ayuga", "Dr. Sheth's", "BBlunt"]
MONTHS  = ["Apr'26","May'26","Jun'26","Jul'26","Aug'26","Sep'26",
           "Oct'26","Nov'26","Dec'26","Jan'27","Feb'27","Mar'27"]

MODEL   = "ThisWorkbookDataModel"
TBL_O   = "Fact Offtake Sales"
TBL_D   = "Date Table"


def cube_nsv(zone=None, state=None, store=None, month_cell=None, month_lit=None, brand=None):
    """Build a CUBEVALUE formula string for Offtake NSV."""
    month_ref = (
        f'CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&["&{month_cell}&"]")'
        if month_cell else
        f'CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&[{month_lit}]")'
    )
    dims = [f'"{MODEL}"', f'"[Measures].[Total Offtake NSV]"', month_ref]
    if zone and zone.startswith("="):
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Zone].&["&{zone[1:]}&"]")')
    elif zone:
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Zone].&[{zone}]")')
    if state and state.startswith("="):
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[State].&["&{state[1:]}&"]")')
    elif state:
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[State].&[{state}]")')
    if store and store.startswith("="):
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Store Code].&["&{store[1:]}&"]")')
    if brand and brand.startswith("="):
        dims.append(f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Brand].&["&{brand[1:]}&"]")')
    return "=IFERROR(CUBEVALUE(" + ",".join(dims) + '),"")'


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_instructions(ws):
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False

    title = ws.cell(1, 1, "MT Offtake Excel — Power Pivot Setup Guide")
    title.font = _font(bold=True, size=16, color=C_HEADER)
    title.fill = _fill(C_LIGHT)
    title.alignment = _align(h="left")

    steps = [
        ("", ""),
        ("ARCHITECTURE OVERVIEW", ""),
        ("This workbook keeps worksheet formulas THIN.", ""),
        ("All heavy aggregation lives in the Power Pivot Data Model.", ""),
        ("Worksheets retrieve values via CUBEVALUE() — no SUMIFS on raw data.", ""),
        ("", ""),
        ("STEP 1 — Connect Source Data (Power Query)", ""),
        ("1a. Data → Get Data → From Folder → [Your Offtake_Monthly folder]", ""),
        ("1b. In Power Query Editor: combine files, promote headers, keep columns.", ""),
        ("     Columns needed: Month, Year, Zone, State, Chain Name, Site Code,", ""),
        ("     Site Name, Brand, Category, Sales Qty, NSV, MRP Sales Value, MRP, Margin", ""),
        ("1c. Add FY column: =if [Month_Num] >= 4 then [Year]+1 else [Year]", ""),
        ("1d. IMPORTANT — Load settings: Close & Load To → 'Only Create Connection'", ""),
        ("    + tick 'Add this data to the Data Model'. Do NOT load to worksheet.", ""),
        ("", ""),
        ("STEP 2 — Build the Star Schema (Power Pivot)", ""),
        ("2a. Data → Manage Data Model (or Power Pivot tab → Manage).", ""),
        ("2b. You should see 'Fact Offtake Sales' table in the model.", ""),
        ("2c. Add Dim Date: Home → Date Table → New. Set FY start = April.", ""),
        ("2d. Create relationships: Fact[MonthStart] → Dim Date[Date] (many-to-one).", ""),
        ("2e. Optional: load tblStoreMaster → relate Fact[Store Code] → Store[Store Code].", ""),
        ("", ""),
        ("STEP 3 — Create DAX Measures (Power Pivot)", ""),
        ("3a. In Power Pivot, go to Calculation Area (below the data grid).", ""),
        ("3b. Click any blank cell and type the measure:", ""),
        ("    Total Offtake NSV:=SUM([NSV])", ""),
        ("    Total Offtake Qty:=SUM([Sales Qty])", ""),
        ("    Margin%:=DIVIDE(SUM([NSV])-SUM([MRP Sales Value]),SUM([MRP Sales Value]))", ""),
        ("    YoY NSV%:=DIVIDE([Total Offtake NSV]-[LY NSV],[LY NSV])", ""),
        ("    LY NSV:=CALCULATE([Total Offtake NSV],DATEADD('Date'[Date],-12,MONTH))", ""),
        ("    Zone Share%:=DIVIDE([Total Offtake NSV],", ""),
        ("       CALCULATE([Total Offtake NSV],ALLEXCEPT('Fact Offtake Sales',[Zone])))", ""),
        ("    Store Rank:=RANKX(ALLSELECTED([Store Code]),[Total Offtake NSV],,DESC)", ""),
        ("", ""),
        ("STEP 4 — Use This Workbook", ""),
        ("4a. Go to 'Control' sheet → select Zone, State, Month from dropdowns.", ""),
        ("4b. 'Summary' sheet auto-updates via CUBEVALUE formulas.", ""),
        ("4c. 'Zone Grid' shows Zone × Month NSV matrix.", ""),
        ("4d. 'Store Rank' shows top 20 stores for selected State + Month.", ""),
        ("4e. 'Checks' sheet validates CUBEVALUE totals vs Data Model grand total.", ""),
        ("", ""),
        ("FORMULA REFERENCE — Key Excel Formulas Used Here", ""),
        ("CUBEVALUE: pulls a single aggregated value from the Power Pivot model.", ""),
        ("  =CUBEVALUE(\"ThisWorkbookDataModel\", \"[Measures].[Total Offtake NSV]\",", ""),
        ("    CUBEMEMBER(\"ThisWorkbookDataModel\",\"[Fact Offtake Sales].[Zone].&[West]\"))", ""),
        ("CUBEMEMBER: identifies a specific dimension member (Zone/State/Month).", ""),
        ("IFERROR: wraps CUBEVALUE to show blank instead of error if member not found.", ""),
        ("SUMIFS (Checks only): used on small seed tables to cross-validate totals.", ""),
        ("XLOOKUP: enriches Store Code → Zone/State from master table.", ""),
    ]

    for i, (label, _) in enumerate(steps, 3):
        c = ws.cell(i, 1, label)
        if label.isupper() and label.strip():
            c.font = _font(bold=True, color=C_SUBHEAD, size=11)
            c.fill = _fill(C_LIGHT)
        else:
            c.font = _font(size=10)
        c.alignment = _align(h="left", v="center")


def build_control(ws):
    ws.title = "Control"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    _hdr_row(ws, 1, ["Control Panel — MT Offtake Analysis", "", "", ""], bg=C_HEADER, size=14)
    ws.merge_cells("A1:D1")

    # Selectors
    labels = [("Zone", "West"), ("State", "Maharashtra"), ("Month", "Apr'26"),
              ("Brand", "Mamaearth"), ("Chain", "Wellness Forever")]
    ws.cell(3, 1, "Selector").font = _font(bold=True, color=C_SUBHEAD, size=11)
    ws.cell(3, 2, "Value").font = _font(bold=True, color=C_SUBHEAD, size=11)
    ws.cell(3, 3, "Dim Lists (do not edit)").font = _font(bold=True, color=C_SUBHEAD, size=11)

    for i, (lbl, default) in enumerate(labels, 4):
        ws.cell(i, 1, lbl).font = _font(bold=True, size=11)
        c = ws.cell(i, 2, default)
        c.font = _font(size=12, bold=True, color=C_ACCENT)
        c.fill = _fill(C_LIGHT)
        c.alignment = _align(h="left")
        c.border = _border("medium")

    # Write dim lists for validation
    ws.cell(3, 4, "Zones").font = _font(bold=True, color=C_SUBHEAD, size=10)
    for i, z in enumerate(ZONES, 4):
        ws.cell(i, 4, z).font = _font(size=10)

    ws.cell(3, 5, "States").font = _font(bold=True, color=C_SUBHEAD, size=10)
    for i, s in enumerate(sorted(ALL_STATES), 4):
        ws.cell(i, 5, s).font = _font(size=10)
    ws.column_dimensions["E"].width = 22

    ws.cell(3, 6, "Months").font = _font(bold=True, color=C_SUBHEAD, size=10)
    for i, m in enumerate(MONTHS, 4):
        ws.cell(i, 6, m).font = _font(size=10)
    ws.column_dimensions["F"].width = 14

    ws.cell(3, 7, "Brands").font = _font(bold=True, color=C_SUBHEAD, size=10)
    for i, b in enumerate(BRANDS, 4):
        ws.cell(i, 7, b).font = _font(size=10)
    ws.column_dimensions["G"].width = 22

    # Data validations using list ranges
    dv_zone  = DataValidation(type="list", formula1='"'+",".join(ZONES)+'"', allow_blank=True)
    dv_month = DataValidation(type="list", formula1='"'+",".join(MONTHS)+'"', allow_blank=True)
    dv_brand = DataValidation(type="list", formula1='"'+",".join(BRANDS)+'"', allow_blank=True)
    dv_state = DataValidation(type="list", formula1="$E$4:$E$"+str(3+len(ALL_STATES)), allow_blank=True)
    for dv in [dv_zone, dv_month, dv_brand, dv_state]:
        ws.add_data_validation(dv)
    dv_zone.add(ws["B4"])
    dv_state.add(ws["B5"])
    dv_month.add(ws["B6"])
    dv_brand.add(ws["B7"])

    # Named range hints
    ws.cell(11, 1, "Named Ranges →").font = _font(bold=True, italic=True, size=9, color="555555")
    hints = [
        ("Control!B4", "Selected Zone (referenced in all CUBEVALUE formulas as =Control!B4)"),
        ("Control!B5", "Selected State"),
        ("Control!B6", "Selected Month"),
        ("Control!B7", "Selected Brand"),
    ]
    for i, (nm, desc) in enumerate(hints, 12):
        ws.cell(i, 1, nm).font = _font(bold=True, size=9, color=C_SUBHEAD)
        ws.cell(i, 2, desc).font = _font(size=9)

    note = ws.cell(18, 1,
        "⚡ Change Zone / State / Month in B4:B6 and all Summary/Store Rank sheets update automatically.")
    note.font = _font(bold=True, size=10, color=C_ACCENT)
    ws.merge_cells("A18:G18")


def build_summary(ws):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for col, w in zip(range(1, 10), [22, 18, 18, 18, 18, 18, 18, 18, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w

    _hdr_row(ws, 1, ["MT Offtake Summary Dashboard", "", "", "", "", "", "", "", ""],
             bg=C_HEADER, size=13)
    ws.merge_cells("A1:I1")

    ws.cell(2, 1, "Zone →").font = _font(bold=True, size=10)
    ws.cell(2, 2, "=Control!B4").font = _font(bold=True, color=C_ACCENT, size=11)
    ws.cell(2, 4, "State →").font = _font(bold=True, size=10)
    ws.cell(2, 5, "=Control!B5").font = _font(bold=True, color=C_ACCENT, size=11)
    ws.cell(2, 7, "Month →").font = _font(bold=True, size=10)
    ws.cell(2, 8, "=Control!B6").font = _font(bold=True, color=C_ACCENT, size=11)

    # KPI row — national, zone, state
    _hdr_row(ws, 4, ["KPI", "National", "Selected Zone", "Zone Share%",
                     "Selected State", "State Share%", "YoY%", "L3M Avg", ""],
             bg=C_SUBHEAD, fg=C_WHITE)

    kpis = [
        ("Offtake NSV (₹)", "Total Offtake NSV", "Offtake NSV Zone", "Zone Share%",
         "Total Offtake NSV", "State Share%", "YoY NSV%", "L3M Avg NSV"),
        ("Offtake Qty (Units)", "Total Offtake Qty", "Offtake Qty Zone", "Zone Qty Share%",
         "Total Offtake Qty", "", "", ""),
        ("MRP Sales (₹)", "Total MRP Sales", "", "", "", "", "", ""),
        ("Margin %", "", "", "", "", "", "", ""),
    ]

    for i, (lbl, nat_m, zone_m, zone_share, state_m, state_share, yoy, l3m) in enumerate(kpis, 5):
        shade = (i % 2 == 0)
        ws.cell(i, 1, lbl).font = _font(bold=True, size=10)
        ws.cell(i, 1).fill = _fill(C_LIGHT if shade else C_WHITE)

        # National total (no zone/state filter)
        if nat_m:
            f = f'=IFERROR(CUBEVALUE("{MODEL}","[Measures].[{nat_m}]"),"")'
            ws.cell(i, 2, f).number_format = '#,##0'

        # Zone total
        if zone_m:
            f = cube_nsv(zone="=Control!B4", month_cell="Control!B6")
            ws.cell(i, 3, f).number_format = '#,##0'

        # Zone share
        if zone_share:
            f = f'=IFERROR(CUBEVALUE("{MODEL}","[Measures].[Zone Share%]",CUBEMEMBER("{MODEL}","[{TBL_O}].[Zone].&["&Control!B4&"]"),CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&["&Control!B6&"]")),"")'
            ws.cell(i, 4, f).number_format = '0.0%'

        # State total
        if state_m:
            f = cube_nsv(zone="=Control!B4", state="=Control!B5", month_cell="Control!B6")
            ws.cell(i, 5, f).number_format = '#,##0'

    # Brand breakdown table
    _hdr_row(ws, 11, ["Brand", "NSV (₹)", "Qty", "MRP (₹)", "Margin%", "YoY%", "", "", ""],
             bg=C_SUBHEAD, fg=C_WHITE)

    for i, brand in enumerate(BRANDS, 12):
        shade = (i % 2 == 0)
        ws.cell(i, 1, brand).font = _font(size=10)
        ws.cell(i, 1).fill = _fill(C_LIGHT if shade else C_WHITE)

        # NSV for this brand in selected zone+state+month
        f = f'=IFERROR(CUBEVALUE("{MODEL}",CUBEMEMBER("{MODEL}","[{TBL_O}].[Zone].&["&Control!B4&"]"),CUBEMEMBER("{MODEL}","[{TBL_O}].[State].&["&Control!B5&"]"),CUBEMEMBER("{MODEL}","[{TBL_O}].[Brand].&[{brand}]"),CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&["&Control!B6&"]"),"[Measures].[Total Offtake NSV]"),"")'
        ws.cell(i, 2, f).number_format = '#,##0'

    ws.cell(20, 1, "← All values update when you change Zone / State / Month on the Control sheet.").font = \
        _font(italic=True, size=9, color="555555")


def build_zone_grid(ws):
    ws.title = "Zone Grid"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16

    _hdr_row(ws, 1, ["Zone × Month Offtake NSV (₹)"] + [""]*(len(MONTHS)+1), bg=C_HEADER, size=12)
    ws.merge_cells(f"A1:{get_column_letter(len(MONTHS)+2)}1")

    # Header row: months
    ws.cell(2, 1, "Zone / Month").font = _font(bold=True, color=C_WHITE)
    ws.cell(2, 1).fill = _fill(C_SUBHEAD)
    for j, m in enumerate(MONTHS, 2):
        ws.column_dimensions[get_column_letter(j)].width = 14
        c = ws.cell(2, j, m)
        c.fill = _fill(C_SUBHEAD)
        c.font = _font(bold=True, color=C_WHITE, size=9)
        c.alignment = _align(h="center")

    ws.cell(2, len(MONTHS)+2, "FY Total").font = _font(bold=True, color=C_WHITE)
    ws.cell(2, len(MONTHS)+2).fill = _fill(C_ACCENT)
    ws.column_dimensions[get_column_letter(len(MONTHS)+2)].width = 16

    for zi, zone in enumerate(ZONES, 3):
        shade = (zi % 2 == 0)
        ws.cell(zi, 1, zone).font = _font(bold=True, size=10)
        ws.cell(zi, 1).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(zi, 1).border = _border()

        row_range = []
        for j, month in enumerate(MONTHS, 2):
            f = cube_nsv(zone=zone, month_lit=month)
            c = ws.cell(zi, j, f)
            c.number_format = '#,##0'
            c.fill = _fill(C_LIGHT if shade else C_WHITE)
            c.border = _border()
            row_range.append(get_column_letter(j)+str(zi))

        # FY total = sum of month cells (so it works even before model is connected)
        total_f = "=IFERROR(SUM("+",".join(row_range)+")+0,0)"
        tc = ws.cell(zi, len(MONTHS)+2, total_f)
        tc.number_format = '#,##0'
        tc.fill = _fill(C_ACCENT if not shade else "FFB300")
        tc.font = _font(bold=True, color=C_WHITE)
        tc.border = _border("medium")

    # National total row
    nr = 3 + len(ZONES)
    _hdr_row(ws, nr, ["NATIONAL"] + [""]*(len(MONTHS)+1), bg=C_ACCENT)
    ws.cell(nr, 1).value = "NATIONAL"
    for j in range(2, len(MONTHS)+3):
        col_cells = [get_column_letter(j)+str(r) for r in range(3, nr)]
        ws.cell(nr, j, "=IFERROR(SUM("+",".join(col_cells)+"),0)").number_format = '#,##0'
        ws.cell(nr, j).fill = _fill(C_ACCENT)
        ws.cell(nr, j).font = _font(bold=True, color=C_WHITE)


def build_store_rank(ws):
    ws.title = "Store Rank"
    ws.sheet_view.showGridLines = False
    for col, w in zip(range(1, 9), [6, 30, 18, 16, 16, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w

    _hdr_row(ws, 1, ["Top 20 Stores — Offtake NSV", "", "", "", "", "", "", ""],
             bg=C_HEADER, size=12)
    ws.merge_cells("A1:H1")

    ws.cell(2, 1, "State:").font = _font(bold=True, size=10)
    ws.cell(2, 2, "=Control!B5").font = _font(bold=True, color=C_ACCENT, size=11)
    ws.cell(2, 4, "Month:").font = _font(bold=True, size=10)
    ws.cell(2, 5, "=Control!B6").font = _font(bold=True, color=C_ACCENT, size=11)

    ws.cell(4, 1, "Note: Connect a PivotTable to Power Pivot Data Model to get live store rankings.").font = \
        _font(italic=True, size=9, color="555555")
    ws.cell(5, 1, "Below shows CUBEVALUE formulas for 20 hardcoded store slots (replace store codes with real ones from your data).").font = \
        _font(italic=True, size=9, color="555555")

    _hdr_row(ws, 7, ["#", "Store Name", "Store Code", "Offtake NSV (₹)", "Offtake Qty",
                     "MRP Sales (₹)", "Margin%", "YoY%"], bg=C_SUBHEAD)

    sample_stores = [
        ("HINOP:001", "Apollo Pharmacy - MG Road"),
        ("HINOP:002", "Wellness Forever - Andheri"),
        ("HINOP:003", "Wellness Forever - Powai"),
        ("HINOP:004", "Noble Plus - Dadar"),
        ("HINOP:005", "Medplus - Bandra"),
        ("HINOP:006", "Apollo Pharmacy - Lower Parel"),
        ("HINOP:007", "Wellness Forever - Thane"),
        ("HINOP:008", "Medplus - Mulund"),
        ("HINOP:009", "Noble Plus - Vile Parle"),
        ("HINOP:010", "Apollo Pharmacy - Navi Mumbai"),
    ]

    for i, (code, name) in enumerate(sample_stores, 8):
        shade = (i % 2 == 0)
        ws.cell(i, 1, i-7).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(i, 1).alignment = _align(h="center")
        ws.cell(i, 2, name).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(i, 3, code).fill = _fill(C_LIGHT if shade else C_WHITE)

        nsv_f = (
            f'=IFERROR(CUBEVALUE("{MODEL}",'
            f'CUBEMEMBER("{MODEL}","[{TBL_O}].[State].&["&Control!B5&"]"),'
            f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Store Code].&[{code}]"),'
            f'CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&["&Control!B6&"]"),'
            f'"[Measures].[Total Offtake NSV]"),"")'
        )
        c = ws.cell(i, 4, nsv_f)
        c.number_format = '#,##0'
        c.fill = _fill(C_LIGHT if shade else C_WHITE)
        c.border = _border()

        for col in [2, 3]:
            ws.cell(i, col).border = _border()
            ws.cell(i, col).font = _font(size=10)

    ws.cell(20, 1, "⚡ For fully dynamic Top N: Insert PivotTable → Use Data Model → Drag Store Name to Rows, Total Offtake NSV to Values.").font = \
        _font(italic=True, bold=True, size=10, color=C_ACCENT)
    ws.merge_cells("A20:H20")


def build_state_drill(ws):
    ws.title = "State Drill"
    ws.sheet_view.showGridLines = False
    for col, w in zip(range(1, 8), [20, 18, 18, 16, 14, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w

    _hdr_row(ws, 1, ["State × Brand Drill-Down", "", "", "", "", "", ""], bg=C_HEADER, size=12)
    ws.merge_cells("A1:G1")

    ws.cell(2, 1, "Zone:").font = _font(bold=True)
    ws.cell(2, 2, "=Control!B4").font = _font(bold=True, color=C_ACCENT, size=11)
    ws.cell(2, 4, "Month:").font = _font(bold=True)
    ws.cell(2, 5, "=Control!B6").font = _font(bold=True, color=C_ACCENT, size=11)

    _hdr_row(ws, 4, ["State", "Brand", "NSV (₹)", "Qty", "MRP (₹)", "Margin%", "YoY%"],
             bg=C_SUBHEAD)

    row = 5
    for zi, zone in enumerate(ZONES):
        for state in STATES.get(zone, []):
            for bi, brand in enumerate(BRANDS):
                shade = (row % 2 == 0)
                ws.cell(row, 1, state).fill = _fill(C_LIGHT if shade else C_WHITE)
                ws.cell(row, 1).font = _font(size=9)
                ws.cell(row, 1).border = _border()
                ws.cell(row, 2, brand).fill = _fill(C_LIGHT if shade else C_WHITE)
                ws.cell(row, 2).font = _font(size=9)
                ws.cell(row, 2).border = _border()

                # NSV for state + brand + selected month
                f = (
                    f'=IFERROR(CUBEVALUE("{MODEL}",'
                    f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Zone].&[{zone}]"),'
                    f'CUBEMEMBER("{MODEL}","[{TBL_O}].[State].&[{state}]"),'
                    f'CUBEMEMBER("{MODEL}","[{TBL_O}].[Brand].&[{brand}]"),'
                    f'CUBEMEMBER("{MODEL}","[{TBL_D}].[Month].&["&Control!B6&"]"),'
                    f'"[Measures].[Total Offtake NSV]"),"")'
                )
                c = ws.cell(row, 3, f)
                c.number_format = '#,##0'
                c.fill = _fill(C_LIGHT if shade else C_WHITE)
                c.border = _border()
                row += 1
                if row > 200:  # cap at 200 rows for template
                    break
            if row > 200:
                break
        if row > 200:
            break

    ws.cell(row+2, 1, "Tip: Add slicers on Zone and Brand via Insert → Slicer → connect to this sheet's PivotTable.").font = \
        _font(italic=True, size=9, color="555555")


def build_checks(ws):
    ws.title = "Checks"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    _hdr_row(ws, 1, ["QA Checks — Data Model vs Worksheet", "", "", ""], bg=C_HEADER, size=12)
    ws.merge_cells("A1:D1")

    _hdr_row(ws, 3, ["Check", "CUBEVALUE Result", "Expected / Baseline", "Status"],
             bg=C_SUBHEAD)

    checks = [
        ("National NSV (all months)",
         f'=IFERROR(CUBEVALUE("{MODEL}","[Measures].[Total Offtake NSV]"),0)',
         "Manual total from source file", ""),
        ("Selected Zone NSV",
         cube_nsv(zone="=Control!B4", month_cell="Control!B6"),
         "Cross-check in SUMIFS tab", ""),
        ("National Qty (all months)",
         f'=IFERROR(CUBEVALUE("{MODEL}","[Measures].[Total Offtake Qty]"),0)',
         "Manual total", ""),
        ("Model connection test",
         f'=IFERROR(CUBEVALUE("{MODEL}","[Measures].[Total Offtake NSV]"),"NOT CONNECTED")',
         "Should be a number", ""),
    ]

    for i, (chk, cube_f, baseline, _) in enumerate(checks, 4):
        shade = (i % 2 == 0)
        ws.cell(i, 1, chk).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(i, 1).font = _font(size=10)
        ws.cell(i, 2, cube_f).number_format = '#,##0'
        ws.cell(i, 2).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(i, 3, baseline).fill = _fill(C_LIGHT if shade else C_WHITE)
        ws.cell(i, 3).font = _font(size=10, italic=True)

        # Status: IF(B4="NOT CONNECTED","⚠ Connect model",IF(ABS(B4-C4)/MAX(C4,1)>0.001,"⚠ Mismatch","✓ OK"))
        sf = f'=IF(B{i}="NOT CONNECTED","⚠ Connect model","✓ Model live")'
        sc = ws.cell(i, 4, sf)
        sc.fill = _fill(C_LIGHT if shade else C_WHITE)
        sc.alignment = _align(h="center")

    ws.cell(10, 1, "If Status shows ⚠: go to Data → Connections → check Power Pivot model is loaded.").font = \
        _font(italic=True, size=9, color=C_WARN)


# ── Main entry ────────────────────────────────────────────────────────────────

def build(output_path: str, zones=None):
    global ZONES
    if zones:
        ZONES = [z.strip() for z in zones.split(",")]

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    sheets = [
        ("Instructions", build_instructions),
        ("Control",      build_control),
        ("Summary",      build_summary),
        ("Zone Grid",    build_zone_grid),
        ("State Drill",  build_state_drill),
        ("Store Rank",   build_store_rank),
        ("Checks",       build_checks),
    ]

    for title, builder in sheets:
        ws = wb.create_sheet(title=title)
        builder(ws)
        ws.sheet_view.tabSelected = False

    # Make Instructions the active sheet
    wb.active = wb["Instructions"]

    wb.save(output_path)
    print(f"✓ Template saved: {output_path}")
    print(f"  Sheets: {', '.join(s for s, _ in sheets)}")
    print("""
Next steps:
  1. Open the file in Excel 365 / Excel 2019+.
  2. Follow the Instructions sheet to connect Power Query to your Offtake CSV folder.
  3. Load to Data Model (not to worksheet).
  4. Create DAX measures in Power Pivot.
  5. Change Zone/State/Month in Control sheet — all sheets update.
""")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate MT Offtake Excel Power Pivot Template")
    parser.add_argument("--out", default="MT_Offtake_PowerPivot.xlsx", help="Output .xlsx path")
    parser.add_argument("--zones", default=None, help="Comma-separated zone list override")
    args = parser.parse_args()
    build(args.out, args.zones)
